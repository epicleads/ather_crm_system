import eventlet
eventlet.monkey_patch()
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, Response, current_app
from supabase.client import create_client, Client
import csv
import openpyxl
import os
from datetime import datetime, timedelta, date
import uuid
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from werkzeug.utils import secure_filename
import random
import string
import io
from dotenv import load_dotenv
from collections import defaultdict, Counter
import json
from auth import AuthManager, require_auth, require_admin, require_cre, require_ps, require_rec
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from security_verification import run_security_verification
import time
import gc
from flask_socketio import SocketIO, emit
import math
# from redis import Redis  # REMOVE this line for local development

# Import optimized operations for faster lead updates
from optimized_lead_operations import create_optimized_operations

# Add this instead:
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from flask import send_file
import tempfile
import matplotlib
matplotlib.use('Agg')  # For headless environments
import matplotlib.pyplot as plt
import seaborn as sns
from matplotlib.backends.backend_pdf import PdfPages
import pytz
import pandas as pd
from werkzeug.security import generate_password_hash, check_password_hash

# Load environment variables from .env file
load_dotenv()

app = Flask(__name__)
# Initialize Flask-SocketIO with better configuration
socketio = SocketIO(
    app, 
    cors_allowed_origins="*",
    async_mode='eventlet',
    logger=True,
    engineio_logger=True,
    ping_timeout=60,
    ping_interval=25,
    max_http_buffer_size=1e8
)

# Routes moved to end of file for proper registration

# Reduce Flask log noise
import logging
log = logging.getLogger('werkzeug')
log.setLevel(logging.WARNING)

# Add custom Jinja2 filters
# Note: Using Flask's built-in tojson filter instead of custom one

# Utility functions
def is_valid_date(date_string):
    """Validate date string format (YYYY-MM-DD)"""
    try:
        datetime.strptime(date_string, '%Y-%m-%d')
        return True
    except ValueError:
        return False

def is_valid_uid(uid):
    """Validate UID format"""
    if not uid:
        return False
    # Add your UID validation logic here
    return True


# Using Flask's built-in tojson filter

# Get environment variables with fallback values for testing
SUPABASE_URL = os.environ.get('SUPABASE_URL')
SUPABASE_KEY = os.environ.get('SUPABASE_ANON_KEY')
SECRET_KEY = os.environ.get('FLASK_SECRET_KEY', 'fallback-secret-key-change-this')

# Email configuration (add these to your .env file)
SMTP_SERVER = os.environ.get('SMTP_SERVER', 'smtp.gmail.com')
SMTP_PORT = int(os.environ.get('SMTP_PORT', '587'))
EMAIL_USER = os.environ.get('EMAIL_USER', '')
EMAIL_PASSWORD = os.environ.get('EMAIL_PASSWORD', '')

# Debug: Print to check if variables are loaded (remove in production)
print(f"SUPABASE_URL loaded: {SUPABASE_URL is not None}")
print(f"SUPABASE_KEY loaded: {SUPABASE_KEY is not None}")
print(f"SUPABASE_URL: {SUPABASE_URL}")
print(f"SUPABASE_ANON_KEY: {SUPABASE_KEY}")

# Validate required environment variables
if not SUPABASE_URL:
    raise ValueError("SUPABASE_URL environment variable is required. Please check your .env file.")
if not SUPABASE_KEY:
    raise ValueError("SUPABASE_ANON_KEY environment variable is required. Please check your .env file.")

app.secret_key = SECRET_KEY
app.permanent_session_lifetime = timedelta(hours=24)

# Initialize Supabase client
try:
    supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
    print("✅ Supabase client initialized successfully")
    # Removed Supabase warm-up query for local development
except Exception as e:
    print(f"❌ Error initializing Supabase client: {e}")
    raise

# Initialize optimized operations for faster lead updates
try:
    optimized_ops = create_optimized_operations(supabase)
    print("✅ Optimized operations initialized successfully")
except Exception as e:
    print(f"❌ Error initializing optimized operations: {e}")
    # Continue without optimized operations if there's an error
    optimized_ops = None

# Initialize AuthManager
auth_manager = AuthManager(supabase)
# Store auth_manager in app config instead of direct attribute
app.config['AUTH_MANAGER'] = auth_manager

# Initialize WebSocket Manager
try:
    from websocket_events import WebSocketManager
    websocket_manager = WebSocketManager(socketio, supabase, auth_manager)
    websocket_manager.register_events()
    app.config['WEBSOCKET_MANAGER'] = websocket_manager
    print("✅ WebSocket manager initialized successfully")
    print(f"✅ WebSocket manager instance: {websocket_manager}")
except Exception as e:
    print(f"❌ Error initializing WebSocket manager: {e}")
    import traceback
    traceback.print_exc()
    websocket_manager = None

# Initialize rate limiter
limiter = Limiter(
    key_func=get_remote_address,
    app=app,
    default_limits=["1000 per minute"]  # Use in-memory backend for local/dev
)

# Upload folder configuration
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max file size

# Create upload folder if it doesn't exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def send_email_to_ps(ps_email, ps_name, lead_data, cre_name):
    """Send email notification to PS when a lead is assigned"""
    try:
        if not EMAIL_USER or not EMAIL_PASSWORD:
            print("Email credentials not configured. Skipping email notification.")
            return False

        # Create message
        msg = MIMEMultipart()
        msg['From'] = EMAIL_USER
        msg['To'] = ps_email
        msg['Subject'] = f"New Lead Assigned - {lead_data['customer_name']}"

        # Email body
        body = f"""
        Dear {ps_name},

        A new lead has been assigned to you by {cre_name}.

        Lead Details:
        - Customer Name: {lead_data['customer_name']}
        - Mobile Number: {lead_data['customer_mobile_number']}
        - Source: {lead_data['source']}
        - Lead Category: {lead_data.get('lead_category', 'Not specified')}
        - Model Interested: {lead_data.get('model_interested', 'Not specified')}
        - Branch: {lead_data.get('branch', 'Not specified')}

        Please log in to the CRM system to view and update this lead.

        Best regards,
        Ather CRM System
        """

        msg.attach(MIMEText(body, 'plain'))

        # Send email
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()
        server.login(EMAIL_USER, EMAIL_PASSWORD)
        text = msg.as_string()
        server.sendmail(EMAIL_USER, ps_email, text)
        server.quit()

        print(f"Email sent successfully to {ps_email}")
        return True

    except Exception as e:
        print(f"Error sending email: {e}")
        return False


def read_csv_file(filepath):
    """Read CSV file and return list of dictionaries with memory optimization"""
    data = []
    try:
        with open(filepath, 'r', encoding='utf-8') as file:
            csv_reader = csv.DictReader(file)
            for row_num, row in enumerate(csv_reader):
                if row_num >= 10000:  # Limit to 10,000 rows
                    print(f"Warning: File contains more than 10,000 rows. Only processing first 10,000.")
                    break

                # Clean and validate row data
                cleaned_row = {}
                for key, value in row.items():
                    if key and value:  # Only include non-empty keys and values
                        cleaned_row[key.strip()] = str(value).strip()

                if cleaned_row:  # Only add non-empty rows
                    data.append(cleaned_row)

                # Memory management for large files
                if row_num % 1000 == 0 and row_num > 0:
                    print(f"Processed {row_num} rows...")

    except Exception as e:
        print(f"Error reading CSV file: {e}")
        raise

    return data


def read_excel_file(filepath):
    """Read Excel file and return list of dictionaries with memory optimization"""
    data = []
    try:
        workbook = openpyxl.load_workbook(filepath, read_only=True)  # Read-only mode for memory efficiency
        sheet = workbook.active

        # Get headers from first row
        headers = []
        if sheet and sheet[1]:
            for cell in sheet[1]:
                if cell and cell.value:
                    headers.append(str(cell.value).strip())
                else:
                    headers.append(None)

        # Read data rows with limit
        row_count = 0
        if sheet:
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if row_count >= 10000:  # Limit to 10,000 rows
                    print(f"Warning: File contains more than 10,000 rows. Only processing first 10,000.")
                    break

                row_data = {}
                has_data = False

                for i, value in enumerate(row):
                    if i < len(headers) and headers[i] and value is not None:
                        row_data[headers[i]] = str(value).strip()
                        has_data = True

                if has_data:  # Only add rows with actual data
                    data.append(row_data)
                    row_count += 1

                # Memory management for large files
                if row_count % 1000 == 0 and row_count > 0:
                    print(f"Processed {row_count} rows...")

        workbook.close()  # Explicitly close workbook

    except Exception as e:
        print(f"Error reading Excel file: {e}")
        raise

    return data


def batch_insert_leads(leads_data, batch_size=100):
    """Insert leads in batches to avoid overwhelming the database"""
    total_inserted = 0
    total_batches = (len(leads_data) + batch_size - 1) // batch_size

    print(f"Starting batch insert: {len(leads_data)} leads in {total_batches} batches")

    for i in range(0, len(leads_data), batch_size):
        batch = leads_data[i:i + batch_size]
        batch_num = (i // batch_size) + 1

        try:
            # Insert batch
            result = supabase.table('lead_master').insert(batch).execute()

            if result.data:
                batch_inserted = len(result.data)
                total_inserted += batch_inserted
                print(f"Batch {batch_num}/{total_batches}: Inserted {batch_inserted} leads")
            else:
                print(f"Batch {batch_num}/{total_batches}: No data returned from insert")

            # Small delay to prevent overwhelming the database
            time.sleep(0.1)  # CHANGED from eventlet.sleep(0.1)

            # Force garbage collection every 10 batches
            if batch_num % 10 == 0:
                gc.collect()

        except Exception as e:
            print(f"Error inserting batch {batch_num}: {e}")
            # Continue with next batch instead of failing completely
            continue

    print(f"Batch insert completed: {total_inserted} total leads inserted")
    return total_inserted


def generate_uid(source, mobile_number, sequence):
    """Generate UID based on source, mobile number, and sequence"""
    source_map = {
        'Google': 'G',
        'Meta': 'M',
        'Affiliate': 'A',
        'Know': 'K',
        'Whatsapp': 'W',
        'Tele': 'T',
        'Activity': 'AC',
        'Walk-in': 'W',  # Walk-in mapping
        'Walkin': 'W'    # Walkin mapping (without hyphen)
    }

    source_char = source_map.get(source, 'X')

    # Get sequence character (A-Z)
    sequence_char = chr(65 + (sequence % 26))  # A=65 in ASCII

    # Get last 4 digits of mobile number
    mobile_str = str(mobile_number).replace(' ', '').replace('-', '')
    mobile_last4 = mobile_str[-4:] if len(mobile_str) >= 4 else mobile_str.zfill(4)

    # Generate sequence number (0001, 0002, etc.)
    seq_num = f"{(sequence % 9999) + 1:04d}"

    return f"{source_char}{sequence_char}-{mobile_last4}-{seq_num}"


def get_next_call_info(lead_data):
    """Determine the next available call number and which calls are completed"""
    call_order = ['first', 'second', 'third', 'fourth', 'fifth', 'sixth', 'seventh']
    completed_calls = []
    next_call = 'first'

    for call_num in call_order:
        call_date_key = f'{call_num}_call_date'
        if lead_data.get(call_date_key):
            completed_calls.append(call_num)
        else:
            next_call = call_num
            break

    return next_call, completed_calls


def get_next_ps_call_info(ps_data):
    """Determine the next available PS call number and which calls are completed (now 7 calls)"""
    call_order = ['first', 'second', 'third', 'fourth', 'fifth', 'sixth', 'seventh']
    completed_calls = []
    next_call = 'first'

    for call_num in call_order:
        call_date_key = f'{call_num}_call_date'
        if ps_data.get(call_date_key):
            completed_calls.append(call_num)
        else:
            next_call = call_num
            break

    return next_call, completed_calls


def get_accurate_count(table_name, filters=None):
    """Get accurate count from Supabase table"""
    try:
        query = supabase.table(table_name).select('id')

        if filters:
            for key, value in filters.items():
                if value is not None:
                    query = query.eq(key, value)

        result = query.execute()

        # Count the returned data
        return len(result.data) if result.data else 0

    except Exception as e:
        print(f"Error getting count from {table_name}: {e}")
        return 0


def safe_get_data(table_name, filters=None, select_fields='*', limit=10000):
    """Safely get data from Supabase with error handling"""
    try:
        query = supabase.table(table_name).select(select_fields)

        if filters:
            for key, value in filters.items():
                if value is not None:
                    query = query.eq(key, value)

        # Add limit to prevent default 1000 row limitation
        if limit:
            query = query.limit(limit)

        result = query.execute()
        return result.data or []
    except Exception as e:
        print(f"Error fetching data from {table_name}: {e}")
        return []


def sync_test_drive_to_alltest_drive(source_table, original_id, lead_data):
    """
    Sync test drive data to alltest_drive table when test_drive_done is updated
    """
    try:
        # Check if test_drive_done is not null and is Yes/No or True/False
        test_drive_done = lead_data.get('test_drive_done')
        
        # Handle both boolean and string values
        if test_drive_done is None:
            return
        
        # Convert boolean to string if needed
        if test_drive_done is True:
            test_drive_done = 'Yes'
        elif test_drive_done is False:
            test_drive_done = 'No'
        elif test_drive_done not in ['Yes', 'No']:
            return
        
        # Check if record already exists in alltest_drive
        existing_record = supabase.table('alltest_drive').select('*').eq('source_table', source_table).eq('original_id', str(original_id)).execute()
        
        # Prepare data for alltest_drive table
        alltest_drive_data = {
            'source_table': source_table,
            'original_id': str(original_id),
            'test_drive_done': test_drive_done,
            'updated_at': datetime.now().isoformat()
        }
        
        # Map fields based on source table
        if source_table == 'walkin_table':
            alltest_drive_data.update({
                'customer_name': lead_data.get('customer_name'),
                'mobile_number': lead_data.get('mobile_number'),
                'lead_status': lead_data.get('lead_status'),
                'lead_category': lead_data.get('lead_category'),
                'model_interested': lead_data.get('model_interested'),
                'final_status': lead_data.get('status'),
                'ps_name': lead_data.get('ps_assigned'),
                'branch': lead_data.get('branch'),
                'created_at': lead_data.get('created_at')
            })
            # For walkin_table, use uid instead of id
            alltest_drive_data['original_id'] = lead_data.get('uid', str(original_id))
        elif source_table == 'ps_followup_master':
            alltest_drive_data.update({
                'customer_name': lead_data.get('customer_name'),
                'mobile_number': lead_data.get('customer_mobile_number'),
                'lead_status': lead_data.get('lead_status'),
                'lead_category': lead_data.get('lead_category'),
                'model_interested': lead_data.get('model_interested'),
                'final_status': lead_data.get('final_status'),
                'ps_name': lead_data.get('ps_name'),
                'branch': lead_data.get('branch') or lead_data.get('ps_branch'),
                'created_at': lead_data.get('created_at'),
                'lead_source': lead_data.get('lead_source'),
                'cre_name': lead_data.get('cre_name')
            })
        elif source_table == 'activity_leads':
            alltest_drive_data.update({
                'customer_name': lead_data.get('customer_name'),
                'mobile_number': lead_data.get('customer_phone_number'),
                'lead_status': lead_data.get('lead_status'),
                'lead_category': lead_data.get('lead_category'),
                'model_interested': lead_data.get('interested_model'),
                'final_status': lead_data.get('final_status'),
                'ps_name': lead_data.get('ps_name'),
                'branch': lead_data.get('location'),
                'created_at': lead_data.get('created_at'),
                'remarks': lead_data.get('remarks'),
                'activity_name': lead_data.get('activity_name'),
                'activity_location': lead_data.get('activity_location'),
                'customer_location': lead_data.get('customer_location'),
                'customer_profession': lead_data.get('customer_profession'),
                'gender': lead_data.get('gender')
            })
        
        # Insert or update record in alltest_drive table
        if existing_record.data:
            # Update existing record
            supabase.table('alltest_drive').update(alltest_drive_data).eq('source_table', source_table).eq('original_id', str(original_id)).execute()
        else:
            # Insert new record
            supabase.table('alltest_drive').insert(alltest_drive_data).execute()
            
        print(f"Successfully synced test drive data for {source_table} - {original_id}")
        
    except Exception as e:
        print(f"Error syncing test drive data to alltest_drive: {e}")


def create_or_update_ps_followup(lead_data, ps_name, ps_branch):
    from datetime import datetime
    try:
        existing = supabase.table('ps_followup_master').select('*').eq('lead_uid', lead_data['uid']).execute()
        ps_followup_data = {
            'lead_uid': lead_data['uid'],
            'ps_name': ps_name,
            'ps_branch': ps_branch,
            'customer_name': lead_data.get('customer_name'),
            'customer_mobile_number': lead_data.get('customer_mobile_number'),
            'source': lead_data.get('source'),
            'cre_name': lead_data.get('cre_name'),
            'lead_category': lead_data.get('lead_category'),
            'model_interested': lead_data.get('model_interested'),
            'final_status': 'Pending',
            'ps_assigned_at': datetime.now().isoformat(),  # Always set when PS is assigned
            'created_at': lead_data.get('created_at') or datetime.now().isoformat(),
            'first_call_date': None  # Ensure fresh leads start without first_call_date
        }
        if existing.data:
            supabase.table('ps_followup_master').update(ps_followup_data).eq('lead_uid', lead_data['uid']).execute()
        else:
            supabase.table('ps_followup_master').insert(ps_followup_data).execute()
    except Exception as e:
        print(f"Error creating/updating PS followup: {e}")

def track_cre_call_attempt(uid, cre_name, call_no, lead_status, call_was_recorded=False, follow_up_date=None, remarks=None):
    """Track CRE call attempt in the history table and update TAT for first attempt"""
    try:
        # Get the next attempt number for this call
        attempt_result = supabase.table('cre_call_attempt_history').select('attempt').eq('uid', uid).eq('call_no', call_no).order('attempt', desc=True).limit(1).execute()
        next_attempt = 1
        if attempt_result.data:
            next_attempt = attempt_result.data[0]['attempt'] + 1

        # Fetch the current final_status from lead_master
        final_status = None
        lead_result = supabase.table('lead_master').select('final_status').eq('uid', uid).limit(1).execute()
        if lead_result.data and 'final_status' in lead_result.data[0]:
            final_status = lead_result.data[0]['final_status']

        # Prepare attempt data
        attempt_data = {
            'uid': uid,
            'call_no': call_no,
            'attempt': next_attempt,
            'status': lead_status,
            'cre_name': cre_name,
            'call_was_recorded': call_was_recorded,
            'follow_up_date': follow_up_date,
            'remarks': remarks,
            'final_status': final_status
        }

        # Insert the attempt record
        insert_result = supabase.table('cre_call_attempt_history').insert(attempt_data).execute()
        print(f"Tracked call attempt: {uid} - {call_no} call, attempt {next_attempt}, status: {lead_status}")
        
        # Send WebSocket notification for real-time updates
        if websocket_manager:
            websocket_manager.broadcast_lead_update(uid, {
                'call_attempt': {
                    'call_no': call_no,
                    'attempt': next_attempt,
                    'status': lead_status,
                    'cre_name': cre_name,
                    'timestamp': datetime.now().isoformat()
                }
            })

        # --- TAT Calculation and Update ---
        if call_no == 'first' and next_attempt == 1:
            # Fetch lead's created_at
            lead_result = supabase.table('lead_master').select('created_at').eq('uid', uid).limit(1).execute()
            if lead_result.data and lead_result.data[0].get('created_at'):
                created_at_str = lead_result.data[0]['created_at']
                from datetime import datetime
                try:
                    if 'T' in created_at_str:
                        created_at = datetime.fromisoformat(created_at_str.replace('Z', '+00:00'))
                    else:
                        created_at = datetime.strptime(created_at_str, '%Y-%m-%d %H:%M:%S')
                except Exception:
                    created_at = None
                # Get updated_at from inserted attempt (if available), else use now
                updated_at_str = None
                if insert_result.data and insert_result.data[0].get('updated_at'):
                    updated_at_str = insert_result.data[0]['updated_at']
                else:
                    from datetime import datetime
                    updated_at_str = datetime.now().isoformat()
                try:
                    if 'T' in updated_at_str:
                        updated_at = datetime.fromisoformat(updated_at_str.replace('Z', '+00:00'))
                    else:
                        updated_at = datetime.strptime(updated_at_str, '%Y-%m-%d %H:%M:%S')
                except Exception:
                    updated_at = datetime.now()
                if created_at:
                    tat_seconds = (updated_at - created_at).total_seconds()
                    # Update lead_master with TAT
                    supabase.table('lead_master').update({'tat': tat_seconds}).eq('uid', uid).execute()
                    print(f"TAT updated for lead {uid}: {tat_seconds} seconds")
    except Exception as e:
        print(f"Error tracking CRE call attempt: {e}")

def track_ps_call_attempt(uid, ps_name, call_no, lead_status, call_was_recorded=False, follow_up_date=None, remarks=None):
    """Track PS call attempt in the history table"""
    try:
        # Get the next attempt number for this call
        attempt_result = supabase.table('ps_call_attempt_history').select('attempt').eq('uid', uid).eq('call_no', call_no).order('attempt', desc=True).limit(1).execute()
        next_attempt = 1
        if attempt_result.data:
            next_attempt = attempt_result.data[0]['attempt'] + 1

        # Fetch the current final_status from multiple tables
        final_status = None
        
        # Try all possible tables and field combinations
        tables_to_check = [
            # (table_name, uid_field, status_field, uid_value)
            ('ps_followup_master', 'lead_uid', 'final_status', uid),
            ('lead_master', 'uid', 'final_status', uid),
            ('walkin_table', 'uid', 'status', uid),
            ('activity_leads', 'activity_uid', 'final_status', uid),
            # Also try with different UID formats for walkin and activity
            ('walkin_table', 'uid', 'status', uid.replace('WB-', 'W')),
            ('activity_leads', 'activity_uid', 'final_status', uid.replace('WB-', 'A'))
        ]
        
        for table_name, uid_field, status_field, uid_value in tables_to_check:
            try:
                result = supabase.table(table_name).select(status_field).eq(uid_field, uid_value).limit(1).execute()
                if result.data and result.data[0].get(status_field):
                    final_status = result.data[0][status_field]
                    print(f"Found final_status '{final_status}' in {table_name} for {uid_value}")
                    break
            except Exception as e:
                print(f"Error checking {table_name}: {e}")
                continue
        
        # If still not found, set a default
        if not final_status:
            final_status = 'Pending'
            print(f"No final_status found for {uid}, defaulting to 'Pending'")

        # Prepare attempt data
        attempt_data = {
            'uid': uid,
            'call_no': call_no,
            'attempt': next_attempt,
            'status': lead_status,
            'ps_name': ps_name,
            'call_was_recorded': call_was_recorded,
            'follow_up_date': follow_up_date,
            'remarks': remarks,
            'final_status': final_status
        }

        # Insert the attempt record
        supabase.table('ps_call_attempt_history').insert(attempt_data).execute()
        print(f"Tracked PS call attempt: {uid} - {call_no} call, attempt {next_attempt}, status: {lead_status}")
        
        # Send WebSocket notification for real-time updates
        if websocket_manager:
            websocket_manager.broadcast_lead_update(uid, {
                'ps_call_attempt': {
                    'call_no': call_no,
                    'attempt': next_attempt,
                    'status': lead_status,
                    'ps_name': ps_name,
                    'timestamp': datetime.now().isoformat()
                }
            })
    except Exception as e:
        print(f"Error tracking PS call attempt: {e}")


def filter_leads_by_date(leads, filter_type, date_field='created_at'):
    """Filter leads based on date range"""
    if filter_type == 'all':
        return leads

    today = datetime.now().date()

    if filter_type == 'today':
        start_date = today
        end_date = today
    elif filter_type == 'mtd':  # Month to Date
        start_date = today.replace(day=1)
        end_date = today
    elif filter_type == 'week':
        start_date = today - timedelta(days=today.weekday())  # Start of current week (Monday)
        end_date = today
    elif filter_type == 'month':
        start_date = today - timedelta(days=30)
        end_date = today
    elif filter_type == 'quarter':
        start_date = today - timedelta(days=90)
        end_date = today
    elif filter_type == 'year':
        start_date = today - timedelta(days=365)
        end_date = today
    else:
        return leads

    filtered_leads = []
    for lead in leads:
        lead_date_str = lead.get(date_field)
        if lead_date_str:
            try:
                # Handle different date formats
                if 'T' in lead_date_str:  # ISO format with time
                    lead_date = datetime.fromisoformat(lead_date_str.replace('Z', '+00:00')).date()
                else:  # Date only format
                    lead_date = datetime.strptime(lead_date_str, '%Y-%m-%d').date()

                if start_date <= lead_date <= end_date:
                    filtered_leads.append(lead)
            except (ValueError, TypeError):
                # If date parsing fails, include the lead
                filtered_leads.append(lead)
        else:
            # If no date field, include the lead
            filtered_leads.append(lead)

    return filtered_leads


def fix_missing_timestamps():
    """
    Fix missing timestamps for existing leads that have final_status but missing won_timestamp or lost_timestamp
    """
    try:
        # Fix lead_master table
        # Get leads with final_status = 'Won' but no won_timestamp
        won_leads = supabase.table('lead_master').select('uid, final_status, won_timestamp, lost_timestamp, updated_at').eq('final_status', 'Won').is_('won_timestamp', 'null').execute()
        
        for lead in won_leads.data:
            supabase.table('lead_master').update({
                'won_timestamp': lead.get('updated_at') or datetime.now().isoformat()
            }).eq('uid', lead['uid']).execute()
        
        # Get leads with final_status = 'Lost' but no lost_timestamp
        lost_leads = supabase.table('lead_master').select('uid, final_status, won_timestamp, lost_timestamp, updated_at').eq('final_status', 'Lost').is_('lost_timestamp', 'null').execute()
        
        for lead in lost_leads.data:
            supabase.table('lead_master').update({
                'lost_timestamp': lead.get('updated_at') or datetime.now().isoformat()
            }).eq('uid', lead['uid']).execute()
        
        # Fix ps_followup_master table
        # Get PS leads with final_status = 'Won' but no won_timestamp
        ps_won_leads = supabase.table('ps_followup_master').select('lead_uid, final_status, won_timestamp, lost_timestamp, updated_at').eq('final_status', 'Won').is_('won_timestamp', 'null').execute()
        
        for lead in ps_won_leads.data:
            supabase.table('ps_followup_master').update({
                'won_timestamp': lead.get('updated_at') or datetime.now().isoformat()
            }).eq('lead_uid', lead['lead_uid']).execute()
        
        # Get PS leads with final_status = 'Lost' but no lost_timestamp
        ps_lost_leads = supabase.table('ps_followup_master').select('lead_uid, final_status, won_timestamp, lost_timestamp, updated_at').eq('final_status', 'Lost').is_('lost_timestamp', 'null').execute()
        
        for lead in ps_lost_leads.data:
            supabase.table('ps_followup_master').update({
                'lost_timestamp': lead.get('updated_at') or datetime.now().isoformat()
            }).eq('lead_uid', lead['lead_uid']).execute()
        
        print(f"Fixed {len(won_leads.data)} won leads, {len(lost_leads.data)} lost leads in lead_master")
        print(f"Fixed {len(ps_won_leads.data)} won leads, {len(ps_lost_leads.data)} lost leads in ps_followup_master")
        
    except Exception as e:
        print(f"Error fixing timestamps: {str(e)}")

# Routes moved to after app initialization


@app.route('/')
def index():
    session.clear()  # Ensure no session data is present
    return render_template('index.html')
@app.route('/unified_login', methods=['POST'])
@limiter.limit("100000 per minute")
def unified_login() -> Response:
    start_time = time.time()
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '').strip()
    user_type = request.form.get('user_type', '').strip().lower()

    valid_user_types = ['admin', 'cre', 'ps', 'branch_head', 'rec']
    if user_type not in valid_user_types:
        flash('Please select a valid role (Admin, CRE, PS, Branch Head, or Receptionist)', 'error')
        return redirect(url_for('index'))

    if user_type == 'branch_head':
        bh = supabase.table('Branch Head').select('*').eq('Username', username).execute().data
        if not bh:
            flash('Invalid username or password', 'error')
            return redirect(url_for('index'))
        bh = bh[0]
        if not bh['Is Active']:
            flash('User is inactive', 'error')
            return redirect(url_for('index'))
        # Use plain text password comparison for Branch Head
        if bh['Password'] != password:
            flash('Incorrect password', 'error')
            return redirect(url_for('index'))
        session['branch_head_id'] = bh['id']
        session['branch_head_name'] = bh['Name']
        session['branch_head_branch'] = bh['Branch']
        session['user_type'] = 'branch_head'  # Add this line to set user_type
        session['username'] = username  # Add username for consistency
        flash('Welcome! Logged in as Branch Head', 'success')
        return redirect('/branch_head_dashboard')

    elif user_type == 'rec':
        # Receptionist authentication
        rec_user = supabase.table('rec_users').select('*').eq('username', username).execute().data
        if not rec_user:
            flash('Invalid username or password', 'error')
            return redirect(url_for('index'))
        rec_user = rec_user[0]
        if not rec_user.get('is_active', True):
            flash('User is inactive', 'error')
            return redirect(url_for('index'))
        # Check password using werkzeug
        if not check_password_hash(rec_user['password_hash'], password):
            flash('Incorrect password', 'error')
            return redirect(url_for('index'))
        session.clear()
        session['rec_user_id'] = rec_user['id']
        session['rec_branch'] = rec_user['branch']
        session['rec_name'] = rec_user.get('name', username)
        session['user_type'] = 'rec'
        session['username'] = username
        flash('Welcome! Logged in as Receptionist', 'success')
        return redirect(url_for('add_walkin_lead'))

    # Existing logic for admin, cre, ps
    t_user = time.time()
    success, message, user_data = auth_manager.authenticate_user(username, password, user_type)
    print(f"[PERF] unified_login: authenticate_user({user_type}) took {time.time() - t_user:.3f} seconds")
    if success:
        t2 = time.time()
        session_id = auth_manager.create_session(user_data['id'], user_type, user_data)
        print(f"DEBUG: Logged in as user_type={user_type}, session.user_type={session.get('user_type')}")
        print(f"[PERF] unified_login: create_session took {time.time() - t2:.3f} seconds")
        if session_id:
            flash(f'Welcome! Logged in as {user_type.upper()}', 'success')
            t3 = time.time()
            # Redirect to appropriate dashboard
            if user_type == 'admin':
                print(f"[PERF] unified_login: redirect to admin_dashboard after {time.time() - t3:.3f} seconds")
                print(f"[PERF] unified_login TOTAL took {time.time() - start_time:.3f} seconds")
                return redirect(url_for('admin_dashboard'))
            elif user_type == 'cre':
                print(f"[PERF] unified_login: redirect to cre_dashboard after {time.time() - t3:.3f} seconds")
                print(f"[PERF] unified_login TOTAL took {time.time() - start_time:.3f} seconds")
                return redirect(url_for('cre_dashboard'))
            elif user_type == 'ps':
                print(f"[PERF] unified_login: redirect to ps_dashboard after {time.time() - t3:.3f} seconds")
                print(f"[PERF] unified_login TOTAL took {time.time() - start_time:.3f} seconds")
                return redirect(url_for('ps_dashboard'))
        else:
            flash('Error creating session', 'error')
            print(f"[PERF] unified_login: session creation failed after {time.time() - t2:.3f} seconds")
            print(f"[PERF] unified_login TOTAL took {time.time() - start_time:.3f} seconds")
            return redirect(url_for('index'))
    else:
        flash('Invalid username or password', 'error')
        print(f"[PERF] unified_login TOTAL (invalid login) took {time.time() - start_time:.3f} seconds")
        return redirect(url_for('index'))
# Keep the old login routes for backward compatibility (redirect to unified login)
@app.route('/admin_login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        return unified_login()
    return redirect(url_for('index'))


@app.route('/cre_login', methods=['GET', 'POST'])
def cre_login():
    if request.method == 'POST':
        return unified_login()
    return redirect(url_for('index'))


@app.route('/ps_login', methods=['GET', 'POST'])
def ps_login():
    if request.method == 'POST':
        return unified_login()
    return redirect(url_for('index'))


@app.route('/password_reset_request', methods=['GET', 'POST'])
@limiter.limit("5 per minute")
def password_reset_request():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        user_type = request.form.get('user_type', '').strip()

        if not username or not user_type:
            flash('Please enter username and select user type', 'error')
            return render_template('password_reset_request.html')

        success, message, token = auth_manager.generate_password_reset_token(username, user_type)

        if success:
            # Send the reset link via email
            # Fetch user email from the appropriate table
            user_email = None
            try:
                if user_type == 'admin':
                    user_result = supabase.table('admin_users').select('email').eq('username', username).execute()
                elif user_type == 'cre':
                    user_result = supabase.table('cre_users').select('email').eq('username', username).execute()
                elif user_type == 'ps':
                    user_result = supabase.table('ps_users').select('email').eq('username', username).execute()
                else:
                    user_result = None
                if user_result and user_result.data and user_result.data[0].get('email'):
                    user_email = user_result.data[0]['email']
            except Exception as e:
                print(f"Error fetching user email for password reset: {e}")
                user_email = None

            reset_url = url_for('password_reset', token=token, _external=True)
            email_sent = False
            if user_email:
                try:
                    msg = MIMEMultipart()
                    msg['From'] = EMAIL_USER
                    msg['To'] = user_email
                    msg['Subject'] = 'Ather CRM Password Reset Request'
                    body = f"""
                    Dear {username},

                    We received a request to reset your password for your Ather CRM account.

                    Please click the link below to reset your password:
                    {reset_url}

                    If you did not request this, please ignore this email.

                    Best regards,\nAther CRM System
                    """
                    msg.attach(MIMEText(body, 'plain'))
                    server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
                    server.starttls()
                    server.login(EMAIL_USER, EMAIL_PASSWORD)
                    text = msg.as_string()
                    server.sendmail(EMAIL_USER, user_email, text)
                    server.quit()
                    print(f"Password reset email sent to {user_email}")
                    email_sent = True
                except Exception as e:
                    print(f"Error sending password reset email: {e}")
                    email_sent = False
            if email_sent:
                flash('If the username exists and is valid, a password reset link has been sent to the registered email address.', 'success')
            else:
                flash('If the username exists and is valid, a password reset link has been sent to the registered email address.', 'success')
                # Optionally, log or alert admin if email sending failed
        else:
            flash(message, 'error')

    return render_template('password_reset_request.html')


@app.route('/password_reset/<token>', methods=['GET', 'POST'])
def password_reset(token):
    if request.method == 'POST':
        new_password = request.form.get('new_password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()

        if not new_password or not confirm_password:
            flash('Please enter and confirm your new password', 'error')
            return render_template('password_reset.html', token=token)

        if new_password != confirm_password:
            flash('Passwords do not match', 'error')
            return render_template('password_reset.html', token=token)

        success, message = auth_manager.reset_password_with_token(token, new_password)

        if success:
            flash('Password reset successfully. Please log in with your new password.', 'success')
            return redirect(url_for('index'))
        else:
            flash(message, 'error')

    return render_template('password_reset.html', token=token)


@app.route('/change_password', methods=['POST'])
@require_auth()
def change_password():
    current_password = request.form.get('current_password', '').strip()
    new_password = request.form.get('new_password', '').strip()
    confirm_password = request.form.get('confirm_password', '').strip()

    if not all([current_password, new_password, confirm_password]):
        flash('All fields are required', 'error')
        return redirect(url_for('security_settings'))

    if new_password != confirm_password:
        flash('New passwords do not match', 'error')
        return redirect(url_for('security_settings'))

    user_id = session.get('user_id')
    user_type = session.get('user_type')

    if not user_id or not user_type:
        flash('Session information not found', 'error')
        return redirect(url_for('security_settings'))

    success, message = auth_manager.change_password(user_id, user_type, current_password, new_password)

    if success:
        flash('Password changed successfully', 'success')
    else:
        flash(message, 'error')

    return redirect(url_for('security_settings'))


@app.route('/change_cre_password', methods=['GET', 'POST'])
@require_cre
def change_cre_password():
    if request.method == 'POST':
        current_password = request.form.get('current_password', '').strip()
        new_password = request.form.get('new_password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()

        if not all([current_password, new_password, confirm_password]):
            flash('All fields are required', 'error')
            return render_template('change_cre_password.html')

        if new_password != confirm_password:
            flash('New passwords do not match', 'error')
            return render_template('change_cre_password.html')

        user_id = session.get('user_id')
        user_type = session.get('user_type')

        if not user_id or not user_type:
            flash('Session information not found', 'error')
            return render_template('change_cre_password.html')

        success, message = auth_manager.change_password(user_id, user_type, current_password, new_password)

        if success:
            flash('Password changed successfully', 'success')
            return redirect(url_for('cre_dashboard'))
        else:
            flash(message, 'error')

    return render_template('change_cre_password.html')


@app.route('/change_ps_password', methods=['GET', 'POST'])
@require_ps
def change_ps_password():
    if request.method == 'POST':
        current_password = request.form.get('current_password', '').strip()
        new_password = request.form.get('new_password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()

        if not all([current_password, new_password, confirm_password]):
            flash('All fields are required', 'error')
            return render_template('change_ps_password.html')

        if new_password != confirm_password:
            flash('New passwords do not match', 'error')
            return render_template('change_ps_password.html')

        user_id = session.get('user_id')
        user_type = session.get('user_type')

        if not user_id or not user_type:
            flash('Session information not found', 'error')
            return render_template('change_ps_password.html')

        success, message = auth_manager.change_password(user_id, user_type, current_password, new_password)

        if success:
            flash('Password changed successfully', 'success')
            return redirect(url_for('ps_dashboard'))
        else:
            flash(message, 'error')

    return render_template('change_ps_password.html')


@app.route('/security_settings')
@require_auth()
def security_settings():
    user_id = session.get('user_id')
    user_type = session.get('user_type')

    if not user_id or not user_type:
        flash('Session information not found', 'error')
        return redirect(url_for('index'))

    # Get active sessions
    sessions = auth_manager.get_user_sessions(user_id, user_type)

    # Get audit logs
    audit_logs = auth_manager.get_audit_logs(user_id, user_type, limit=20)

    return render_template('security_settings.html', sessions=sessions, audit_logs=audit_logs)


# WebSocket user info route moved to top of file to avoid conflicts with parameterized routes

@app.route('/security_audit')
@require_admin
def security_audit():
    """Security audit dashboard"""
    return render_template('security_audit.html')


@app.route('/run_security_audit', methods=['POST'])
@require_admin
def run_security_audit():
    """Run comprehensive security audit"""
    try:
        # Run security verification
        audit_results = run_security_verification(supabase)

        # Log the security audit
        user_id = session.get('user_id')
        user_type = session.get('user_type')
        
        if user_id and user_type:
            auth_manager.log_audit_event(
                user_id=user_id,
                user_type=user_type,
                action='SECURITY_AUDIT_RUN',
                resource='security_audit',
                details={'overall_score': audit_results.get('overall_score', 0)}
            )

        return jsonify({
            'success': True,
            'results': audit_results
        })
    except Exception as e:
        print(f"Error running security audit: {e}")
        return jsonify({
            'success': False,
            'message': str(e)
        })


@app.route('/terminate_session', methods=['POST'])
@require_auth()
def terminate_session():
    try:
        data = request.get_json()
        session_id = data.get('session_id')

        if not session_id:
            return jsonify({'success': False, 'message': 'Session ID required'})

        auth_manager.deactivate_session(session_id)

        user_id = session.get('user_id')
        user_type = session.get('user_type')
        
        if user_id and user_type:
            auth_manager.log_audit_event(
                user_id=user_id,
                user_type=user_type,
                action='SESSION_TERMINATED',
                details={'terminated_session': session_id}
            )

        return jsonify({'success': True, 'message': 'Session terminated successfully'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/terminate_all_sessions', methods=['POST'])
@require_auth()
def terminate_all_sessions():
    try:
        user_id = session.get('user_id')
        user_type = session.get('user_type')
        current_session = session.get('session_id')

        if user_id and user_type and current_session:
            auth_manager.deactivate_all_user_sessions(user_id, user_type, current_session)
            auth_manager.log_audit_event(
                user_id=user_id,
                user_type=user_type,
                action='ALL_SESSIONS_TERMINATED',
                details={'except_session': current_session}
            )
        else:
            return jsonify({'success': False, 'message': 'Session information not found'})

        return jsonify({'success': True, 'message': 'All other sessions terminated successfully'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/admin_dashboard')
@require_admin
def admin_dashboard():
    # Get counts for dashboard with better error handling and actual queries
    try:
        # Get actual counts from database with proper queries
        cre_count = get_accurate_count('cre_users')
        ps_count = get_accurate_count('ps_users')
        leads_count = get_accurate_count('lead_master')
        unassigned_leads = get_accurate_count('lead_master', {'assigned': 'No'})

        print(
            f"Dashboard counts - CRE: {cre_count}, PS: {ps_count}, Total Leads: {leads_count}, Unassigned: {unassigned_leads}")

    except Exception as e:
        print(f"Error getting dashboard counts: {e}")
        cre_count = ps_count = leads_count = unassigned_leads = 0

    # Log dashboard access
    user_id = session.get('user_id')
    user_type = session.get('user_type')
    if user_id and user_type:
        auth_manager.log_audit_event(
            user_id=user_id,
            user_type=user_type,
            action='DASHBOARD_ACCESS',
            resource='admin_dashboard'
        )

    return render_template('admin_dashboard.html',
                           cre_count=cre_count,
                           ps_count=ps_count,
                           leads_count=leads_count,
                           unassigned_leads=unassigned_leads)


@app.route('/upload_data', methods=['GET', 'POST'])
@require_admin
def upload_data():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file selected', 'error')
            return redirect(request.url)

        file = request.files['file']
        source = request.form.get('source', '').strip()

        if not source:
            flash('Please select a data source', 'error')
            return redirect(request.url)

        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(request.url)

        if file and file.filename and allowed_file(file.filename):
            filename = secure_filename(str(file.filename))
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

            try:
                file.save(filepath)

                # Check file size
                file_size = os.path.getsize(filepath)
                if file_size > 50 * 1024 * 1024:  # 50MB limit
                    flash('File too large. Maximum size is 50MB.', 'error')
                    os.remove(filepath)
                    return redirect(request.url)

                print(f"Processing file: {filename} ({file_size / 1024 / 1024:.2f} MB)")

                # Read the file based on extension
                if filename.lower().endswith('.csv'):
                    data = read_csv_file(filepath)
                else:
                    data = read_excel_file(filepath)

                if not data:
                    flash('No valid data found in file', 'error')
                    os.remove(filepath)
                    return redirect(request.url)

                print(f"Read {len(data)} rows from file")

                # Get current sequence number for UID generation
                result = supabase.table('lead_master').select('uid').execute()
                current_count = len(result.data) if result.data else 0

                # Prepare leads data for batch insert
                leads_to_insert = []
                skipped_rows = 0

                for index, row in enumerate(data):
                    try:
                        # Validate required fields
                        required_fields = ['customer_name', 'customer_mobile_number', 'date']
                        if not all(key in row and str(row[key]).strip() for key in required_fields):
                            skipped_rows += 1
                            continue

                        uid = generate_uid(source, row['customer_mobile_number'],
                                           current_count + len(leads_to_insert) + 1)

                        lead_data = {
                            'uid': uid,
                            'date': str(row['date']).strip(),
                            'customer_name': str(row['customer_name']).strip(),
                            'customer_mobile_number': str(row['customer_mobile_number']).strip(),
                            'source': source,
                            'assigned': 'No',
                            'final_status': 'Pending'
                        }

                        leads_to_insert.append(lead_data)


                    except Exception as e:
                        print(f"Error processing row {index}: {e}")
                        skipped_rows += 1
                        continue

                if not leads_to_insert:
                    flash('No valid leads found to insert', 'error')
                    os.remove(filepath)
                    return redirect(request.url)

                print(f"Prepared {len(leads_to_insert)} leads for insertion")

                # Batch insert leads
                success_count = batch_insert_leads(leads_to_insert)

                # Log data upload
                auth_manager.log_audit_event(
                    user_id=session.get('user_id'),
                    user_type=session.get('user_type'),
                    action='DATA_UPLOAD',
                    resource='lead_master',
                    details={
                        'source': source,
                        'records_uploaded': success_count,
                        'filename': filename,
                        'file_size_mb': round(file_size / 1024 / 1024, 2),
                        'skipped_rows': skipped_rows
                    }
                )

                # Create success message
                message = f'Successfully uploaded {success_count} leads'
                if skipped_rows > 0:
                    message += f' ({skipped_rows} rows skipped due to missing data)'
                message += '. Please go to "Assign Leads" to assign them to CREs.'

                flash(message, 'success')

                # Clean up uploaded file
                os.remove(filepath)

            except Exception as e:
                print(f"Error processing file: {e}")
                flash(f'Error processing file: {str(e)}', 'error')
                if os.path.exists(filepath):
                    os.remove(filepath)
        else:
            flash('Invalid file format. Please upload CSV or Excel files only.', 'error')

    return render_template('upload_data.html')

@app.route('/assign_leads')
@require_admin
def assign_leads():
    try:
        # Fetch all unassigned leads in batches of 1000
        all_unassigned_leads = []
        batch_size = 1000
        offset = 0
        while True:
            result = supabase.table('lead_master').select('*').eq('assigned', 'No').range(offset, offset + batch_size - 1).execute()
            batch = result.data or []
            all_unassigned_leads.extend(batch)
            if len(batch) < batch_size:
                break
            offset += batch_size

        # Organize by source
        leads_by_source = {}
        for lead in all_unassigned_leads:
            source = lead.get('source', 'Unknown')
            leads_by_source.setdefault(source, []).append(lead)

        # Get CREs
        cres = safe_get_data('cre_users')

        # Get accurate total unassigned count
        actual_unassigned_count = get_accurate_count('lead_master', {'assigned': 'No'})

        # Get accurate per-source unassigned counts
        source_unassigned_counts = {}
        for source in leads_by_source.keys():
            source_unassigned_counts[source] = get_accurate_count('lead_master', {'assigned': 'No', 'source': source})

        return render_template('assign_leads.html',
                               unassigned_leads=all_unassigned_leads,
                               actual_unassigned_count=actual_unassigned_count,
                               cres=cres,
                               leads_by_source=leads_by_source,
                               source_unassigned_counts=source_unassigned_counts)
    except Exception as e:
        print(f"Error loading assign leads data: {e}")
        flash(f'Error loading data: {str(e)}', 'error')
        return redirect(url_for('admin_dashboard'))


@app.route('/assign_leads_dynamic_action', methods=['POST'])
@require_admin
def assign_leads_dynamic_action():
    try:
        data = request.get_json()
        assignments = data.get('assignments', [])

        if not assignments:
            return jsonify({'success': False, 'message': 'No assignments provided'}), 400

        # Fetch all unassigned leads in batches
        all_unassigned = []
        batch_size = 1000
        offset = 0
        while True:
            result = supabase.table('lead_master').select('*').eq('assigned', 'No').range(offset, offset + batch_size - 1).execute()
            batch = result.data or []
            all_unassigned.extend(batch)
            if len(batch) < batch_size:
                break
            offset += batch_size

        leads_by_source = {}
        for lead in all_unassigned:
            source = lead.get('source', 'Unknown')
            leads_by_source.setdefault(source, []).append(lead)

        total_assigned = 0

        for assignment in assignments:
            cre_id = assignment.get('cre_id')
            source = assignment.get('source')
            quantity = assignment.get('quantity')

            if not cre_id or not source or not quantity:
                continue

            # Get CRE details
            cre_data = supabase.table('cre_users').select('*').eq('id', cre_id).execute()
            if not cre_data.data:
                continue

            cre = cre_data.data[0]
            leads = leads_by_source.get(source, [])

            if not leads:
                print(f"No unassigned leads found for source {source}")
                continue

            random.shuffle(leads)
            leads_to_assign = leads[:quantity]
            leads_by_source[source] = leads[quantity:]  # Remove assigned leads

            for lead in leads_to_assign:
                update_data = {
                    'cre_name': cre['name'],
                    'assigned': 'Yes',
                    'cre_assigned_at': datetime.now().isoformat()

                }

                try:
                    supabase.table('lead_master').update(update_data).eq('uid', lead['uid']).execute()
                    total_assigned += 1
                    print(f"Assigned lead {lead['uid']} to CRE {cre['name']} for source {source}")
                    
                    # Send WebSocket notification for real-time updates
                    if websocket_manager:
                        websocket_manager.notify_user(
                            user_id=cre.get('id'),
                            user_type='cre',
                            event='lead_assigned_notification',
                            data={
                                'lead_uid': lead['uid'],
                                'source': source,
                                'assigned_at': datetime.now().isoformat(),
                                'message': f'New lead {lead["uid"]} assigned to you from {source}'
                            }
                        )
                    
                    if total_assigned % 100 == 0:
                        time.sleep(0.1)
                except Exception as e:
                    print(f"Error assigning lead {lead['uid']}: {e}")

        print(f"Total leads assigned: {total_assigned}")
        return jsonify({'success': True, 'message': f'Total {total_assigned} leads assigned successfully'})

    except Exception as e:
        print(f"Error in dynamic lead assignment: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500
@app.route('/add_cre', methods=['GET', 'POST'])
@require_admin
def add_cre():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        phone = request.form.get('phone', '').strip()
        email = request.form.get('email', '').strip()

        if not all([name, username, password, phone, email]):
            flash('All fields are required', 'error')
            return render_template('add_cre.html')

        if not auth_manager.validate_password_strength(password):
            flash(
                'Password must be at least 8 characters long and contain uppercase, lowercase, number, and special character',
                'error')
            return render_template('add_cre.html')

        try:
            # Check if username already exists
            existing = supabase.table('cre_users').select('username').eq('username', username).execute()
            if existing.data:
                flash('Username already exists', 'error')
                return render_template('add_cre.html')

            # Hash password
            password_hash, salt = auth_manager.hash_password(password)

            # Replace the existing cre_data creation with this:
            cre_data = {
                'name': name,
                'username': username,
                'password': password,  # Keep for backward compatibility
                'password_hash': password_hash,
                'salt': salt,
                'phone': phone,
                'email': email,
                'is_active': True,
                'role': 'cre',
                'failed_login_attempts': 0
            }

            result = supabase.table('cre_users').insert(cre_data).execute()

            # Log CRE creation
            auth_manager.log_audit_event(
                user_id=session.get('user_id'),
                user_type=session.get('user_type'),
                action='CRE_CREATED',
                resource='cre_users',
                resource_id=str(result.data[0]['id']) if result.data else None,
                details={'cre_name': name, 'username': username}
            )

            flash('CRE added successfully', 'success')
            return redirect(url_for('manage_cre'))
        except Exception as e:
            flash(f'Error adding CRE: {str(e)}', 'error')

    return render_template('add_cre.html')
@app.route('/add_ps', methods=['GET', 'POST'])
@require_admin
def add_ps():
    branches = get_system_branches()

    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        phone = request.form.get('phone', '').strip()
        email = request.form.get('email', '').strip()
        branch = request.form.get('branch', '').strip()

        if not all([name, username, password, phone, email, branch]):
            flash('All fields are required', 'error')
            return render_template('add_ps.html', branches=branches)

        if not auth_manager.validate_password_strength(password):
            flash(
                'Password must be at least 8 characters long and contain uppercase, lowercase, number, and special character',
                'error')
            return render_template('add_ps.html', branches=branches)

        try:
            # Check if username already exists
            existing = supabase.table('ps_users').select('username').eq('username', username).execute()
            if existing.data:
                flash('Username already exists', 'error')
                return render_template('add_ps.html', branches=branches)

            # Hash password
            password_hash, salt = auth_manager.hash_password(password)

            # Replace the existing ps_data creation with this:
            ps_data = {
                'name': name,
                'username': username,
                'password': password,  # Keep for backward compatibility
                'password_hash': password_hash,
                'salt': salt,
                'phone': phone,
                'email': email,
                'branch': branch,
                'is_active': True,
                'role': 'ps',
                'failed_login_attempts': 0
            }

            result = supabase.table('ps_users').insert(ps_data).execute()

            # Log PS creation
            auth_manager.log_audit_event(
                user_id=session.get('user_id'),
                user_type=session.get('user_type'),
                action='PS_CREATED',
                resource='ps_users',
                resource_id=str(result.data[0]['id']) if result.data else None,
                details={'ps_name': name, 'username': username, 'branch': branch}
            )

            flash('Product Specialist added successfully', 'success')
            return redirect(url_for('manage_ps'))
        except Exception as e:
            flash(f'Error adding Product Specialist: {str(e)}', 'error')

    return render_template('add_ps.html', branches=branches)


@app.route('/manage_cre')
@require_admin
def manage_cre():
    try:
        cre_users = safe_get_data('cre_users')
        return render_template('manage_cre.html', cre_users=cre_users)
    except Exception as e:
        flash(f'Error loading CRE users: {str(e)}', 'error')
        return render_template('manage_cre.html', cre_users=[])


@app.route('/manage_ps')
@require_admin
def manage_ps():
    try:
        ps_users = safe_get_data('ps_users')
        return render_template('manage_ps.html', ps_users=ps_users)
    except Exception as e:
        flash(f'Error loading PS users: {str(e)}', 'error')
        return render_template('manage_ps.html', ps_users=[])


@app.route('/toggle_ps_status/<int:ps_id>', methods=['POST'])
@require_admin
def toggle_ps_status(ps_id):
    try:
        data = request.get_json()
        active_status = data.get('active', True)

        # Update PS status
        result = supabase.table('ps_users').update({
            'is_active': active_status
        }).eq('id', ps_id).execute()

        if result.data:
            # Log status change
            auth_manager.log_audit_event(
                user_id=session.get('user_id'),
                user_type=session.get('user_type'),
                action='PS_STATUS_CHANGED',
                resource='ps_users',
                resource_id=str(ps_id),
                details={'new_status': 'active' if active_status else 'inactive'}
            )

            return jsonify({'success': True, 'message': 'PS status updated successfully'})
        else:
            return jsonify({'success': False, 'message': 'PS not found'})

    except Exception as e:
        print(f"Error toggling PS status: {e}")
        return jsonify({'success': False, 'message': str(e)})


@app.route('/manage_leads')
@require_admin
def manage_leads():
    try:
        cres = safe_get_data('cre_users')
        cre_id = request.args.get('cre_id')
        source = request.args.get('source')
        qualification = request.args.get('qualification', 'all')
        date_filter = request.args.get('date_filter', 'all')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        final_status = request.args.get('final_status', '')
        page = int(request.args.get('page', 1))
        per_page = 50
        search_uid = request.args.get('search_uid', '').strip()
        selected_cre = None
        leads = []
        sources = []
        total_leads = 0
        if cre_id:
            # Find selected CRE
            selected_cre = next((cre for cre in cres if str(cre.get('id')) == str(cre_id)), None)
            # Fetch leads for this CRE
            filters = {'cre_name': selected_cre['name']} if selected_cre else {}
            if source:
                filters['source'] = source
            leads = safe_get_data('lead_master', filters)
            # UID substring search for this CRE
            if search_uid:
                leads = [lead for lead in leads if search_uid.lower() in str(lead.get('uid', '')).lower()]
            # Qualification filter
            if qualification == 'qualified':
                leads = [lead for lead in leads if lead.get('first_call_date')]
            elif qualification == 'unqualified':
                leads = [lead for lead in leads if not lead.get('first_call_date')]
            # Final status filter
            if final_status:
                leads = [lead for lead in leads if (lead.get('final_status') or '') == final_status]
            # Date filtering
            if date_filter == 'today':
                today_str = datetime.now().strftime('%Y-%m-%d')
                leads = [lead for lead in leads if lead.get('cre_assigned_at') and str(lead.get('cre_assigned_at')).startswith(today_str)]
            elif date_filter == 'range' and start_date and end_date:
                def in_range(ld):
                    dt = ld.get('cre_assigned_at')
                    if not dt:
                        return False
                    try:
                        dt_val = dt[:10]
                        return start_date <= dt_val <= end_date
                    except Exception:
                        return False
                leads = [lead for lead in leads if in_range(lead)]
            # Get all unique sources for this CRE's leads
            sources = sorted(list(set(lead.get('source', 'Unknown') for lead in leads)))
            # Pagination
            total_leads = len(leads)
            total_pages = (total_leads + per_page - 1) // per_page
            start_idx = (page - 1) * per_page
            end_idx = start_idx + per_page
            leads = leads[start_idx:end_idx]
        else:
            if search_uid:
                # Search by UID substring across all leads
                all_leads = safe_get_data('lead_master')
                leads = [lead for lead in all_leads if search_uid.lower() in str(lead.get('uid', '')).lower()]
                sources = sorted(list(set(lead.get('source', 'Unknown') for lead in leads)))
                total_leads = len(leads)
                total_pages = 1
                page = 1
            else:
                total_pages = 1
                page = 1
        # Add formatted CRE TAT for display
        def format_cre_tat(tat):
            try:
                tat = float(tat)
            except (TypeError, ValueError):
                return 'N/A'
            if tat < 60:
                return f"{int(tat)}s"
            elif tat < 3600:
                m = int(tat // 60)
                s = int(tat % 60)
                return f"{m}m {s}s"
            elif tat < 86400:
                h = int(tat // 3600)
                m = int((tat % 3600) // 60)
                s = int(tat % 60)
                return f"{h}h {m}m {s}s"
            else:
                d = int(tat // 86400)
                h = int((tat % 86400) // 3600)
                return f"{d} Days {h}h"
        for lead in leads:
            lead['cre_tat_display'] = format_cre_tat(lead.get('tat'))
        return render_template('manage_leads.html', cres=cres, selected_cre=selected_cre, leads=leads, sources=sources, selected_source=source, qualification=qualification, date_filter=date_filter, start_date=start_date, end_date=end_date, page=page, total_pages=total_pages, total_leads=total_leads, final_status=final_status)
    except Exception as e:
        flash(f'Error loading leads: {str(e)}', 'error')
        return render_template('manage_leads.html', cres=[], selected_cre=None, leads=[], sources=[], selected_source=None, qualification='all', date_filter='all', start_date=None, end_date=None, page=1, total_pages=1, total_leads=0, final_status='')


@app.route('/delete_leads', methods=['POST'])
@require_admin
def delete_leads():
    try:
        delete_type = request.form.get('delete_type')

        if delete_type == 'single':
            uid = request.form.get('uid')
            if not uid:
                return jsonify({'success': False, 'message': 'No UID provided'})

            # Delete from ps_followup_master first (foreign key constraint)
            supabase.table('ps_followup_master').delete().eq('lead_uid', uid).execute()

            # Delete from lead_master
            result = supabase.table('lead_master').delete().eq('uid', uid).execute()

            # Log deletion
            auth_manager.log_audit_event(
                user_id=session.get('user_id'),
                user_type=session.get('user_type'),
                action='LEAD_DELETED',
                resource='lead_master',
                resource_id=uid,
                details={'delete_type': 'single'}
            )

            return jsonify({'success': True, 'message': 'Lead deleted successfully'})

        elif delete_type == 'bulk':
            uids = request.form.getlist('uids')
            if not uids:
                return jsonify({'success': False, 'message': 'No leads selected'})

            # Delete from ps_followup_master first
            for uid in uids:
                supabase.table('ps_followup_master').delete().eq('lead_uid', uid).execute()

            # Delete from lead_master
            for uid in uids:
                supabase.table('lead_master').delete().eq('uid', uid).execute()

            # Log bulk deletion
            auth_manager.log_audit_event(
                user_id=session.get('user_id'),
                user_type=session.get('user_type'),
                action='LEADS_BULK_DELETED',
                resource='lead_master',
                details={'delete_type': 'bulk', 'count': len(uids), 'uids': uids}
            )

            return jsonify({'success': True, 'message': f'{len(uids)} leads deleted successfully'})

        else:
            return jsonify({'success': False, 'message': 'Invalid delete type'})

    except Exception as e:
        print(f"Error deleting leads: {e}")
        return jsonify({'success': False, 'message': f'Error deleting leads: {str(e)}'})


@app.route('/bulk_unassign_leads', methods=['POST'])
@require_admin
def bulk_unassign_leads():
    try:
        uids = request.form.getlist('uids')
        if not uids:
            return jsonify({'success': False, 'message': 'No leads selected'})

        # Update leads to unassigned
        for uid in uids:
            supabase.table('lead_master').update({
                'cre_name': None,
                'assigned': 'No',
                'ps_name': None
            }).eq('uid', uid).execute()

            # Also remove from PS followup if exists
            supabase.table('ps_followup_master').delete().eq('lead_uid', uid).execute()

        # Log bulk unassignment
        auth_manager.log_audit_event(
            user_id=session.get('user_id'),
            user_type=session.get('user_type'),
            action='LEADS_BULK_UNASSIGNED',
            resource='lead_master',
            details={'count': len(uids), 'uids': uids}
        )

        return jsonify({'success': True, 'message': f'{len(uids)} leads unassigned successfully'})

    except Exception as e:
        print(f"Error unassigning leads: {e}")
        return jsonify({'success': False, 'message': f'Error unassigning leads: {str(e)}'})


@app.route('/delete_cre/<int:cre_id>')
@require_admin
def delete_cre(cre_id):
    try:
        print(f"[DEBUG] delete_cre function called with cre_id: {cre_id}")
        
        # Get the CRE details first
        cre_result = supabase.table('cre_users').select('*').eq('id', cre_id).execute()
        print(f"[DEBUG] CRE query result: {cre_result}")
        
        if not cre_result.data:
            print(f"[DEBUG] No CRE found with id: {cre_id}")
            flash('CRE not found', 'error')
            return redirect(url_for('manage_cre'))
        
        cre = cre_result.data[0]
        cre_name = cre.get('name')
        print(f"[DEBUG] Found CRE: {cre_name}")
        
        print(f"[DEBUG] Checking pending leads for CRE: {cre_name}")
        
        # Check if CRE has any pending leads in lead_master (only leads with completed qualifying calls)
        try:
            pending_leads_result = supabase.table('lead_master').select('id').eq('cre_name', cre_name).eq('final_status', 'Pending').not_.is_('first_call_date', 'null').execute()
            pending_count = len(pending_leads_result.data) if pending_leads_result.data else 0
            print(f"[DEBUG] lead_master pending count (qualifying call completed): {pending_count}")
        except Exception as e:
            print(f"[DEBUG] Error checking lead_master: {str(e)}")
            pending_count = 0
        
        # Check if CRE has any pending leads in ps_followup_master
        try:
            ps_pending_result = supabase.table('ps_followup_master').select('id').eq('cre_name', cre_name).eq('final_status', 'Pending').execute()
            ps_pending_count = len(ps_pending_result.data) if ps_pending_result.data else 0
            print(f"[DEBUG] ps_followup_master pending count: {ps_pending_count}")
        except Exception as e:
            print(f"[DEBUG] Error checking ps_followup_master: {str(e)}")
            ps_pending_count = 0
        
        total_pending = pending_count + ps_pending_count
        print(f"[DEBUG] Total pending count: {total_pending}")
        
        if total_pending > 0:
            print(f"[DEBUG] Blocking deletion - {total_pending} pending leads found")
            flash(f'Cannot delete CRE {cre_name}. They have {total_pending} pending leads ({pending_count} in lead_master, {ps_pending_count} in ps_followup). Please transfer or close these leads first.', 'error')
            return redirect(url_for('manage_cre'))
        
        print(f"[DEBUG] No pending leads found, proceeding with deletion for CRE: {cre_name}")
        
        # If no pending leads, proceed with deletion
        # Update leads assigned to this CRE to unassigned
        try:
            supabase.table('lead_master').update({
                'cre_name': None,
                'assigned': 'No'
            }).eq('cre_name', cre_name).execute()
            print(f"[DEBUG] Updated lead_master")
        except Exception as e:
            print(f"[DEBUG] Error updating lead_master: {str(e)}")
        
        # Update ps_followup_master leads
        try:
            supabase.table('ps_followup_master').update({
                'cre_name': None
            }).eq('cre_name', cre_name).execute()
            print(f"[DEBUG] Updated ps_followup_master")
        except Exception as e:
            print(f"[DEBUG] Error updating ps_followup_master: {str(e)}")

        # Delete the CRE user
        try:
            supabase.table('cre_users').delete().eq('id', cre_id).execute()
            print(f"[DEBUG] Deleted CRE user from cre_users table")
        except Exception as e:
            print(f"[DEBUG] Error deleting CRE user: {str(e)}")
            raise e

        # Log CRE deletion
        try:
            auth_manager.log_audit_event(
                user_id=session.get('user_id'),
                user_type=session.get('user_type'),
                action='CRE_DELETED',
                resource='cre_users',
                resource_id=str(cre_id),
                details={'cre_name': cre_name}
            )
            print(f"[DEBUG] Logged audit event")
        except Exception as e:
            print(f"[DEBUG] Error logging audit event: {str(e)}")

        flash(f'CRE {cre_name} has been deleted successfully', 'success')
        print(f"[DEBUG] CRE deletion completed successfully")
        
    except Exception as e:
        print(f"[DEBUG] Error deleting CRE: {str(e)}")
        print(f"[DEBUG] Exception type: {type(e)}")
        import traceback
        print(f"[DEBUG] Full traceback: {traceback.format_exc()}")
        flash('Error deleting CRE', 'error')

    return redirect(url_for('manage_cre'))


@app.route('/delete_ps/<int:ps_id>')
@require_admin
def delete_ps(ps_id):
    try:
        print(f"[DEBUG] delete_ps function called with ps_id: {ps_id}")
        
        # Get the PS details first
        ps_result = supabase.table('ps_users').select('*').eq('id', ps_id).execute()
        print(f"[DEBUG] PS query result: {ps_result}")
        
        if not ps_result.data:
            print(f"[DEBUG] No PS found with id: {ps_id}")
            flash('PS not found', 'error')
            return redirect(url_for('manage_ps'))
        
        ps = ps_result.data[0]
        ps_name = ps.get('name')
        print(f"[DEBUG] Found PS: {ps_name}")
        
        print(f"[DEBUG] Checking pending leads for PS: {ps_name}")
        
        # Check if PS has any pending leads in ps_followup_master
        try:
            ps_pending_result = supabase.table('ps_followup_master').select('id').eq('ps_name', ps_name).eq('final_status', 'Pending').execute()
            ps_pending_count = len(ps_pending_result.data) if ps_pending_result.data else 0
            print(f"[DEBUG] ps_followup_master pending count: {ps_pending_count}")
        except Exception as e:
            print(f"[DEBUG] Error checking ps_followup_master: {str(e)}")
            ps_pending_count = 0
        
        # Check if PS has any pending leads in walkin_table
        try:
            walkin_pending_result = supabase.table('walkin_table').select('id').eq('ps_assigned', ps_name).eq('status', 'pending').execute()
            walkin_pending_count = len(walkin_pending_result.data) if walkin_pending_result.data else 0
            print(f"[DEBUG] walkin_table pending count: {walkin_pending_count}")
        except Exception as e:
            print(f"[DEBUG] Error checking walkin_table: {str(e)}")
            walkin_pending_count = 0
        
        # Check if PS has any pending leads in activity_leads
        try:
            activity_pending_result = supabase.table('activity_leads').select('id').eq('ps_name', ps_name).eq('final_status', 'Pending').execute()
            activity_pending_count = len(activity_pending_result.data) if activity_pending_result.data else 0
            print(f"[DEBUG] activity_leads pending count: {activity_pending_count}")
        except Exception as e:
            print(f"[DEBUG] Error checking activity_leads: {str(e)}")
            activity_pending_count = 0
        
        total_pending = ps_pending_count + walkin_pending_count + activity_pending_count
        print(f"[DEBUG] Total pending count: {total_pending}")
        
        if total_pending > 0:
            print(f"[DEBUG] Blocking deletion - {total_pending} pending leads found")
            flash(f'Cannot delete PS {ps_name}. They have {total_pending} pending leads ({ps_pending_count} in ps_followup, {walkin_pending_count} in walkin, {activity_pending_count} in activity). Please transfer or close these leads first.', 'error')
            return redirect(url_for('manage_ps'))
        
        print(f"[DEBUG] No pending leads found, proceeding with deletion for PS: {ps_name}")
        
        # If no pending leads, proceed with deletion
        # Update leads assigned to this PS to unassigned
        try:
            supabase.table('lead_master').update({
                'ps_name': None
            }).eq('ps_name', ps_name).execute()
            print(f"[DEBUG] Updated lead_master")
        except Exception as e:
            print(f"[DEBUG] Error updating lead_master: {str(e)}")
        
        # Update ps_followup_master leads
        try:
            supabase.table('ps_followup_master').update({
                'ps_name': None
            }).eq('ps_name', ps_name).execute()
            print(f"[DEBUG] Updated ps_followup_master")
        except Exception as e:
            print(f"[DEBUG] Error updating ps_followup_master: {str(e)}")
        
        # Update walkin_table leads
        try:
            supabase.table('walkin_table').update({
                'ps_assigned': None
            }).eq('ps_assigned', ps_name).execute()
            print(f"[DEBUG] Updated walkin_table")
        except Exception as e:
            print(f"[DEBUG] Error updating walkin_table: {str(e)}")
        
        # Update activity_leads
        try:
            supabase.table('activity_leads').update({
                'ps_name': None
            }).eq('ps_name', ps_name).execute()
            print(f"[DEBUG] Updated activity_leads")
        except Exception as e:
            print(f"[DEBUG] Error updating activity_leads: {str(e)}")

        # Delete the PS user
        try:
            supabase.table('ps_users').delete().eq('id', ps_id).execute()
            print(f"[DEBUG] Deleted PS user from ps_users table")
        except Exception as e:
            print(f"[DEBUG] Error deleting PS user: {str(e)}")
            raise e

        # Log PS deletion
        try:
            auth_manager.log_audit_event(
                user_id=session.get('user_id'),
                user_type=session.get('user_type'),
                action='PS_DELETED',
                resource='ps_users',
                resource_id=str(ps_id),
                details={'ps_name': ps_name}
            )
            print(f"[DEBUG] Logged audit event")
        except Exception as e:
            print(f"[DEBUG] Error logging audit event: {str(e)}")

        flash(f'PS {ps_name} has been deleted successfully', 'success')
        print(f"[DEBUG] PS deletion completed successfully")
        
    except Exception as e:
        print(f"[DEBUG] Error deleting PS: {str(e)}")
        print(f"[DEBUG] Exception type: {type(e)}")
        import traceback
        print(f"[DEBUG] Full traceback: {traceback.format_exc()}")
        flash('Error deleting PS', 'error')

    return redirect(url_for('manage_ps'))


@app.route('/edit_cre/<int:cre_id>', methods=['GET', 'POST'])
@require_admin
def edit_cre(cre_id):
    """Edit CRE user details"""
    try:
        if request.method == 'POST':
            email = request.form.get('email')
            phone = request.form.get('phone')
            
            if not email or not phone:
                flash('Email and phone are required', 'error')
                return redirect(url_for('edit_cre', cre_id=cre_id))
            
            # Update the CRE
            supabase.table('cre_users').update({
                'email': email,
                'phone': phone
            }).eq('id', cre_id).execute()
            
            flash('CRE details updated successfully', 'success')
            return redirect(url_for('manage_cre'))
        
        # Get CRE details for editing
        cre_result = supabase.table('cre_users').select('*').eq('id', cre_id).execute()
        if not cre_result.data:
            flash('CRE not found', 'error')
            return redirect(url_for('manage_cre'))
        
        cre = cre_result.data[0]
        return render_template('edit_cre.html', cre=cre)
        
    except Exception as e:
        print(f"Error editing CRE: {str(e)}")
        flash('Error editing CRE', 'error')
        return redirect(url_for('manage_cre'))


@app.route('/edit_ps/<int:ps_id>', methods=['GET', 'POST'])
@require_admin
def edit_ps(ps_id):
    """Edit PS user details"""
    try:
        if request.method == 'POST':
            email = request.form.get('email')
            phone = request.form.get('phone')
            
            if not email or not phone:
                flash('Email and phone are required', 'error')
                return redirect(url_for('edit_ps', ps_id=ps_id))
            
            # Update the PS
            supabase.table('ps_users').update({
                'email': email,
                'phone': phone
            }).eq('id', ps_id).execute()
            
            flash('PS details updated successfully', 'success')
            return redirect(url_for('manage_ps'))
        
        # Get PS details for editing
        ps_result = supabase.table('ps_users').select('*').eq('id', ps_id).execute()
        if not ps_result.data:
            flash('PS not found', 'error')
            return redirect(url_for('manage_ps'))
        
        ps = ps_result.data[0]
        return render_template('edit_ps.html', ps=ps)
        
    except Exception as e:
        print(f"Error editing PS: {str(e)}")
        flash('Error editing PS', 'error')
        return redirect(url_for('manage_ps'))


@app.route('/manage_rec', methods=['GET', 'POST'])
@require_admin
def manage_rec():
    branches = get_system_branches()
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        name = request.form.get('name', '').strip()
        branch = request.form.get('branch', '').strip()
        if not all([username, password, name, branch]):
            flash('All fields are required', 'error')
        else:
            existing = supabase.table('rec_users').select('username').eq('username', username).execute()
            if existing.data:
                flash('Username already exists', 'error')
            else:
                password_hash = generate_password_hash(password)
                rec_data = {
                    'username': username,
                    'password_hash': password_hash,
                    'password': password,  # Store plain password as well
                    'name': name,
                    'branch': branch,
                    'is_active': True
                }
                try:
                    supabase.table('rec_users').insert(rec_data).execute()
                    flash('Receptionist user added successfully', 'success')
                except Exception as e:
                    flash(f'Error adding receptionist: {str(e)}', 'error')
    rec_users = safe_get_data('rec_users')
    return render_template('manage_rec.html', rec_users=rec_users, branches=branches)


@app.route('/delete_rec/<int:rec_id>')
@require_admin
def delete_rec(rec_id):
    try:
        supabase.table('rec_users').delete().eq('id', rec_id).execute()
        flash('Receptionist user deleted successfully', 'success')
    except Exception as e:
        flash(f'Error deleting receptionist: {str(e)}', 'error')
    return redirect(url_for('manage_rec'))


@app.route('/check_duplicate_lead', methods=['POST'])
@require_auth(['admin', 'cre'])
def check_duplicate_lead():
    """
    Check if a lead with the same phone number already exists.
    Returns duplicate information if found.
    """
    try:
        data = request.get_json()
        phone_number = data.get('phone_number', '').strip()
        source = data.get('source', '').strip()
        subsource = data.get('subsource', '').strip()
        
        if not phone_number:
            return jsonify({'success': False, 'message': 'Phone number is required'}), 400
        
        # Normalize phone number (remove all non-digits)
        normalized_phone = ''.join(filter(str.isdigit, phone_number))
        
        # Check in lead_master table
        result = supabase.table('lead_master').select('*').eq('customer_mobile_number', normalized_phone).execute()
        existing_leads = result.data or []
        
        if existing_leads:
            # Found existing lead(s)
            existing_lead = existing_leads[0]  # Get the first one
            
            # Check if this exact source-subsource combination already exists
            exact_match = any(
                lead.get('source') == source and lead.get('sub_source') == subsource 
                for lead in existing_leads
            )
            
            if exact_match:
                # This is a true duplicate - same phone, same source, same subsource
                return jsonify({
                    'success': True,
                    'is_duplicate': True,
                    'existing_lead': existing_lead,
                    'duplicate_type': 'exact_match',
                    'message': 'Lead with this phone number and source-subsource combination already exists'
                })
            else:
                # Phone exists but with different source/subsource
                # Get all existing sources for this phone number
                existing_sources = []
                for lead in existing_leads:
                    if lead.get('source') and lead.get('sub_source'):
                        existing_sources.append({
                            'source': lead.get('source'),
                            'sub_source': lead.get('sub_source')
                        })
                
                return jsonify({
                    'success': True,
                    'is_duplicate': True,
                    'existing_lead': existing_lead,
                    'existing_sources': existing_sources,
                    'duplicate_type': 'new_source',
                    'message': 'Phone number exists with different sources'
                })
        
        # Check in duplicate_leads table
        duplicate_result = supabase.table('duplicate_leads').select('*').eq('customer_mobile_number', normalized_phone).execute()
        duplicate_leads = duplicate_result.data or []
        
        if duplicate_leads:
            # Found in duplicate_leads table
            duplicate_lead = duplicate_leads[0]
            
            # Check if this exact source-subsource combination already exists in any slot
            exact_match = False
            existing_sources = []
            
            # Check all source slots (source1 to source10)
            for i in range(1, 11):
                source_field = f'source{i}'
                sub_source_field = f'sub_source{i}'
                
                if duplicate_lead.get(source_field) and duplicate_lead.get(sub_source_field):
                    existing_sources.append({
                        'source': duplicate_lead.get(source_field),
                        'sub_source': duplicate_lead.get(sub_source_field)
                    })
                    
                    # Check if this slot matches the new source/subsource
                    if (duplicate_lead.get(source_field) == source and 
                        duplicate_lead.get(sub_source_field) == subsource):
                        exact_match = True
            
            if exact_match:
                return jsonify({
                    'success': True,
                    'is_duplicate': True,
                    'existing_lead': duplicate_lead,
                    'duplicate_type': 'exact_match',
                    'message': 'Lead with this phone number and source-subsource combination already exists in duplicates'
                })
            else:
                return jsonify({
                    'success': True,
                    'is_duplicate': True,
                    'existing_lead': duplicate_lead,
                    'existing_sources': existing_sources,
                    'duplicate_type': 'new_source',
                    'message': 'Phone number exists in duplicates with different sources'
                })
        
        # No duplicate found
        return jsonify({
            'success': True,
            'is_duplicate': False,
            'message': 'No duplicate found'
        })
        
    except Exception as e:
        print(f"Error checking duplicate lead: {e}")
        return jsonify({'success': False, 'message': f'Error checking duplicate: {str(e)}'}), 500

@app.route('/check_duplicate_walkin_lead', methods=['POST'])
@require_rec
def check_duplicate_walkin_lead():
    """
    Check if a walkin lead with the same phone number already exists.
    Returns duplicate information if found.
    """
    try:
        data = request.get_json()
        phone_number = data.get('phone_number', '').strip()
        
        if not phone_number:
            return jsonify({'success': False, 'message': 'Phone number is required'}), 400
        
        # Normalize phone number (remove all non-digits)
        normalized_phone = ''.join(filter(str.isdigit, phone_number))
        
        print(f"🔍 Checking for duplicates for phone: {normalized_phone}")
        
        # 1. Check in walkin_table (most important for walkin leads)
        walkin_result = supabase.table('walkin_table').select('*').eq('mobile_number', normalized_phone).execute()
        existing_walkin_leads = walkin_result.data or []
        
        if existing_walkin_leads:
            print(f"📱 Found {len(existing_walkin_leads)} existing walkin leads")
            # Found existing walkin lead(s) - this is an exact duplicate
            existing_lead = existing_walkin_leads[0]
            return jsonify({
                'success': True,
                'is_duplicate': True,
                'existing_lead': existing_lead,
                'duplicate_type': 'exact_match',
                'message': 'Lead with this phone number already exists as Walk-in',
                'source': 'walkin_table'
            })
        
        # 2. Check in activity_leads table
        activity_result = supabase.table('activity_leads').select('*').eq('customer_phone_number', normalized_phone).execute()
        existing_activity_leads = activity_result.data or []
        
        if existing_activity_leads:
            print(f"🎯 Found {len(existing_activity_leads)} existing activity leads")
            # Found existing activity lead(s)
            existing_lead = existing_activity_leads[0]
            existing_sources = []
            for lead in existing_activity_leads:
                existing_sources.append({
                    'source': 'Event',
                    'sub_source': lead.get('location', 'N/A')
                })
            
            return jsonify({
                'success': True,
                'is_duplicate': True,
                'existing_lead': existing_lead,
                'existing_sources': existing_sources,
                'duplicate_type': 'new_source',
                'message': 'Phone number exists with Event source',
                'source': 'activity_leads'
            })
        
        # 3. Check in lead_master table
        lead_master_result = supabase.table('lead_master').select('*').eq('customer_mobile_number', normalized_phone).execute()
        existing_leads = lead_master_result.data or []
        
        if existing_leads:
            print(f"📋 Found {len(existing_leads)} existing leads in lead_master")
            # Found existing lead(s)
            existing_lead = existing_leads[0]
            
            # Check if Walk-in source already exists
            walkin_exists = any(
                lead.get('source') == 'Walk-in' 
                for lead in existing_leads
            )
            
            if walkin_exists:
                # This is a true duplicate - same phone, same source (Walk-in)
                return jsonify({
                    'success': True,
                    'is_duplicate': True,
                    'existing_lead': existing_lead,
                    'duplicate_type': 'exact_match',
                    'message': 'Lead with this phone number already exists as Walk-in',
                    'source': 'lead_master'
                })
            else:
                # Phone exists but with different source
                # Get all existing sources for this phone number
                existing_sources = []
                for lead in existing_leads:
                    if lead.get('source'):
                        existing_sources.append({
                            'source': lead.get('source'),
                            'sub_source': lead.get('sub_source')
                        })
                
                return jsonify({
                    'success': True,
                    'is_duplicate': True,
                    'existing_lead': existing_lead,
                    'existing_sources': existing_sources,
                    'duplicate_type': 'new_source',
                    'message': 'Phone number exists with different sources',
                    'source': 'lead_master'
                })
        
        # 4. Check in duplicate_leads table
        duplicate_result = supabase.table('duplicate_leads').select('*').eq('customer_mobile_number', normalized_phone).execute()
        duplicate_leads = duplicate_result.data or []
        
        if duplicate_leads:
            print(f"🔄 Found {len(duplicate_leads)} existing duplicate leads")
            # Found in duplicate_leads table
            duplicate_lead = duplicate_leads[0]
            
            # Check if Walk-in source already exists in any slot
            walkin_exists = False
            for duplicate_lead in duplicate_leads:
                # Check all source slots (source1 to source10)
                for i in range(1, 11):
                    source_field = f'source{i}'
                    
                    if duplicate_lead.get(source_field) == 'Walk-in':
                        walkin_exists = True
                        break
                if walkin_exists:
                    break
            
            if walkin_exists:
                return jsonify({
                    'success': True,
                    'is_duplicate': True,
                    'existing_lead': duplicate_lead,
                    'duplicate_type': 'exact_match',
                    'message': 'Lead with this phone number already exists as Walk-in in duplicates',
                    'source': 'duplicate_leads'
                })
            else:
                # Phone exists in duplicates but not as Walk-in
                existing_sources = []
                for duplicate_lead in duplicate_leads:
                    for i in range(1, 11):
                        source_field = f'source{i}'
                        sub_source_field = f'sub_source{i}'
                        source_value = duplicate_lead.get(source_field)
                        
                        if source_value:
                            existing_sources.append({
                                'source': source_value,
                                'sub_source': duplicate_lead.get(sub_source_field)
                            })
                
                return jsonify({
                    'success': True,
                    'is_duplicate': True,
                    'existing_lead': duplicate_lead,
                    'existing_sources': existing_sources,
                    'duplicate_type': 'new_source',
                    'message': 'Phone number exists in duplicates with different sources',
                    'source': 'duplicate_leads'
                })
        
        # No duplicate found
        print(f"✅ No duplicates found for phone: {normalized_phone}")
        return jsonify({
            'success': True,
            'is_duplicate': False,
            'message': 'No duplicate found'
        })
        
    except Exception as e:
        print(f"❌ Error checking duplicate walkin lead: {e}")
        return jsonify({'success': False, 'message': f'Error checking duplicate: {str(e)}'}), 500

@app.route('/check_duplicate_event_lead', methods=['POST'])
@require_ps
def check_duplicate_event_lead():
    """
    Check if an event lead with the same phone number already exists.
    Returns duplicate information if found.
    """
    try:
        data = request.get_json()
        phone_number = data.get('phone_number', '').strip()
        
        if not phone_number:
            return jsonify({'success': False, 'message': 'Phone number is required'}), 400
        
        # Normalize phone number (remove all non-digits)
        normalized_phone = ''.join(filter(str.isdigit, phone_number))
        
        print(f"🔍 Checking for duplicates for event lead phone: {normalized_phone}")
        
        # 1. Check in activity_leads table FIRST (most important for event leads)
        activity_result = supabase.table('activity_leads').select('*').eq('customer_phone_number', normalized_phone).execute()
        existing_activity_leads = activity_result.data or []
        
        if existing_activity_leads:
            print(f"🎯 Found {len(existing_activity_leads)} existing activity leads")
            # Found existing activity lead(s) - this is an exact duplicate
            existing_lead = existing_activity_leads[0]
            return jsonify({
                'success': True,
                'is_duplicate': True,
                'existing_lead': existing_lead,
                'duplicate_type': 'exact_match',
                'message': 'Lead with this phone number already exists as Event lead',
                'source': 'activity_leads'
            })
        
        # 2. Check in walkin_table
        walkin_result = supabase.table('walkin_table').select('*').eq('mobile_number', normalized_phone).execute()
        existing_walkin_leads = walkin_result.data or []
        
        if existing_walkin_leads:
            print(f"📱 Found {len(existing_walkin_leads)} existing walkin leads")
            # Found existing walkin lead(s)
            existing_lead = existing_walkin_leads[0]
            existing_sources = []
            for lead in existing_walkin_leads:
                existing_sources.append({
                    'source': 'Walk-in',
                    'sub_source': lead.get('branch', 'N/A')
                })
            
            return jsonify({
                'success': True,
                'is_duplicate': True,
                'existing_lead': existing_lead,
                'existing_sources': existing_sources,
                'duplicate_type': 'new_source',
                'message': 'Phone number exists with Walk-in source',
                'source': 'walkin_table'
            })
        
        # 3. Check in lead_master table
        lead_master_result = supabase.table('lead_master').select('*').eq('customer_mobile_number', normalized_phone).execute()
        existing_leads = lead_master_result.data or []
        
        if existing_leads:
            print(f"📋 Found {len(existing_leads)} existing leads in lead_master")
            # Found existing lead(s)
            existing_lead = existing_leads[0]
            
            # Check if Event source already exists
            event_exists = any(
                lead.get('source') == 'Event' 
                for lead in existing_leads
            )
            
            if event_exists:
                # This is a true duplicate - same phone, same source (Event)
                return jsonify({
                    'success': True,
                    'is_duplicate': True,
                    'existing_lead': existing_lead,
                    'duplicate_type': 'exact_match',
                    'message': 'Lead with this phone number already exists as Event',
                    'source': 'lead_master'
                })
            else:
                # Phone exists but with different source
                # Get all existing sources for this phone number
                existing_sources = []
                for lead in existing_leads:
                    if lead.get('source'):
                        existing_sources.append({
                            'source': lead.get('source'),
                            'sub_source': lead.get('sub_source')
                        })
                
                return jsonify({
                    'success': True,
                    'is_duplicate': True,
                    'existing_lead': existing_lead,
                    'existing_sources': existing_sources,
                    'duplicate_type': 'new_source',
                    'message': 'Phone number exists with different sources',
                    'source': 'lead_master'
                })
        
        # 4. Check in duplicate_leads table
        duplicate_result = supabase.table('duplicate_leads').select('*').eq('customer_mobile_number', normalized_phone).execute()
        duplicate_leads = duplicate_result.data or []
        
        if duplicate_leads:
            print(f"🔄 Found {len(duplicate_leads)} existing duplicate leads")
            # Found in duplicate_leads table
            duplicate_lead = duplicate_leads[0]
            
            # Check if Event source already exists in any slot
            event_exists = False
            for duplicate_lead in duplicate_leads:
                # Check all source slots (source1 to source10)
                for i in range(1, 11):
                    source_field = f'source{i}'
                    
                    if duplicate_lead.get(source_field) == 'Event':
                        event_exists = True
                        break
                if event_exists:
                    break
            
            if event_exists:
                return jsonify({
                    'success': True,
                    'is_duplicate': True,
                    'existing_lead': duplicate_lead,
                    'duplicate_type': 'exact_match',
                    'message': 'Lead with this phone number already exists as Event in duplicates',
                    'source': 'duplicate_leads'
                })
            else:
                # Phone exists in duplicates but not as Event
                existing_sources = []
                for duplicate_lead in duplicate_leads:
                    for i in range(1, 11):
                        source_field = f'source{i}'
                        sub_source_field = f'sub_source{i}'
                        source_value = duplicate_lead.get(source_field)
                        
                        if source_value:
                            existing_sources.append({
                                'source': source_value,
                                'sub_source': duplicate_lead.get(sub_source_field)
                            })
                
                return jsonify({
                    'success': True,
                    'is_duplicate': True,
                    'existing_lead': duplicate_lead,
                    'existing_sources': existing_sources,
                    'duplicate_type': 'new_source',
                    'message': 'Phone number exists in duplicates with different sources',
                    'source': 'duplicate_leads'
                })
        
        # No duplicate found
        print(f"✅ No duplicates found for event lead phone: {normalized_phone}")
        return jsonify({
            'success': True,
            'is_duplicate': False,
            'message': 'No duplicate found'
        })
        
    except Exception as e:
        print(f"❌ Error checking duplicate event lead: {e}")
        return jsonify({'success': False, 'message': f'Error checking duplicate: {str(e)}'}), 500

@app.route('/add_lead', methods=['GET', 'POST'])
@require_cre
def add_lead():
    from datetime import datetime, date
    branches = get_system_branches()
    ps_users = safe_get_data('ps_users')
    if request.method == 'POST':
        customer_name = request.form.get('customer_name', '').strip()
        customer_mobile_number = request.form.get('customer_mobile_number', '').strip()
        source = request.form.get('source', '').strip()
        subsource = request.form.get('subsource', '').strip()
        lead_status = request.form.get('lead_status', '').strip()
        lead_category = request.form.get('lead_category', '').strip()
        model_interested = request.form.get('model_interested', '').strip()
        branch = request.form.get('branch', '').strip()
        ps_name = request.form.get('ps_name', '').strip()
        final_status = request.form.get('final_status', 'Pending').strip()
        follow_up_date = request.form.get('follow_up_date', '').strip()
        remark = request.form.get('remark', '').strip()
        is_duplicate_new_source = request.form.get('is_duplicate_new_source', '').strip()
        date_now = datetime.now().strftime('%Y-%m-%d')
        
        # Validation
        if not customer_name or not customer_mobile_number or not source or not subsource:
            flash('Please fill all required fields', 'error')
            return render_template('add_lead.html', branches=branches, ps_users=ps_users)
        
        # Validate follow_up_date is required when final_status is Pending
        if final_status == 'Pending' and not follow_up_date:
            flash('Follow-up date is required when final status is Pending', 'error')
            return render_template('add_lead.html', branches=branches, ps_users=ps_users)
        
        # Normalize phone number
        normalized_phone = ''.join(filter(str.isdigit, customer_mobile_number))
        
        # Check for duplicates if not already confirmed as new source
        if not is_duplicate_new_source:
            try:
                # Check in lead_master
                result = supabase.table('lead_master').select('*').eq('customer_mobile_number', normalized_phone).execute()
                existing_leads = result.data or []
                
                if existing_leads:
                    # Check for exact source-subsource match
                    exact_match = any(
                        lead.get('source') == source and lead.get('sub_source') == subsource 
                        for lead in existing_leads
                    )
                    
                    if exact_match:
                        flash('Lead with this phone number and source-subsource combination already exists!', 'error')
                        return render_template('add_lead.html', branches=branches, ps_users=ps_users)
                
                # Check in duplicate_leads
                duplicate_result = supabase.table('duplicate_leads').select('*').eq('customer_mobile_number', normalized_phone).execute()
                duplicate_leads = duplicate_result.data or []
                
                if duplicate_leads:
                    # Check if this exact source-subsource combination already exists in any slot
                    exact_match = False
                    for duplicate_lead in duplicate_leads:
                        # Check all source slots (source1 to source10)
                        for i in range(1, 11):
                            source_field = f'source{i}'
                            sub_source_field = f'sub_source{i}'
                            
                            if (duplicate_lead.get(source_field) == source and 
                                duplicate_lead.get(sub_source_field) == subsource):
                                exact_match = True
                                break
                        if exact_match:
                            break
                    
                    if exact_match:
                        flash('Lead with this phone number and source-subsource combination already exists in duplicates!', 'error')
                        return render_template('add_lead.html', branches=branches, ps_users=ps_users)
                        
            except Exception as e:
                print(f"Error checking duplicates: {e}")
                flash('Error checking for duplicates. Please try again.', 'error')
                return render_template('add_lead.html', branches=branches, ps_users=ps_users)
        
        # UID: Source initial (uppercase) + '-' + first 5 letters of name (no spaces, uppercase) + last 5 digits of phone
        src_initial = source[0].upper() if source else 'X'
        name_part = ''.join(customer_name.split()).upper()[:5]
        phone_part = normalized_phone[-5:] if len(normalized_phone) >= 5 else normalized_phone
        uid = f"{src_initial}-{name_part}{phone_part}"
        
        # CRE name from session
        cre_name = session.get('cre_name')
        
        # Prepare lead data
        lead_data = {
            'uid': uid,
            'date': date_now,
            'customer_name': customer_name,
            'customer_mobile_number': normalized_phone,
            'source': source,
            'sub_source': subsource,
            'lead_status': lead_status,
            'lead_category': lead_category,
            'model_interested': model_interested,
            'branch': branch,
            'ps_name': ps_name if ps_name else None,
            'final_status': final_status,
            'follow_up_date': follow_up_date if follow_up_date else None,
            'assigned': 'Yes' if cre_name else 'No',  # Set to Yes if CRE is adding the lead
            'cre_assigned_at': datetime.now().isoformat() if cre_name else None,
            'ps_assigned_at': datetime.now().isoformat() if ps_name else None,
            'created_at': datetime.now().isoformat(),
            'updated_at': datetime.now().isoformat(),
            'first_remark': remark,
            'cre_name': cre_name,
            'first_call_date': date_now
        }
        try:
            # If this is a duplicate with new source, add to duplicate_leads table
            if is_duplicate_new_source:
                # Check if there's already a duplicate record
                existing_duplicate = supabase.table('duplicate_leads').select('*').eq('customer_mobile_number', normalized_phone).execute()
                
                if existing_duplicate.data:
                    # Add to existing duplicate record
                    duplicate_record = existing_duplicate.data[0]
                    # Find next available slot
                    next_slot = None
                    for i in range(1, 11):
                        source_field = f'source{i}'
                        if not duplicate_record.get(source_field):
                            next_slot = i
                            break
                    
                    if next_slot:
                        # Update the existing duplicate record
                        update_data = {
                            f'source{next_slot}': source,
                            f'sub_source{next_slot}': subsource,
                            f'date{next_slot}': date_now,
                            'duplicate_count': duplicate_record.get('duplicate_count', 0) + 1,
                            'updated_at': datetime.now().isoformat()
                        }
                        supabase.table('duplicate_leads').update(update_data).eq('id', duplicate_record['id']).execute()
                        flash(f'Lead added to existing duplicate record with new source: {source} - {subsource}', 'success')
                    else:
                        flash('Error: Duplicate record is full (max 10 sources reached)', 'error')
                else:
                    # Create new duplicate record
                    original_lead = supabase.table('lead_master').select('*').eq('customer_mobile_number', normalized_phone).execute()
                    if original_lead.data:
                        original = original_lead.data[0]
                        # Create duplicate record with proper structure
                        duplicate_data = {
                            'uid': uid,
                            'customer_mobile_number': normalized_phone,
                            'customer_name': customer_name,
                            'original_lead_id': original['id'],
                            'source1': original['source'],
                            'sub_source1': original.get('sub_source'),
                            'date1': original['date'],
                            'source2': source,
                            'sub_source2': subsource,
                            'date2': date_now,
                            'duplicate_count': 2,
                            'created_at': datetime.now().isoformat(),
                            'updated_at': datetime.now().isoformat()
                        }
                        supabase.table('duplicate_leads').insert(duplicate_data).execute()
                        flash(f'Lead added to duplicates with new source: {source} - {subsource}', 'success')
                    else:
                        flash('Error: Original lead not found for duplicate creation', 'error')
            else:
                supabase.table('lead_master').insert(lead_data).execute()
                
                # Track the initial call attempt for fresh leads
                if lead_status:
                    track_cre_call_attempt(
                        uid=uid,
                        cre_name=cre_name,
                        call_no='first',
                        lead_status=lead_status,
                        call_was_recorded=True,  # Fresh leads always have first_call_date recorded
                        follow_up_date=follow_up_date if follow_up_date else None,
                        remarks=remark if remark else None
                    )
                
                # Create PS followup if PS is assigned during lead creation
                if ps_name:
                    ps_user = next((ps for ps in ps_users if ps['name'] == ps_name), None)
                    if ps_user:
                        create_or_update_ps_followup(lead_data, ps_name, ps_user['branch'])
                        
                        # Send email notification to PS
                        try:
                            socketio.start_background_task(send_email_to_ps, ps_user['email'], ps_user['name'], lead_data, cre_name)
                            flash(f'Lead added successfully and assigned to {ps_name}! Email notification sent.', 'success')
                        except Exception as e:
                            print(f"Error sending email: {e}")
                            flash(f'Lead added successfully and assigned to {ps_name}! (Email notification failed)', 'warning')
                    else:
                        flash('Lead added successfully! (PS assignment failed)', 'warning')
                else:
                    flash('Lead added successfully!', 'success')
            
            return redirect(url_for('cre_dashboard'))
        except Exception as e:
            flash(f'Error adding lead: {str(e)}', 'error')
            return render_template('add_lead.html', branches=branches, ps_users=ps_users)
    return render_template('add_lead.html', branches=branches, ps_users=ps_users)

# Add import for optimized operations
from optimized_lead_operations import create_optimized_operations

@app.route('/add_lead_optimized', methods=['POST'])
@require_cre
def add_lead_optimized():
    """
    Optimized lead creation endpoint with improved performance
    """
    try:
        from datetime import datetime
        
        # Get form data
        customer_name = request.form.get('customer_name', '').strip()
        customer_mobile_number = request.form.get('customer_mobile_number', '').strip()
        source = request.form.get('source', '').strip()
        subsource = request.form.get('subsource', '').strip()
        lead_status = request.form.get('lead_status', '').strip()
        lead_category = request.form.get('lead_category', '').strip()
        model_interested = request.form.get('model_interested', '').strip()
        branch = request.form.get('branch', '').strip()
        ps_name = request.form.get('ps_name', '').strip()
        final_status = request.form.get('final_status', 'Pending').strip()
        follow_up_date = request.form.get('follow_up_date', '').strip()
        remark = request.form.get('remark', '').strip()
        date_now = datetime.now().strftime('%Y-%m-%d')
        
        # Validation
        if not customer_name or not customer_mobile_number or not source or not subsource:
            return jsonify({
                'success': False,
                'message': 'Please fill all required fields'
            })
        
        # Validate follow_up_date is required when final_status is Pending
        if final_status == 'Pending' and not follow_up_date:
            return jsonify({
                'success': False,
                'message': 'Follow-up date is required when final status is Pending'
            })
        
        # Normalize phone number
        normalized_phone = ''.join(filter(str.isdigit, customer_mobile_number))
        
        # CRE name from session
        cre_name = session.get('cre_name')
        
        # Prepare lead data
        lead_data = {
            'date': date_now,
            'customer_name': customer_name,
            'customer_mobile_number': normalized_phone,
            'source': source,
            'sub_source': subsource,
            'lead_status': lead_status,
            'lead_category': lead_category,
            'model_interested': model_interested,
            'branch': branch,
            'ps_name': ps_name if ps_name else None,
            'final_status': final_status,
            'follow_up_date': follow_up_date if follow_up_date else None,
            'assigned': 'Yes' if cre_name else 'No',
            'cre_assigned_at': datetime.now().isoformat() if cre_name else None,
            'ps_assigned_at': datetime.now().isoformat() if ps_name else None,
            'created_at': datetime.now().isoformat(),
            'updated_at': datetime.now().isoformat(),
            'first_remark': remark,
            'cre_name': cre_name,
            'first_call_date': date_now
        }
        
        # Get PS branch if PS is assigned
        ps_branch = None
        if ps_name:
            ps_users = safe_get_data('ps_users')
            ps_user = next((ps for ps in ps_users if ps['name'] == ps_name), None)
            if ps_user:
                ps_branch = ps_user['branch']
        
        # Use optimized operations
        optimized_ops = create_optimized_operations(supabase)
        result = optimized_ops.create_lead_optimized(lead_data, cre_name, ps_name, ps_branch)
        
        if result['success']:
            return jsonify({
                'success': True,
                'message': result['message'],
                'uid': result['uid'],
                'execution_time': f"{result['execution_time']:.3f}s"
            })
        else:
            return jsonify({
                'success': False,
                'message': result['message'],
                'execution_time': f"{result['execution_time']:.3f}s"
            })
            
    except Exception as e:
        print(f"Error in optimized lead creation: {e}")
        return jsonify({
            'success': False,
            'message': f'Error creating lead: {str(e)}'
        })

@app.route('/add_lead_with_cre', methods=['POST'])
@require_admin  # <-- Change to @require_cre if you want CREs to use it, or create a custom decorator for both
def add_lead_with_cre():
    """
    Add a new lead with minimal required columns.
    - Checks for duplicate by last 10 digits of phone number.
    - If duplicate, returns UID of existing lead.
    - Only fills: uid, customer_name, customer_mobile_number, source, date, assigned.
    - Source is always 'Google(Web)', UID uses 'G' as the source character.
    """
    try:
        # Check if this is an AJAX request
        is_ajax = request.headers.get('Content-Type') == 'application/json'
        
        if is_ajax:
            # Handle JSON request
            data = request.get_json()
            customer_name = data.get('customer_name', '').strip()
            customer_mobile_number = data.get('customer_mobile_number', '').strip()
            source = data.get('source', 'GOOGLE').strip()
            subsource = data.get('subsource', '').strip()
            assigned_cre_id = data.get('assigned_cre')
        else:
            # Handle form request
            customer_name = request.form.get('customer_name', '').strip()
            customer_mobile_number = request.form.get('customer_mobile_number', '').strip()
            source = request.form.get('source', 'GOOGLE').strip()
            subsource = request.form.get('subsource', '').strip()
            assigned_cre_id = request.form.get('assigned_cre')
        
        assigned = "Yes"
        date_now = datetime.now().strftime('%Y-%m-%d')

        # Validate required fields
        if not customer_name or not customer_mobile_number or not source or not subsource:
            if is_ajax:
                return jsonify({
                    'success': False,
                    'message': 'Customer name, mobile number, source, and subsource are required'
                })
            else:
                flash('Customer name, mobile number, source, and subsource are required', 'error')
                return redirect('/assign_leads')

        # Normalize phone number to last 10 digits
        mobile_digits = ''.join(filter(str.isdigit, customer_mobile_number))[-10:]
        
        if len(mobile_digits) != 10:
            if is_ajax:
                return jsonify({
                    'success': False,
                    'message': 'Invalid mobile number. Please provide a 10-digit number.'
                })
            else:
                flash('Invalid mobile number. Please provide a 10-digit number.', 'error')
                return redirect('/assign_leads')

        # Check if this is a duplicate with new source from form
        if is_ajax:
            is_duplicate_new_source = data.get('is_duplicate_new_source', '').strip() == 'true'
        else:
            is_duplicate_new_source = request.form.get('is_duplicate_new_source', '').strip() == 'true'
        
        # Check for duplicate by phone number and source-subsource combination
        existing_leads = supabase.table('lead_master').select('*').eq('customer_mobile_number', mobile_digits).execute()
        duplicate_leads = supabase.table('duplicate_leads').select('*').eq('customer_mobile_number', mobile_digits).execute()
        
        original_lead = None
        
        # Only check for duplicates if not already confirmed as duplicate with new source
        if not is_duplicate_new_source:
            # Check in lead_master table
            if existing_leads.data:
                original_lead = existing_leads.data[0]
                # Check if this exact source-subsource combination already exists
                if original_lead.get('source') == source and original_lead.get('sub_source') == subsource:
                    if is_ajax:
                        return jsonify({
                            'success': False,
                            'message': f'Lead with this phone number and source-subsource combination already exists. UID: {original_lead["uid"]}',
                            'uid': original_lead["uid"]
                        })
                    else:
                        flash(f'Lead with this phone number and source-subsource combination already exists. UID: {original_lead["uid"]}', 'error')
                        return redirect('/assign_leads')
                else:
                    # Phone exists but with different source/subsource - this is a duplicate with new source
                    is_duplicate_new_source = True
            
            # Check in duplicate_leads table
            if duplicate_leads.data:
                duplicate_lead = duplicate_leads.data[0]
                # Check if this exact source-subsource combination already exists in any slot
                exact_match = False
                for i in range(1, 11):
                    source_field = f'source{i}'
                    sub_source_field = f'sub_source{i}'
                    
                    if (duplicate_lead.get(source_field) == source and 
                        duplicate_lead.get(sub_source_field) == subsource):
                        exact_match = True
                        break
                
                if exact_match:
                    if is_ajax:
                        return jsonify({
                            'success': False,
                            'message': f'Lead with this phone number and source-subsource combination already exists in duplicates. UID: {duplicate_lead["uid"]}',
                            'uid': duplicate_lead["uid"]
                        })
                    else:
                        flash(f'Lead with this phone number and source-subsource combination already exists in duplicates. UID: {duplicate_lead["uid"]}', 'error')
                        return redirect('/assign_leads')
                else:
                    # Phone exists in duplicates but with different source/subsource
                    is_duplicate_new_source = True
        else:
            # If is_duplicate_new_source is true, we need to get the original lead
            if existing_leads.data:
                original_lead = existing_leads.data[0]

        # Generate UID using the correct function based on source
        # Map source to UID source character
        source_mapping = {
            'GOOGLE': 'Google',
            'META': 'Meta',
            'BTL': 'BTL',
            'OEM': 'OEM'
        }
        uid_source = source_mapping.get(source, 'Google')
        
        sequence = 1
        uid = generate_uid(uid_source, mobile_digits, sequence)
        # Ensure UID is unique
        while supabase.table('lead_master').select('uid').eq('uid', uid).execute().data:
            sequence += 1
            uid = generate_uid(uid_source, mobile_digits, sequence)

        # Get assigned CRE ID from form
        if is_ajax:
            assigned_cre_id = data.get('assigned_cre')
        else:
            assigned_cre_id = request.form.get('assigned_cre')
        print(f"🔍 Raw assigned_cre_id from form: '{assigned_cre_id}' (type: {type(assigned_cre_id)})")
        
        cre_name = None
        if assigned_cre_id and assigned_cre_id.strip():
            try:
                # Convert to integer if it's a string
                cre_id = int(assigned_cre_id) if isinstance(assigned_cre_id, str) else assigned_cre_id
                print(f"🔍 Looking up CRE with ID: {cre_id}")
                
                cre_data = supabase.table('cre_users').select('name').eq('id', cre_id).execute()
                if cre_data.data:
                    cre_name = cre_data.data[0]['name']
                    print(f"✅ Found CRE: {cre_name} for ID: {cre_id}")
                else:
                    print(f"❌ No CRE found for ID: {cre_id}")
                    print(f"Available CRE IDs: {[cre['id'] for cre in supabase.table('cre_users').select('id,name').execute().data]}")
            except Exception as e:
                print(f"❌ Error fetching CRE data: {e}")
                import traceback
                traceback.print_exc()
        else:
            print("❌ No assigned_cre_id provided in form or it's empty")
            print(f"All form fields: {dict(request.form)}")

        # Prepare lead data (only required columns)
        # Set assigned based on whether CRE is assigned
        assigned_status = "Yes" if cre_name else "No"
        
        lead_data = {
            'uid': uid,
            'customer_name': customer_name,
            'customer_mobile_number': mobile_digits,
            'source': source,
            'sub_source': subsource,
            'date': date_now,   
            'assigned': assigned_status,
            'final_status': 'Pending',
            'cre_name': cre_name,
            'lead_status': 'Pending',
            'lead_category': None,  # Keep as null for assign leads page
            'cre_assigned_at': datetime.now().isoformat() if cre_name else None,
            'ps_assigned_at': None,  # Will be set when PS is assigned
            'created_at': datetime.now().isoformat(),
            'updated_at': datetime.now().isoformat(),
            'ps_assigned_at': None  # Added this line to ensure ps_assigned_at is set
        }
        print('=== DEBUG INFO ===')
        print('Form data:', dict(request.form))
        print('Assigned CRE ID:', assigned_cre_id)
        print('CRE name fetched:', cre_name)
        print('Is duplicate new source:', is_duplicate_new_source)
        print('Lead data to insert:', lead_data)
        print('==================')

        # Insert lead based on whether it's a duplicate with new source
        if is_duplicate_new_source:
            print("=== DUPLICATE HANDLING ===")
            print(f"Original lead: {original_lead}")
            print(f"Duplicate leads: {duplicate_leads.data}")
            
            if original_lead:
                # Create new duplicate record
                duplicate_data = {
                    'uid': uid,
                    'customer_mobile_number': mobile_digits,
                    'customer_name': customer_name,
                    'original_lead_id': original_lead['id'],
                    'source1': original_lead['source'],
                    'sub_source1': original_lead.get('sub_source'),
                    'date1': original_lead['date'],
                    'source2': source,
                    'sub_source2': subsource,
                    'date2': date_now,
                    'duplicate_count': 2,
                    'created_at': datetime.now().isoformat(),
                    'updated_at': datetime.now().isoformat()
                }
                print(f"Creating duplicate record: {duplicate_data}")
                result = supabase.table('duplicate_leads').insert(duplicate_data).execute()
                if result.data:
                    print("Duplicate record created successfully")
                    if is_ajax:
                        return jsonify({'success': True, 'message': 'Lead added to duplicates with new source', 'uid': uid})
                    else:
                        flash('Lead added to duplicates with new source', 'success')
                        return redirect('/assign_leads')
                else:
                    print("Failed to create duplicate record")
                    if is_ajax:
                        return jsonify({'success': False, 'message': 'Failed to add duplicate lead'})
                    else:
                        flash('Failed to add duplicate lead', 'error')
                        return redirect('/assign_leads')
            elif duplicate_leads.data:
                # Add to existing duplicate record
                duplicate_record = duplicate_leads.data[0]
                # Find next available slot
                next_slot = None
                for i in range(1, 11):
                    source_field = f'source{i}'
                    if not duplicate_record.get(source_field):
                        next_slot = i
                        break
                
                if next_slot:
                    # Update the existing duplicate record
                    update_data = {
                        f'source{next_slot}': source,
                        f'sub_source{next_slot}': subsource,
                        f'date{next_slot}': date_now,
                        'duplicate_count': duplicate_record.get('duplicate_count', 0) + 1,
                        'updated_at': datetime.now().isoformat()
                    }
                    result = supabase.table('duplicate_leads').update(update_data).eq('id', duplicate_record['id']).execute()
                    if result.data:
                        if is_ajax:
                            return jsonify({'success': True, 'message': 'Lead added to existing duplicate record', 'uid': uid})
                        else:
                            flash('Lead added to existing duplicate record', 'success')
                            return redirect('/assign_leads')
                    else:
                        if is_ajax:
                            return jsonify({'success': False, 'message': 'Failed to update duplicate record'})
                        else:
                            flash('Failed to update duplicate record', 'error')
                            return redirect('/assign_leads')
                else:
                    if is_ajax:
                        return jsonify({'success': False, 'message': 'Duplicate record is full (max 10 sources reached)'})
                    else:
                        flash('Duplicate record is full (max 10 sources reached)', 'error')
                        return redirect('/assign_leads')
            else:
                # Phone exists in duplicate_leads but not in lead_master - this shouldn't happen
                if is_ajax:
                    return jsonify({'success': False, 'message': 'Error: Original lead not found for duplicate creation'})
                else:
                    flash('Error: Original lead not found for duplicate creation', 'error')
                    return redirect('/assign_leads')
        else:
            # Insert as new lead
            print("=== FRESH LEAD INSERTION ===")
            print(f"Inserting fresh lead: {lead_data}")
            result = supabase.table('lead_master').insert(lead_data).execute()
            if result.data:
                print("Fresh lead inserted successfully")
                if is_ajax:
                    return jsonify({'success': True, 'message': 'Lead added successfully', 'uid': uid})
                else:
                    flash('Lead added successfully', 'success')
                    return redirect('/assign_leads')
            else:
                print("Failed to insert fresh lead")
                if is_ajax:
                    return jsonify({'success': False, 'message': 'Failed to add lead'})
                else:
                    flash('Failed to add lead', 'error')
                    return redirect('/assign_leads')

    except Exception as e:
        print(f"Error adding lead with CRE: {e}")
        import traceback
        traceback.print_exc()
        if is_ajax:
            return jsonify({'success': False, 'message': f'Error adding lead: {str(e)}'})
        else:
            flash(f'Error adding lead: {str(e)}', 'error')
            return redirect('/assign_leads')
@app.route('/cre_dashboard')
@require_cre
def cre_dashboard():
    import time
    from datetime import datetime, date
    start_time = time.time()
    cre_name = session.get('cre_name')
    
    # Get status parameter for Won/Lost toggle
    status = request.args.get('status', 'lost')
    # Get tab and sub_tab parameters for redirection
    tab = request.args.get('tab', '')
    sub_tab = request.args.get('sub_tab', '')
    
    today = date.today()
    today_str = today.isoformat()

    # --- AUTO-INCREMENT LOGIC FOR UNATTENDED LEADS (CRE) ---
    # 1. Regular leads (lead_master)
    all_leads = safe_get_data('lead_master', {'cre_name': cre_name})
    for lead in all_leads:
        follow_up_date = lead.get('follow_up_date')
        final_status = lead.get('final_status')
        if follow_up_date and str(follow_up_date) < today_str and final_status not in ['Won', 'Lost']:
            supabase.table('lead_master').update({'follow_up_date': today_str}).eq('uid', lead.get('uid')).execute()



    # 3. Event leads (activity_leads)
    event_event_leads = safe_get_data('activity_leads', {'cre_assigned': cre_name})
    for lead in event_event_leads:
        cre_followup_date = lead.get('cre_followup_date')
        final_status = lead.get('final_status')
        if cre_followup_date and str(cre_followup_date)[:10] < today_str and final_status not in ['Won', 'Lost']:
            supabase.table('activity_leads').update({'cre_followup_date': today_str}).eq('activity_uid', lead.get('activity_uid')).execute()

    # --- Date Filter Logic ---
    filter_type = request.args.get('filter_type', 'all')
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')

    all_leads = safe_get_data('lead_master', {'cre_name': cre_name})
    
    # Apply date filtering using the 'date' column
    if filter_type == 'today':
        today_str = date.today().isoformat()
        all_leads = [lead for lead in all_leads if lead.get('date') == today_str]
    elif filter_type == 'range' and start_date_str and end_date_str:
        # Filter leads where lead['date'] falls within the range
        all_leads = [
            lead for lead in all_leads 
            if lead.get('date') and start_date_str <= lead['date'] <= end_date_str
        ]
    # For 'all' filter type, we keep all leads without date filtering

    print(f"Date filter applied - Type: {filter_type}, Leads count: {len(all_leads)}")
    if filter_type == 'range':
        print(f"Date range: {start_date_str} to {end_date_str}")

    # Fetch event leads assigned to this CRE
    event_event_leads = safe_get_data('activity_leads', {'cre_assigned': cre_name})

    print("Fetched leads for CRE:", cre_name, "Count:", len(all_leads))
    print("Status parameter from URL:", status)

    # Initialize buckets for leads for mutual exclusivity
    untouched_leads = []
    called_leads = []
    follow_up_leads = []  # New list for "Call me Back" leads
    attended_leads = []
    assigned_to_ps = []
    won_leads = []
    lost_leads = []

    non_contact_statuses = ['RNR', 'Busy on another Call', 'Call me Back', 'Call Disconnected', 'Call not Connected']

    for lead in all_leads:
        lead_status = (lead.get('lead_status') or '').strip()
        final_status = lead.get('final_status')
        has_first_call = lead.get('first_call_date') is not None

        if final_status == 'Won':
            won_leads.append(lead)
            continue
        if final_status == 'Lost':
            lost_leads.append(lead)
            continue

        # PS Assigned: any lead with a PS assigned
        if lead.get('ps_name'):
            assigned_to_ps.append(lead)

        # Pending Leads: only leads with final_status == 'Pending' AND qualifying call completed (first_call_date IS NOT NULL)
        if final_status == 'Pending' and has_first_call:
            attended_leads.append(lead)

        # Untouched Leads (Fresh leads that are still pending)
        if not has_first_call and lead_status == 'Pending':
            untouched_leads.append(lead)
            continue

        # Called Fresh Leads: Non-contact status on FIRST update only
        if lead_status in non_contact_statuses and not has_first_call:
            # Separate "Call me Back" leads into follow_up_leads
            if lead_status == 'Call me Back':
                follow_up_leads.append(lead)
            else:
                called_leads.append(lead)
            continue

    print(f"Won leads count: {len(won_leads)}")
    print(f"Lost leads count: {len(lost_leads)}")
    
    untouched_count = len(untouched_leads)
    called_count = len(called_leads)
    follow_up_count = len(follow_up_leads)
    total_fresh_leads = untouched_count + called_count + follow_up_count

    fresh_leads_sorted = sorted(
        untouched_leads + called_leads + follow_up_leads,
        key=lambda l: l.get('date') or '',  # Sort by 'date' column
        reverse=True
    )

    # Get today's followups
    today = date.today()
    today_str = today.isoformat()
    # Get all leads with follow-up date <= today, final_status = 'Pending', qualifying call completed (first_call_date IS NOT NULL), 
    # AND NOT qualifying call leads (must have follow-up call data beyond first call)
    todays_followups = []
    for lead in all_leads:
        follow_up_date = lead.get('follow_up_date')
        final_status = lead.get('final_status')
        has_first_call = lead.get('first_call_date') is not None
        
        # Check if this is NOT just a qualifying call (must have follow-up call data)
        has_followup_calls = any([
            lead.get('second_call_date'),
            lead.get('third_call_date'),
            lead.get('fourth_call_date'),
            lead.get('fifth_call_date'),
            lead.get('sixth_call_date'),
            lead.get('seventh_call_date')
        ])
        
        if follow_up_date and final_status == 'Pending' and has_first_call and has_followup_calls:
            # Parse follow_up_date and check if it's <= today
            try:
                if 'T' in str(follow_up_date):
                    followup_date_parsed = datetime.fromisoformat(str(follow_up_date).replace('Z', '+00:00')).date()
                else:
                    followup_date_parsed = datetime.strptime(str(follow_up_date)[:10], '%Y-%m-%d').date()
                
                if followup_date_parsed <= datetime.now().date():
                    # Add overdue flag for highlighting
                    lead['is_overdue'] = followup_date_parsed < datetime.now().date()
                    lead['overdue_days'] = (datetime.now().date() - followup_date_parsed).days
                    todays_followups.append(lead)
            except (ValueError, TypeError):
                # If date parsing fails, include the lead for manual review
                lead['is_overdue'] = False
                lead['overdue_days'] = 0
                todays_followups.append(lead)

    # Add event leads with today's cre_followup_date to the follow-up list
    event_leads_today = []
    for lead in event_event_leads:
        cre_followup_date = lead.get('cre_followup_date')
        if cre_followup_date and str(cre_followup_date)[:10] == today_str:
            event_lead_row = {
                'is_event_lead': True,
                'activity_uid': lead.get('activity_uid'),
                'customer_name': lead.get('customer_name'),
                'customer_phone_number': lead.get('customer_phone_number'),
                'lead_status': lead.get('lead_status'),
                'location': lead.get('location'),
                'activity_name': lead.get('activity_name'),
            }
            event_leads_today.append(event_lead_row)
    
    todays_followups.extend(event_leads_today)

    print(f"[PERF] cre_dashboard TOTAL took {time.time() - start_time:.3f} seconds")
    return render_template(
        'cre_dashboard.html',
        untouched_count=untouched_count,
        called_count=called_count,
        follow_up_count=follow_up_count,
        total_fresh_leads=total_fresh_leads,
        fresh_leads_sorted=fresh_leads_sorted,
        untouched_leads=untouched_leads,
        called_leads=called_leads,
        follow_up_leads=follow_up_leads,
        pending_leads=attended_leads,
        todays_followups=todays_followups,
        attended_leads=attended_leads,
        assigned_to_ps=assigned_to_ps,
        won_leads=won_leads,
        lost_leads=lost_leads,
        event_event_leads=event_event_leads,
        filter_type=filter_type,  # Pass filter type to template
        start_date=start_date_str,  # Pass start date to template
        end_date=end_date_str,  # Pass end date to template
        status=status,  # Pass status parameter to template
        return_tab=tab,  # Pass tab parameter to template
        return_sub_tab=sub_tab  # Pass sub_tab parameter to template
    )

import eventlet
eventlet.monkey_patch()
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, Response, current_app
from supabase.client import create_client, Client
import csv
import openpyxl
import os
from datetime import datetime, timedelta, date
import uuid
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from werkzeug.utils import secure_filename
import random
import string
import io
from dotenv import load_dotenv
from collections import defaultdict, Counter
import json
from auth import AuthManager, require_auth, require_admin, require_cre, require_ps, require_rec
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from security_verification import run_security_verification
import time
import gc
from flask_socketio import SocketIO, emit
import math
# from redis import Redis  # REMOVE this line for local development

# Import optimized operations for faster lead updates
from optimized_lead_operations import create_optimized_operations

# Add this instead:
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from flask import send_file
import tempfile
import matplotlib
matplotlib.use('Agg')  # For headless environments
import matplotlib.pyplot as plt
import seaborn as sns
from matplotlib.backends.backend_pdf import PdfPages
import pytz
import pandas as pd
from werkzeug.security import generate_password_hash, check_password_hash

# Load environment variables from .env file
load_dotenv()

app = Flask(__name__)
socketio = SocketIO(app, cors_allowed_origins="*")

# Reduce Flask log noise
import logging
log = logging.getLogger('werkzeug')
log.setLevel(logging.WARNING)

# Add custom Jinja2 filters
# Note: Using Flask's built-in tojson filter instead of custom one

# Utility functions
def is_valid_date(date_string):
    """Validate date string format (YYYY-MM-DD)"""
    try:
        datetime.strptime(date_string, '%Y-%m-%d')
        return True
    except ValueError:
        return False

def is_valid_uid(uid):
    """Validate UID format"""
    if not uid:
        return False
    # Add your UID validation logic here
    return True


# Using Flask's built-in tojson filter

# Get environment variables with fallback values for testing
SUPABASE_URL = os.environ.get('SUPABASE_URL')
SUPABASE_KEY = os.environ.get('SUPABASE_ANON_KEY')
SECRET_KEY = os.environ.get('FLASK_SECRET_KEY', 'fallback-secret-key-change-this')

# Email configuration (add these to your .env file)
SMTP_SERVER = os.environ.get('SMTP_SERVER', 'smtp.gmail.com')
SMTP_PORT = int(os.environ.get('SMTP_PORT', '587'))
EMAIL_USER = os.environ.get('EMAIL_USER', '')
EMAIL_PASSWORD = os.environ.get('EMAIL_PASSWORD', '')

# Debug: Print to check if variables are loaded (remove in production)
print(f"SUPABASE_URL loaded: {SUPABASE_URL is not None}")
print(f"SUPABASE_KEY loaded: {SUPABASE_KEY is not None}")
print(f"SUPABASE_URL: {SUPABASE_URL}")
print(f"SUPABASE_ANON_KEY: {SUPABASE_KEY}")

# Validate required environment variables
if not SUPABASE_URL:
    raise ValueError("SUPABASE_URL environment variable is required. Please check your .env file.")
if not SUPABASE_KEY:
    raise ValueError("SUPABASE_ANON_KEY environment variable is required. Please check your .env file.")

app.secret_key = SECRET_KEY
app.permanent_session_lifetime = timedelta(hours=24)

# Initialize Supabase client
try:
    supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
    print("✅ Supabase client initialized successfully")
    # Removed Supabase warm-up query for local development
except Exception as e:
    print(f"❌ Error initializing Supabase client: {e}")
    raise

# Initialize optimized operations for faster lead updates
try:
    optimized_ops = create_optimized_operations(supabase)
    print("✅ Optimized operations initialized successfully")
except Exception as e:
    print(f"❌ Error initializing optimized operations: {e}")
    # Continue without optimized operations if there's an error
    optimized_ops = None

# Initialize AuthManager
auth_manager = AuthManager(supabase)
# Store auth_manager in app config instead of direct attribute
app.config['AUTH_MANAGER'] = auth_manager

# Initialize rate limiter
limiter = Limiter(
    key_func=get_remote_address,
    app=app,
    default_limits=["1000 per minute"]  # Use in-memory backend for local/dev
)

# Upload folder configuration
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max file size

# Create upload folder if it doesn't exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def send_email_to_ps(ps_email, ps_name, lead_data, cre_name):
    """Send email notification to PS when a lead is assigned"""
    try:
        if not EMAIL_USER or not EMAIL_PASSWORD:
            print("Email credentials not configured. Skipping email notification.")
            return False

        # Create message
        msg = MIMEMultipart()
        msg['From'] = EMAIL_USER
        msg['To'] = ps_email
        msg['Subject'] = f"New Lead Assigned - {lead_data['customer_name']}"

        # Email body
        body = f"""
        Dear {ps_name},

        A new lead has been assigned to you by {cre_name}.

        Lead Details:
        - Customer Name: {lead_data['customer_name']}
        - Mobile Number: {lead_data['customer_mobile_number']}
        - Source: {lead_data['source']}
        - Lead Category: {lead_data.get('lead_category', 'Not specified')}
        - Model Interested: {lead_data.get('model_interested', 'Not specified')}
        - Branch: {lead_data.get('branch', 'Not specified')}

        Please log in to the CRM system to view and update this lead.

        Best regards,
        Ather CRM System
        """

        msg.attach(MIMEText(body, 'plain'))

        # Send email
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()
        server.login(EMAIL_USER, EMAIL_PASSWORD)
        text = msg.as_string()
        server.sendmail(EMAIL_USER, ps_email, text)
        server.quit()

        print(f"Email sent successfully to {ps_email}")
        return True

    except Exception as e:
        print(f"Error sending email: {e}")
        return False


def read_csv_file(filepath):
    """Read CSV file and return list of dictionaries with memory optimization"""
    data = []
    try:
        with open(filepath, 'r', encoding='utf-8') as file:
            csv_reader = csv.DictReader(file)
            for row_num, row in enumerate(csv_reader):
                if row_num >= 10000:  # Limit to 10,000 rows
                    print(f"Warning: File contains more than 10,000 rows. Only processing first 10,000.")
                    break

                # Clean and validate row data
                cleaned_row = {}
                for key, value in row.items():
                    if key and value:  # Only include non-empty keys and values
                        cleaned_row[key.strip()] = str(value).strip()

                if cleaned_row:  # Only add non-empty rows
                    data.append(cleaned_row)

                # Memory management for large files
                if row_num % 1000 == 0 and row_num > 0:
                    print(f"Processed {row_num} rows...")

    except Exception as e:
        print(f"Error reading CSV file: {e}")
        raise

    return data


def read_excel_file(filepath):
    """Read Excel file and return list of dictionaries with memory optimization"""
    data = []
    try:
        workbook = openpyxl.load_workbook(filepath, read_only=True)  # Read-only mode for memory efficiency
        sheet = workbook.active

        # Get headers from first row
        headers = []
        if sheet and sheet[1]:
            for cell in sheet[1]:
                if cell and cell.value:
                    headers.append(str(cell.value).strip())
                else:
                    headers.append(None)

        # Read data rows with limit
        row_count = 0
        if sheet:
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if row_count >= 10000:  # Limit to 10,000 rows
                    print(f"Warning: File contains more than 10,000 rows. Only processing first 10,000.")
                    break

                row_data = {}
                has_data = False

                for i, value in enumerate(row):
                    if i < len(headers) and headers[i] and value is not None:
                        row_data[headers[i]] = str(value).strip()
                        has_data = True

                if has_data:  # Only add rows with actual data
                    data.append(row_data)
                    row_count += 1

                # Memory management for large files
                if row_count % 1000 == 0 and row_count > 0:
                    print(f"Processed {row_count} rows...")

        workbook.close()  # Explicitly close workbook

    except Exception as e:
        print(f"Error reading Excel file: {e}")
        raise

    return data


def batch_insert_leads(leads_data, batch_size=100):
    """Insert leads in batches to avoid overwhelming the database"""
    total_inserted = 0
    total_batches = (len(leads_data) + batch_size - 1) // batch_size

    print(f"Starting batch insert: {len(leads_data)} leads in {total_batches} batches")

    for i in range(0, len(leads_data), batch_size):
        batch = leads_data[i:i + batch_size]
        batch_num = (i // batch_size) + 1

        try:
            # Insert batch
            result = supabase.table('lead_master').insert(batch).execute()

            if result.data:
                batch_inserted = len(result.data)
                total_inserted += batch_inserted
                print(f"Batch {batch_num}/{total_batches}: Inserted {batch_inserted} leads")
            else:
                print(f"Batch {batch_num}/{total_batches}: No data returned from insert")

            # Small delay to prevent overwhelming the database
            time.sleep(0.1)  # CHANGED from eventlet.sleep(0.1)

            # Force garbage collection every 10 batches
            if batch_num % 10 == 0:
                gc.collect()

        except Exception as e:
            print(f"Error inserting batch {batch_num}: {e}")
            # Continue with next batch instead of failing completely
            continue

    print(f"Batch insert completed: {total_inserted} total leads inserted")
    return total_inserted


def generate_uid(source, mobile_number, sequence):
    """Generate UID based on source, mobile number, and sequence"""
    source_map = {
        'Google': 'G',
        'Meta': 'M',
        'Affiliate': 'A',
        'Know': 'K',
        'Whatsapp': 'W',
        'Tele': 'T',
        'Activity': 'AC',
        'Walk-in': 'W',  # Walk-in mapping
        'Walkin': 'W'    # Walkin mapping (without hyphen)
    }

    source_char = source_map.get(source, 'X')

    # Get sequence character (A-Z)
    sequence_char = chr(65 + (sequence % 26))  # A=65 in ASCII

    # Get last 4 digits of mobile number
    mobile_str = str(mobile_number).replace(' ', '').replace('-', '')
    mobile_last4 = mobile_str[-4:] if len(mobile_str) >= 4 else mobile_str.zfill(4)

    # Generate sequence number (0001, 0002, etc.)
    seq_num = f"{(sequence % 9999) + 1:04d}"

    return f"{source_char}{sequence_char}-{mobile_last4}-{seq_num}"


def get_next_call_info(lead_data):
    """Determine the next available call number and which calls are completed"""
    call_order = ['first', 'second', 'third', 'fourth', 'fifth', 'sixth', 'seventh']
    completed_calls = []
    next_call = 'first'

    for call_num in call_order:
        call_date_key = f'{call_num}_call_date'
        if lead_data.get(call_date_key):
            completed_calls.append(call_num)
        else:
            next_call = call_num
            break

    return next_call, completed_calls


def get_next_ps_call_info(ps_data):
    """Determine the next available PS call number and which calls are completed (now 7 calls)"""
    call_order = ['first', 'second', 'third', 'fourth', 'fifth', 'sixth', 'seventh']
    completed_calls = []
    next_call = 'first'

    for call_num in call_order:
        call_date_key = f'{call_num}_call_date'
        if ps_data.get(call_date_key):
            completed_calls.append(call_num)
        else:
            next_call = call_num
            break

    return next_call, completed_calls


def get_accurate_count(table_name, filters=None):
    """Get accurate count from Supabase table"""
    try:
        query = supabase.table(table_name).select('id')

        if filters:
            for key, value in filters.items():
                if value is not None:
                    query = query.eq(key, value)

        result = query.execute()

        # Count the returned data
        return len(result.data) if result.data else 0

    except Exception as e:
        print(f"Error getting count from {table_name}: {e}")
        return 0


def safe_get_data(table_name, filters=None, select_fields='*', limit=10000):
    """Safely get data from Supabase with error handling"""
    try:
        query = supabase.table(table_name).select(select_fields)

        if filters:
            for key, value in filters.items():
                if value is not None:
                    query = query.eq(key, value)

        # Add limit to prevent default 1000 row limitation
        if limit:
            query = query.limit(limit)

        result = query.execute()
        return result.data or []
    except Exception as e:
        print(f"Error fetching data from {table_name}: {e}")
        return []


def sync_test_drive_to_alltest_drive(source_table, original_id, lead_data):
    """
    Sync test drive data to alltest_drive table when test_drive_done is updated
    """
    try:
        # Check if test_drive_done is not null and is Yes/No or True/False
        test_drive_done = lead_data.get('test_drive_done')
        
        # Handle both boolean and string values
        if test_drive_done is None:
            return
        
        # Convert boolean to string if needed
        if test_drive_done is True:
            test_drive_done = 'Yes'
        elif test_drive_done is False:
            test_drive_done = 'No'
        elif test_drive_done not in ['Yes', 'No']:
            return
        
        # Check if record already exists in alltest_drive
        existing_record = supabase.table('alltest_drive').select('*').eq('source_table', source_table).eq('original_id', str(original_id)).execute()
        
        # Prepare data for alltest_drive table
        alltest_drive_data = {
            'source_table': source_table,
            'original_id': str(original_id),
            'test_drive_done': test_drive_done,
            'updated_at': datetime.now().isoformat()
        }
        
        # Map fields based on source table
        if source_table == 'walkin_table':
            alltest_drive_data.update({
                'customer_name': lead_data.get('customer_name'),
                'mobile_number': lead_data.get('mobile_number'),
                'lead_status': lead_data.get('lead_status'),
                'lead_category': lead_data.get('lead_category'),
                'model_interested': lead_data.get('model_interested'),
                'final_status': lead_data.get('status'),
                'ps_name': lead_data.get('ps_assigned'),
                'branch': lead_data.get('branch'),
                'created_at': lead_data.get('created_at')
            })
            # For walkin_table, use uid instead of id
            alltest_drive_data['original_id'] = lead_data.get('uid', str(original_id))
        elif source_table == 'ps_followup_master':
            alltest_drive_data.update({
                'customer_name': lead_data.get('customer_name'),
                'mobile_number': lead_data.get('customer_mobile_number'),
                'lead_status': lead_data.get('lead_status'),
                'lead_category': lead_data.get('lead_category'),
                'model_interested': lead_data.get('model_interested'),
                'final_status': lead_data.get('final_status'),
                'ps_name': lead_data.get('ps_name'),
                'branch': lead_data.get('branch') or lead_data.get('ps_branch'),
                'created_at': lead_data.get('created_at'),
                'lead_source': lead_data.get('lead_source'),
                'cre_name': lead_data.get('cre_name')
            })
        elif source_table == 'activity_leads':
            alltest_drive_data.update({
                'customer_name': lead_data.get('customer_name'),
                'mobile_number': lead_data.get('customer_phone_number'),
                'lead_status': lead_data.get('lead_status'),
                'lead_category': lead_data.get('lead_category'),
                'model_interested': lead_data.get('interested_model'),
                'final_status': lead_data.get('final_status'),
                'ps_name': lead_data.get('ps_name'),
                'branch': lead_data.get('location'),
                'created_at': lead_data.get('created_at'),
                'remarks': lead_data.get('remarks'),
                'activity_name': lead_data.get('activity_name'),
                'activity_location': lead_data.get('activity_location'),
                'customer_location': lead_data.get('customer_location'),
                'customer_profession': lead_data.get('customer_profession'),
                'gender': lead_data.get('gender')
            })
        
        # Insert or update record in alltest_drive table
        if existing_record.data:
            # Update existing record
            supabase.table('alltest_drive').update(alltest_drive_data).eq('source_table', source_table).eq('original_id', str(original_id)).execute()
        else:
            # Insert new record
            supabase.table('alltest_drive').insert(alltest_drive_data).execute()
            
        print(f"Successfully synced test drive data for {source_table} - {original_id}")
        
    except Exception as e:
        print(f"Error syncing test drive data to alltest_drive: {e}")


def create_or_update_ps_followup(lead_data, ps_name, ps_branch):
    from datetime import datetime
    try:
        existing = supabase.table('ps_followup_master').select('*').eq('lead_uid', lead_data['uid']).execute()
        ps_followup_data = {
            'lead_uid': lead_data['uid'],
            'ps_name': ps_name,
            'ps_branch': ps_branch,
            'customer_name': lead_data.get('customer_name'),
            'customer_mobile_number': lead_data.get('customer_mobile_number'),
            'source': lead_data.get('source'),
            'cre_name': lead_data.get('cre_name'),
            'lead_category': lead_data.get('lead_category'),
            'model_interested': lead_data.get('model_interested'),
            'final_status': 'Pending',
            'ps_assigned_at': datetime.now().isoformat(),  # Always set when PS is assigned
            'created_at': lead_data.get('created_at') or datetime.now().isoformat(),
            'first_call_date': None  # Ensure fresh leads start without first_call_date
        }
        if existing.data:
            supabase.table('ps_followup_master').update(ps_followup_data).eq('lead_uid', lead_data['uid']).execute()
        else:
            supabase.table('ps_followup_master').insert(ps_followup_data).execute()
    except Exception as e:
        print(f"Error creating/updating PS followup: {e}")

def track_cre_call_attempt(uid, cre_name, call_no, lead_status, call_was_recorded=False, follow_up_date=None, remarks=None):
    """Track CRE call attempt in the history table and update TAT for first attempt"""
    try:
        # Get the next attempt number for this call
        attempt_result = supabase.table('cre_call_attempt_history').select('attempt').eq('uid', uid).eq('call_no', call_no).order('attempt', desc=True).limit(1).execute()
        next_attempt = 1
        if attempt_result.data:
            next_attempt = attempt_result.data[0]['attempt'] + 1

        # Fetch the current final_status from lead_master
        final_status = None
        lead_result = supabase.table('lead_master').select('final_status').eq('uid', uid).limit(1).execute()
        if lead_result.data and 'final_status' in lead_result.data[0]:
            final_status = lead_result.data[0]['final_status']

        # Prepare attempt data
        attempt_data = {
            'uid': uid,
            'call_no': call_no,
            'attempt': next_attempt,
            'status': lead_status,
            'cre_name': cre_name,
            'call_was_recorded': call_was_recorded,
            'follow_up_date': follow_up_date,
            'remarks': remarks,
            'final_status': final_status
        }

        # Insert the attempt record
        insert_result = supabase.table('cre_call_attempt_history').insert(attempt_data).execute()
        print(f"Tracked call attempt: {uid} - {call_no} call, attempt {next_attempt}, status: {lead_status}")

        # --- TAT Calculation and Update ---
        if call_no == 'first' and next_attempt == 1:
            # Fetch lead's created_at
            lead_result = supabase.table('lead_master').select('created_at').eq('uid', uid).limit(1).execute()
            if lead_result.data and lead_result.data[0].get('created_at'):
                created_at_str = lead_result.data[0]['created_at']
                from datetime import datetime
                try:
                    if 'T' in created_at_str:
                        created_at = datetime.fromisoformat(created_at_str.replace('Z', '+00:00'))
                    else:
                        created_at = datetime.strptime(created_at_str, '%Y-%m-%d %H:%M:%S')
                except Exception:
                    created_at = None
                # Get updated_at from inserted attempt (if available), else use now
                updated_at_str = None
                if insert_result.data and insert_result.data[0].get('updated_at'):
                    updated_at_str = insert_result.data[0]['updated_at']
                else:
                    from datetime import datetime
                    updated_at_str = datetime.now().isoformat()
                try:
                    if 'T' in updated_at_str:
                        updated_at = datetime.fromisoformat(updated_at_str.replace('Z', '+00:00'))
                    else:
                        updated_at = datetime.strptime(updated_at_str, '%Y-%m-%d %H:%M:%S')
                except Exception:
                    updated_at = datetime.now()
                if created_at:
                    tat_seconds = (updated_at - created_at).total_seconds()
                    # Update lead_master with TAT
                    supabase.table('lead_master').update({'tat': tat_seconds}).eq('uid', uid).execute()
                    print(f"TAT updated for lead {uid}: {tat_seconds} seconds")
    except Exception as e:
        print(f"Error tracking CRE call attempt: {e}")

def track_ps_call_attempt(uid, ps_name, call_no, lead_status, call_was_recorded=False, follow_up_date=None, remarks=None):
    """Track PS call attempt in the history table"""
    try:
        # Get the next attempt number for this call
        attempt_result = supabase.table('ps_call_attempt_history').select('attempt').eq('uid', uid).eq('call_no', call_no).order('attempt', desc=True).limit(1).execute()
        next_attempt = 1
        if attempt_result.data:
            next_attempt = attempt_result.data[0]['attempt'] + 1

        # Fetch the current final_status from multiple tables
        final_status = None
        
        # Try all possible tables and field combinations
        tables_to_check = [
            # (table_name, uid_field, status_field, uid_value)
            ('ps_followup_master', 'lead_uid', 'final_status', uid),
            ('lead_master', 'uid', 'final_status', uid),
            ('walkin_table', 'uid', 'status', uid),
            ('activity_leads', 'activity_uid', 'final_status', uid),
            # Also try with different UID formats for walkin and activity
            ('walkin_table', 'uid', 'status', uid.replace('WB-', 'W')),
            ('activity_leads', 'activity_uid', 'final_status', uid.replace('WB-', 'A'))
        ]
        
        for table_name, uid_field, status_field, uid_value in tables_to_check:
            try:
                result = supabase.table(table_name).select(status_field).eq(uid_field, uid_value).limit(1).execute()
                if result.data and result.data[0].get(status_field):
                    final_status = result.data[0][status_field]
                    print(f"Found final_status '{final_status}' in {table_name} for {uid_value}")
                    break
            except Exception as e:
                print(f"Error checking {table_name}: {e}")
                continue
        
        # If still not found, set a default
        if not final_status:
            final_status = 'Pending'
            print(f"No final_status found for {uid}, defaulting to 'Pending'")

        # Prepare attempt data
        attempt_data = {
            'uid': uid,
            'call_no': call_no,
            'attempt': next_attempt,
            'status': lead_status,
            'ps_name': ps_name,
            'call_was_recorded': call_was_recorded,
            'follow_up_date': follow_up_date,
            'remarks': remarks,
            'final_status': final_status
        }

        # Insert the attempt record
        supabase.table('ps_call_attempt_history').insert(attempt_data).execute()
        print(f"Tracked PS call attempt: {uid} - {call_no} call, attempt {next_attempt}, status: {lead_status}")
    except Exception as e:
        print(f"Error tracking PS call attempt: {e}")


def filter_leads_by_date(leads, filter_type, date_field='created_at'):
    """Filter leads based on date range"""
    if filter_type == 'all':
        return leads

    today = datetime.now().date()

    if filter_type == 'today':
        start_date = today
        end_date = today
    elif filter_type == 'mtd':  # Month to Date
        start_date = today.replace(day=1)
        end_date = today
    elif filter_type == 'week':
        start_date = today - timedelta(days=today.weekday())  # Start of current week (Monday)
        end_date = today
    elif filter_type == 'month':
        start_date = today - timedelta(days=30)
        end_date = today
    elif filter_type == 'quarter':
        start_date = today - timedelta(days=90)
        end_date = today
    elif filter_type == 'year':
        start_date = today - timedelta(days=365)
        end_date = today
    else:
        return leads

    filtered_leads = []
    for lead in leads:
        lead_date_str = lead.get(date_field)
        if lead_date_str:
            try:
                # Handle different date formats
                if 'T' in lead_date_str:  # ISO format with time
                    lead_date = datetime.fromisoformat(lead_date_str.replace('Z', '+00:00')).date()
                else:  # Date only format
                    lead_date = datetime.strptime(lead_date_str, '%Y-%m-%d').date()

                if start_date <= lead_date <= end_date:
                    filtered_leads.append(lead)
            except (ValueError, TypeError):
                # If date parsing fails, include the lead
                filtered_leads.append(lead)
        else:
            # If no date field, include the lead
            filtered_leads.append(lead)

    return filtered_leads


def fix_missing_timestamps():
    """
    Fix missing timestamps for existing leads that have final_status but missing won_timestamp or lost_timestamp
    """
    try:
        # Fix lead_master table
        # Get leads with final_status = 'Won' but no won_timestamp
        won_leads = supabase.table('lead_master').select('uid, final_status, won_timestamp, lost_timestamp, updated_at').eq('final_status', 'Won').is_('won_timestamp', 'null').execute()
        
        for lead in won_leads.data:
            supabase.table('lead_master').update({
                'won_timestamp': lead.get('updated_at') or datetime.now().isoformat()
            }).eq('uid', lead['uid']).execute()
        
        # Get leads with final_status = 'Lost' but no lost_timestamp
        lost_leads = supabase.table('lead_master').select('uid, final_status, won_timestamp, lost_timestamp, updated_at').eq('final_status', 'Lost').is_('lost_timestamp', 'null').execute()
        
        for lead in lost_leads.data:
            supabase.table('lead_master').update({
                'lost_timestamp': lead.get('updated_at') or datetime.now().isoformat()
            }).eq('uid', lead['uid']).execute()
        
        # Fix ps_followup_master table
        # Get PS leads with final_status = 'Won' but no won_timestamp
        ps_won_leads = supabase.table('ps_followup_master').select('lead_uid, final_status, won_timestamp, lost_timestamp, updated_at').eq('final_status', 'Won').is_('won_timestamp', 'null').execute()
        
        for lead in ps_won_leads.data:
            supabase.table('ps_followup_master').update({
                'won_timestamp': lead.get('updated_at') or datetime.now().isoformat()
            }).eq('lead_uid', lead['lead_uid']).execute()
        
        # Get PS leads with final_status = 'Lost' but no lost_timestamp
        ps_lost_leads = supabase.table('ps_followup_master').select('lead_uid, final_status, won_timestamp, lost_timestamp, updated_at').eq('final_status', 'Lost').is_('lost_timestamp', 'null').execute()
        
        for lead in ps_lost_leads.data:
            supabase.table('ps_followup_master').update({
                'lost_timestamp': lead.get('updated_at') or datetime.now().isoformat()
            }).eq('lead_uid', lead['lead_uid']).execute()
        
        print(f"Fixed {len(won_leads.data)} won leads, {len(lost_leads.data)} lost leads in lead_master")
        print(f"Fixed {len(ps_won_leads.data)} won leads, {len(ps_lost_leads.data)} lost leads in ps_followup_master")
        
    except Exception as e:
        print(f"Error fixing timestamps: {str(e)}")


@app.route('/')
def index():
    session.clear()  # Ensure no session data is present
    return render_template('index.html')
@app.route('/unified_login', methods=['POST'])
@limiter.limit("100000 per minute")
def unified_login() -> Response:
    start_time = time.time()
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '').strip()
    user_type = request.form.get('user_type', '').strip().lower()

    valid_user_types = ['admin', 'cre', 'ps', 'branch_head', 'rec']
    if user_type not in valid_user_types:
        flash('Please select a valid role (Admin, CRE, PS, Branch Head, or Receptionist)', 'error')
        return redirect(url_for('index'))

    if user_type == 'branch_head':
        bh = supabase.table('Branch Head').select('*').eq('Username', username).execute().data
        if not bh:
            flash('Invalid username or password', 'error')
            return redirect(url_for('index'))
        bh = bh[0]
        if not bh['Is Active']:
            flash('User is inactive', 'error')
            return redirect(url_for('index'))
        # Use plain text password comparison for Branch Head
        if bh['Password'] != password:
            flash('Incorrect password', 'error')
            return redirect(url_for('index'))
        session['branch_head_id'] = bh['id']
        session['branch_head_name'] = bh['Name']
        session['branch_head_branch'] = bh['Branch']
        session['user_type'] = 'branch_head'  # Add this line to set user_type
        session['username'] = username  # Add username for consistency
        flash('Welcome! Logged in as Branch Head', 'success')
        return redirect('/branch_head_dashboard')

    elif user_type == 'rec':
        # Receptionist authentication
        rec_user = supabase.table('rec_users').select('*').eq('username', username).execute().data
        if not rec_user:
            flash('Invalid username or password', 'error')
            return redirect(url_for('index'))
        rec_user = rec_user[0]
        if not rec_user.get('is_active', True):
            flash('User is inactive', 'error')
            return redirect(url_for('index'))
        # Check password using werkzeug
        if not check_password_hash(rec_user['password_hash'], password):
            flash('Incorrect password', 'error')
            return redirect(url_for('index'))
        session.clear()
        session['rec_user_id'] = rec_user['id']
        session['rec_branch'] = rec_user['branch']
        session['rec_name'] = rec_user.get('name', username)
        session['user_type'] = 'rec'
        session['username'] = username
        flash('Welcome! Logged in as Receptionist', 'success')
        return redirect(url_for('add_walkin_lead'))

    # Existing logic for admin, cre, ps
    t_user = time.time()
    success, message, user_data = auth_manager.authenticate_user(username, password, user_type)
    print(f"[PERF] unified_login: authenticate_user({user_type}) took {time.time() - t_user:.3f} seconds")
    if success:
        t2 = time.time()
        session_id = auth_manager.create_session(user_data['id'], user_type, user_data)
        print(f"DEBUG: Logged in as user_type={user_type}, session.user_type={session.get('user_type')}")
        print(f"[PERF] unified_login: create_session took {time.time() - t2:.3f} seconds")
        if session_id:
            flash(f'Welcome! Logged in as {user_type.upper()}', 'success')
            t3 = time.time()
            # Redirect to appropriate dashboard
            if user_type == 'admin':
                print(f"[PERF] unified_login: redirect to admin_dashboard after {time.time() - t3:.3f} seconds")
                print(f"[PERF] unified_login TOTAL took {time.time() - start_time:.3f} seconds")
                return redirect(url_for('admin_dashboard'))
            elif user_type == 'cre':
                print(f"[PERF] unified_login: redirect to cre_dashboard after {time.time() - t3:.3f} seconds")
                print(f"[PERF] unified_login TOTAL took {time.time() - start_time:.3f} seconds")
                return redirect(url_for('cre_dashboard'))
            elif user_type == 'ps':
                print(f"[PERF] unified_login: redirect to ps_dashboard after {time.time() - t3:.3f} seconds")
                print(f"[PERF] unified_login TOTAL took {time.time() - start_time:.3f} seconds")
                return redirect(url_for('ps_dashboard'))
        else:
            flash('Error creating session', 'error')
            print(f"[PERF] unified_login: session creation failed after {time.time() - t2:.3f} seconds")
            print(f"[PERF] unified_login TOTAL took {time.time() - start_time:.3f} seconds")
            return redirect(url_for('index'))
    else:
        flash('Invalid username or password', 'error')
        print(f"[PERF] unified_login TOTAL (invalid login) took {time.time() - start_time:.3f} seconds")
        return redirect(url_for('index'))
# Keep the old login routes for backward compatibility (redirect to unified login)
@app.route('/admin_login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        return unified_login()
    return redirect(url_for('index'))


@app.route('/cre_login', methods=['GET', 'POST'])
def cre_login():
    if request.method == 'POST':
        return unified_login()
    return redirect(url_for('index'))


@app.route('/ps_login', methods=['GET', 'POST'])
def ps_login():
    if request.method == 'POST':
        return unified_login()
    return redirect(url_for('index'))


@app.route('/password_reset_request', methods=['GET', 'POST'])
@limiter.limit("5 per minute")
def password_reset_request():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        user_type = request.form.get('user_type', '').strip()

        if not username or not user_type:
            flash('Please enter username and select user type', 'error')
            return render_template('password_reset_request.html')

        success, message, token = auth_manager.generate_password_reset_token(username, user_type)

        if success:
            # Send the reset link via email
            # Fetch user email from the appropriate table
            user_email = None
            try:
                if user_type == 'admin':
                    user_result = supabase.table('admin_users').select('email').eq('username', username).execute()
                elif user_type == 'cre':
                    user_result = supabase.table('cre_users').select('email').eq('username', username).execute()
                elif user_type == 'ps':
                    user_result = supabase.table('ps_users').select('email').eq('username', username).execute()
                else:
                    user_result = None
                if user_result and user_result.data and user_result.data[0].get('email'):
                    user_email = user_result.data[0]['email']
            except Exception as e:
                print(f"Error fetching user email for password reset: {e}")
                user_email = None

            reset_url = url_for('password_reset', token=token, _external=True)
            email_sent = False
            if user_email:
                try:
                    msg = MIMEMultipart()
                    msg['From'] = EMAIL_USER
                    msg['To'] = user_email
                    msg['Subject'] = 'Ather CRM Password Reset Request'
                    body = f"""
                    Dear {username},

                    We received a request to reset your password for your Ather CRM account.

                    Please click the link below to reset your password:
                    {reset_url}

                    If you did not request this, please ignore this email.

                    Best regards,\nAther CRM System
                    """
                    msg.attach(MIMEText(body, 'plain'))
                    server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
                    server.starttls()
                    server.login(EMAIL_USER, EMAIL_PASSWORD)
                    text = msg.as_string()
                    server.sendmail(EMAIL_USER, user_email, text)
                    server.quit()
                    print(f"Password reset email sent to {user_email}")
                    email_sent = True
                except Exception as e:
                    print(f"Error sending password reset email: {e}")
                    email_sent = False
            if email_sent:
                flash('If the username exists and is valid, a password reset link has been sent to the registered email address.', 'success')
            else:
                flash('If the username exists and is valid, a password reset link has been sent to the registered email address.', 'success')
                # Optionally, log or alert admin if email sending failed
        else:
            flash(message, 'error')

    return render_template('password_reset_request.html')


@app.route('/password_reset/<token>', methods=['GET', 'POST'])
def password_reset(token):
    if request.method == 'POST':
        new_password = request.form.get('new_password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()

        if not new_password or not confirm_password:
            flash('Please enter and confirm your new password', 'error')
            return render_template('password_reset.html', token=token)

        if new_password != confirm_password:
            flash('Passwords do not match', 'error')
            return render_template('password_reset.html', token=token)

        success, message = auth_manager.reset_password_with_token(token, new_password)

        if success:
            flash('Password reset successfully. Please log in with your new password.', 'success')
            return redirect(url_for('index'))
        else:
            flash(message, 'error')

    return render_template('password_reset.html', token=token)


@app.route('/change_password', methods=['POST'])
@require_auth()
def change_password():
    current_password = request.form.get('current_password', '').strip()
    new_password = request.form.get('new_password', '').strip()
    confirm_password = request.form.get('confirm_password', '').strip()

    if not all([current_password, new_password, confirm_password]):
        flash('All fields are required', 'error')
        return redirect(url_for('security_settings'))

    if new_password != confirm_password:
        flash('New passwords do not match', 'error')
        return redirect(url_for('security_settings'))

    user_id = session.get('user_id')
    user_type = session.get('user_type')

    if not user_id or not user_type:
        flash('Session information not found', 'error')
        return redirect(url_for('security_settings'))

    success, message = auth_manager.change_password(user_id, user_type, current_password, new_password)

    if success:
        flash('Password changed successfully', 'success')
    else:
        flash(message, 'error')

    return redirect(url_for('security_settings'))


@app.route('/change_cre_password', methods=['GET', 'POST'])
@require_cre
def change_cre_password():
    if request.method == 'POST':
        current_password = request.form.get('current_password', '').strip()
        new_password = request.form.get('new_password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()

        if not all([current_password, new_password, confirm_password]):
            flash('All fields are required', 'error')
            return render_template('change_cre_password.html')

        if new_password != confirm_password:
            flash('New passwords do not match', 'error')
            return render_template('change_cre_password.html')

        user_id = session.get('user_id')
        user_type = session.get('user_type')

        if not user_id or not user_type:
            flash('Session information not found', 'error')
            return render_template('change_cre_password.html')

        success, message = auth_manager.change_password(user_id, user_type, current_password, new_password)

        if success:
            flash('Password changed successfully', 'success')
            return redirect(url_for('cre_dashboard'))
        else:
            flash(message, 'error')

    return render_template('change_cre_password.html')


@app.route('/change_ps_password', methods=['GET', 'POST'])
@require_ps
def change_ps_password():
    if request.method == 'POST':
        current_password = request.form.get('current_password', '').strip()
        new_password = request.form.get('new_password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()

        if not all([current_password, new_password, confirm_password]):
            flash('All fields are required', 'error')
            return render_template('change_ps_password.html')

        if new_password != confirm_password:
            flash('New passwords do not match', 'error')
            return render_template('change_ps_password.html')

        user_id = session.get('user_id')
        user_type = session.get('user_type')

        if not user_id or not user_type:
            flash('Session information not found', 'error')
            return render_template('change_ps_password.html')

        success, message = auth_manager.change_password(user_id, user_type, current_password, new_password)

        if success:
            flash('Password changed successfully', 'success')
            return redirect(url_for('ps_dashboard'))
        else:
            flash(message, 'error')

    return render_template('change_ps_password.html')


@app.route('/security_settings')
@require_auth()
def security_settings():
    user_id = session.get('user_id')
    user_type = session.get('user_type')

    if not user_id or not user_type:
        flash('Session information not found', 'error')
        return redirect(url_for('index'))

    # Get active sessions
    sessions = auth_manager.get_user_sessions(user_id, user_type)

    # Get audit logs
    audit_logs = auth_manager.get_audit_logs(user_id, user_type, limit=20)

    return render_template('security_settings.html', sessions=sessions, audit_logs=audit_logs)


@app.route('/security_audit')
@require_admin
def security_audit():
    """Security audit dashboard"""
    return render_template('security_audit.html')


@app.route('/run_security_audit', methods=['POST'])
@require_admin
def run_security_audit():
    """Run comprehensive security audit"""
    try:
        # Run security verification
        audit_results = run_security_verification(supabase)

        # Log the security audit
        user_id = session.get('user_id')
        user_type = session.get('user_type')
        
        if user_id and user_type:
            auth_manager.log_audit_event(
                user_id=user_id,
                user_type=user_type,
                action='SECURITY_AUDIT_RUN',
                resource='security_audit',
                details={'overall_score': audit_results.get('overall_score', 0)}
            )

        return jsonify({
            'success': True,
            'results': audit_results
        })
    except Exception as e:
        print(f"Error running security audit: {e}")
        return jsonify({
            'success': False,
            'message': str(e)
        })


@app.route('/terminate_session', methods=['POST'])
@require_auth()
def terminate_session():
    try:
        data = request.get_json()
        session_id = data.get('session_id')

        if not session_id:
            return jsonify({'success': False, 'message': 'Session ID required'})

        auth_manager.deactivate_session(session_id)

        user_id = session.get('user_id')
        user_type = session.get('user_type')
        
        if user_id and user_type:
            auth_manager.log_audit_event(
                user_id=user_id,
                user_type=user_type,
                action='SESSION_TERMINATED',
                details={'terminated_session': session_id}
            )

        return jsonify({'success': True, 'message': 'Session terminated successfully'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/terminate_all_sessions', methods=['POST'])
@require_auth()
def terminate_all_sessions():
    try:
        user_id = session.get('user_id')
        user_type = session.get('user_type')
        current_session = session.get('session_id')

        if user_id and user_type and current_session:
            auth_manager.deactivate_all_user_sessions(user_id, user_type, current_session)
            auth_manager.log_audit_event(
                user_id=user_id,
                user_type=user_type,
                action='ALL_SESSIONS_TERMINATED',
                details={'except_session': current_session}
            )
        else:
            return jsonify({'success': False, 'message': 'Session information not found'})

        return jsonify({'success': True, 'message': 'All other sessions terminated successfully'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/admin_dashboard')
@require_admin
def admin_dashboard():
    # Get counts for dashboard with better error handling and actual queries
    try:
        # Get actual counts from database with proper queries
        cre_count = get_accurate_count('cre_users')
        ps_count = get_accurate_count('ps_users')
        leads_count = get_accurate_count('lead_master')
        unassigned_leads = get_accurate_count('lead_master', {'assigned': 'No'})

        print(
            f"Dashboard counts - CRE: {cre_count}, PS: {ps_count}, Total Leads: {leads_count}, Unassigned: {unassigned_leads}")

    except Exception as e:
        print(f"Error getting dashboard counts: {e}")
        cre_count = ps_count = leads_count = unassigned_leads = 0

    # Log dashboard access
    user_id = session.get('user_id')
    user_type = session.get('user_type')
    if user_id and user_type:
        auth_manager.log_audit_event(
            user_id=user_id,
            user_type=user_type,
            action='DASHBOARD_ACCESS',
            resource='admin_dashboard'
        )

    return render_template('admin_dashboard.html',
                           cre_count=cre_count,
                           ps_count=ps_count,
                           leads_count=leads_count,
                           unassigned_leads=unassigned_leads)


@app.route('/upload_data', methods=['GET', 'POST'])
@require_admin
def upload_data():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file selected', 'error')
            return redirect(request.url)

        file = request.files['file']
        source = request.form.get('source', '').strip()

        if not source:
            flash('Please select a data source', 'error')
            return redirect(request.url)

        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(request.url)

        if file and file.filename and allowed_file(file.filename):
            filename = secure_filename(str(file.filename))
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

            try:
                file.save(filepath)

                # Check file size
                file_size = os.path.getsize(filepath)
                if file_size > 50 * 1024 * 1024:  # 50MB limit
                    flash('File too large. Maximum size is 50MB.', 'error')
                    os.remove(filepath)
                    return redirect(request.url)

                print(f"Processing file: {filename} ({file_size / 1024 / 1024:.2f} MB)")

                # Read the file based on extension
                if filename.lower().endswith('.csv'):
                    data = read_csv_file(filepath)
                else:
                    data = read_excel_file(filepath)

                if not data:
                    flash('No valid data found in file', 'error')
                    os.remove(filepath)
                    return redirect(request.url)

                print(f"Read {len(data)} rows from file")

                # Get current sequence number for UID generation
                result = supabase.table('lead_master').select('uid').execute()
                current_count = len(result.data) if result.data else 0

                # Prepare leads data for batch insert
                leads_to_insert = []
                skipped_rows = 0

                for index, row in enumerate(data):
                    try:
                        # Validate required fields
                        required_fields = ['customer_name', 'customer_mobile_number', 'date']
                        if not all(key in row and str(row[key]).strip() for key in required_fields):
                            skipped_rows += 1
                            continue

                        uid = generate_uid(source, row['customer_mobile_number'],
                                           current_count + len(leads_to_insert) + 1)

                        lead_data = {
                            'uid': uid,
                            'date': str(row['date']).strip(),
                            'customer_name': str(row['customer_name']).strip(),
                            'customer_mobile_number': str(row['customer_mobile_number']).strip(),
                            'source': source,
                            'assigned': 'No',
                            'final_status': 'Pending'
                        }

                        leads_to_insert.append(lead_data)


                    except Exception as e:
                        print(f"Error processing row {index}: {e}")
                        skipped_rows += 1
                        continue

                if not leads_to_insert:
                    flash('No valid leads found to insert', 'error')
                    os.remove(filepath)
                    return redirect(request.url)

                print(f"Prepared {len(leads_to_insert)} leads for insertion")

                # Batch insert leads
                success_count = batch_insert_leads(leads_to_insert)

                # Log data upload
                auth_manager.log_audit_event(
                    user_id=session.get('user_id'),
                    user_type=session.get('user_type'),
                    action='DATA_UPLOAD',
                    resource='lead_master',
                    details={
                        'source': source,
                        'records_uploaded': success_count,
                        'filename': filename,
                        'file_size_mb': round(file_size / 1024 / 1024, 2),
                        'skipped_rows': skipped_rows
                    }
                )

                # Create success message
                message = f'Successfully uploaded {success_count} leads'
                if skipped_rows > 0:
                    message += f' ({skipped_rows} rows skipped due to missing data)'
                message += '. Please go to "Assign Leads" to assign them to CREs.'

                flash(message, 'success')

                # Clean up uploaded file
                os.remove(filepath)

            except Exception as e:
                print(f"Error processing file: {e}")
                flash(f'Error processing file: {str(e)}', 'error')
                if os.path.exists(filepath):
                    os.remove(filepath)
        else:
            flash('Invalid file format. Please upload CSV or Excel files only.', 'error')

    return render_template('upload_data.html')

@app.route('/assign_leads')
@require_admin
def assign_leads():
    try:
        # Fetch all unassigned leads in batches of 1000
        all_unassigned_leads = []
        batch_size = 1000
        offset = 0
        while True:
            result = supabase.table('lead_master').select('*').eq('assigned', 'No').range(offset, offset + batch_size - 1).execute()
            batch = result.data or []
            all_unassigned_leads.extend(batch)
            if len(batch) < batch_size:
                break
            offset += batch_size

        # Organize by source
        leads_by_source = {}
        for lead in all_unassigned_leads:
            source = lead.get('source', 'Unknown')
            leads_by_source.setdefault(source, []).append(lead)

        # Get CREs
        cres = safe_get_data('cre_users')

        # Get accurate total unassigned count
        actual_unassigned_count = get_accurate_count('lead_master', {'assigned': 'No'})

        # Get accurate per-source unassigned counts
        source_unassigned_counts = {}
        for source in leads_by_source.keys():
            source_unassigned_counts[source] = get_accurate_count('lead_master', {'assigned': 'No', 'source': source})

        return render_template('assign_leads.html',
                               unassigned_leads=all_unassigned_leads,
                               actual_unassigned_count=actual_unassigned_count,
                               cres=cres,
                               leads_by_source=leads_by_source,
                               source_unassigned_counts=source_unassigned_counts)
    except Exception as e:
        print(f"Error loading assign leads data: {e}")
        flash(f'Error loading data: {str(e)}', 'error')
        return redirect(url_for('admin_dashboard'))


@app.route('/assign_leads_dynamic_action', methods=['POST'])
@require_admin
def assign_leads_dynamic_action():
    try:
        data = request.get_json()
        assignments = data.get('assignments', [])

        if not assignments:
            return jsonify({'success': False, 'message': 'No assignments provided'}), 400

        # Fetch all unassigned leads in batches
        all_unassigned = []
        batch_size = 1000
        offset = 0
        while True:
            result = supabase.table('lead_master').select('*').eq('assigned', 'No').range(offset, offset + batch_size - 1).execute()
            batch = result.data or []
            all_unassigned.extend(batch)
            if len(batch) < batch_size:
                break
            offset += batch_size

        leads_by_source = {}
        for lead in all_unassigned:
            source = lead.get('source', 'Unknown')
            leads_by_source.setdefault(source, []).append(lead)

        total_assigned = 0

        for assignment in assignments:
            cre_id = assignment.get('cre_id')
            source = assignment.get('source')
            quantity = assignment.get('quantity')

            if not cre_id or not source or not quantity:
                continue

            # Get CRE details
            cre_data = supabase.table('cre_users').select('*').eq('id', cre_id).execute()
            if not cre_data.data:
                continue

            cre = cre_data.data[0]
            leads = leads_by_source.get(source, [])

            if not leads:
                print(f"No unassigned leads found for source {source}")
                continue

            random.shuffle(leads)
            leads_to_assign = leads[:quantity]
            leads_by_source[source] = leads[quantity:]  # Remove assigned leads

            for lead in leads_to_assign:
                update_data = {
                    'cre_name': cre['name'],
                    'assigned': 'Yes',
                    'cre_assigned_at': datetime.now().isoformat()

                }

                try:
                    supabase.table('lead_master').update(update_data).eq('uid', lead['uid']).execute()
                    total_assigned += 1
                    print(f"Assigned lead {lead['uid']} to CRE {cre['name']} for source {source}")
                    if total_assigned % 100 == 0:
                        time.sleep(0.1)
                except Exception as e:
                    print(f"Error assigning lead {lead['uid']}: {e}")

        print(f"Total leads assigned: {total_assigned}")
        return jsonify({'success': True, 'message': f'Total {total_assigned} leads assigned successfully'})

    except Exception as e:
        print(f"Error in dynamic lead assignment: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500
@app.route('/add_cre', methods=['GET', 'POST'])
@require_admin
def add_cre():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        phone = request.form.get('phone', '').strip()
        email = request.form.get('email', '').strip()

        if not all([name, username, password, phone, email]):
            flash('All fields are required', 'error')
            return render_template('add_cre.html')

        if not auth_manager.validate_password_strength(password):
            flash(
                'Password must be at least 8 characters long and contain uppercase, lowercase, number, and special character',
                'error')
            return render_template('add_cre.html')

        try:
            # Check if username already exists
            existing = supabase.table('cre_users').select('username').eq('username', username).execute()
            if existing.data:
                flash('Username already exists', 'error')
                return render_template('add_cre.html')

            # Hash password
            password_hash, salt = auth_manager.hash_password(password)

            # Replace the existing cre_data creation with this:
            cre_data = {
                'name': name,
                'username': username,
                'password': password,  # Keep for backward compatibility
                'password_hash': password_hash,
                'salt': salt,
                'phone': phone,
                'email': email,
                'is_active': True,
                'role': 'cre',
                'failed_login_attempts': 0
            }

            result = supabase.table('cre_users').insert(cre_data).execute()

            # Log CRE creation
            auth_manager.log_audit_event(
                user_id=session.get('user_id'),
                user_type=session.get('user_type'),
                action='CRE_CREATED',
                resource='cre_users',
                resource_id=str(result.data[0]['id']) if result.data else None,
                details={'cre_name': name, 'username': username}
            )

            flash('CRE added successfully', 'success')
            return redirect(url_for('manage_cre'))
        except Exception as e:
            flash(f'Error adding CRE: {str(e)}', 'error')

    return render_template('add_cre.html')
@app.route('/add_ps', methods=['GET', 'POST'])
@require_admin
def add_ps():
    branches = get_system_branches()

    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        phone = request.form.get('phone', '').strip()
        email = request.form.get('email', '').strip()
        branch = request.form.get('branch', '').strip()

        if not all([name, username, password, phone, email, branch]):
            flash('All fields are required', 'error')
            return render_template('add_ps.html', branches=branches)

        if not auth_manager.validate_password_strength(password):
            flash(
                'Password must be at least 8 characters long and contain uppercase, lowercase, number, and special character',
                'error')
            return render_template('add_ps.html', branches=branches)

        try:
            # Check if username already exists
            existing = supabase.table('ps_users').select('username').eq('username', username).execute()
            if existing.data:
                flash('Username already exists', 'error')
                return render_template('add_ps.html', branches=branches)

            # Hash password
            password_hash, salt = auth_manager.hash_password(password)

            # Replace the existing ps_data creation with this:
            ps_data = {
                'name': name,
                'username': username,
                'password': password,  # Keep for backward compatibility
                'password_hash': password_hash,
                'salt': salt,
                'phone': phone,
                'email': email,
                'branch': branch,
                'is_active': True,
                'role': 'ps',
                'failed_login_attempts': 0
            }

            result = supabase.table('ps_users').insert(ps_data).execute()

            # Log PS creation
            auth_manager.log_audit_event(
                user_id=session.get('user_id'),
                user_type=session.get('user_type'),
                action='PS_CREATED',
                resource='ps_users',
                resource_id=str(result.data[0]['id']) if result.data else None,
                details={'ps_name': name, 'username': username, 'branch': branch}
            )

            flash('Product Specialist added successfully', 'success')
            return redirect(url_for('manage_ps'))
        except Exception as e:
            flash(f'Error adding Product Specialist: {str(e)}', 'error')

    return render_template('add_ps.html', branches=branches)


@app.route('/manage_cre')
@require_admin
def manage_cre():
    try:
        cre_users = safe_get_data('cre_users')
        return render_template('manage_cre.html', cre_users=cre_users)
    except Exception as e:
        flash(f'Error loading CRE users: {str(e)}', 'error')
        return render_template('manage_cre.html', cre_users=[])


@app.route('/manage_ps')
@require_admin
def manage_ps():
    try:
        ps_users = safe_get_data('ps_users')
        return render_template('manage_ps.html', ps_users=ps_users)
    except Exception as e:
        flash(f'Error loading PS users: {str(e)}', 'error')
        return render_template('manage_ps.html', ps_users=[])


@app.route('/toggle_ps_status/<int:ps_id>', methods=['POST'])
@require_admin
def toggle_ps_status(ps_id):
    try:
        data = request.get_json()
        active_status = data.get('active', True)

        # Update PS status
        result = supabase.table('ps_users').update({
            'is_active': active_status
        }).eq('id', ps_id).execute()

        if result.data:
            # Log status change
            auth_manager.log_audit_event(
                user_id=session.get('user_id'),
                user_type=session.get('user_type'),
                action='PS_STATUS_CHANGED',
                resource='ps_users',
                resource_id=str(ps_id),
                details={'new_status': 'active' if active_status else 'inactive'}
            )

            return jsonify({'success': True, 'message': 'PS status updated successfully'})
        else:
            return jsonify({'success': False, 'message': 'PS not found'})

    except Exception as e:
        print(f"Error toggling PS status: {e}")
        return jsonify({'success': False, 'message': str(e)})


@app.route('/manage_leads')
@require_admin
def manage_leads():
    try:
        cres = safe_get_data('cre_users')
        cre_id = request.args.get('cre_id')
        source = request.args.get('source')
        qualification = request.args.get('qualification', 'all')
        date_filter = request.args.get('date_filter', 'all')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        final_status = request.args.get('final_status', '')
        page = int(request.args.get('page', 1))
        per_page = 50
        search_uid = request.args.get('search_uid', '').strip()
        selected_cre = None
        leads = []
        sources = []
        total_leads = 0
        if cre_id:
            # Find selected CRE
            selected_cre = next((cre for cre in cres if str(cre.get('id')) == str(cre_id)), None)
            # Fetch leads for this CRE
            filters = {'cre_name': selected_cre['name']} if selected_cre else {}
            if source:
                filters['source'] = source
            leads = safe_get_data('lead_master', filters)
            # UID substring search for this CRE
            if search_uid:
                leads = [lead for lead in leads if search_uid.lower() in str(lead.get('uid', '')).lower()]
            # Qualification filter
            if qualification == 'qualified':
                leads = [lead for lead in leads if lead.get('first_call_date')]
            elif qualification == 'unqualified':
                leads = [lead for lead in leads if not lead.get('first_call_date')]
            # Final status filter
            if final_status:
                leads = [lead for lead in leads if (lead.get('final_status') or '') == final_status]
            # Date filtering
            if date_filter == 'today':
                today_str = datetime.now().strftime('%Y-%m-%d')
                leads = [lead for lead in leads if lead.get('cre_assigned_at') and str(lead.get('cre_assigned_at')).startswith(today_str)]
            elif date_filter == 'range' and start_date and end_date:
                def in_range(ld):
                    dt = ld.get('cre_assigned_at')
                    if not dt:
                        return False
                    try:
                        dt_val = dt[:10]
                        return start_date <= dt_val <= end_date
                    except Exception:
                        return False
                leads = [lead for lead in leads if in_range(lead)]
            # Get all unique sources for this CRE's leads
            sources = sorted(list(set(lead.get('source', 'Unknown') for lead in leads)))
            # Pagination
            total_leads = len(leads)
            total_pages = (total_leads + per_page - 1) // per_page
            start_idx = (page - 1) * per_page
            end_idx = start_idx + per_page
            leads = leads[start_idx:end_idx]
        else:
            if search_uid:
                # Search by UID substring across all leads
                all_leads = safe_get_data('lead_master')
                leads = [lead for lead in all_leads if search_uid.lower() in str(lead.get('uid', '')).lower()]
                sources = sorted(list(set(lead.get('source', 'Unknown') for lead in leads)))
                total_leads = len(leads)
                total_pages = 1
                page = 1
            else:
                total_pages = 1
                page = 1
        # Add formatted CRE TAT for display
        def format_cre_tat(tat):
            try:
                tat = float(tat)
            except (TypeError, ValueError):
                return 'N/A'
            if tat < 60:
                return f"{int(tat)}s"
            elif tat < 3600:
                m = int(tat // 60)
                s = int(tat % 60)
                return f"{m}m {s}s"
            elif tat < 86400:
                h = int(tat // 3600)
                m = int((tat % 3600) // 60)
                s = int(tat % 60)
                return f"{h}h {m}m {s}s"
            else:
                d = int(tat // 86400)
                h = int((tat % 86400) // 3600)
                return f"{d} Days {h}h"
        for lead in leads:
            lead['cre_tat_display'] = format_cre_tat(lead.get('tat'))
        return render_template('manage_leads.html', cres=cres, selected_cre=selected_cre, leads=leads, sources=sources, selected_source=source, qualification=qualification, date_filter=date_filter, start_date=start_date, end_date=end_date, page=page, total_pages=total_pages, total_leads=total_leads, final_status=final_status)
    except Exception as e:
        flash(f'Error loading leads: {str(e)}', 'error')
        return render_template('manage_leads.html', cres=[], selected_cre=None, leads=[], sources=[], selected_source=None, qualification='all', date_filter='all', start_date=None, end_date=None, page=1, total_pages=1, total_leads=0, final_status='')


@app.route('/delete_leads', methods=['POST'])
@require_admin
def delete_leads():
    try:
        delete_type = request.form.get('delete_type')

        if delete_type == 'single':
            uid = request.form.get('uid')
            if not uid:
                return jsonify({'success': False, 'message': 'No UID provided'})

            # Delete from ps_followup_master first (foreign key constraint)
            supabase.table('ps_followup_master').delete().eq('lead_uid', uid).execute()

            # Delete from lead_master
            result = supabase.table('lead_master').delete().eq('uid', uid).execute()

            # Log deletion
            auth_manager.log_audit_event(
                user_id=session.get('user_id'),
                user_type=session.get('user_type'),
                action='LEAD_DELETED',
                resource='lead_master',
                resource_id=uid,
                details={'delete_type': 'single'}
            )

            return jsonify({'success': True, 'message': 'Lead deleted successfully'})

        elif delete_type == 'bulk':
            uids = request.form.getlist('uids')
            if not uids:
                return jsonify({'success': False, 'message': 'No leads selected'})

            # Delete from ps_followup_master first
            for uid in uids:
                supabase.table('ps_followup_master').delete().eq('lead_uid', uid).execute()

            # Delete from lead_master
            for uid in uids:
                supabase.table('lead_master').delete().eq('uid', uid).execute()

            # Log bulk deletion
            auth_manager.log_audit_event(
                user_id=session.get('user_id'),
                user_type=session.get('user_type'),
                action='LEADS_BULK_DELETED',
                resource='lead_master',
                details={'delete_type': 'bulk', 'count': len(uids), 'uids': uids}
            )

            return jsonify({'success': True, 'message': f'{len(uids)} leads deleted successfully'})

        else:
            return jsonify({'success': False, 'message': 'Invalid delete type'})

    except Exception as e:
        print(f"Error deleting leads: {e}")
        return jsonify({'success': False, 'message': f'Error deleting leads: {str(e)}'})


@app.route('/bulk_unassign_leads', methods=['POST'])
@require_admin
def bulk_unassign_leads():
    try:
        uids = request.form.getlist('uids')
        if not uids:
            return jsonify({'success': False, 'message': 'No leads selected'})

        # Update leads to unassigned
        for uid in uids:
            supabase.table('lead_master').update({
                'cre_name': None,
                'assigned': 'No',
                'ps_name': None
            }).eq('uid', uid).execute()

            # Also remove from PS followup if exists
            supabase.table('ps_followup_master').delete().eq('lead_uid', uid).execute()

        # Log bulk unassignment
        auth_manager.log_audit_event(
            user_id=session.get('user_id'),
            user_type=session.get('user_type'),
            action='LEADS_BULK_UNASSIGNED',
            resource='lead_master',
            details={'count': len(uids), 'uids': uids}
        )

        return jsonify({'success': True, 'message': f'{len(uids)} leads unassigned successfully'})

    except Exception as e:
        print(f"Error unassigning leads: {e}")
        return jsonify({'success': False, 'message': f'Error unassigning leads: {str(e)}'})


@app.route('/delete_cre/<int:cre_id>')
@require_admin
def delete_cre(cre_id):
    try:
        print(f"[DEBUG] delete_cre function called with cre_id: {cre_id}")
        
        # Get the CRE details first
        cre_result = supabase.table('cre_users').select('*').eq('id', cre_id).execute()
        print(f"[DEBUG] CRE query result: {cre_result}")
        
        if not cre_result.data:
            print(f"[DEBUG] No CRE found with id: {cre_id}")
            flash('CRE not found', 'error')
            return redirect(url_for('manage_cre'))
        
        cre = cre_result.data[0]
        cre_name = cre.get('name')
        print(f"[DEBUG] Found CRE: {cre_name}")
        
        print(f"[DEBUG] Checking pending leads for CRE: {cre_name}")
        
        # Check if CRE has any pending leads in lead_master (only leads with completed qualifying calls)
        try:
            pending_leads_result = supabase.table('lead_master').select('id').eq('cre_name', cre_name).eq('final_status', 'Pending').not_.is_('first_call_date', 'null').execute()
            pending_count = len(pending_leads_result.data) if pending_leads_result.data else 0
            print(f"[DEBUG] lead_master pending count (qualifying call completed): {pending_count}")
        except Exception as e:
            print(f"[DEBUG] Error checking lead_master: {str(e)}")
            pending_count = 0
        
        # Check if CRE has any pending leads in ps_followup_master
        try:
            ps_pending_result = supabase.table('ps_followup_master').select('id').eq('cre_name', cre_name).eq('final_status', 'Pending').execute()
            ps_pending_count = len(ps_pending_result.data) if ps_pending_result.data else 0
            print(f"[DEBUG] ps_followup_master pending count: {ps_pending_count}")
        except Exception as e:
            print(f"[DEBUG] Error checking ps_followup_master: {str(e)}")
            ps_pending_count = 0
        
        total_pending = pending_count + ps_pending_count
        print(f"[DEBUG] Total pending count: {total_pending}")
        
        if total_pending > 0:
            print(f"[DEBUG] Blocking deletion - {total_pending} pending leads found")
            flash(f'Cannot delete CRE {cre_name}. They have {total_pending} pending leads ({pending_count} in lead_master, {ps_pending_count} in ps_followup). Please transfer or close these leads first.', 'error')
            return redirect(url_for('manage_cre'))
        
        print(f"[DEBUG] No pending leads found, proceeding with deletion for CRE: {cre_name}")
        
        # If no pending leads, proceed with deletion
        # Update leads assigned to this CRE to unassigned
        try:
            supabase.table('lead_master').update({
                'cre_name': None,
                'assigned': 'No'
            }).eq('cre_name', cre_name).execute()
            print(f"[DEBUG] Updated lead_master")
        except Exception as e:
            print(f"[DEBUG] Error updating lead_master: {str(e)}")
        
        # Update ps_followup_master leads
        try:
            supabase.table('ps_followup_master').update({
                'cre_name': None
            }).eq('cre_name', cre_name).execute()
            print(f"[DEBUG] Updated ps_followup_master")
        except Exception as e:
            print(f"[DEBUG] Error updating ps_followup_master: {str(e)}")

        # Delete the CRE user
        try:
            supabase.table('cre_users').delete().eq('id', cre_id).execute()
            print(f"[DEBUG] Deleted CRE user from cre_users table")
        except Exception as e:
            print(f"[DEBUG] Error deleting CRE user: {str(e)}")
            raise e

        # Log CRE deletion
        try:
            auth_manager.log_audit_event(
                user_id=session.get('user_id'),
                user_type=session.get('user_type'),
                action='CRE_DELETED',
                resource='cre_users',
                resource_id=str(cre_id),
                details={'cre_name': cre_name}
            )
            print(f"[DEBUG] Logged audit event")
        except Exception as e:
            print(f"[DEBUG] Error logging audit event: {str(e)}")

        flash(f'CRE {cre_name} has been deleted successfully', 'success')
        print(f"[DEBUG] CRE deletion completed successfully")
        
    except Exception as e:
        print(f"[DEBUG] Error deleting CRE: {str(e)}")
        print(f"[DEBUG] Exception type: {type(e)}")
        import traceback
        print(f"[DEBUG] Full traceback: {traceback.format_exc()}")
        flash('Error deleting CRE', 'error')

    return redirect(url_for('manage_cre'))


@app.route('/delete_ps/<int:ps_id>')
@require_admin
def delete_ps(ps_id):
    try:
        print(f"[DEBUG] delete_ps function called with ps_id: {ps_id}")
        
        # Get the PS details first
        ps_result = supabase.table('ps_users').select('*').eq('id', ps_id).execute()
        print(f"[DEBUG] PS query result: {ps_result}")
        
        if not ps_result.data:
            print(f"[DEBUG] No PS found with id: {ps_id}")
            flash('PS not found', 'error')
            return redirect(url_for('manage_ps'))
        
        ps = ps_result.data[0]
        ps_name = ps.get('name')
        print(f"[DEBUG] Found PS: {ps_name}")
        
        print(f"[DEBUG] Checking pending leads for PS: {ps_name}")
        
        # Check if PS has any pending leads in ps_followup_master
        try:
            ps_pending_result = supabase.table('ps_followup_master').select('id').eq('ps_name', ps_name).eq('final_status', 'Pending').execute()
            ps_pending_count = len(ps_pending_result.data) if ps_pending_result.data else 0
            print(f"[DEBUG] ps_followup_master pending count: {ps_pending_count}")
        except Exception as e:
            print(f"[DEBUG] Error checking ps_followup_master: {str(e)}")
            ps_pending_count = 0
        
        # Check if PS has any pending leads in walkin_table
        try:
            walkin_pending_result = supabase.table('walkin_table').select('id').eq('ps_assigned', ps_name).eq('status', 'pending').execute()
            walkin_pending_count = len(walkin_pending_result.data) if walkin_pending_result.data else 0
            print(f"[DEBUG] walkin_table pending count: {walkin_pending_count}")
        except Exception as e:
            print(f"[DEBUG] Error checking walkin_table: {str(e)}")
            walkin_pending_count = 0
        
        # Check if PS has any pending leads in activity_leads
        try:
            activity_pending_result = supabase.table('activity_leads').select('id').eq('ps_name', ps_name).eq('final_status', 'Pending').execute()
            activity_pending_count = len(activity_pending_result.data) if activity_pending_result.data else 0
            print(f"[DEBUG] activity_leads pending count: {activity_pending_count}")
        except Exception as e:
            print(f"[DEBUG] Error checking activity_leads: {str(e)}")
            activity_pending_count = 0
        
        total_pending = ps_pending_count + walkin_pending_count + activity_pending_count
        print(f"[DEBUG] Total pending count: {total_pending}")
        
        if total_pending > 0:
            print(f"[DEBUG] Blocking deletion - {total_pending} pending leads found")
            flash(f'Cannot delete PS {ps_name}. They have {total_pending} pending leads ({ps_pending_count} in ps_followup, {walkin_pending_count} in walkin, {activity_pending_count} in activity). Please transfer or close these leads first.', 'error')
            return redirect(url_for('manage_ps'))
        
        print(f"[DEBUG] No pending leads found, proceeding with deletion for PS: {ps_name}")
        
        # If no pending leads, proceed with deletion
        # Update leads assigned to this PS to unassigned
        try:
            supabase.table('lead_master').update({
                'ps_name': None
            }).eq('ps_name', ps_name).execute()
            print(f"[DEBUG] Updated lead_master")
        except Exception as e:
            print(f"[DEBUG] Error updating lead_master: {str(e)}")
        
        # Update ps_followup_master leads
        try:
            supabase.table('ps_followup_master').update({
                'ps_name': None
            }).eq('ps_name', ps_name).execute()
            print(f"[DEBUG] Updated ps_followup_master")
        except Exception as e:
            print(f"[DEBUG] Error updating ps_followup_master: {str(e)}")
        
        # Update walkin_table leads
        try:
            supabase.table('walkin_table').update({
                'ps_assigned': None
            }).eq('ps_assigned', ps_name).execute()
            print(f"[DEBUG] Updated walkin_table")
        except Exception as e:
            print(f"[DEBUG] Error updating walkin_table: {str(e)}")
        
        # Update activity_leads
        try:
            supabase.table('activity_leads').update({
                'ps_name': None
            }).eq('ps_name', ps_name).execute()
            print(f"[DEBUG] Updated activity_leads")
        except Exception as e:
            print(f"[DEBUG] Error updating activity_leads: {str(e)}")

        # Delete the PS user
        try:
            supabase.table('ps_users').delete().eq('id', ps_id).execute()
            print(f"[DEBUG] Deleted PS user from ps_users table")
        except Exception as e:
            print(f"[DEBUG] Error deleting PS user: {str(e)}")
            raise e

        # Log PS deletion
        try:
            auth_manager.log_audit_event(
                user_id=session.get('user_id'),
                user_type=session.get('user_type'),
                action='PS_DELETED',
                resource='ps_users',
                resource_id=str(ps_id),
                details={'ps_name': ps_name}
            )
            print(f"[DEBUG] Logged audit event")
        except Exception as e:
            print(f"[DEBUG] Error logging audit event: {str(e)}")

        flash(f'PS {ps_name} has been deleted successfully', 'success')
        print(f"[DEBUG] PS deletion completed successfully")
        
    except Exception as e:
        print(f"[DEBUG] Error deleting PS: {str(e)}")
        print(f"[DEBUG] Exception type: {type(e)}")
        import traceback
        print(f"[DEBUG] Full traceback: {traceback.format_exc()}")
        flash('Error deleting PS', 'error')

    return redirect(url_for('manage_ps'))


@app.route('/edit_cre/<int:cre_id>', methods=['GET', 'POST'])
@require_admin
def edit_cre(cre_id):
    """Edit CRE user details"""
    try:
        if request.method == 'POST':
            email = request.form.get('email')
            phone = request.form.get('phone')
            
            if not email or not phone:
                flash('Email and phone are required', 'error')
                return redirect(url_for('edit_cre', cre_id=cre_id))
            
            # Update the CRE
            supabase.table('cre_users').update({
                'email': email,
                'phone': phone
            }).eq('id', cre_id).execute()
            
            flash('CRE details updated successfully', 'success')
            return redirect(url_for('manage_cre'))
        
        # Get CRE details for editing
        cre_result = supabase.table('cre_users').select('*').eq('id', cre_id).execute()
        if not cre_result.data:
            flash('CRE not found', 'error')
            return redirect(url_for('manage_cre'))
        
        cre = cre_result.data[0]
        return render_template('edit_cre.html', cre=cre)
        
    except Exception as e:
        print(f"Error editing CRE: {str(e)}")
        flash('Error editing CRE', 'error')
        return redirect(url_for('manage_cre'))


@app.route('/edit_ps/<int:ps_id>', methods=['GET', 'POST'])
@require_admin
def edit_ps(ps_id):
    """Edit PS user details"""
    try:
        if request.method == 'POST':
            email = request.form.get('email')
            phone = request.form.get('phone')
            
            if not email or not phone:
                flash('Email and phone are required', 'error')
                return redirect(url_for('edit_ps', ps_id=ps_id))
            
            # Update the PS
            supabase.table('ps_users').update({
                'email': email,
                'phone': phone
            }).eq('id', ps_id).execute()
            
            flash('PS details updated successfully', 'success')
            return redirect(url_for('manage_ps'))
        
        # Get PS details for editing
        ps_result = supabase.table('ps_users').select('*').eq('id', ps_id).execute()
        if not ps_result.data:
            flash('PS not found', 'error')
            return redirect(url_for('manage_ps'))
        
        ps = ps_result.data[0]
        return render_template('edit_ps.html', ps=ps)
        
    except Exception as e:
        print(f"Error editing PS: {str(e)}")
        flash('Error editing PS', 'error')
        return redirect(url_for('manage_ps'))


@app.route('/manage_rec', methods=['GET', 'POST'])
@require_admin
def manage_rec():
    branches = get_system_branches()
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        name = request.form.get('name', '').strip()
        branch = request.form.get('branch', '').strip()
        if not all([username, password, name, branch]):
            flash('All fields are required', 'error')
        else:
            existing = supabase.table('rec_users').select('username').eq('username', username).execute()
            if existing.data:
                flash('Username already exists', 'error')
            else:
                password_hash = generate_password_hash(password)
                rec_data = {
                    'username': username,
                    'password_hash': password_hash,
                    'password': password,  # Store plain password as well
                    'name': name,
                    'branch': branch,
                    'is_active': True
                }
                try:
                    supabase.table('rec_users').insert(rec_data).execute()
                    flash('Receptionist user added successfully', 'success')
                except Exception as e:
                    flash(f'Error adding receptionist: {str(e)}', 'error')
    rec_users = safe_get_data('rec_users')
    return render_template('manage_rec.html', rec_users=rec_users, branches=branches)


@app.route('/delete_rec/<int:rec_id>')
@require_admin
def delete_rec(rec_id):
    try:
        supabase.table('rec_users').delete().eq('id', rec_id).execute()
        flash('Receptionist user deleted successfully', 'success')
    except Exception as e:
        flash(f'Error deleting receptionist: {str(e)}', 'error')
    return redirect(url_for('manage_rec'))


@app.route('/check_duplicate_lead', methods=['POST'])
@require_auth(['admin', 'cre'])
def check_duplicate_lead():
    """
    Check if a lead with the same phone number already exists.
    Returns duplicate information if found.
    """
    try:
        data = request.get_json()
        phone_number = data.get('phone_number', '').strip()
        source = data.get('source', '').strip()
        subsource = data.get('subsource', '').strip()
        
        if not phone_number:
            return jsonify({'success': False, 'message': 'Phone number is required'}), 400
        
        # Normalize phone number (remove all non-digits)
        normalized_phone = ''.join(filter(str.isdigit, phone_number))
        
        # Check in lead_master table
        result = supabase.table('lead_master').select('*').eq('customer_mobile_number', normalized_phone).execute()
        existing_leads = result.data or []
        
        if existing_leads:
            # Found existing lead(s)
            existing_lead = existing_leads[0]  # Get the first one
            
            # Check if this exact source-subsource combination already exists
            exact_match = any(
                lead.get('source') == source and lead.get('sub_source') == subsource 
                for lead in existing_leads
            )
            
            if exact_match:
                # This is a true duplicate - same phone, same source, same subsource
                return jsonify({
                    'success': True,
                    'is_duplicate': True,
                    'existing_lead': existing_lead,
                    'duplicate_type': 'exact_match',
                    'message': 'Lead with this phone number and source-subsource combination already exists'
                })
            else:
                # Phone exists but with different source/subsource
                # Get all existing sources for this phone number
                existing_sources = []
                for lead in existing_leads:
                    if lead.get('source') and lead.get('sub_source'):
                        existing_sources.append({
                            'source': lead.get('source'),
                            'sub_source': lead.get('sub_source')
                        })
                
                return jsonify({
                    'success': True,
                    'is_duplicate': True,
                    'existing_lead': existing_lead,
                    'existing_sources': existing_sources,
                    'duplicate_type': 'new_source',
                    'message': 'Phone number exists with different sources'
                })
        
        # Check in duplicate_leads table
        duplicate_result = supabase.table('duplicate_leads').select('*').eq('customer_mobile_number', normalized_phone).execute()
        duplicate_leads = duplicate_result.data or []
        
        if duplicate_leads:
            # Found in duplicate_leads table
            duplicate_lead = duplicate_leads[0]
            
            # Check if this exact source-subsource combination already exists in any slot
            exact_match = False
            existing_sources = []
            
            # Check all source slots (source1 to source10)
            for i in range(1, 11):
                source_field = f'source{i}'
                sub_source_field = f'sub_source{i}'
                
                if duplicate_lead.get(source_field) and duplicate_lead.get(sub_source_field):
                    existing_sources.append({
                        'source': duplicate_lead.get(source_field),
                        'sub_source': duplicate_lead.get(sub_source_field)
                    })
                    
                    # Check if this slot matches the new source/subsource
                    if (duplicate_lead.get(source_field) == source and 
                        duplicate_lead.get(sub_source_field) == subsource):
                        exact_match = True
            
            if exact_match:
                return jsonify({
                    'success': True,
                    'is_duplicate': True,
                    'existing_lead': duplicate_lead,
                    'duplicate_type': 'exact_match',
                    'message': 'Lead with this phone number and source-subsource combination already exists in duplicates'
                })
            else:
                return jsonify({
                    'success': True,
                    'is_duplicate': True,
                    'existing_lead': duplicate_lead,
                    'existing_sources': existing_sources,
                    'duplicate_type': 'new_source',
                    'message': 'Phone number exists in duplicates with different sources'
                })
        
        # No duplicate found
        return jsonify({
            'success': True,
            'is_duplicate': False,
            'message': 'No duplicate found'
        })
        
    except Exception as e:
        print(f"Error checking duplicate lead: {e}")
        return jsonify({'success': False, 'message': f'Error checking duplicate: {str(e)}'}), 500

@app.route('/check_duplicate_walkin_lead', methods=['POST'])
@require_rec
def check_duplicate_walkin_lead():
    """
    Check if a walkin lead with the same phone number already exists.
    Returns duplicate information if found.
    """
    try:
        data = request.get_json()
        phone_number = data.get('phone_number', '').strip()
        
        if not phone_number:
            return jsonify({'success': False, 'message': 'Phone number is required'}), 400
        
        # Normalize phone number (remove all non-digits)
        normalized_phone = ''.join(filter(str.isdigit, phone_number))
        
        print(f"🔍 Checking for duplicates for phone: {normalized_phone}")
        
        # 1. Check in walkin_table (most important for walkin leads)
        walkin_result = supabase.table('walkin_table').select('*').eq('mobile_number', normalized_phone).execute()
        existing_walkin_leads = walkin_result.data or []
        
        if existing_walkin_leads:
            print(f"📱 Found {len(existing_walkin_leads)} existing walkin leads")
            # Found existing walkin lead(s) - this is an exact duplicate
            existing_lead = existing_walkin_leads[0]
            return jsonify({
                'success': True,
                'is_duplicate': True,
                'existing_lead': existing_lead,
                'duplicate_type': 'exact_match',
                'message': 'Lead with this phone number already exists as Walk-in',
                'source': 'walkin_table'
            })
        
        # 2. Check in activity_leads table
        activity_result = supabase.table('activity_leads').select('*').eq('customer_phone_number', normalized_phone).execute()
        existing_activity_leads = activity_result.data or []
        
        if existing_activity_leads:
            print(f"🎯 Found {len(existing_activity_leads)} existing activity leads")
            # Found existing activity lead(s)
            existing_lead = existing_activity_leads[0]
            existing_sources = []
            for lead in existing_activity_leads:
                existing_sources.append({
                    'source': 'Event',
                    'sub_source': lead.get('location', 'N/A')
                })
            
            return jsonify({
                'success': True,
                'is_duplicate': True,
                'existing_lead': existing_lead,
                'existing_sources': existing_sources,
                'duplicate_type': 'new_source',
                'message': 'Phone number exists with Event source',
                'source': 'activity_leads'
            })
        
        # 3. Check in lead_master table
        lead_master_result = supabase.table('lead_master').select('*').eq('customer_mobile_number', normalized_phone).execute()
        existing_leads = lead_master_result.data or []
        
        if existing_leads:
            print(f"📋 Found {len(existing_leads)} existing leads in lead_master")
            # Found existing lead(s)
            existing_lead = existing_leads[0]
            
            # Check if Walk-in source already exists
            walkin_exists = any(
                lead.get('source') == 'Walk-in' 
                for lead in existing_leads
            )
            
            if walkin_exists:
                # This is a true duplicate - same phone, same source (Walk-in)
                return jsonify({
                    'success': True,
                    'is_duplicate': True,
                    'existing_lead': existing_lead,
                    'duplicate_type': 'exact_match',
                    'message': 'Lead with this phone number already exists as Walk-in',
                    'source': 'lead_master'
                })
            else:
                # Phone exists but with different source
                # Get all existing sources for this phone number
                existing_sources = []
                for lead in existing_leads:
                    if lead.get('source'):
                        existing_sources.append({
                            'source': lead.get('source'),
                            'sub_source': lead.get('sub_source')
                        })
                
                return jsonify({
                    'success': True,
                    'is_duplicate': True,
                    'existing_lead': existing_lead,
                    'existing_sources': existing_sources,
                    'duplicate_type': 'new_source',
                    'message': 'Phone number exists with different sources',
                    'source': 'lead_master'
                })
        
        # 4. Check in duplicate_leads table
        duplicate_result = supabase.table('duplicate_leads').select('*').eq('customer_mobile_number', normalized_phone).execute()
        duplicate_leads = duplicate_result.data or []
        
        if duplicate_leads:
            print(f"🔄 Found {len(duplicate_leads)} existing duplicate leads")
            # Found in duplicate_leads table
            duplicate_lead = duplicate_leads[0]
            
            # Check if Walk-in source already exists in any slot
            walkin_exists = False
            for duplicate_lead in duplicate_leads:
                # Check all source slots (source1 to source10)
                for i in range(1, 11):
                    source_field = f'source{i}'
                    
                    if duplicate_lead.get(source_field) == 'Walk-in':
                        walkin_exists = True
                        break
                if walkin_exists:
                    break
            
            if walkin_exists:
                return jsonify({
                    'success': True,
                    'is_duplicate': True,
                    'existing_lead': duplicate_lead,
                    'duplicate_type': 'exact_match',
                    'message': 'Lead with this phone number already exists as Walk-in in duplicates',
                    'source': 'duplicate_leads'
                })
            else:
                # Phone exists in duplicates but not as Walk-in
                existing_sources = []
                for duplicate_lead in duplicate_leads:
                    for i in range(1, 11):
                        source_field = f'source{i}'
                        sub_source_field = f'sub_source{i}'
                        source_value = duplicate_lead.get(source_field)
                        
                        if source_value:
                            existing_sources.append({
                                'source': source_value,
                                'sub_source': duplicate_lead.get(sub_source_field)
                            })
                
                return jsonify({
                    'success': True,
                    'is_duplicate': True,
                    'existing_lead': duplicate_lead,
                    'existing_sources': existing_sources,
                    'duplicate_type': 'new_source',
                    'message': 'Phone number exists in duplicates with different sources',
                    'source': 'duplicate_leads'
                })
        
        # No duplicate found
        print(f"✅ No duplicates found for phone: {normalized_phone}")
        return jsonify({
            'success': True,
            'is_duplicate': False,
            'message': 'No duplicate found'
        })
        
    except Exception as e:
        print(f"❌ Error checking duplicate walkin lead: {e}")
        return jsonify({'success': False, 'message': f'Error checking duplicate: {str(e)}'}), 500

@app.route('/check_duplicate_event_lead', methods=['POST'])
@require_ps
def check_duplicate_event_lead():
    """
    Check if an event lead with the same phone number already exists.
    Returns duplicate information if found.
    """
    try:
        data = request.get_json()
        phone_number = data.get('phone_number', '').strip()
        
        if not phone_number:
            return jsonify({'success': False, 'message': 'Phone number is required'}), 400
        
        # Normalize phone number (remove all non-digits)
        normalized_phone = ''.join(filter(str.isdigit, phone_number))
        
        print(f"🔍 Checking for duplicates for event lead phone: {normalized_phone}")
        
        # 1. Check in activity_leads table FIRST (most important for event leads)
        activity_result = supabase.table('activity_leads').select('*').eq('customer_phone_number', normalized_phone).execute()
        existing_activity_leads = activity_result.data or []
        
        if existing_activity_leads:
            print(f"🎯 Found {len(existing_activity_leads)} existing activity leads")
            # Found existing activity lead(s) - this is an exact duplicate
            existing_lead = existing_activity_leads[0]
            return jsonify({
                'success': True,
                'is_duplicate': True,
                'existing_lead': existing_lead,
                'duplicate_type': 'exact_match',
                'message': 'Lead with this phone number already exists as Event lead',
                'source': 'activity_leads'
            })
        
        # 2. Check in walkin_table
        walkin_result = supabase.table('walkin_table').select('*').eq('mobile_number', normalized_phone).execute()
        existing_walkin_leads = walkin_result.data or []
        
        if existing_walkin_leads:
            print(f"📱 Found {len(existing_walkin_leads)} existing walkin leads")
            # Found existing walkin lead(s)
            existing_lead = existing_walkin_leads[0]
            existing_sources = []
            for lead in existing_walkin_leads:
                existing_sources.append({
                    'source': 'Walk-in',
                    'sub_source': lead.get('branch', 'N/A')
                })
            
            return jsonify({
                'success': True,
                'is_duplicate': True,
                'existing_lead': existing_lead,
                'existing_sources': existing_sources,
                'duplicate_type': 'new_source',
                'message': 'Phone number exists with Walk-in source',
                'source': 'walkin_table'
            })
        
        # 3. Check in lead_master table
        lead_master_result = supabase.table('lead_master').select('*').eq('customer_mobile_number', normalized_phone).execute()
        existing_leads = lead_master_result.data or []
        
        if existing_leads:
            print(f"📋 Found {len(existing_leads)} existing leads in lead_master")
            # Found existing lead(s)
            existing_lead = existing_leads[0]
            
            # Check if Event source already exists
            event_exists = any(
                lead.get('source') == 'Event' 
                for lead in existing_leads
            )
            
            if event_exists:
                # This is a true duplicate - same phone, same source (Event)
                return jsonify({
                    'success': True,
                    'is_duplicate': True,
                    'existing_lead': existing_lead,
                    'duplicate_type': 'exact_match',
                    'message': 'Lead with this phone number already exists as Event',
                    'source': 'lead_master'
                })
            else:
                # Phone exists but with different source
                # Get all existing sources for this phone number
                existing_sources = []
                for lead in existing_leads:
                    if lead.get('source'):
                        existing_sources.append({
                            'source': lead.get('source'),
                            'sub_source': lead.get('sub_source')
                        })
                
                return jsonify({
                    'success': True,
                    'is_duplicate': True,
                    'existing_lead': existing_lead,
                    'existing_sources': existing_sources,
                    'duplicate_type': 'new_source',
                    'message': 'Phone number exists with different sources',
                    'source': 'lead_master'
                })
        
        # 4. Check in duplicate_leads table
        duplicate_result = supabase.table('duplicate_leads').select('*').eq('customer_mobile_number', normalized_phone).execute()
        duplicate_leads = duplicate_result.data or []
        
        if duplicate_leads:
            print(f"🔄 Found {len(duplicate_leads)} existing duplicate leads")
            # Found in duplicate_leads table
            duplicate_lead = duplicate_leads[0]
            
            # Check if Event source already exists in any slot
            event_exists = False
            for duplicate_lead in duplicate_leads:
                # Check all source slots (source1 to source10)
                for i in range(1, 11):
                    source_field = f'source{i}'
                    
                    if duplicate_lead.get(source_field) == 'Event':
                        event_exists = True
                        break
                if event_exists:
                    break
            
            if event_exists:
                return jsonify({
                    'success': True,
                    'is_duplicate': True,
                    'existing_lead': duplicate_lead,
                    'duplicate_type': 'exact_match',
                    'message': 'Lead with this phone number already exists as Event in duplicates',
                    'source': 'duplicate_leads'
                })
            else:
                # Phone exists in duplicates but not as Event
                existing_sources = []
                for duplicate_lead in duplicate_leads:
                    for i in range(1, 11):
                        source_field = f'source{i}'
                        sub_source_field = f'sub_source{i}'
                        source_value = duplicate_lead.get(source_field)
                        
                        if source_value:
                            existing_sources.append({
                                'source': source_value,
                                'sub_source': duplicate_lead.get(sub_source_field)
                            })
                
                return jsonify({
                    'success': True,
                    'is_duplicate': True,
                    'existing_lead': duplicate_lead,
                    'existing_sources': existing_sources,
                    'duplicate_type': 'new_source',
                    'message': 'Phone number exists in duplicates with different sources',
                    'source': 'duplicate_leads'
                })
        
        # No duplicate found
        print(f"✅ No duplicates found for event lead phone: {normalized_phone}")
        return jsonify({
            'success': True,
            'is_duplicate': False,
            'message': 'No duplicate found'
        })
        
    except Exception as e:
        print(f"❌ Error checking duplicate event lead: {e}")
        return jsonify({'success': False, 'message': f'Error checking duplicate: {str(e)}'}), 500

@app.route('/add_lead', methods=['GET', 'POST'])
@require_cre
def add_lead():
    from datetime import datetime, date
    branches = get_system_branches()
    ps_users = safe_get_data('ps_users')
    if request.method == 'POST':
        customer_name = request.form.get('customer_name', '').strip()
        customer_mobile_number = request.form.get('customer_mobile_number', '').strip()
        source = request.form.get('source', '').strip()
        subsource = request.form.get('subsource', '').strip()
        lead_status = request.form.get('lead_status', '').strip()
        lead_category = request.form.get('lead_category', '').strip()
        model_interested = request.form.get('model_interested', '').strip()
        branch = request.form.get('branch', '').strip()
        ps_name = request.form.get('ps_name', '').strip()
        final_status = request.form.get('final_status', 'Pending').strip()
        follow_up_date = request.form.get('follow_up_date', '').strip()
        remark = request.form.get('remark', '').strip()
        is_duplicate_new_source = request.form.get('is_duplicate_new_source', '').strip()
        date_now = datetime.now().strftime('%Y-%m-%d')
        
        # Validation
        if not customer_name or not customer_mobile_number or not source or not subsource:
            flash('Please fill all required fields', 'error')
            return render_template('add_lead.html', branches=branches, ps_users=ps_users)
        
        # Validate follow_up_date is required when final_status is Pending
        if final_status == 'Pending' and not follow_up_date:
            flash('Follow-up date is required when final status is Pending', 'error')
            return render_template('add_lead.html', branches=branches, ps_users=ps_users)
        
        # Normalize phone number
        normalized_phone = ''.join(filter(str.isdigit, customer_mobile_number))
        
        # Check for duplicates if not already confirmed as new source
        if not is_duplicate_new_source:
            try:
                # Check in lead_master
                result = supabase.table('lead_master').select('*').eq('customer_mobile_number', normalized_phone).execute()
                existing_leads = result.data or []
                
                if existing_leads:
                    # Check for exact source-subsource match
                    exact_match = any(
                        lead.get('source') == source and lead.get('sub_source') == subsource 
                        for lead in existing_leads
                    )
                    
                    if exact_match:
                        flash('Lead with this phone number and source-subsource combination already exists!', 'error')
                        return render_template('add_lead.html', branches=branches, ps_users=ps_users)
                
                # Check in duplicate_leads
                duplicate_result = supabase.table('duplicate_leads').select('*').eq('customer_mobile_number', normalized_phone).execute()
                duplicate_leads = duplicate_result.data or []
                
                if duplicate_leads:
                    # Check if this exact source-subsource combination already exists in any slot
                    exact_match = False
                    for duplicate_lead in duplicate_leads:
                        # Check all source slots (source1 to source10)
                        for i in range(1, 11):
                            source_field = f'source{i}'
                            sub_source_field = f'sub_source{i}'
                            
                            if (duplicate_lead.get(source_field) == source and 
                                duplicate_lead.get(sub_source_field) == subsource):
                                exact_match = True
                                break
                        if exact_match:
                            break
                    
                    if exact_match:
                        flash('Lead with this phone number and source-subsource combination already exists in duplicates!', 'error')
                        return render_template('add_lead.html', branches=branches, ps_users=ps_users)
                        
            except Exception as e:
                print(f"Error checking duplicates: {e}")
                flash('Error checking for duplicates. Please try again.', 'error')
                return render_template('add_lead.html', branches=branches, ps_users=ps_users)
        
        # UID: Source initial (uppercase) + '-' + first 5 letters of name (no spaces, uppercase) + last 5 digits of phone
        src_initial = source[0].upper() if source else 'X'
        name_part = ''.join(customer_name.split()).upper()[:5]
        phone_part = normalized_phone[-5:] if len(normalized_phone) >= 5 else normalized_phone
        uid = f"{src_initial}-{name_part}{phone_part}"
        
        # CRE name from session
        cre_name = session.get('cre_name')
        
        # Prepare lead data
        lead_data = {
            'uid': uid,
            'date': date_now,
            'customer_name': customer_name,
            'customer_mobile_number': normalized_phone,
            'source': source,
            'sub_source': subsource,
            'lead_status': lead_status,
            'lead_category': lead_category,
            'model_interested': model_interested,
            'branch': branch,
            'ps_name': ps_name if ps_name else None,
            'final_status': final_status,
            'follow_up_date': follow_up_date if follow_up_date else None,
            'assigned': 'Yes' if cre_name else 'No',  # Set to Yes if CRE is adding the lead
            'cre_assigned_at': datetime.now().isoformat() if cre_name else None,
            'ps_assigned_at': datetime.now().isoformat() if ps_name else None,
            'created_at': datetime.now().isoformat(),
            'updated_at': datetime.now().isoformat(),
            'first_remark': remark,
            'cre_name': cre_name,
            'first_call_date': date_now
        }
        try:
            # If this is a duplicate with new source, add to duplicate_leads table
            if is_duplicate_new_source:
                # Check if there's already a duplicate record
                existing_duplicate = supabase.table('duplicate_leads').select('*').eq('customer_mobile_number', normalized_phone).execute()
                
                if existing_duplicate.data:
                    # Add to existing duplicate record
                    duplicate_record = existing_duplicate.data[0]
                    # Find next available slot
                    next_slot = None
                    for i in range(1, 11):
                        source_field = f'source{i}'
                        if not duplicate_record.get(source_field):
                            next_slot = i
                            break
                    
                    if next_slot:
                        # Update the existing duplicate record
                        update_data = {
                            f'source{next_slot}': source,
                            f'sub_source{next_slot}': subsource,
                            f'date{next_slot}': date_now,
                            'duplicate_count': duplicate_record.get('duplicate_count', 0) + 1,
                            'updated_at': datetime.now().isoformat()
                        }
                        supabase.table('duplicate_leads').update(update_data).eq('id', duplicate_record['id']).execute()
                        flash(f'Lead added to existing duplicate record with new source: {source} - {subsource}', 'success')
                    else:
                        flash('Error: Duplicate record is full (max 10 sources reached)', 'error')
                else:
                    # Create new duplicate record
                    original_lead = supabase.table('lead_master').select('*').eq('customer_mobile_number', normalized_phone).execute()
                    if original_lead.data:
                        original = original_lead.data[0]
                        # Create duplicate record with proper structure
                        duplicate_data = {
                            'uid': uid,
                            'customer_mobile_number': normalized_phone,
                            'customer_name': customer_name,
                            'original_lead_id': original['id'],
                            'source1': original['source'],
                            'sub_source1': original.get('sub_source'),
                            'date1': original['date'],
                            'source2': source,
                            'sub_source2': subsource,
                            'date2': date_now,
                            'duplicate_count': 2,
                            'created_at': datetime.now().isoformat(),
                            'updated_at': datetime.now().isoformat()
                        }
                        supabase.table('duplicate_leads').insert(duplicate_data).execute()
                        flash(f'Lead added to duplicates with new source: {source} - {subsource}', 'success')
                    else:
                        flash('Error: Original lead not found for duplicate creation', 'error')
            else:
                supabase.table('lead_master').insert(lead_data).execute()
                
                # Track the initial call attempt for fresh leads
                if lead_status:
                    track_cre_call_attempt(
                        uid=uid,
                        cre_name=cre_name,
                        call_no='first',
                        lead_status=lead_status,
                        call_was_recorded=True,  # Fresh leads always have first_call_date recorded
                        follow_up_date=follow_up_date if follow_up_date else None,
                        remarks=remark if remark else None
                    )
                
                # Create PS followup if PS is assigned during lead creation
                if ps_name:
                    ps_user = next((ps for ps in ps_users if ps['name'] == ps_name), None)
                    if ps_user:
                        create_or_update_ps_followup(lead_data, ps_name, ps_user['branch'])
                        
                        # Send email notification to PS
                        try:
                            socketio.start_background_task(send_email_to_ps, ps_user['email'], ps_user['name'], lead_data, cre_name)
                            flash(f'Lead added successfully and assigned to {ps_name}! Email notification sent.', 'success')
                        except Exception as e:
                            print(f"Error sending email: {e}")
                            flash(f'Lead added successfully and assigned to {ps_name}! (Email notification failed)', 'warning')
                    else:
                        flash('Lead added successfully! (PS assignment failed)', 'warning')
                else:
                    flash('Lead added successfully!', 'success')
            
            return redirect(url_for('cre_dashboard'))
        except Exception as e:
            flash(f'Error adding lead: {str(e)}', 'error')
            return render_template('add_lead.html', branches=branches, ps_users=ps_users)
    return render_template('add_lead.html', branches=branches, ps_users=ps_users)

# Add import for optimized operations
from optimized_lead_operations import create_optimized_operations

@app.route('/add_lead_optimized', methods=['POST'])
@require_cre
def add_lead_optimized():
    """
    Optimized lead creation endpoint with improved performance
    """
    try:
        from datetime import datetime
        
        # Get form data
        customer_name = request.form.get('customer_name', '').strip()
        customer_mobile_number = request.form.get('customer_mobile_number', '').strip()
        source = request.form.get('source', '').strip()
        subsource = request.form.get('subsource', '').strip()
        lead_status = request.form.get('lead_status', '').strip()
        lead_category = request.form.get('lead_category', '').strip()
        model_interested = request.form.get('model_interested', '').strip()
        branch = request.form.get('branch', '').strip()
        ps_name = request.form.get('ps_name', '').strip()
        final_status = request.form.get('final_status', 'Pending').strip()
        follow_up_date = request.form.get('follow_up_date', '').strip()
        remark = request.form.get('remark', '').strip()
        date_now = datetime.now().strftime('%Y-%m-%d')
        
        # Validation
        if not customer_name or not customer_mobile_number or not source or not subsource:
            return jsonify({
                'success': False,
                'message': 'Please fill all required fields'
            })
        
        # Validate follow_up_date is required when final_status is Pending
        if final_status == 'Pending' and not follow_up_date:
            return jsonify({
                'success': False,
                'message': 'Follow-up date is required when final status is Pending'
            })
        
        # Normalize phone number
        normalized_phone = ''.join(filter(str.isdigit, customer_mobile_number))
        
        # CRE name from session
        cre_name = session.get('cre_name')
        
        # Prepare lead data
        lead_data = {
            'date': date_now,
            'customer_name': customer_name,
            'customer_mobile_number': normalized_phone,
            'source': source,
            'sub_source': subsource,
            'lead_status': lead_status,
            'lead_category': lead_category,
            'model_interested': model_interested,
            'branch': branch,
            'ps_name': ps_name if ps_name else None,
            'final_status': final_status,
            'follow_up_date': follow_up_date if follow_up_date else None,
            'assigned': 'Yes' if cre_name else 'No',
            'cre_assigned_at': datetime.now().isoformat() if cre_name else None,
            'ps_assigned_at': datetime.now().isoformat() if ps_name else None,
            'created_at': datetime.now().isoformat(),
            'updated_at': datetime.now().isoformat(),
            'first_remark': remark,
            'cre_name': cre_name,
            'first_call_date': date_now
        }
        
        # Get PS branch if PS is assigned
        ps_branch = None
        if ps_name:
            ps_users = safe_get_data('ps_users')
            ps_user = next((ps for ps in ps_users if ps['name'] == ps_name), None)
            if ps_user:
                ps_branch = ps_user['branch']
        
        # Use optimized operations
        optimized_ops = create_optimized_operations(supabase)
        result = optimized_ops.create_lead_optimized(lead_data, cre_name, ps_name, ps_branch)
        
        if result['success']:
            return jsonify({
                'success': True,
                'message': result['message'],
                'uid': result['uid'],
                'execution_time': f"{result['execution_time']:.3f}s"
            })
        else:
            return jsonify({
                'success': False,
                'message': result['message'],
                'execution_time': f"{result['execution_time']:.3f}s"
            })
            
    except Exception as e:
        print(f"Error in optimized lead creation: {e}")
        return jsonify({
            'success': False,
            'message': f'Error creating lead: {str(e)}'
        })

@app.route('/add_lead_with_cre', methods=['POST'])
@require_admin  # <-- Change to @require_cre if you want CREs to use it, or create a custom decorator for both
def add_lead_with_cre():
    """
    Add a new lead with minimal required columns.
    - Checks for duplicate by last 10 digits of phone number.
    - If duplicate, returns UID of existing lead.
    - Only fills: uid, customer_name, customer_mobile_number, source, date, assigned.
    - Source is always 'Google(Web)', UID uses 'G' as the source character.
    """
    try:
        # Check if this is an AJAX request
        is_ajax = request.headers.get('Content-Type') == 'application/json'
        
        if is_ajax:
            # Handle JSON request
            data = request.get_json()
            customer_name = data.get('customer_name', '').strip()
            customer_mobile_number = data.get('customer_mobile_number', '').strip()
            source = data.get('source', 'GOOGLE').strip()
            subsource = data.get('subsource', '').strip()
            assigned_cre_id = data.get('assigned_cre')
        else:
            # Handle form request
            customer_name = request.form.get('customer_name', '').strip()
            customer_mobile_number = request.form.get('customer_mobile_number', '').strip()
            source = request.form.get('source', 'GOOGLE').strip()
            subsource = request.form.get('subsource', '').strip()
            assigned_cre_id = request.form.get('assigned_cre')
        
        assigned = "Yes"
        date_now = datetime.now().strftime('%Y-%m-%d')

        # Validate required fields
        if not customer_name or not customer_mobile_number or not source or not subsource:
            if is_ajax:
                return jsonify({
                    'success': False,
                    'message': 'Customer name, mobile number, source, and subsource are required'
                })
            else:
                flash('Customer name, mobile number, source, and subsource are required', 'error')
                return redirect('/assign_leads')

        # Normalize phone number to last 10 digits
        mobile_digits = ''.join(filter(str.isdigit, customer_mobile_number))[-10:]
        
        if len(mobile_digits) != 10:
            if is_ajax:
                return jsonify({
                    'success': False,
                    'message': 'Invalid mobile number. Please provide a 10-digit number.'
                })
            else:
                flash('Invalid mobile number. Please provide a 10-digit number.', 'error')
                return redirect('/assign_leads')

        # Check if this is a duplicate with new source from form
        if is_ajax:
            is_duplicate_new_source = data.get('is_duplicate_new_source', '').strip() == 'true'
        else:
            is_duplicate_new_source = request.form.get('is_duplicate_new_source', '').strip() == 'true'
        
        # Check for duplicate by phone number and source-subsource combination
        existing_leads = supabase.table('lead_master').select('*').eq('customer_mobile_number', mobile_digits).execute()
        duplicate_leads = supabase.table('duplicate_leads').select('*').eq('customer_mobile_number', mobile_digits).execute()
        
        original_lead = None
        
        # Only check for duplicates if not already confirmed as duplicate with new source
        if not is_duplicate_new_source:
            # Check in lead_master table
            if existing_leads.data:
                original_lead = existing_leads.data[0]
                # Check if this exact source-subsource combination already exists
                if original_lead.get('source') == source and original_lead.get('sub_source') == subsource:
                    if is_ajax:
                        return jsonify({
                            'success': False,
                            'message': f'Lead with this phone number and source-subsource combination already exists. UID: {original_lead["uid"]}',
                            'uid': original_lead["uid"]
                        })
                    else:
                        flash(f'Lead with this phone number and source-subsource combination already exists. UID: {original_lead["uid"]}', 'error')
                        return redirect('/assign_leads')
                else:
                    # Phone exists but with different source/subsource - this is a duplicate with new source
                    is_duplicate_new_source = True
            
            # Check in duplicate_leads table
            if duplicate_leads.data:
                duplicate_lead = duplicate_leads.data[0]
                # Check if this exact source-subsource combination already exists in any slot
                exact_match = False
                for i in range(1, 11):
                    source_field = f'source{i}'
                    sub_source_field = f'sub_source{i}'
                    
                    if (duplicate_lead.get(source_field) == source and 
                        duplicate_lead.get(sub_source_field) == subsource):
                        exact_match = True
                        break
                
                if exact_match:
                    if is_ajax:
                        return jsonify({
                            'success': False,
                            'message': f'Lead with this phone number and source-subsource combination already exists in duplicates. UID: {duplicate_lead["uid"]}',
                            'uid': duplicate_lead["uid"]
                        })
                    else:
                        flash(f'Lead with this phone number and source-subsource combination already exists in duplicates. UID: {duplicate_lead["uid"]}', 'error')
                        return redirect('/assign_leads')
                else:
                    # Phone exists in duplicates but with different source/subsource
                    is_duplicate_new_source = True
        else:
            # If is_duplicate_new_source is true, we need to get the original lead
            if existing_leads.data:
                original_lead = existing_leads.data[0]

        # Generate UID using the correct function based on source
        # Map source to UID source character
        source_mapping = {
            'GOOGLE': 'Google',
            'META': 'Meta',
            'BTL': 'BTL',
            'OEM': 'OEM'
        }
        uid_source = source_mapping.get(source, 'Google')
        
        sequence = 1
        uid = generate_uid(uid_source, mobile_digits, sequence)
        # Ensure UID is unique
        while supabase.table('lead_master').select('uid').eq('uid', uid).execute().data:
            sequence += 1
            uid = generate_uid(uid_source, mobile_digits, sequence)

        # Get assigned CRE ID from form
        if is_ajax:
            assigned_cre_id = data.get('assigned_cre')
        else:
            assigned_cre_id = request.form.get('assigned_cre')
        print(f"🔍 Raw assigned_cre_id from form: '{assigned_cre_id}' (type: {type(assigned_cre_id)})")
        
        cre_name = None
        if assigned_cre_id and assigned_cre_id.strip():
            try:
                # Convert to integer if it's a string
                cre_id = int(assigned_cre_id) if isinstance(assigned_cre_id, str) else assigned_cre_id
                print(f"🔍 Looking up CRE with ID: {cre_id}")
                
                cre_data = supabase.table('cre_users').select('name').eq('id', cre_id).execute()
                if cre_data.data:
                    cre_name = cre_data.data[0]['name']
                    print(f"✅ Found CRE: {cre_name} for ID: {cre_id}")
                else:
                    print(f"❌ No CRE found for ID: {cre_id}")
                    print(f"Available CRE IDs: {[cre['id'] for cre in supabase.table('cre_users').select('id,name').execute().data]}")
            except Exception as e:
                print(f"❌ Error fetching CRE data: {e}")
                import traceback
                traceback.print_exc()
        else:
            print("❌ No assigned_cre_id provided in form or it's empty")
            print(f"All form fields: {dict(request.form)}")

        # Prepare lead data (only required columns)
        # Set assigned based on whether CRE is assigned
        assigned_status = "Yes" if cre_name else "No"
        
        lead_data = {
            'uid': uid,
            'customer_name': customer_name,
            'customer_mobile_number': mobile_digits,
            'source': source,
            'sub_source': subsource,
            'date': date_now,   
            'assigned': assigned_status,
            'final_status': 'Pending',
            'cre_name': cre_name,
            'lead_status': 'Pending',
            'lead_category': None,  # Keep as null for assign leads page
            'cre_assigned_at': datetime.now().isoformat() if cre_name else None,
            'ps_assigned_at': None,  # Will be set when PS is assigned
            'created_at': datetime.now().isoformat(),
            'updated_at': datetime.now().isoformat(),
            'ps_assigned_at': None  # Added this line to ensure ps_assigned_at is set
        }
        print('=== DEBUG INFO ===')
        print('Form data:', dict(request.form))
        print('Assigned CRE ID:', assigned_cre_id)
        print('CRE name fetched:', cre_name)
        print('Is duplicate new source:', is_duplicate_new_source)
        print('Lead data to insert:', lead_data)
        print('==================')

        # Insert lead based on whether it's a duplicate with new source
        if is_duplicate_new_source:
            print("=== DUPLICATE HANDLING ===")
            print(f"Original lead: {original_lead}")
            print(f"Duplicate leads: {duplicate_leads.data}")
            
            if original_lead:
                # Create new duplicate record
                duplicate_data = {
                    'uid': uid,
                    'customer_mobile_number': mobile_digits,
                    'customer_name': customer_name,
                    'original_lead_id': original_lead['id'],
                    'source1': original_lead['source'],
                    'sub_source1': original_lead.get('sub_source'),
                    'date1': original_lead['date'],
                    'source2': source,
                    'sub_source2': subsource,
                    'date2': date_now,
                    'duplicate_count': 2,
                    'created_at': datetime.now().isoformat(),
                    'updated_at': datetime.now().isoformat()
                }
                print(f"Creating duplicate record: {duplicate_data}")
                result = supabase.table('duplicate_leads').insert(duplicate_data).execute()
                if result.data:
                    print("Duplicate record created successfully")
                    if is_ajax:
                        return jsonify({'success': True, 'message': 'Lead added to duplicates with new source', 'uid': uid})
                    else:
                        flash('Lead added to duplicates with new source', 'success')
                        return redirect('/assign_leads')
                else:
                    print("Failed to create duplicate record")
                    if is_ajax:
                        return jsonify({'success': False, 'message': 'Failed to add duplicate lead'})
                    else:
                        flash('Failed to add duplicate lead', 'error')
                        return redirect('/assign_leads')
            elif duplicate_leads.data:
                # Add to existing duplicate record
                duplicate_record = duplicate_leads.data[0]
                # Find next available slot
                next_slot = None
                for i in range(1, 11):
                    source_field = f'source{i}'
                    if not duplicate_record.get(source_field):
                        next_slot = i
                        break
                
                if next_slot:
                    # Update the existing duplicate record
                    update_data = {
                        f'source{next_slot}': source,
                        f'sub_source{next_slot}': subsource,
                        f'date{next_slot}': date_now,
                        'duplicate_count': duplicate_record.get('duplicate_count', 0) + 1,
                        'updated_at': datetime.now().isoformat()
                    }
                    result = supabase.table('duplicate_leads').update(update_data).eq('id', duplicate_record['id']).execute()
                    if result.data:
                        if is_ajax:
                            return jsonify({'success': True, 'message': 'Lead added to existing duplicate record', 'uid': uid})
                        else:
                            flash('Lead added to existing duplicate record', 'success')
                            return redirect('/assign_leads')
                    else:
                        if is_ajax:
                            return jsonify({'success': False, 'message': 'Failed to update duplicate record'})
                        else:
                            flash('Failed to update duplicate record', 'error')
                            return redirect('/assign_leads')
                else:
                    if is_ajax:
                        return jsonify({'success': False, 'message': 'Duplicate record is full (max 10 sources reached)'})
                    else:
                        flash('Duplicate record is full (max 10 sources reached)', 'error')
                        return redirect('/assign_leads')
            else:
                # Phone exists in duplicate_leads but not in lead_master - this shouldn't happen
                if is_ajax:
                    return jsonify({'success': False, 'message': 'Error: Original lead not found for duplicate creation'})
                else:
                    flash('Error: Original lead not found for duplicate creation', 'error')
                    return redirect('/assign_leads')
        else:
            # Insert as new lead
            print("=== FRESH LEAD INSERTION ===")
            print(f"Inserting fresh lead: {lead_data}")
            result = supabase.table('lead_master').insert(lead_data).execute()
            if result.data:
                print("Fresh lead inserted successfully")
                if is_ajax:
                    return jsonify({'success': True, 'message': 'Lead added successfully', 'uid': uid})
                else:
                    flash('Lead added successfully', 'success')
                    return redirect('/assign_leads')
            else:
                print("Failed to insert fresh lead")
                if is_ajax:
                    return jsonify({'success': False, 'message': 'Failed to add lead'})
                else:
                    flash('Failed to add lead', 'error')
                    return redirect('/assign_leads')

    except Exception as e:
        print(f"Error adding lead with CRE: {e}")
        import traceback
        traceback.print_exc()
        if is_ajax:
            return jsonify({'success': False, 'message': f'Error adding lead: {str(e)}'})
        else:
            flash(f'Error adding lead: {str(e)}', 'error')
            return redirect('/assign_leads')
@app.route('/cre_dashboard')
@require_cre
def cre_dashboard():
    import time
    from datetime import datetime, date
    start_time = time.time()
    cre_name = session.get('cre_name')
    
    # Get status parameter for Won/Lost toggle
    status = request.args.get('status', 'lost')
    # Get tab and sub_tab parameters for redirection
    tab = request.args.get('tab', '')
    sub_tab = request.args.get('sub_tab', '')
    
    today = date.today()
    today_str = today.isoformat()

    # --- AUTO-INCREMENT LOGIC FOR UNATTENDED LEADS (CRE) ---
    # 1. Regular leads (lead_master)
    all_leads = safe_get_data('lead_master', {'cre_name': cre_name})
    for lead in all_leads:
        follow_up_date = lead.get('follow_up_date')
        final_status = lead.get('final_status')
        if follow_up_date and str(follow_up_date) < today_str and final_status not in ['Won', 'Lost']:
            supabase.table('lead_master').update({'follow_up_date': today_str}).eq('uid', lead.get('uid')).execute()



    # 3. Event leads (activity_leads)
    event_event_leads = safe_get_data('activity_leads', {'cre_assigned': cre_name})
    for lead in event_event_leads:
        cre_followup_date = lead.get('cre_followup_date')
        final_status = lead.get('final_status')
        if cre_followup_date and str(cre_followup_date)[:10] < today_str and final_status not in ['Won', 'Lost']:
            supabase.table('activity_leads').update({'cre_followup_date': today_str}).eq('activity_uid', lead.get('activity_uid')).execute()

    # --- Date Filter Logic ---
    filter_type = request.args.get('filter_type', 'all')
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')

    all_leads = safe_get_data('lead_master', {'cre_name': cre_name})
    
    # Apply date filtering using the 'date' column
    if filter_type == 'today':
        today_str = date.today().isoformat()
        all_leads = [lead for lead in all_leads if lead.get('date') == today_str]
    elif filter_type == 'range' and start_date_str and end_date_str:
        # Filter leads where lead['date'] falls within the range
        all_leads = [
            lead for lead in all_leads 
            if lead.get('date') and start_date_str <= lead['date'] <= end_date_str
        ]
    # For 'all' filter type, we keep all leads without date filtering

    print(f"Date filter applied - Type: {filter_type}, Leads count: {len(all_leads)}")
    if filter_type == 'range':
        print(f"Date range: {start_date_str} to {end_date_str}")

    # Fetch event leads assigned to this CRE
    event_event_leads = safe_get_data('activity_leads', {'cre_assigned': cre_name})

    print("Fetched leads for CRE:", cre_name, "Count:", len(all_leads))
    print("Status parameter from URL:", status)

    # Initialize buckets for leads for mutual exclusivity
    untouched_leads = []
    called_leads = []
    follow_up_leads = []  # New list for "Call me Back" leads
    attended_leads = []
    assigned_to_ps = []
    won_leads = []
    lost_leads = []

    non_contact_statuses = ['RNR', 'Busy on another Call', 'Call me Back', 'Call Disconnected', 'Call not Connected']

    for lead in all_leads:
        lead_status = (lead.get('lead_status') or '').strip()
        final_status = lead.get('final_status')
        has_first_call = lead.get('first_call_date') is not None

        if final_status == 'Won':
            won_leads.append(lead)
            continue
        if final_status == 'Lost':
            lost_leads.append(lead)
            continue

        # PS Assigned: any lead with a PS assigned
        if lead.get('ps_name'):
            assigned_to_ps.append(lead)

        # Pending Leads: only leads with final_status == 'Pending' AND qualifying call completed (first_call_date IS NOT NULL)
        if final_status == 'Pending' and has_first_call:
            attended_leads.append(lead)

        # Untouched Leads (Fresh leads that are still pending)
        if not has_first_call and lead_status == 'Pending':
            untouched_leads.append(lead)
            continue

        # Called Fresh Leads: Non-contact status on FIRST update only
        if lead_status in non_contact_statuses and not has_first_call:
            # Separate "Call me Back" leads into follow_up_leads
            if lead_status == 'Call me Back':
                follow_up_leads.append(lead)
            else:
                called_leads.append(lead)
            continue

    print(f"Won leads count: {len(won_leads)}")
    print(f"Lost leads count: {len(lost_leads)}")
    
    untouched_count = len(untouched_leads)
    called_count = len(called_leads)
    follow_up_count = len(follow_up_leads)
    total_fresh_leads = untouched_count + called_count + follow_up_count

    fresh_leads_sorted = sorted(
        untouched_leads + called_leads + follow_up_leads,
        key=lambda l: l.get('date') or '',  # Sort by 'date' column
        reverse=True
    )

    # Get today's followups
    today = date.today()
    today_str = today.isoformat()
    # Get all leads with follow-up date <= today, final_status = 'Pending', qualifying call completed (first_call_date IS NOT NULL), 
    # AND NOT qualifying call leads (must have follow-up call data beyond first call)
    todays_followups = []
    for lead in all_leads:
        follow_up_date = lead.get('follow_up_date')
        final_status = lead.get('final_status')
        has_first_call = lead.get('first_call_date') is not None
        
        # Check if this is NOT just a qualifying call (must have follow-up call data)
        has_followup_calls = any([
            lead.get('second_call_date'),
            lead.get('third_call_date'),
            lead.get('fourth_call_date'),
            lead.get('fifth_call_date'),
            lead.get('sixth_call_date'),
            lead.get('seventh_call_date')
        ])
        
        if follow_up_date and final_status == 'Pending' and has_first_call and has_followup_calls:
            # Parse follow_up_date and check if it's <= today
            try:
                if 'T' in str(follow_up_date):
                    followup_date_parsed = datetime.fromisoformat(str(follow_up_date).replace('Z', '+00:00')).date()
                else:
                    followup_date_parsed = datetime.strptime(str(follow_up_date)[:10], '%Y-%m-%d').date()
                
                if followup_date_parsed <= datetime.now().date():
                    # Add overdue flag for highlighting
                    lead['is_overdue'] = followup_date_parsed < datetime.now().date()
                    lead['overdue_days'] = (datetime.now().date() - followup_date_parsed).days
                    todays_followups.append(lead)
            except (ValueError, TypeError):
                # If date parsing fails, include the lead for manual review
                lead['is_overdue'] = False
                lead['overdue_days'] = 0
                todays_followups.append(lead)

    # Add event leads with today's cre_followup_date to the follow-up list
    event_leads_today = []
    for lead in event_event_leads:
        cre_followup_date = lead.get('cre_followup_date')
        if cre_followup_date and str(cre_followup_date)[:10] == today_str:
            event_lead_row = {
                'is_event_lead': True,
                'activity_uid': lead.get('activity_uid'),
                'customer_name': lead.get('customer_name'),
                'customer_phone_number': lead.get('customer_phone_number'),
                'lead_status': lead.get('lead_status'),
                'location': lead.get('location'),
                'activity_name': lead.get('activity_name'),
            }
            event_leads_today.append(event_lead_row)
    
    todays_followups.extend(event_leads_today)

    print(f"[PERF] cre_dashboard TOTAL took {time.time() - start_time:.3f} seconds")
    return render_template(
        'cre_dashboard.html',
        untouched_count=untouched_count,
        called_count=called_count,
        follow_up_count=follow_up_count,
        total_fresh_leads=total_fresh_leads,
        fresh_leads_sorted=fresh_leads_sorted,
        untouched_leads=untouched_leads,
        called_leads=called_leads,
        follow_up_leads=follow_up_leads,
        pending_leads=attended_leads,
        todays_followups=todays_followups,
        attended_leads=attended_leads,
        assigned_to_ps=assigned_to_ps,
        won_leads=won_leads,
        lost_leads=lost_leads,
        event_event_leads=event_event_leads,
        filter_type=filter_type,  # Pass filter type to template
        start_date=start_date_str,  # Pass start date to template
        end_date=end_date_str,  # Pass end date to template
        status=status,  # Pass status parameter to template
        return_tab=tab,  # Pass tab parameter to template
        return_sub_tab=sub_tab  # Pass sub_tab parameter to template
    )

@app.route('/update_lead/<uid>', methods=['GET', 'POST'])
@require_cre
def update_lead(uid):
    # Get return_tab parameter for redirection
    return_tab = request.args.get('return_tab', '')
    try:
        # Get lead data
        lead_result = supabase.table('lead_master').select('*').eq('uid', uid).execute()
        if not lead_result.data:
            flash('Lead not found', 'error')
            return redirect(url_for('cre_dashboard'))

        lead_data = lead_result.data[0]

        # Verify this lead belongs to the current CRE
        if lead_data.get('cre_name') != session.get('cre_name'):
            flash('Access denied - This lead is not assigned to you', 'error')
            return redirect(url_for('cre_dashboard'))

        # Get next call info
        next_call, completed_calls = get_next_call_info(lead_data)

        # Get PS users for branch selection
        ps_users = safe_get_data('ps_users')

        # Model options
        rizta_models = [
            'Rizta S Mono (2.9 kWh)',
            'Rizta S Super Matte (2.9 kWh)',
            'Rizta Z Mono (2.9 kWh)',
            'Rizta Z Duo (2.9 kWh)',
            'Rizta Z Super Matte (2.9 kWh)',
            'Rizta Z Mono (3.7 kWh)',
            'Rizta Z Duo (3.7 kWh)',
            'Rizta Z Super Matte (3.7 kWh)'
        ]

        x450_models = [
            '450 X (2.9 kWh)',
            '450 X (3.7 kWh)',
            '450 X (2.9 kWh) Pro Pack',
            '450 X (3.7 kWh) Pro Pack',
            '450 Apex STD'
        ]

        branches = get_system_branches()

        lead_statuses = [
            'Busy on another Call', 'RNR', 'Call me Back', 'Interested',
            'Not Interested', 'Did Not Inquire', 'Lost to Competition',
            'Lost to Co Dealer', 'Call Disconnected', 'Wrong Number'
        ]

        if request.method == 'POST':
            update_data = {}
            lead_status = request.form.get('lead_status', '')
            follow_up_date = request.form.get('follow_up_date', '')
            call_remark = request.form.get('call_remark', '')

            # Lock follow_up_date and set final_status for certain statuses
            lock_statuses = ['Booked', 'Retailed']
            lost_statuses = ['Not Interested', 'Lost to Codealer', 'Lost to Competition']
            if lead_status in lock_statuses:
                update_data['lead_status'] = lead_status
                update_data['follow_up_date'] = follow_up_date or lead_data.get('follow_up_date')
                update_data['final_status'] = 'Won' if lead_status in ['Booked', 'Retailed'] else lead_data.get('final_status')
            elif lead_status in lost_statuses:
                update_data['lead_status'] = lead_status
                update_data['follow_up_date'] = follow_up_date or lead_data.get('follow_up_date')
                update_data['final_status'] = 'Lost'
            else:
                if lead_status:
                    update_data['lead_status'] = lead_status

            if request.form.get('customer_name'):
                update_data['customer_name'] = request.form['customer_name']

            if request.form.get('source'):
                update_data['source'] = request.form['source']

            if request.form.get('lead_category'):
                update_data['lead_category'] = request.form['lead_category']

            if request.form.get('model_interested'):
                update_data['model_interested'] = request.form['model_interested']

            if request.form.get('branch'):
                update_data['branch'] = request.form['branch']

            # Handle PS assignment and email notification
            ps_name = request.form.get('ps_name')
            if ps_name and ps_name != lead_data.get('ps_name'):
                update_data['ps_name'] = ps_name
                update_data['assigned'] = 'Yes'
                update_data['ps_assigned_at'] = datetime.now().isoformat()

                # Get PS details for followup creation
                ps_user = next((ps for ps in ps_users if ps['name'] == ps_name), None)
                if ps_user:
                    # Create PS followup record
                    updated_lead_data = {**lead_data, **update_data}
                    create_or_update_ps_followup(updated_lead_data, ps_name, ps_user['branch'])

                # Send email to PS
                try:
                    if ps_user:
                        lead_data_for_email = {**lead_data, **update_data}
                        socketio.start_background_task(send_email_to_ps, ps_user['email'], ps_user['name'], lead_data_for_email, session.get('cre_name'))
                        flash(f'Lead assigned to {ps_name} and email notification sent', 'success')
                    else:
                        flash(f'Lead assigned to {ps_name}', 'success')
                except Exception as e:
                    print(f"Error sending email: {e}")
                    flash(f'Lead assigned to {ps_name} (email notification failed)', 'warning')

            if follow_up_date:
                update_data['follow_up_date'] = follow_up_date

            if request.form.get('final_status'):
                final_status = request.form['final_status']
                update_data['final_status'] = final_status
                if final_status == 'Won':
                    update_data['won_timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                elif final_status == 'Lost':
                    update_data['lost_timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            # Handle call dates and remarks - record for all statuses including RNR
            if request.form.get('call_date') and call_remark:
                combined_remark = f"{lead_status}, {call_remark}"
                # For all 7 calls, use the correct schema columns
                call_names = ['first', 'second', 'third', 'fourth', 'fifth', 'sixth', 'seventh']
                if next_call in call_names:
                    update_data[f'{next_call}_call_date'] = request.form['call_date']
                    update_data[f'{next_call}_remark'] = combined_remark
                # Notification removed - no longer sending notifications between PS and CRE

            try:
                # Track the call attempt before updating the lead - ALWAYS track regardless of status
                if lead_status:
                    # Determine if this attempt resulted in a recorded call
                    call_was_recorded = bool(request.form.get('call_date') and call_remark)
                    # Track the attempt - this ensures RNR and other statuses are properly tracked
                    track_cre_call_attempt(
                        uid=uid,
                        cre_name=session.get('cre_name'),
                        call_no=next_call,
                        lead_status=lead_status,
                        call_was_recorded=call_was_recorded,
                        follow_up_date=follow_up_date if follow_up_date else None,
                        remarks=call_remark if call_remark else None
                    )

                if update_data:
                    supabase.table('lead_master').update(update_data).eq('uid', uid).execute()

                    # Keep ps_followup_master.final_status in sync if final_status is updated and PS followup exists
                    if 'final_status' in update_data and lead_data.get('ps_name'):
                        ps_result = supabase.table('ps_followup_master').select('lead_uid').eq('lead_uid', uid).execute()
                        if ps_result.data:
                            supabase.table('ps_followup_master').update({'final_status': update_data['final_status']}).eq('lead_uid', uid).execute()

                    # Log lead update
                    auth_manager.log_audit_event(
                        user_id=session.get('user_id'),
                        user_type=session.get('user_type'),
                        action='LEAD_UPDATED',
                        resource='lead_master',
                        resource_id=uid,
                        details={'updated_fields': list(update_data.keys())}
                    )

                    flash('Lead updated successfully', 'success')
                else:
                    flash('No changes to update', 'info')
                
                # Redirect based on return_tab parameter
                if return_tab:
                    if return_tab in ['untouched-leads', 'called-leads', 'follow-up-leads']:
                        return redirect(url_for('cre_dashboard', tab='fresh-leads', sub_tab=return_tab))
                    elif return_tab == 'followups':
                        return redirect(url_for('cre_dashboard', tab='followups'))
                    elif return_tab == 'pending':
                        return redirect(url_for('cre_dashboard', tab='pending'))
                    elif return_tab == 'ps-assigned':
                        return redirect(url_for('cre_dashboard', tab='ps-assigned'))
                    elif return_tab == 'won-leads':
                        return redirect(url_for('cre_dashboard', tab='won-leads'))
                    else:
                        return redirect(url_for('cre_dashboard'))
                else:
                    return redirect(url_for('cre_dashboard'))
            except Exception as e:
                flash(f'Error updating lead: {str(e)}', 'error')

        # Fetch PS call summary from ps_followup_master
        ps_call_summary = {}
        if lead_data.get('ps_name'):
            ps_result = supabase.table('ps_followup_master').select('*').eq('lead_uid', uid).execute()
            if ps_result.data:
                ps_followup = ps_result.data[0]
                call_order = ['first', 'second', 'third']
                for call in call_order:
                    date_key = f"{call}_call_date"
                    remark_key = f"{call}_call_remark"
                    ps_call_summary[call] = {
                        "date": ps_followup.get(date_key),
                        "remark": ps_followup.get(remark_key)
                    }

        return render_template('update_lead.html',
                               lead=lead_data,
                               ps_users=ps_users,
                               rizta_models=rizta_models,
                               x450_models=x450_models,
                               branches=branches,
                               lead_statuses=lead_statuses,
                               next_call=next_call,
                               completed_calls=completed_calls,
                               today=date.today(),
                               ps_call_summary=ps_call_summary)

    except Exception as e:
        flash(f'Error loading lead: {str(e)}', 'error')
        return redirect(url_for('cre_dashboard'))
@app.route('/ps_dashboard')
@require_ps
def ps_dashboard():
    start_time = time.time()
    ps_name = session.get('ps_name')
    
    # Debug session data
    print(f"[DEBUG] PS Dashboard - Session data: {dict(session)}")
    print(f"[DEBUG] PS Dashboard - ps_name from session: {ps_name}")
    
    # Fallback to username if ps_name is not available
    if not ps_name:
        ps_name = session.get('username')
        print(f"[DEBUG] PS Dashboard - Using username as fallback: {ps_name}")
    
    # If still no name, try to get from database
    if not ps_name:
        user_id = session.get('user_id')
        if user_id:
            try:
                ps_user = supabase.table('ps_users').select('name,username').eq('id', user_id).execute()
                if ps_user.data:
                    ps_name = ps_user.data[0].get('name') or ps_user.data[0].get('username')
                    print(f"[DEBUG] PS Dashboard - Retrieved from DB: {ps_name}")
            except Exception as e:
                print(f"[DEBUG] PS Dashboard - Error fetching PS data: {e}")
    
    if not ps_name:
        flash('Error: Could not determine PS name. Please log in again.', 'error')
        return redirect(url_for('logout'))
    
    today = datetime.now().date()
    today_str = today.isoformat()

    # --- AUTO-INCREMENT LOGIC FOR UNATTENDED LEADS ---
    # 1. ps_followup_master
    assigned_leads = safe_get_data('ps_followup_master', {'ps_name': ps_name})
    for lead in assigned_leads:
        follow_up_date = lead.get('follow_up_date')
        final_status = lead.get('final_status')
        if follow_up_date and str(follow_up_date) < today_str and final_status not in ['Won', 'Lost']:
            supabase.table('ps_followup_master').update({'follow_up_date': today_str}).eq('lead_uid', lead.get('lead_uid')).execute()



    # 3. activity_leads
    event_leads = safe_get_data('activity_leads', {'ps_name': ps_name})
    for lead in event_leads:
        ps_followup_date_ts = lead.get('ps_followup_date_ts')
        final_status = lead.get('final_status')
        if ps_followup_date_ts and str(ps_followup_date_ts)[:10] < today_str and final_status not in ['Won', 'Lost']:
            if not lead.get('ps_first_call_date'):
                supabase.table('activity_leads').update({'ps_followup_date_ts': today_str}).eq('activity_uid', lead.get('activity_uid')).execute()

    # Get filter parameters from query string
    filter_type = request.args.get('filter_type', 'all')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    status_filter = request.args.get('status_filter', '')
    status = request.args.get('status', 'lost')  # <-- Add this line
    tab = request.args.get('tab', 'fresh-leads')  # <-- Add tab parameter
    category_filter = request.args.get('category_filter', '')  # <-- Add category filter parameter

    try:
        t0 = time.time()
        # Get all assigned leads for this PS
        print(f"[DEBUG] PS Dashboard - Fetching leads for PS: {ps_name}")
        assigned_leads = safe_get_data('ps_followup_master', {'ps_name': ps_name})
        print(f"[DEBUG] PS Dashboard - Assigned leads count: {len(assigned_leads) if assigned_leads else 0}")

        # Apply date filtering to assigned leads using ps_assigned_at for ps_followup_master table
        filtered_leads = filter_leads_by_date(assigned_leads, filter_type, 'ps_assigned_at')
        print(f"[DEBUG] PS Dashboard - Filtered leads count: {len(filtered_leads) if filtered_leads else 0}")

        # Initialize lists for different lead categories
        todays_followups_regular = []

        todays_followups_event = []  # Separate list for event followups
        fresh_leads = []
        # Initialize three subsections for fresh leads (similar to CRE dashboard)
        untouched_leads = []
        called_leads = []
        follow_up_leads = []
        pending_leads = []
        attended_leads = []
        won_leads = []
        lost_leads = []
        event_leads = []  # Separate list for event leads

        # Define statuses that should be excluded from Today's Follow-up and Pending Leads
        excluded_statuses = ['Lost to Codealer', 'Lost to Competition', 'Dropped', 'Booked', 'Retailed']



        print(f"[PERF] ps_dashboard: data fetching took {time.time() - t0:.3f} seconds")
        print(f"[DEBUG] Total assigned leads fetched: {len(assigned_leads)}")
        print(f"[DEBUG] Total filtered leads: {len(filtered_leads)}")
        print(f"[DEBUG] PS Name: {ps_name}")
        print(f"[DEBUG] Filter type: {filter_type}")

        t1 = time.time()
        # --- Process Regular Assigned Leads ---
        for lead in filtered_leads:
            lead_dict = dict(lead)  # Make a copy to avoid mutating the original
            lead_dict['lead_uid'] = lead.get('lead_uid')  # Ensure lead_uid for template compatibility

            final_status = lead.get('final_status')
            lead_status = lead.get('lead_status')

            # DEBUG: Print lead details for troubleshooting
            print(f"[DEBUG] Processing lead: {lead.get('lead_uid')} | final_status: {final_status} | lead_status: {lead_status} | first_call_date: {lead.get('first_call_date')} | ps_name: {lead.get('ps_name')}")

            # Categorize leads based on final_status and first_call_date
            if final_status == 'Won':
                won_leads.append(lead_dict)
            elif final_status == 'Lost':
                lost_leads.append(lead_dict)
            elif final_status == 'Pending':
                # Check if lead has been called (has first_call_date)
                if lead.get('first_call_date'):
                    # Lead has been called, goes to pending_leads
                    print(f"[DEBUG] Found lead with final_status == 'Pending' and first_call_date: {lead.get('lead_uid')} | first_call_date: {lead.get('first_call_date')} | lead_status: {lead_status}")
                    if not lead_status or lead_status not in excluded_statuses:
                        # Set default lead_category if missing
                        if 'lead_category' not in lead_dict or not lead_dict['lead_category']:
                            lead_dict['lead_category'] = 'Not Set'
                        pending_leads.append(lead_dict)
                        print(f"[DEBUG] Added to pending_leads (Pending status with first_call_date): {lead.get('lead_uid')}")
                    else:
                        print(f"[DEBUG] Lead {lead.get('lead_uid')} excluded from pending_leads due to lead_status: {lead_status}")
                else:
                    # Lead hasn't been called yet, categorize into fresh leads subsections
                    print(f"[DEBUG] Found lead with final_status == 'Pending' but no first_call_date: {lead.get('lead_uid')} | lead_status: {lead_status}")
                    if not lead_status or lead_status not in excluded_statuses:
                        # Set default lead_category if missing
                        if 'lead_category' not in lead_dict or not lead_dict['lead_category']:
                            lead_dict['lead_category'] = 'Not Set'
                        
                        # Categorize into fresh leads subsections based on lead_status
                        if lead_status == 'Call me Back':
                            follow_up_leads.append(lead_dict)
                            print(f"[DEBUG] Added to follow_up_leads (Call me Back status): {lead.get('lead_uid')}")
                        elif lead_status in ['RNR', 'Busy on another Call', 'Call Disconnected', 'Call not Connected']:
                            called_leads.append(lead_dict)
                            print(f"[DEBUG] Added to called_leads (Non-contact status): {lead.get('lead_uid')}")
                        else:
                            untouched_leads.append(lead_dict)
                            print(f"[DEBUG] Added to untouched_leads (Pending status): {lead.get('lead_uid')}")
                        
                        # Also add to fresh_leads for backward compatibility
                        fresh_leads.append(lead_dict)
                    else:
                        print(f"[DEBUG] Lead {lead.get('lead_uid')} excluded from fresh_leads due to lead_status: {lead_status}")
            elif not final_status:  # Include leads with no final_status
                if lead.get('first_call_date'):
                    # Lead has been called, goes to pending_leads
                    print(f"[DEBUG] Found lead with no final_status but has first_call_date: {lead.get('lead_uid')} | first_call_date: {lead.get('first_call_date')} | lead_status: {lead_status}")
                    if not lead_status or lead_status not in excluded_statuses:
                        # Set default lead_category if missing
                        if 'lead_category' not in lead_dict or not lead_dict['lead_category']:
                            lead_dict['lead_category'] = 'Not Set'
                        pending_leads.append(lead_dict)
                        print(f"[DEBUG] Added to pending_leads (no final_status with first_call_date): {lead.get('lead_uid')}")
                    else:
                        print(f"[DEBUG] Lead {lead.get('lead_uid')} excluded from pending_leads due to lead_status: {lead_status}")
                else:
                    # Lead hasn't been called yet, categorize into fresh leads subsections
                    print(f"[DEBUG] Found lead with no final_status and no first_call_date: {lead.get('lead_uid')} | lead_status: {lead_status}")
                    if not lead_status or lead_status not in excluded_statuses:
                        # Set default lead_category if missing
                        if 'lead_category' not in lead_dict or not lead_dict['lead_category']:
                            lead_dict['lead_category'] = 'Not Set'
                        
                        # Categorize into fresh leads subsections based on lead_status
                        if lead_status == 'Call me Back':
                            follow_up_leads.append(lead_dict)
                            print(f"[DEBUG] Added to follow_up_leads (Call me Back status): {lead.get('lead_uid')}")
                        elif lead_status in ['RNR', 'Busy on another Call', 'Call Disconnected', 'Call not Connected']:
                            called_leads.append(lead_dict)
                            print(f"[DEBUG] Added to called_leads (Non-contact status): {lead.get('lead_uid')}")
                        else:
                            untouched_leads.append(lead_dict)
                            print(f"[DEBUG] Added to untouched_leads (Pending status): {lead.get('lead_uid')}")
                        
                        # Also add to fresh_leads for backward compatibility
                        fresh_leads.append(lead_dict)
                    else:
                        print(f"[DEBUG] Lead {lead.get('lead_uid')} excluded from fresh_leads due to lead_status: {lead_status}")

            # Add to today's followups if applicable (exclude Won/Lost and specific statuses)
            # ONLY show leads where qualifying call is completed (first_call_date is NOT NULL)
            follow_up_date = lead.get('follow_up_date')
            if (follow_up_date and final_status == 'Pending' and lead.get('first_call_date') and
                (not lead_status or lead_status not in excluded_statuses)):
                # Parse follow_up_date and check if it's <= today
                try:
                    if 'T' in str(follow_up_date):
                        followup_date_parsed = datetime.fromisoformat(str(follow_up_date).replace('Z', '+00:00')).date()
                    else:
                        followup_date_parsed = datetime.strptime(str(follow_up_date)[:10], '%Y-%m-%d').date()
                    
                    if followup_date_parsed <= datetime.now().date():
                        # Add overdue flag for highlighting
                        lead_dict['is_overdue'] = followup_date_parsed < datetime.now().date()
                        lead_dict['overdue_days'] = (datetime.now().date() - followup_date_parsed).days
                        todays_followups_regular.append(lead_dict)
                        print(f"[DEBUG] Added to today's followups (qualifying call completed): {lead.get('lead_uid')}")
                except (ValueError, TypeError):
                    # If date parsing fails, include the lead for manual review
                    lead_dict['is_overdue'] = False
                    lead_dict['overdue_days'] = 0
                    todays_followups_regular.append(lead_dict)
                    print(f"[DEBUG] Added to today's followups (date parse failed, qualifying call completed): {lead.get('lead_uid')}")



        print(f"[PERF] ps_dashboard: regular leads processing took {time.time() - t1:.3f} seconds")
        print(f"[DEBUG] Fresh leads count: {len(fresh_leads)}")
        print(f"[DEBUG] Attended leads count: {len(attended_leads)}")
        print(f"[DEBUG] Won leads count: {len(won_leads)}")
        print(f"[DEBUG] Lost leads count: {len(lost_leads)}")
        print(f"[DEBUG] Today's followups (regular) count: {len(todays_followups_regular)}")
        
        # Debug: Check lead_category values in attended_leads
        print(f"[DEBUG] Attended leads with lead_category:")
        for lead in attended_leads:
            category = lead.get('lead_category')
            source = lead.get('source', 'Unknown')
            print(f"  - {lead.get('lead_uid')}: {category} (Source: {source})")
            
        # Debug: Check raw lead_category values from database
        print(f"[DEBUG] Raw lead_category values from database:")
        for lead in assigned_leads:
            category = lead.get('lead_category')
            print(f"  - {lead.get('lead_uid')}: {category} (Type: {type(category)})")





        t3 = time.time()
        # Fetch event leads for this PS
        fetched_event_leads = safe_get_data('activity_leads', {'ps_name': ps_name})
        
        # Fetch walk-in leads for this PS
        print(f"[DEBUG] PS Dashboard - Fetching walk-in leads for PS: {ps_name}")
        walkin_leads = safe_get_data('walkin_table', {'ps_assigned': ps_name})
        print(f"[DEBUG] PS Dashboard - Walk-in leads count: {len(walkin_leads) if walkin_leads else 0}")
        if walkin_leads:
            print(f"[DEBUG] Walk-in leads details:")
            for lead in walkin_leads:
                print(f"  - ID: {lead.get('id')}, UID: {lead.get('uid')}, Status: {lead.get('status')}, First Call: {lead.get('first_call_date')}")



        # --- Process Event Leads ---
        print(f"[DEBUG] Processing {len(fetched_event_leads)} event leads")
        for lead in fetched_event_leads:
            lead_dict = dict(lead)  # Make a copy to avoid mutating the original
            lead_dict['lead_uid'] = lead.get('activity_uid') or lead.get('uid')
            lead_dict['customer_mobile_number'] = lead.get('customer_phone_number')  # For template compatibility
            
            # For event leads, use 'final_status' instead of 'ps_final_status'
            final_status = lead.get('final_status')
            lead_status = lead.get('lead_status')
            ps_first_call_date = lead.get('ps_first_call_date')
            
            print(f"[DEBUG] Event lead: {lead_dict['lead_uid']} | final_status: {final_status} | lead_status: {lead_status} | ps_first_call_date: {ps_first_call_date}")

            # Always add event leads to event_leads list first
            if 'lead_category' not in lead_dict or not lead_dict['lead_category']:
                lead_dict['lead_category'] = 'Not Set'
            event_leads.append(lead_dict)
            print(f"[DEBUG] Event lead {lead_dict['lead_uid']} added to event_leads list")
            
            # Also categorize event leads for other tabs based on final_status
            if final_status == 'Won':
                won_leads.append(lead_dict)
                print(f"[DEBUG] Event lead {lead_dict['lead_uid']} also added to won_leads")
            elif final_status == 'Lost':
                lost_leads.append(lead_dict)
                print(f"[DEBUG] Event lead {lead_dict['lead_uid']} also added to lost_leads")
            elif final_status == 'Pending':
                # Check if event lead has been called (has ps_first_call_date)
                if ps_first_call_date:
                    # Event lead has been called, goes to pending_leads
                    print(f"[DEBUG] Event lead with final_status == 'Pending' and ps_first_call_date: {lead_dict['lead_uid']} | ps_first_call_date: {ps_first_call_date} | lead_status: {lead_status}")
                    if not lead_status or lead_status not in excluded_statuses:
                        pending_leads.append(lead_dict)
                        print(f"[DEBUG] Event lead {lead_dict['lead_uid']} also added to pending_leads (Pending status with ps_first_call_date)")
                    else:
                        print(f"[DEBUG] Event lead {lead_dict['lead_uid']} excluded from pending_leads due to lead_status: {lead_status}")
                else:
                    # Event lead hasn't been called yet, categorize into fresh leads subsections
                    print(f"[DEBUG] Event lead with final_status == 'Pending' but no ps_first_call_date: {lead_dict['lead_uid']} | lead_status: {lead_status}")
                    if not lead_status or lead_status not in excluded_statuses:
                        # Categorize into fresh leads subsections based on lead_status
                        if lead_status == 'Call me Back':
                            follow_up_leads.append(lead_dict)
                            print(f"[DEBUG] Event lead {lead_dict['lead_uid']} added to follow_up_leads (Call me Back status)")
                        elif lead_status in ['RNR', 'Busy on another Call', 'Call Disconnected', 'Call not Connected']:
                            called_leads.append(lead_dict)
                            print(f"[DEBUG] Event lead {lead_dict['lead_uid']} added to called_leads (Non-contact status)")
                        else:
                            untouched_leads.append(lead_dict)
                            print(f"[DEBUG] Event lead {lead_dict['lead_uid']} added to untouched_leads (Pending status)")
                        
                        # Also add to fresh_leads for backward compatibility
                        fresh_leads.append(lead_dict)
                    else:
                        print(f"[DEBUG] Event lead {lead_dict['lead_uid']} excluded from fresh_leads due to lead_status: {lead_status}")
            elif not final_status:  # Include leads with no final_status
                if ps_first_call_date:
                    # Event lead has been called, goes to pending_leads
                    print(f"[DEBUG] Event lead with no final_status but has ps_first_call_date: {lead_dict['lead_uid']} | ps_first_call_date: {ps_first_call_date} | lead_status: {lead_status}")
                    if not lead_status or lead_status not in excluded_statuses:
                        pending_leads.append(lead_dict)
                        print(f"[DEBUG] Event lead {lead_dict['lead_uid']} also added to pending_leads (no final_status with ps_first_call_date)")
                    else:
                        print(f"[DEBUG] Event lead {lead_dict['lead_uid']} excluded from pending_leads due to lead_status: {lead_status}")
                else:
                    # Event lead hasn't been called yet, categorize into fresh leads subsections
                    print(f"[DEBUG] Event lead with no final_status and no ps_first_call_date: {lead_dict['lead_uid']} | lead_status: {lead_status}")
                    if not lead_status or lead_status not in excluded_statuses:
                        # Categorize into fresh leads subsections based on lead_status
                        if lead_status == 'Call me Back':
                            follow_up_leads.append(lead_dict)
                            print(f"[DEBUG] Event lead {lead_dict['lead_uid']} added to follow_up_leads (Call me Back status)")
                        elif lead_status in ['RNR', 'Busy on another Call', 'Call Disconnected', 'Call not Connected']:
                            called_leads.append(lead_dict)
                            print(f"[DEBUG] Event lead {lead_dict['lead_uid']} added to called_leads (Non-contact status)")
                        else:
                            untouched_leads.append(lead_dict)
                            print(f"[DEBUG] Event lead {lead_dict['lead_uid']} added to untouched_leads (Pending status)")
                        
                        # Also add to fresh_leads for backward compatibility
                        fresh_leads.append(lead_dict)
                    else:
                        print(f"[DEBUG] Event lead {lead_dict['lead_uid']} excluded from fresh_leads due to lead_status: {lead_status}")

            # Add event leads with today's ps_followup_date_ts to today's followups
            # ONLY show leads where qualifying call is completed (ps_first_call_date is NOT NULL)
            ps_followup_date_ts = lead.get('ps_followup_date_ts')
            if (ps_followup_date_ts and
                str(ps_followup_date_ts)[:10] == today_str and
                final_status not in ['Won', 'Lost'] and
                lead.get('ps_first_call_date') and  # Ensure qualifying call is completed
                (not lead_status or lead_status not in excluded_statuses)):
                # Set follow_up_date for template compatibility
                lead_dict['follow_up_date'] = str(ps_followup_date_ts)
                todays_followups_event.append(lead_dict)
                print(f"[DEBUG] Event lead {lead_dict['lead_uid']} added to today's followups (qualifying call completed)")

        print(f"[PERF] ps_dashboard: event leads processing took {time.time() - t3:.3f} seconds")

        t3_5 = time.time()
        # --- Process Walk-in Leads ---
        print(f"[DEBUG] Processing {len(walkin_leads)} walk-in leads")
        todays_followups_walkin = []
        
        for lead in walkin_leads:
            lead_dict = dict(lead)  # Make a copy to avoid mutating the original
            
            # Ensure UID is properly set for walkin leads
            if lead.get('uid'):
                lead_dict['lead_uid'] = lead.get('uid')
            else:
                # Generate UID if not present (for older records)
                lead_dict['lead_uid'] = f"W{lead.get('id')}"
            
            # Also set the uid field for consistency
            lead_dict['uid'] = lead_dict['lead_uid']
            print(f"[DEBUG] Walkin lead UID: {lead_dict['lead_uid']} for ID: {lead.get('id')}")
            lead_dict['customer_mobile_number'] = lead.get('mobile_number', '')  # For template compatibility
            lead_dict['customer_name'] = lead.get('customer_name', '')  # Ensure customer_name is set
            lead_dict['source'] = 'Walk-in'  # Set source for consistency
            lead_dict['is_walkin'] = True  # Flag to identify walk-in leads
            lead_dict['walkin_id'] = lead.get('id')  # Store the walk-in ID for URL generation
            lead_dict['cre_name'] = ''  # Set empty cre_name for walk-in leads
            lead_dict['created_at'] = lead.get('created_at', '')  # Add created_at for template compatibility
            lead_dict['lead_category'] = lead.get('lead_category', 'Not Set')  # Add lead_category for template compatibility
            
            final_status = lead.get('status')
            lead_status = lead.get('lead_status')  # Add lead_status for walk-in leads
            followup_no = lead.get('followup_no', 1)
            next_followup_date = lead.get('next_followup_date')
            
            print(f"[DEBUG] Walk-in lead: {lead_dict['lead_uid']} | status: {final_status} | lead_status: {lead_status} | followup_no: {followup_no} | next_followup_date: {next_followup_date}")

            # Categorize walk-in leads based on status
            if final_status == 'Won':
                # Add won timestamp for filtering
                lead_dict['won_timestamp'] = lead.get('updated_at') or datetime.now().isoformat()
                won_leads.append(lead_dict)
                print(f"[DEBUG] Walk-in lead {lead_dict['lead_uid']} added to won_leads")
            elif final_status == 'Lost':
                # Add lost timestamp for filtering
                lead_dict['lost_timestamp'] = lead.get('updated_at') or datetime.now().isoformat()
                lost_leads.append(lead_dict)
                print(f"[DEBUG] Walk-in lead {lead_dict['lead_uid']} added to lost_leads")
            elif final_status == 'Pending' or not final_status:
                # Add to today's followups if next_followup_date is <= today
                # ONLY show leads where qualifying call is completed (first_call_date is NOT NULL)
                if next_followup_date and lead.get('first_call_date'):
                    try:
                        if 'T' in str(next_followup_date):
                            followup_date_parsed = datetime.fromisoformat(str(next_followup_date).replace('Z', '+00:00')).date()
                        else:
                            followup_date_parsed = datetime.strptime(str(next_followup_date)[:10], '%Y-%m-%d').date()
                        
                        if followup_date_parsed <= datetime.now().date():
                            # Add overdue flag for highlighting
                            lead_dict['is_overdue'] = followup_date_parsed < datetime.now().date()
                            lead_dict['overdue_days'] = (datetime.now().date() - followup_date_parsed).days
                            lead_dict['follow_up_date'] = str(next_followup_date)
                            todays_followups_walkin.append(lead_dict)
                            print(f"[DEBUG] Walk-in lead {lead_dict['lead_uid']} added to today's followups (qualifying call completed)")
                    except (ValueError, TypeError):
                        # If date parsing fails, include the lead for manual review
                        lead_dict['is_overdue'] = False
                        lead_dict['overdue_days'] = 0
                        lead_dict['follow_up_date'] = str(next_followup_date)
                        todays_followups_walkin.append(lead_dict)
                        print(f"[DEBUG] Walk-in lead {lead_dict['lead_uid']} added to today's followups (date parse failed, qualifying call completed)")
                
                # Add to pending leads if no first call has been made yet
                first_call_date = lead.get('first_call_date')
                # Get lead_status for walk-in leads (use status field or default to None)
                lead_status = lead.get('status') or lead.get('lead_status')
                print(f"[DEBUG] Walk-in lead {lead_dict['lead_uid']} - first_call_date: {first_call_date}, lead_status: {lead_status}")
                
                # For walk-in leads, check if any call dates exist
                has_any_call = any(lead.get(f'{i}_call_date') for i in range(1, 8))
                print(f"[DEBUG] Walk-in lead {lead_dict['lead_uid']} - has_any_call: {has_any_call}")
                
                if not has_any_call and (not lead_status or lead_status not in excluded_statuses):
                    pending_leads.append(lead_dict)
                    print(f"[DEBUG] Walk-in lead {lead_dict['lead_uid']} added to pending_leads")
                else:
                    # Add to attended leads if any call has been made
                    attended_leads.append(lead_dict)
                    print(f"[DEBUG] Walk-in lead {lead_dict['lead_uid']} added to attended_leads")

        # Update the walkin_leads list with processed data for template
        processed_walkin_leads = []
        for lead in walkin_leads:
            lead_dict = dict(lead)  # Make a copy to avoid mutating the original
            
            # Ensure UID is properly set for walkin leads
            if lead.get('uid'):
                lead_dict['lead_uid'] = lead.get('uid')
            else:
                # Generate UID if not present (for older records)
                lead_dict['lead_uid'] = f"W{lead.get('id')}"
            
            # Also set the uid field for consistency
            lead_dict['uid'] = lead_dict['lead_uid']
            lead_dict['customer_mobile_number'] = lead.get('mobile_number', '')
            lead_dict['customer_name'] = lead.get('customer_name', '')
            lead_dict['source'] = 'Walk-in'
            lead_dict['is_walkin'] = True
            lead_dict['walkin_id'] = lead.get('id')
            lead_dict['cre_name'] = ''
            lead_dict['created_at'] = lead.get('created_at', '')
            lead_dict['lead_category'] = lead.get('lead_category', 'Not Set')
            
            processed_walkin_leads.append(lead_dict)
        
        # Replace the original walkin_leads with processed data
        walkin_leads = processed_walkin_leads
        
        print(f"[PERF] ps_dashboard: walk-in leads processing took {time.time() - t3_5:.3f} seconds")

        t4 = time.time()
        # Apply date filtering to lost leads using lost_timestamp
        def filter_lost_by_timestamp(leads_list):
            if filter_type == 'all':
                return leads_list
            elif filter_type == 'today':
                return [lead for lead in leads_list if lead.get('lost_timestamp') and str(lead.get('lost_timestamp')).startswith(today_str)]
            elif filter_type == 'range' and start_date and end_date:
                filtered = []
                for lead in leads_list:
                    ts = lead.get('lost_timestamp')
                    if ts:
                        try:
                            date_val = ts[:10]
                            if start_date <= date_val <= end_date:
                                filtered.append(lead)
                        except Exception:
                            continue
                return filtered
            else:
                return leads_list

        # Apply timestamp filtering to lost leads
        lost_leads = filter_lost_by_timestamp(lost_leads)

        print(f"[PERF] ps_dashboard: lost leads timestamp filtering took {time.time() - t4:.3f} seconds")

        t5 = time.time()
        # Merge today's followup lists
        todays_followups = todays_followups_regular + todays_followups_event + todays_followups_walkin

        print(f"[PERF] ps_dashboard: final processing took {time.time() - t5:.3f} seconds")

        t6 = time.time()
        # Render template with all lead categories and the new status variable
        result = render_template('ps_dashboard.html',
                               assigned_leads=assigned_leads,
                               fresh_leads=fresh_leads,
                               pending_leads=pending_leads,
                               todays_followups=todays_followups,
                               attended_leads=attended_leads,
                               won_leads=won_leads,
                               lost_leads=lost_leads,
                               event_leads=event_leads,
                               walkin_leads=walkin_leads,
                               filter_type=filter_type,
                               start_date=start_date,
                               end_date=end_date,
                               status_filter=status_filter,
                               status=status)  # <-- Add status here

        print(f"[PERF] ps_dashboard: render_template took {time.time() - t6:.3f} seconds")
        # Calculate counts for fresh leads subsections
        untouched_count = len(untouched_leads)
        called_count = len(called_leads)
        follow_up_count = len(follow_up_leads)
        total_fresh_leads = untouched_count + called_count + follow_up_count
        
        print(f"[DEBUG] FINAL COUNTS - Fresh: {len(fresh_leads)}, Pending: {len(pending_leads)}, Attended: {len(attended_leads)}, Won: {len(won_leads)}, Lost: {len(lost_leads)}, Event: {len(event_leads)}")
        print(f"[DEBUG] Fresh leads subsections - Untouched: {untouched_count}, Called: {called_count}, Follow Up: {follow_up_count}")
        print(f"[DEBUG] Fresh leads being sent to template: {[lead.get('lead_uid') for lead in fresh_leads]}")
        print(f"[DEBUG] Pending leads being sent to template: {[lead.get('lead_uid') for lead in pending_leads]}")
        print(f"[DEBUG] Won leads being sent to template: {[lead.get('lead_uid') for lead in won_leads]}")
        print(f"[DEBUG] Lost leads being sent to template: {[lead.get('lead_uid') for lead in lost_leads]}")
        print(f"[DEBUG] Event leads being sent to template: {[lead.get('lead_uid') for lead in event_leads]}")
        print(f"[DEBUG] Walk-in leads count: {len(walkin_leads) if walkin_leads else 0}")
        print(f"[DEBUG] Today's followups count: {len(todays_followups)}")
        print(f"[PERF] ps_dashboard TOTAL took {time.time() - start_time:.3f} seconds")
        
        # Add session data to template context for debugging
        template_data = {
            'assigned_leads': assigned_leads,
            'fresh_leads': fresh_leads,
            'untouched_leads': untouched_leads,
            'called_leads': called_leads,
            'follow_up_leads': follow_up_leads,
            'untouched_count': untouched_count,
            'called_count': called_count,
            'follow_up_count': follow_up_count,
            'total_fresh_leads': total_fresh_leads,
            'pending_leads': pending_leads,
            'todays_followups': todays_followups,
            'attended_leads': attended_leads,
            'won_leads': won_leads,
            'lost_leads': lost_leads,
            'event_leads': event_leads,
            'walkin_leads': walkin_leads,
            'filter_type': filter_type,
            'start_date': start_date,
            'end_date': end_date,
            'status_filter': status_filter,
            'status': status,
            'tab': tab,  # Add tab parameter
            'category_filter': category_filter,  # Add category filter parameter
            'ps_name': ps_name,  # Pass ps_name directly to template
            'debug_info': {
                'session_data': dict(session),
                'ps_name': ps_name,
                'total_leads': len(assigned_leads) if assigned_leads else 0
            }
        }
        
        result = render_template('ps_dashboard.html', **template_data)
        return result

    except Exception as e:
        print(f"[PERF] ps_dashboard failed after {time.time() - start_time:.3f} seconds: {e}")
        flash(f'Error loading dashboard: {str(e)}', 'error')
        return render_template('ps_dashboard.html',
                             assigned_leads=[],
                             fresh_leads=[],
                             pending_leads=[],
                             todays_followups=[],
                             attended_leads=[],
                             won_leads=[],
                             lost_leads=[],
                             event_leads=[],
                             walkin_leads=[],
                             filter_type=filter_type,
                             start_date=start_date,
                             end_date=end_date,
                             status_filter=status_filter,
                             status=status,
                             ps_name=ps_name if 'ps_name' in locals() else 'Unknown',
                             debug_info={
                                 'session_data': dict(session),
                                 'ps_name': ps_name if 'ps_name' in locals() else 'Unknown',
                                 'error': str(e)
                             })



@app.route('/update_walkin_lead/<int:walkin_id>', methods=['GET', 'POST'])
@require_ps
def update_walkin_lead(walkin_id):
    return_tab = request.args.get('return_tab', 'walkin-leads')
    try:
        # Get the walk-in lead data
        walkin_result = supabase.table('walkin_table').select('*').eq('id', walkin_id).execute()
        if not walkin_result.data:
            flash('Walk-in lead not found', 'error')
            return redirect(url_for('ps_dashboard'))
        
        walkin_lead = walkin_result.data[0]
        ps_name = session.get('ps_name')
        
        # Check if this PS is assigned to this walk-in lead
        if walkin_lead.get('ps_assigned') != ps_name:
            flash('You are not authorized to update this walk-in lead', 'error')
            return redirect(url_for('ps_dashboard'))
        
        if request.method == 'POST':
            # Get form data
            submitted_followup_no = int(request.form.get('followup_no', 1))
            call_date = request.form.get('call_date')
            call_remark = request.form.get('call_remark')
            status = request.form.get('status', 'Pending')
            lead_category = request.form.get('lead_category', 'Warm')
            lead_status = request.form.get('lead_status', '')
            model_interested = request.form.get('model_interested', '')
            test_drive_done = request.form.get('test_drive_done', '')
            next_followup_date = request.form.get('next_followup_date')
            
            # Validate that the submitted followup number is the correct next call
            next_call, completed_calls = get_next_ps_call_info(walkin_lead)
            call_order = ['first', 'second', 'third', 'fourth', 'fifth', 'sixth', 'seventh']
            expected_followup_no = call_order.index(next_call) + 1
            
            if submitted_followup_no != expected_followup_no:
                flash(f'Invalid follow-up number. Only call #{expected_followup_no} can be recorded next.', 'error')
                redirect_url = url_for('update_walkin_lead', walkin_id=walkin_id, return_tab=return_tab)
                return redirect(redirect_url)
            
            # Validate next follow-up date is required unless status is Won/Lost
            if status not in ['Won', 'Lost'] and not next_followup_date:
                flash('Next follow-up date is required for Pending status.', 'error')
                redirect_url = url_for('update_walkin_lead', walkin_id=walkin_id, return_tab=return_tab)
                return redirect(redirect_url)
            
            # Validate test_drive_done is required
            if not test_drive_done:
                flash('Test Drive Done field is required.', 'error')
                redirect_url = url_for('update_walkin_lead', walkin_id=walkin_id, return_tab=return_tab)
                return redirect(redirect_url)
            
            followup_no = submitted_followup_no
            
            # Update the specific call field based on followup number
            update_data = {
                'status': status,
                'lead_category': lead_category,
                'lead_status': lead_status,
                'model_interested': model_interested,
                'test_drive_done': test_drive_done,
                'followup_no': followup_no,
                'updated_at': datetime.now().isoformat()
            }
            
            # Set the specific call date and remark fields
            if followup_no == 1:
                update_data['first_call_date'] = call_date
                update_data['first_call_remark'] = call_remark
            elif followup_no == 2:
                update_data['second_call_date'] = call_date
                update_data['second_call_remark'] = call_remark
            elif followup_no == 3:
                update_data['third_call_date'] = call_date
                update_data['third_call_remark'] = call_remark
            elif followup_no == 4:
                update_data['fourth_call_date'] = call_date
                update_data['fourth_call_remark'] = call_remark
            elif followup_no == 5:
                update_data['fifth_call_date'] = call_date
                update_data['fifth_call_remark'] = call_remark
            elif followup_no == 6:
                update_data['sixth_call_date'] = call_date
                update_data['sixth_call_remark'] = call_remark
            elif followup_no == 7:
                update_data['seventh_call_date'] = call_date
                update_data['seventh_call_remark'] = call_remark
            
            # Set next followup date if provided
            if next_followup_date:
                update_data['next_followup_date'] = next_followup_date
            
            # Update the walk-in lead
            supabase.table('walkin_table').update(update_data).eq('id', walkin_id).execute()
            
            # Track the PS call attempt in ps_call_attempt_history
            if lead_status:
                call_was_recorded = bool(call_date and call_remark)
                track_ps_call_attempt(
                    uid=walkin_lead['uid'],
                    ps_name=ps_name,
                    call_no=next_call,
                    lead_status=lead_status,
                    call_was_recorded=call_was_recorded,
                    follow_up_date=next_followup_date if next_followup_date else None,
                    remarks=call_remark if call_remark else None
                )
            
            # Sync to alltest_drive table if test_drive_done is set
            if test_drive_done in ['Yes', 'No', True, False]:
                # Get the updated walkin lead data for syncing
                updated_walkin_result = supabase.table('walkin_table').select('*').eq('id', walkin_id).execute()
                if updated_walkin_result.data:
                    updated_walkin_data = updated_walkin_result.data[0]
                    sync_test_drive_to_alltest_drive('walkin_table', walkin_id, updated_walkin_data)
            
            flash(f'Walk-in lead updated successfully! Follow-up {followup_no} recorded.', 'success')
            redirect_url = url_for('ps_dashboard', tab=return_tab)
            return redirect(redirect_url)
        
        # For GET request, render the update form
        # Determine next call number and completed calls
        next_call, completed_calls = get_next_ps_call_info(walkin_lead)
        
        # Create a list of available call numbers for the dropdown
        # Only allow the next call number to be selected
        available_calls = []
        call_order = ['first', 'second', 'third', 'fourth', 'fifth', 'sixth', 'seventh']
        
        # Find the index of the next call
        next_call_index = call_order.index(next_call)
        
        # Only add the next call as an option
        if next_call_index < len(call_order):
            call_number = next_call_index + 1
            available_calls.append({
                'number': call_number,
                'name': next_call,
                'display': f"{call_number}{'st' if call_number == 1 else 'nd' if call_number == 2 else 'rd' if call_number == 3 else 'th'} Call"
            })
        
        models = [
            "450X", "450S", "Rizta S Mono (2.9 kWh)", "Rizta S Super Matte (2.9 kWh)",
            "Rizta Z Mono (2.9 kWh)", "Rizta Z Duo (2.9 kWh)", "Rizta Z Super Matte (2.9 kWh)",
            "Rizta Z Mono (3.7 kWh)", "Rizta Z Duo (3.7 kWh)", "Rizta Z Super Matte (3.7 kWh)",
            "450 X (2.9 kWh)", "450 X (3.7 kWh)", "450 X (2.9 kWh) Pro Pack", "450 X (3.7 kWh) Pro Pack",
            "450 Apex STD"
        ]
        return render_template(
            'update_walkin_lead.html',
            walkin_lead=walkin_lead,
            next_call=next_call,
            completed_calls=completed_calls,
            available_calls=available_calls,
            models=models
        )
        
    except Exception as e:
        flash(f'Error updating walk-in lead: {str(e)}', 'error')
        redirect_url = url_for('ps_dashboard', tab=return_tab)
        return redirect(redirect_url)
@app.route('/update_ps_lead/<uid>', methods=['GET', 'POST'])
@require_ps
def update_ps_lead(uid):
    return_tab = request.args.get('return_tab', 'fresh-leads')
    status_filter = request.args.get('status_filter', '')
    category_filter = request.args.get('category_filter', '')
    try:
        # Try to get PS followup data (regular leads)
        ps_result = supabase.table('ps_followup_master').select('*').eq('lead_uid', uid).execute()
        if ps_result.data:
            ps_data = ps_result.data[0]
            # Fetch CRE call summary from lead_master
            lead_result = supabase.table('lead_master').select('*').eq('uid', uid).execute()
            cre_call_summary = {}
            if lead_result.data:
                lead = lead_result.data[0]
                call_order = ['first', 'second', 'third', 'fourth', 'fifth', 'sixth', 'seventh']
                for call in call_order:
                    date_key = f"{call}_call_date"
                    remark_key = f"{call}_remark"
                    cre_call_summary[call] = {
                        "date": lead.get(date_key),
                        "remark": lead.get(remark_key)
                    }
            else:
                cre_call_summary = {}
            # Verify this lead belongs to the current PS
            if ps_data.get('ps_name') != session.get('ps_name'):
                flash('Access denied - This lead is not assigned to you', 'error')
                redirect_url = url_for('ps_dashboard', tab=return_tab)
                if status_filter:
                    redirect_url += f'&status_filter={status_filter}'
                return redirect(redirect_url)
            # Get next call info for PS (only 7 calls)
            next_call, completed_calls = get_next_ps_call_info(ps_data)
            if request.method == 'POST':
                update_data = {}
                lead_status = request.form.get('lead_status', '')
                follow_up_date = request.form.get('follow_up_date', '')
                call_remark = request.form.get('call_remark', '')
                lead_category = request.form.get('lead_category', '')
                model_interested = request.form.get('model_interested', '')
                test_drive_done = request.form.get('test_drive_done', '')
                
                # Always update lead_status from the form
                if lead_status:
                    update_data['lead_status'] = lead_status
                    
                    # Auto-set final_status to Won for Booked with another number
                    if lead_status == 'Booked with another number':
                        update_data['final_status'] = 'Won'
                        update_data['won_timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        
                if lead_category:
                    update_data['lead_category'] = lead_category
                if model_interested:
                    update_data['model_interested'] = model_interested
                # Test Drive Done is now required
                if not test_drive_done:
                    flash('Test Drive Done field is required', 'error')
                    return redirect(url_for('update_ps_lead', uid=uid, return_tab=return_tab))
                update_data['test_drive_done'] = test_drive_done
                if request.form.get('follow_up_date'):
                    update_data['follow_up_date'] = follow_up_date
                if request.form.get('final_status'):
                    final_status = request.form['final_status']
                    update_data['final_status'] = final_status
                    if final_status == 'Won':
                        update_data['won_timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    elif final_status == 'Lost':
                        update_data['lost_timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                # Handle call dates and remarks for the next available call
                skip_first_call_statuses = [
                    'Call not Connected',
                    'Busy on another call',
                    'RNR',
                    'Call me Back'
                ]
                if request.form.get('call_date') and lead_status not in skip_first_call_statuses:
                    update_data[f'{next_call}_call_date'] = request.form['call_date']
                if call_remark:
                    combined_remark = f"{lead_status}, {call_remark}"
                    update_data[f'{next_call}_call_remark'] = combined_remark
                    # Notification removed - no longer sending notifications between PS and CRE
                try:
                    # Track the PS call attempt before updating the lead
                    if lead_status:
                        call_was_recorded = bool(request.form.get('call_date') and call_remark)
                        track_ps_call_attempt(
                            uid=uid,
                            ps_name=session.get('ps_name'),
                            call_no=next_call,
                            lead_status=lead_status,
                            call_was_recorded=call_was_recorded,
                            follow_up_date=follow_up_date if follow_up_date else None,
                            remarks=call_remark if call_remark else None
                        )
                    if update_data:
                        supabase.table('ps_followup_master').update(update_data).eq('lead_uid', uid).execute()
                        
                        # Sync to alltest_drive table if test_drive_done is set
                        if test_drive_done in ['Yes', 'No', True, False]:
                            # Get the updated ps_followup data for syncing
                            updated_ps_result = supabase.table('ps_followup_master').select('*').eq('lead_uid', uid).execute()
                            if updated_ps_result.data:
                                updated_ps_data = updated_ps_result.data[0]
                                sync_test_drive_to_alltest_drive('ps_followup_master', uid, updated_ps_data)
                        
                        # Also update the main lead table final status
                        if request.form.get('final_status'):
                            final_status = request.form['final_status']
                            main_update_data = {'final_status': final_status}
                            if final_status == 'Won':
                                main_update_data['won_timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            elif final_status == 'Lost':
                                main_update_data['lost_timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            supabase.table('lead_master').update(main_update_data).eq('uid', uid).execute()
                        # Log PS lead update
                        auth_manager.log_audit_event(
                            user_id=session.get('user_id'),
                            user_type=session.get('user_type'),
                            action='PS_LEAD_UPDATED',
                            resource='ps_followup_master',
                            resource_id=uid,
                            details={'updated_fields': list(update_data.keys())}
                        )
                        flash('Lead updated successfully', 'success')
                    else:
                        flash('No changes to update', 'info')
                    redirect_url = url_for('ps_dashboard', tab=return_tab)
                    if status_filter:
                        redirect_url += f'&status_filter={status_filter}'
                    if category_filter:
                        redirect_url += f'&category_filter={category_filter}'
                    return redirect(redirect_url)
                except Exception as e:
                    flash(f'Error updating lead: {str(e)}', 'error')
            else:
                return render_template('update_ps_lead.html',
                                       lead=ps_data,
                                       next_call=next_call,
                                       completed_calls=completed_calls,
                                       today=date.today(),
                                       cre_call_summary=cre_call_summary)
        else:
            flash('Lead not found', 'error')
            redirect_url = url_for('ps_dashboard', tab=return_tab)
            if status_filter:
                redirect_url += f'&status_filter={status_filter}'
            return redirect(redirect_url)
    except Exception as e:
        flash(f'Error loading lead: {str(e)}', 'error')
        return redirect(url_for('ps_dashboard'))


@app.route('/logout')
@require_auth()
def logout():
    # Log logout
    auth_manager.log_audit_event(
        user_id=session.get('user_id'),
        user_type=session.get('user_type'),
        action='LOGOUT',
        details={'session_id': session.get('session_id')}
    )

    # Deactivate session
    if session.get('session_id'):
        auth_manager.deactivate_session(session.get('session_id'))

    session.clear()
    flash('You have been logged out successfully', 'info')
    return redirect(url_for('index'))

# =====================================================
# OPTIMIZED LEAD UPDATE ENDPOINTS
# =====================================================

@app.route('/update_lead_optimized/<uid>', methods=['GET', 'POST'])
@require_cre
def update_lead_optimized(uid):
    """
    Optimized lead update endpoint for CRE users
    """
    return_tab = request.args.get('return_tab', '')

    try:
        # Get lead data with optimized query
        lead_data_result = optimized_ops.get_lead_with_related_data(uid)

        if not lead_data_result or not lead_data_result['lead']:
            flash('Lead not found', 'error')
            return redirect(url_for('cre_dashboard'))

        lead_data = lead_data_result['lead']
        ps_followup = lead_data_result.get('ps_followup')

        # Verify this lead belongs to the current CRE
        if lead_data.get('cre_name') != session.get('cre_name'):
            flash('Access denied - This lead is not assigned to you', 'error')
            return redirect(url_for('cre_dashboard'))

        # Get next call info
        next_call, completed_calls = get_next_call_info(lead_data)

        # Get PS users for branch selection
        ps_users = safe_get_data('ps_users')

        # Model options and other data (same as original)
        rizta_models = [
            'Rizta S Mono (2.9 kWh)',
            'Rizta S Super Matte (2.9 kWh)',
            'Rizta Z Mono (2.9 kWh)',
            'Rizta Z Duo (2.9 kWh)',
            'Rizta Z Super Matte (2.9 kWh)',
            'Rizta Z Mono (3.7 kWh)',
            'Rizta Z Duo (3.7 kWh)',
            'Rizta Z Super Matte (3.7 kWh)'
        ]

        x450_models = [
            '450 X (2.9 kWh)',
            '450 X (3.7 kWh)',
            '450 X (2.9 kWh) Pro Pack',
            '450 X (3.7 kWh) Pro Pack',
            '450 Apex STD'
        ]

        branches = get_system_branches()

        lead_statuses = [
            'Busy on another Call', 'RNR', 'Call me Back', 'Interested',
            'Not Interested', 'Did Not Inquire', 'Lost to Competition',
            'Lost to Co Dealer', 'Call Disconnected', 'Wrong Number'
        ]

        if request.method == 'POST':
            # Prepare update data
            update_data = {}

            # Process form data (same logic as original)
            lead_status = request.form.get('lead_status', '')
            follow_up_date = request.form.get('follow_up_date', '')
            call_remark = request.form.get('call_remark', '')

            # Lock follow_up_date and set final_status for certain statuses
            lock_statuses = ['Booked', 'Retailed']
            lost_statuses = ['Not Interested', 'Lost to Codealer', 'Lost to Competition']

            if lead_status in lock_statuses:
                update_data['lead_status'] = lead_status
                update_data['follow_up_date'] = follow_up_date or lead_data.get('follow_up_date')
                update_data['final_status'] = 'Won' if lead_status in ['Booked', 'Retailed'] else lead_data.get('final_status')
            elif lead_status in lost_statuses:
                update_data['lead_status'] = lead_status
                update_data['follow_up_date'] = follow_up_date or lead_data.get('follow_up_date')
                update_data['final_status'] = 'Lost'
            else:
                if lead_status:
                    update_data['lead_status'] = lead_status

            # Process other form fields
            if request.form.get('customer_name'):
                update_data['customer_name'] = request.form['customer_name']

            if request.form.get('source'):
                update_data['source'] = request.form['source']

            if request.form.get('lead_category'):
                update_data['lead_category'] = request.form['lead_category']

            if request.form.get('model_interested'):
                update_data['model_interested'] = request.form['model_interested']

            if request.form.get('branch'):
                update_data['branch'] = request.form['branch']

            # Handle PS assignment
            ps_name = request.form.get('ps_name')
            if ps_name and ps_name != lead_data.get('ps_name'):
                update_data['ps_name'] = ps_name
                update_data['assigned'] = 'Yes'
                update_data['ps_assigned_at'] = datetime.now().isoformat()

                # Create PS followup record
                ps_user = next((ps for ps in ps_users if ps['name'] == ps_name), None)
                if ps_user:
                    create_or_update_ps_followup(lead_data, ps_name, ps_user['branch'])

                    # Send email notification (non-blocking)
                    try:
                        lead_data_for_email = {**lead_data, **update_data}
                        socketio.start_background_task(send_email_to_ps, ps_user['email'], ps_user['name'], lead_data_for_email, session.get('cre_name'))
                        flash(f'Lead assigned to {ps_name} and email notification sent', 'success')
                    except Exception as e:
                        print(f"Error sending email: {e}")
                        flash(f'Lead assigned to {ps_name} (email notification failed)', 'warning')

            if follow_up_date:
                update_data['follow_up_date'] = follow_up_date

            if request.form.get('final_status'):
                final_status = request.form['final_status']
                update_data['final_status'] = final_status
                if final_status == 'Won':
                    update_data['won_timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                elif final_status == 'Lost':
                    update_data['lost_timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            # Handle call dates and remarks
            if request.form.get('call_date') and call_remark:
                combined_remark = f"{lead_status}, {call_remark}"
                call_names = ['first', 'second', 'third', 'fourth', 'fifth', 'sixth', 'seventh']
                if next_call in call_names:
                    update_data[f'{next_call}_call_date'] = request.form['call_date']
                    update_data[f'{next_call}_remark'] = combined_remark

            # Track call attempt (non-blocking)
            if lead_status:
                call_was_recorded = bool(request.form.get('call_date') and call_remark)
                try:
                    track_cre_call_attempt(
                        uid=uid,
                        cre_name=session.get('cre_name'),
                        call_no=next_call,
                        lead_status=lead_status,
                        call_was_recorded=call_was_recorded,
                        follow_up_date=follow_up_date if follow_up_date else None,
                        remarks=call_remark if call_remark else None
                    )
                except Exception as e:
                    print(f"Error tracking call attempt: {e}")

                # Also mirror the record into ps_call_attempt_history for unified analytics
                try:
                    ps_name_for_log = None
                    try:
                        psq = supabase.table('ps_followup_master').select('ps_name').eq('lead_uid', uid).limit(1).execute()
                        if psq.data:
                            ps_name_for_log = psq.data[0].get('ps_name')
                    except Exception:
                        ps_name_for_log = None
                    if not ps_name_for_log:
                        try:
                            lm = supabase.table('lead_master').select('ps_name').eq('uid', uid).limit(1).execute()
                            if lm.data:
                                ps_name_for_log = lm.data[0].get('ps_name')
                        except Exception:
                            ps_name_for_log = None
                    track_ps_call_attempt(
                        uid=uid,
                        ps_name=ps_name_for_log,
                        call_no=next_call,
                        lead_status=lead_status,
                        call_was_recorded=call_was_recorded,
                        follow_up_date=follow_up_date if follow_up_date else None,
                        remarks=call_remark if call_remark else None
                    )
                except Exception as e:
                    print(f"Error mirroring into ps_call_attempt_history: {e}")

            # Use optimized update operation
            if update_data:
                result = optimized_ops.update_lead_optimized(
                    uid=uid,
                    update_data=update_data,
                    user_type='cre',
                    user_name=session.get('cre_name')
                )

                if result['success']:
                    flash('Lead updated successfully', 'success')
                else:
                    flash(f'Error updating lead: {result.get("error", "Unknown error")}', 'error')
            else:
                flash('No changes to update', 'info')

            # Redirect based on return_tab parameter
            if return_tab:
                if return_tab in ['untouched-leads', 'called-leads', 'follow-up-leads']:
                    return redirect(url_for('cre_dashboard', tab='fresh-leads', sub_tab=return_tab))
                elif return_tab == 'followups':
                    return redirect(url_for('cre_dashboard', tab='followups'))
                elif return_tab == 'pending':
                    return redirect(url_for('cre_dashboard', tab='pending'))
                elif return_tab == 'ps-assigned':
                    return redirect(url_for('cre_dashboard', tab='ps-assigned'))
                elif return_tab == 'won-leads':
                    return redirect(url_for('cre_dashboard', tab='won-leads'))
                else:
                    return redirect(url_for('cre_dashboard'))
            else:
                return redirect(url_for('cre_dashboard'))

        # Fetch PS call summary from ps_followup_master
        ps_call_summary = {}
        if ps_followup:
            call_order = ['first', 'second', 'third']
            for call in call_order:
                date_key = f"{call}_call_date"
                remark_key = f"{call}_call_remark"
                ps_call_summary[call] = {
                    "date": ps_followup.get(date_key),
                    "remark": ps_followup.get(remark_key)
                }

        return render_template('update_lead.html',
                               lead=lead_data,
                               ps_users=ps_users,
                               rizta_models=rizta_models,
                               x450_models=x450_models,
                               branches=branches,
                               lead_statuses=lead_statuses,
                               next_call=next_call,
                               completed_calls=completed_calls,
                               today=date.today(),
                               ps_call_summary=ps_call_summary)

    except Exception as e:
        flash(f'Error loading lead: {str(e)}', 'error')
        return redirect(url_for('cre_dashboard'))

@app.route('/update_ps_lead_optimized/<uid>', methods=['GET', 'POST'])
@require_ps
def update_ps_lead_optimized(uid):
    """
    Optimized PS lead update endpoint
    """
    return_tab = request.args.get('return_tab', 'fresh-leads')
    status_filter = request.args.get('status_filter', '')
    category_filter = request.args.get('category_filter', '')

    try:
        # Get PS followup data with optimized query
        ps_result = optimized_ops.supabase.table('ps_followup_master').select('*').eq('lead_uid', uid).execute()

        if not ps_result.data:
            flash('Lead not found', 'error')
            redirect_url = url_for('ps_dashboard', tab=return_tab)
            if status_filter:
                redirect_url += f'&status_filter={status_filter}'
            return redirect(redirect_url)

        ps_data = ps_result.data[0]

        # Verify this lead belongs to the current PS
        if ps_data.get('ps_name') != session.get('ps_name'):
            flash('Access denied - This lead is not assigned to you', 'error')
            redirect_url = url_for('ps_dashboard', tab=return_tab)
            if status_filter:
                redirect_url += f'&status_filter={status_filter}'
            return redirect(redirect_url)

        # Get next call info for PS
        next_call, completed_calls = get_next_ps_call_info(ps_data)

        # Fetch CRE call summary from lead_master
        cre_call_summary = {}
        lead_result = optimized_ops.supabase.table('lead_master').select('*').eq('uid', uid).execute()
        if lead_result.data:
            lead = lead_result.data[0]
            call_order = ['first', 'second', 'third', 'fourth', 'fifth', 'sixth', 'seventh']
            for call in call_order:
                date_key = f"{call}_call_date"
                remark_key = f"{call}_remark"
                cre_call_summary[call] = {
                    "date": lead.get(date_key),
                    "remark": lead.get(remark_key)
                }

        if request.method == 'POST':
            # Prepare update data
            update_data = {}

            # Process form data
            lead_status = request.form.get('lead_status', '')
            follow_up_date = request.form.get('follow_up_date', '')
            call_remark = request.form.get('call_remark', '')
            lead_category = request.form.get('lead_category', '')
            model_interested = request.form.get('model_interested', '')

            # Always update lead_status from the form
            if lead_status:
                update_data['lead_status'] = lead_status
                
                # Auto-set final_status to Won for Booked with another number
                if lead_status == 'Booked with another number':
                    update_data['final_status'] = 'Won'
                    update_data['won_timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    
            if lead_category:
                update_data['lead_category'] = lead_category
            if model_interested:
                update_data['model_interested'] = model_interested
            if request.form.get('follow_up_date'):
                update_data['follow_up_date'] = follow_up_date
            if request.form.get('final_status'):
                final_status = request.form['final_status']
                update_data['final_status'] = final_status
                if final_status == 'Won':
                    update_data['won_timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                elif final_status == 'Lost':
                    update_data['lost_timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            # Handle call dates and remarks for the next available call
            skip_first_call_statuses = [
                'Call not Connected',
                'Busy on another call',
                'RNR',
                'Call me Back'
            ]

            if request.form.get('call_date') and lead_status not in skip_first_call_statuses:
                update_data[f'{next_call}_call_date'] = request.form['call_date']
            if call_remark:
                combined_remark = f"{lead_status}, {call_remark}"
                update_data[f'{next_call}_call_remark'] = combined_remark

            # Track call attempt (non-blocking)
            if lead_status:
                call_was_recorded = bool(request.form.get('call_date') and call_remark)
                try:
                    track_ps_call_attempt(
                        uid=uid,
                        ps_name=session.get('ps_name'),
                        call_no=next_call,
                        lead_status=lead_status,
                        call_was_recorded=call_was_recorded,
                        follow_up_date=follow_up_date if follow_up_date else None,
                        remarks=call_remark if call_remark else None
                    )
                except Exception as e:
                    print(f"Error tracking PS call attempt: {e}")

            # Use optimized update operation
            if update_data:
                result = optimized_ops.update_ps_lead_optimized(
                    uid=uid,
                    update_data=update_data,
                    ps_name=session.get('ps_name')
                )

                if result['success']:
                    flash('Lead updated successfully', 'success')
                else:
                    flash(f'Error updating lead: {result.get("error", "Unknown error")}', 'error')
            else:
                flash('No changes to update', 'info')

            # Redirect with filters
            redirect_url = url_for('ps_dashboard', tab=return_tab)
            if status_filter:
                redirect_url += f'&status_filter={status_filter}'
            if category_filter:
                redirect_url += f'&category_filter={category_filter}'
            return redirect(redirect_url)

        return render_template('update_ps_lead.html',
                               lead=ps_data,
                               next_call=next_call,
                               completed_calls=completed_calls,
                               today=date.today(),
                               cre_call_summary=cre_call_summary)

    except Exception as e:
        flash(f'Error loading lead: {str(e)}', 'error')
        return redirect(url_for('ps_dashboard'))

@app.route('/dashboard_leads_optimized')
@require_auth(['cre', 'ps'])
def dashboard_leads_optimized():
    """
    Optimized dashboard leads endpoint for AJAX loading
    """
    try:
        user_type = session.get('user_type')
        user_name = session.get('cre_name') if user_type == 'cre' else session.get('ps_name')

        # Get filters from request
        filters = {
            'final_status': request.args.get('final_status'),
            'lead_status': request.args.get('lead_status'),
            'page': int(request.args.get('page', 1)),
            'per_page': int(request.args.get('per_page', 50))
        }

        # Add date range filter if provided
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        if start_date and end_date:
            filters['date_range'] = (start_date, end_date)

        # Get optimized dashboard data
        result = optimized_ops.get_dashboard_leads_optimized(user_type, user_name, filters)

        if 'error' in result:
            return jsonify({'error': result['error']}), 500

        return jsonify({
            'success': True,
            'leads': result['leads'],
            'total_count': result['total_count'],
            'page': result['page'],
            'per_page': result['per_page']
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/batch_update_leads', methods=['POST'])
@require_auth(['cre', 'ps'])
def batch_update_leads():
    """
    Batch update multiple leads for better performance
    """
    try:
        updates = request.json.get('updates', [])

        if not updates:
            return jsonify({'error': 'No updates provided'}), 400

        # Add user information to each update
        user_type = session.get('user_type')
        user_name = session.get('cre_name') if user_type == 'cre' else session.get('ps_name')

        for update in updates:
            update['user_type'] = user_type
            update['user_name'] = user_name

        # Perform batch update
        result = optimized_ops.batch_update_leads(updates)

        return jsonify({
            'success': True,
            'successful': result['successful'],
            'failed': result['failed'],
            'errors': result['errors']
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/performance_metrics')
@require_admin
def performance_metrics():
    """
    Get performance metrics for monitoring
    """
    try:
        # Get basic metrics
        metrics = {
            'total_leads': len(safe_get_data('lead_master')),
            'total_ps_followups': len(safe_get_data('ps_followup_master')),
            'active_cre_users': len([u for u in safe_get_data('cre_users') if u.get('is_active')]),
            'active_ps_users': len([u for u in safe_get_data('ps_users') if u.get('is_active')])
        }

        return jsonify({
            'success': True,
            'metrics': metrics,
            'timestamp': datetime.now().isoformat()
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500










def ensure_static_directories():
    """Ensure static directories exist"""
    try:
        static_dir = os.path.join(os.path.dirname(__file__), 'static')
        images_dir = os.path.join(static_dir, 'images')

        os.makedirs(static_dir, exist_ok=True)
        os.makedirs(images_dir, exist_ok=True)

        print(f"Static directories ensured: {images_dir}")
        return images_dir
    except Exception as e:
        print(f"Error creating static directories: {e}")
        return None
@app.route('/analytics')
@require_auth(['admin'])
def analytics():
    try:
        # Get filter period or custom date range from query parameter
        period = request.args.get('period', '30')
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        today_filter = request.args.get('today', 'false') == 'true'

        today = datetime.now().date()
        start_date = None
        end_date = None
        
        if today_filter:
            # Today filter - set both start and end to today
            start_date = today
            end_date = today
        elif start_date_str and end_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
                # Ensure end_date includes the full day by setting it to end of day
                end_date = end_date
            except Exception:
                start_date = None
                end_date = None
        elif period == 'all':
            start_date = None
        else:
            days = int(period) if period.isdigit() else 30
            start_date = today - timedelta(days=days)
            end_date = today

        # Get all data
        all_leads = safe_get_data('lead_master')
        all_cres = safe_get_data('cre_users')
        all_ps = safe_get_data('ps_users')
        ps_followups = safe_get_data('ps_followup_master')
        all_walkins = safe_get_data('walkin_table')  # Add walkin table data

        # Filter leads by date if needed
        filtered_leads = []
        for lead in all_leads:
            lead_date_str = lead.get('created_at') or lead.get('date')
            if lead_date_str:
                try:
                    if 'T' in lead_date_str:
                        lead_date = datetime.fromisoformat(lead_date_str.replace('Z', '+00:00')).date()
                    else:
                        lead_date = datetime.strptime(lead_date_str, '%Y-%m-%d').date()
                except (ValueError, TypeError):
                    filtered_leads.append(lead)
                    continue
                if start_date and end_date:
                    # Include both start and end dates (inclusive)
                    if start_date <= lead_date <= end_date:
                        filtered_leads.append(lead)
                elif start_date:
                    if lead_date >= start_date:
                        filtered_leads.append(lead)
                else:
                    filtered_leads.append(lead)
            else:
                filtered_leads.append(lead)
        leads = filtered_leads

        # Filter walkins by date if needed
        filtered_walkins = []
        for walkin in all_walkins:
            walkin_date_str = walkin.get('created_at')
            if walkin_date_str:
                try:
                    if 'T' in walkin_date_str:
                        walkin_date = datetime.fromisoformat(walkin_date_str.replace('Z', '+00:00')).date()
                    else:
                        walkin_date = datetime.strptime(walkin_date_str, '%Y-%m-%d').date()
                except (ValueError, TypeError):
                    filtered_walkins.append(walkin)
                    continue
                if start_date and end_date:
                    # Include both start and end dates (inclusive)
                    if start_date <= walkin_date <= end_date:
                        filtered_walkins.append(walkin)
                elif start_date:
                    if walkin_date >= start_date:
                        filtered_walkins.append(walkin)
                else:
                    filtered_walkins.append(walkin)
            else:
                filtered_walkins.append(walkin)
        walkins = filtered_walkins

        # Calculate KPIs
        total_leads = len(leads)
        won_leads = len([l for l in leads if l.get('final_status') == 'Won'])
        conversion_rate = round((won_leads / total_leads * 100) if total_leads > 0 else 0, 1)

        # Calculate walkin KPIs
        total_walkins = len(walkins)
        won_walkins = len([w for w in walkins if w.get('status') == 'Won'])
        walkin_conversion_rate = round((won_walkins / total_walkins * 100) if total_walkins > 0 else 0, 1)



        # Calculate average response time (days to first call)
        response_times = []
        for lead in leads:
            if lead.get('first_call_date') and lead.get('date'):
                try:
                    lead_date = datetime.strptime(lead['date'], '%Y-%m-%d').date()
                    call_date = datetime.strptime(lead['first_call_date'], '%Y-%m-%d').date()
                    response_times.append((call_date - lead_date).days)
                except (ValueError, TypeError):
                    continue
        avg_response_time = f"{round(sum(response_times) / len(response_times), 1)} days" if response_times else "N/A"

        # Active CREs (CREs with leads assigned)
        active_cres = len(set([l.get('cre_name') for l in leads if l.get('cre_name')]))
        total_cres = len(all_cres)

        # Source distribution
        source_counts = Counter([l.get('source', 'Unknown') for l in leads])
        source_labels = list(source_counts.keys())
        source_data = list(source_counts.values())

        # Lead trends (last 30 days)
        trend_data = []
        trend_labels = []
        for i in range(29, -1, -1):
            date = today - timedelta(days=i)
            count = len([l for l in leads if l.get('date') == str(date)])
            trend_data.append(count)
            trend_labels.append(date.strftime('%m/%d'))

        # Top performing CREs with new parameters
        cre_performance = defaultdict(lambda: {'total': 0, 'hot': 0, 'warm': 0, 'cold': 0, 'won': 0, 'lost': 0, 'calls': []})
        for lead in leads:
            cre_name = lead.get('cre_name')
            if cre_name:
                cre_performance[cre_name]['total'] += 1
                # Hot/Warm/Cold by lead_category
                category = (lead.get('lead_category') or '').strip().lower()
                if category == 'hot':
                    cre_performance[cre_name]['hot'] += 1
                elif category == 'warm':
                    cre_performance[cre_name]['warm'] += 1
                elif category == 'cold':
                    cre_performance[cre_name]['cold'] += 1
                # Won/Lost by final_status
                if (lead.get('final_status') or '').strip().lower() == 'won':
                    cre_performance[cre_name]['won'] += 1
                elif (lead.get('final_status') or '').strip().lower() == 'lost':
                    cre_performance[cre_name]['lost'] += 1
                # Count calls made
                call_count = 0
                call_fields = ['first', 'second', 'third', 'fourth', 'fifth', 'sixth', 'seventh']
                for call_field in call_fields:
                    if lead.get(f'{call_field}_call_date'):
                        call_count += 1
                cre_performance[cre_name]['calls'].append(call_count)

        # Calculate conversion rates and average calls for CREs
        top_cres = []
        for cre_name, data in cre_performance.items():
            total_calls = sum(data['calls'])
            avg_calls = round(total_calls / len(data['calls']), 1) if data['calls'] else 0
            conversion_rate_cre = round((data['won'] / data['total'] * 100) if data['total'] > 0 else 0, 1)
            top_cres.append({
                'name': cre_name,
                'total_leads': data['total'],
                'hot': data['hot'],
                'warm': data['warm'],
                'cold': data['cold'],
                'won_leads': data['won'],
                'lost_leads': data['lost'],
                'conversion_rate': conversion_rate_cre,
                'avg_calls': avg_calls
            })
        # Sort by won leads and take top 5
        top_cres = sorted(top_cres, key=lambda x: x['won_leads'], reverse=True)[:5]

        # Lead categories (add Won and Lost as categories)
        category_counts = Counter([l.get('lead_category', 'None') for l in leads])
        won_count = len([l for l in leads if (l.get('final_status') or '').strip().lower() == 'won'])
        lost_count = len([l for l in leads if (l.get('final_status') or '').strip().lower() == 'lost'])
        total_leads = len(leads)
        lead_categories = []
        # Add regular categories
        for category, count in category_counts.items():
            percentage = round((count / total_leads * 100) if total_leads > 0 else 0, 1)
            lead_categories.append({
                'name': category,
                'count': count,
                'percentage': percentage
            })
        # Add Won and Lost as categories
        if won_count > 0:
            lead_categories.append({
                'name': 'Won',
                'count': won_count,
                'percentage': round((won_count / total_leads * 100) if total_leads > 0 else 0, 1)
            })
        if lost_count > 0:
            lead_categories.append({
                'name': 'Lost',
                'count': lost_count,
                'percentage': round((lost_count / total_leads * 100) if total_leads > 0 else 0, 1)
            })

        # Model interest
        model_counts = Counter([l.get('model_interested', 'Not specified') for l in leads])
        model_interest = []
        for model, count in model_counts.items():
            percentage = round((count / total_leads * 100) if total_leads > 0 else 0, 1)
            model_interest.append({
                'name': model,
                'count': count,
                'percentage': percentage
            })

        # Branch performance
        branch_performance = []
        branches = set([ps.get('branch') for ps in all_ps if ps.get('branch')])
        for branch in branches:
            branch_ps = [ps for ps in all_ps if ps.get('branch') == branch]
            ps_names = [ps['name'] for ps in branch_ps]
            branch_leads = [l for l in leads if l.get('ps_name') in ps_names]
            branch_won = len([l for l in branch_leads if l.get('final_status') == 'Won'])
            success_rate = round((branch_won / len(branch_leads) * 100) if branch_leads else 0, 1)
            branch_performance.append({
                'name': branch,
                'ps_count': len(branch_ps),
                'assigned_leads': len(branch_leads),
                'won_leads': branch_won,
                'success_rate': success_rate
            })

        # Funnel data
        assigned_cre = len([l for l in leads if l.get('cre_name')])
        first_call = len([l for l in leads if l.get('first_call_date')])
        assigned_ps = len([l for l in leads if l.get('ps_name')])
        funnel = {
            'total': total_leads,
            'assigned_cre': assigned_cre,
            'assigned_cre_percent': round((assigned_cre / total_leads * 100) if total_leads > 0 else 0, 1),
            'first_call': first_call,
            'first_call_percent': round((first_call / total_leads * 100) if total_leads > 0 else 0, 1),
            'assigned_ps': assigned_ps,
            'assigned_ps_percent': round((assigned_ps / total_leads * 100) if total_leads > 0 else 0, 1),
            'won': won_leads,
            'won_percent': round((won_leads / total_leads * 100) if total_leads > 0 else 0, 1)
        }

        # Recent activities (mock data for now)
        recent_activities = [
            {
                'title': 'New Lead Assigned',
                'description': 'Lead assigned to CRE for follow-up',
                'time': '2 hours ago'
            },
            {
                'title': 'Lead Converted',
                'description': 'Customer purchased Rizta Z model',
                'time': '4 hours ago'
            },
            {
                'title': 'Follow-up Scheduled',
                'description': 'PS scheduled follow-up call',
                'time': '6 hours ago'
            }
        ]

        # Calculate leads growth (mock calculation)
        leads_growth = 15

        # Calculate Campaign & Platform Lead Counts
        def parse_timestamp(timestamp_str):
            """Parse timestamp string to datetime object"""
            if not timestamp_str:
                return None
            try:
                if 'T' in timestamp_str:
                    return datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))
                else:
                    return datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S')
            except (ValueError, TypeError):
                return None

        def is_within_date_filter(target_date, start_date, end_date):
            """Check if target date is within the filter range"""
            if not target_date:
                return False
            if start_date and end_date:
                return start_date <= target_date.date() <= end_date
            elif start_date:
                return target_date.date() >= start_date
            return True

        # Get today's date for "Today's Leads" calculation
        today = datetime.now().date()
        
        # Group leads by campaign and source
        campaign_platform_data = {}
        
        for lead in all_leads:
            campaign = lead.get('campaign')
            source = lead.get('source')
            # Normalize and check for None, empty, whitespace, or 'none' (case-insensitive)
            campaign_clean = str(campaign).strip().lower() if campaign is not None else ''
            source_clean = str(source).strip().lower() if source is not None else ''
            if not campaign_clean or campaign_clean == 'none' or not source_clean or source_clean == 'none':
                continue
            key = f"{campaign}|{source}"
            
            if key not in campaign_platform_data:
                campaign_platform_data[key] = {
                    'campaign': campaign,
                    'platform': source,
                    'total_leads': 0,
                    'todays_leads': 0,
                    'lost': 0,
                    'pending': 0,
                    'won': 0
                }
            
            # Parse timestamps
            created_at = parse_timestamp(lead.get('created_at'))
            won_timestamp = parse_timestamp(lead.get('won_timestamp'))
            lost_timestamp = parse_timestamp(lead.get('lost_timestamp'))
            final_status = lead.get('final_status', '').strip()
            
            # Total Leads (respects date filter)
            if created_at and is_within_date_filter(created_at, start_date, end_date):
                campaign_platform_data[key]['total_leads'] += 1
            
            # Today's Leads (ignores date filter)
            if created_at and created_at.date() == today:
                campaign_platform_data[key]['todays_leads'] += 1
            
            # Lost leads (lost_timestamp within date filter)
            if lost_timestamp and is_within_date_filter(lost_timestamp, start_date, end_date):
                campaign_platform_data[key]['lost'] += 1
            
            # Pending leads (final_status = 'Pending', no won/lost timestamp, created_at within filter)
            if (final_status == 'Pending' and 
                not won_timestamp and 
                not lost_timestamp and 
                created_at and 
                is_within_date_filter(created_at, start_date, end_date)):
                campaign_platform_data[key]['pending'] += 1
            
            # Won leads (won_timestamp within date filter)
            if won_timestamp and is_within_date_filter(won_timestamp, start_date, end_date):
                campaign_platform_data[key]['won'] += 1
        
        # Calculate conversion rates and format data
        campaign_platform_counts = []
        for key, data in campaign_platform_data.items():
            conversion_rate = round((data['won'] / data['total_leads'] * 100) if data['total_leads'] > 0 else 0, 1)
            
            campaign_platform_counts.append({
                'campaign': data['campaign'],
                'platform': data['platform'],
                'total_leads': data['total_leads'],
                'todays_leads': data['todays_leads'],
                'lost': data['lost'],
                'pending': data['pending'],
                'won': data['won'],
                'conversion_rate': conversion_rate
            })
        
        # Sort by total leads descending
        campaign_platform_counts.sort(key=lambda x: x['total_leads'], reverse=True)

        # Create analytics object that matches template expectations
        analytics = {
            'total_leads': total_leads,
            'leads_growth': leads_growth,
            'won_leads': won_leads,  # Add won leads count
            'conversion_rate': conversion_rate,
            'total_walkins': total_walkins,  # Add walkin data
            'won_walkins': won_walkins,  # Add won walkins count
            'walkin_conversion_rate': walkin_conversion_rate,  # Add walkin conversion rate
            'avg_response_time': avg_response_time,
            'active_cres': active_cres,
            'total_cres': total_cres,
            'source_labels': source_labels,
            'source_data': source_data,
            'trend_labels': trend_labels,
            'trend_data': trend_data,
            'top_cres': top_cres,
            'lead_categories': lead_categories,
            'model_interest': model_interest,
            'branch_performance': branch_performance,
            'funnel': funnel,
            'recent_activities': recent_activities,
            'campaign_platform_counts': campaign_platform_counts,
            'all_leads_count': len(all_leads)
        }

        # Log analytics access
        auth_manager.log_audit_event(
            user_id=session.get('user_id'),
            user_type=session.get('user_type'),
            action='ANALYTICS_ACCESS',
            resource='analytics',
            details={'period': period, 'start_date': start_date_str, 'end_date': end_date_str}
        )

        return render_template('analytics.html', analytics=analytics)

    except Exception as e:
        print(f"Error in analytics: {e}")
        flash(f'Error loading analytics: {str(e)}', 'error')
        # Return empty analytics object to prevent template errors
        empty_analytics = {
            'total_leads': 0,
            'leads_growth': 0,
            'won_leads': 0,  # Add won leads count
            'conversion_rate': 0,
            'total_walkins': 0,  # Add walkin data
            'won_walkins': 0,  # Add won walkins count
            'walkin_conversion_rate': 0,  # Add walkin conversion rate
            'avg_response_time': "N/A",
            'active_cres': 0,
            'total_cres': 0,
            'source_labels': [],
            'source_data': [],
            'trend_labels': [],
            'trend_data': [],
            'top_cres': [],
            'lead_categories': [],
            'model_interest': [],
            'branch_performance': [],
            'funnel': {
                'total': 0,
                'assigned_cre': 0,
                'assigned_cre_percent': 0,
                'first_call': 0,
                'first_call_percent': 0,
                'assigned_ps': 0,
                'assigned_ps_percent': 0,
                'won': 0,
                'won_percent': 0
            },
            'recent_activities': [],
            'campaign_platform_counts': [],
            'all_leads_count': 0
        }
        return render_template('analytics.html', analytics=empty_analytics)


@app.route('/branch_head_dashboard')
def branch_head_dashboard():
    if 'branch_head_id' not in session:
        return redirect(url_for('index'))
    
    # Set user type for dashboard API access
    session['user_type'] = 'branch_head'
    
    branch = session.get('branch_head_branch')
    ps_users = supabase.table('ps_users').select('*').eq('branch', branch).execute().data or []
    
    # Calculate initial KPI counts for the dashboard
    try:
        # Fresh leads count
        fresh_leads = supabase.table('ps_followup_master').select('*').eq('ps_branch', branch).execute().data or []
        fresh_leads_count = len([
            row for row in fresh_leads
            if (not row.get('first_call_date') or str(row.get('first_call_date')).strip() == '')
                and row.get('final_status', '').strip().lower() == 'pending'
        ])
        
        # Walk-in leads count
        walkin_leads = supabase.table('walkin_table').select('*').eq('branch', branch).execute().data or []
        walkin_leads_count = len(walkin_leads)
        
        # CRE assigned leads count
        cre_assigned_leads = supabase.table('ps_followup_master').select('*').eq('ps_branch', branch).execute().data or []
        cre_assigned_leads_count = len(cre_assigned_leads)
        
        # Pending leads count (from multiple sources)
        pending_followup = supabase.table('ps_followup_master').select('*').eq('ps_branch', branch).eq('final_status', 'Pending').execute().data or []
        pending_event = supabase.table('activity_leads').select('*').eq('location', branch).eq('final_status', 'Pending').execute().data or []
        pending_walkin = supabase.table('walkin_table').select('*').eq('branch', branch).eq('status', 'Pending').execute().data or []
        pending_leads_count = len(pending_followup) + len(pending_event) + len(pending_walkin)
        
        # Today's follow-ups count
        today_str = datetime.now().strftime('%Y-%m-%d')
        # Get followups with date <= today and status = Pending
        followup_today = supabase.table('ps_followup_master').select('*').eq('ps_branch', branch).eq('final_status', 'Pending').lte('follow_up_date', today_str).execute().data or []
        event_today = supabase.table('activity_leads').select('*').eq('location', branch).eq('final_status', 'Pending').lte('ps_followup_date_ts', today_str).execute().data or []
        walkin_today = supabase.table('walkin_table').select('*').eq('branch', branch).eq('status', 'Pending').lte('next_followup_date', today_str).execute().data or []
        followup_leads_count = len(followup_today) + len(event_today) + len(walkin_today)
        
        # Event leads count
        event_leads = supabase.table('activity_leads').select('*').eq('location', branch).execute().data or []
        event_leads_count = len(event_leads)
        
        # Won leads count
        won_leads = supabase.table('ps_followup_master').select('*').eq('ps_branch', branch).eq('final_status', 'Won').execute().data or []
        won_leads_count = len(won_leads)
        
        # Lost leads count
        lost_leads = supabase.table('ps_followup_master').select('*').eq('ps_branch', branch).eq('final_status', 'Lost').execute().data or []
        lost_leads_count = len(lost_leads)
        
    except Exception as e:
        print(f"Error calculating KPI counts: {str(e)}")
        fresh_leads_count = walkin_leads_count = cre_assigned_leads_count = pending_leads_count = 0
        followup_leads_count = event_leads_count = won_leads_count = lost_leads_count = 0
    
    return render_template('branch_head_dashboard.html', 
                         ps_users=ps_users,
                         fresh_leads_count=fresh_leads_count,
                         walkin_leads_count=walkin_leads_count,
                         cre_assigned_leads_count=cre_assigned_leads_count,
                         pending_leads_count=pending_leads_count,
                         followup_leads_count=followup_leads_count,
                         event_leads_count=event_leads_count,
                         won_leads_count=won_leads_count,
                         lost_leads_count=lost_leads_count)
@app.route('/api/branch_head_dashboard_data')
def api_branch_head_dashboard_data():
    """Get data for branch head dashboard"""
    # Get query parameters
    section = request.args.get('section', 'fresh_leads')
    # Note: All filtering (date, search, etc.) is handled client-side
    # No backend filtering parameters needed

    # Get branch from session
    branch = session.get('branch_head_branch')

    # OPTIMIZED: Calculate KPI counts using efficient count queries with caching
    try:
        # Cache key for KPI counts
        cache_key = f"branch_kpi_{branch}_{datetime.now().strftime('%Y-%m-%d')}"
        
        # Check if we have cached KPI counts (cache for 5 minutes)
        cached_kpis = session.get(cache_key)
        if cached_kpis and (datetime.now() - cached_kpis.get('timestamp', datetime.now())).seconds < 300:
            fresh_leads_count = cached_kpis.get('fresh_leads_count', 0)
            followup_leads_count = cached_kpis.get('followup_leads_count', 0)
        else:
            # Fresh leads count - EXCLUDE those with today's follow-up date
            today_str = datetime.now().strftime('%Y-%m-%d')
            
            # PS fresh leads count - ONLY those with NULL follow_up_date AND final_status = 'Pending'
            ps_fresh_count = supabase.table('ps_followup_master').select('*', count='exact').eq('ps_branch', branch).eq('final_status', 'Pending').is_('follow_up_date', 'null').execute().count or 0
            
            # Walk-in fresh leads count - ONLY those with NULL next_followup_date AND status = 'Pending'
            walkin_fresh_count = supabase.table('walkin_table').select('*', count='exact').eq('branch', branch).eq('status', 'Pending').is_('next_followup_date', 'null').execute().count or 0
            
            # Event fresh leads count - ONLY those with NULL ps_followup_date_ts AND final_status = 'Pending'
            event_fresh_count = supabase.table('activity_leads').select('*', count='exact').eq('location', branch).eq('final_status', 'Pending').is_('ps_followup_date_ts', 'null').execute().count or 0
            
            fresh_leads_count = ps_fresh_count + walkin_fresh_count + event_fresh_count
            
            # Today's follow-ups count (using exact date matching like your SQL query)
            today_str = datetime.now().strftime('%Y-%m-%d')
            
            # Use the same logic as data loading for accurate counts - followup_date <= today
            followup_today_count = supabase.table('ps_followup_master').select('*', count='exact').eq('ps_branch', branch).eq('final_status', 'Pending').lte('follow_up_date', today_str).execute().count or 0
            
            # Event count with ps_followup_date_ts <= today AND final_status = 'Pending'
            event_today_count = supabase.table('activity_leads').select('*', count='exact').eq('location', branch).eq('final_status', 'Pending').lte('ps_followup_date_ts', today_str).execute().count or 0
            
            # Walk-in count with next_followup_date <= today AND status = 'Pending'
            walkin_today_count = supabase.table('walkin_table').select('*', count='exact').eq('branch', branch).eq('status', 'Pending').lte('next_followup_date', today_str).execute().count or 0
            
            followup_leads_count = followup_today_count + event_today_count + walkin_today_count
            
            # Cache the results
            session[cache_key] = {
                'fresh_leads_count': fresh_leads_count,
                'followup_leads_count': followup_leads_count,
                'timestamp': datetime.now()
            }
        
    except Exception as e:
        print(f"Error calculating KPI counts: {str(e)}")
        fresh_leads_count = followup_leads_count = 0

    # Initialize query - no backend filtering needed
    query = supabase.table('ps_followup_master').select('*').eq('ps_branch', branch)

    # OPTIMIZED: Server-side filtering function
    def apply_server_filters(query, table_type='ps'):
        """Apply server-side filters to database query for better performance"""
        if ps_name:
            if table_type == 'ps':
                query = query.eq('ps_name', ps_name)
            elif table_type == 'walkin':
                query = query.eq('ps_assigned', ps_name)
            elif table_type == 'event':
                query = query.eq('ps_name', ps_name)
        
        if final_status:
            if table_type == 'ps':
                query = query.eq('final_status', final_status)
            elif table_type == 'walkin':
                query = query.eq('status', final_status)
            elif table_type == 'event':
                query = query.eq('final_status', final_status)
        
        if source:
            if table_type == 'ps':
                # For PS table, filter by source field
                query = query.eq('source', source)
            elif table_type == 'walkin':
                # For walkin table, only include if source is 'Walk In' or no source filter
                if source != 'Walk In':
                    # Return empty result for non-walkin sources
                    return query.limit(0)
            elif table_type == 'event':
                # For event table, only include if source is 'Event' or no source filter
                if source != 'Event':
                    # Return empty result for non-event sources
                    return query.limit(0)
        
        # OPTIMIZED: Date filtering for ps_assigned_at column
        if start_date and end_date:
            if table_type == 'ps':
                # Filter by ps_assigned_at timestamp for PS leads
                query = query.gte('ps_assigned_at', f'{start_date} 00:00:00').lte('ps_assigned_at', f'{end_date} 23:59:59')
            elif table_type == 'walkin':
                # Filter by created_at for walk-in leads (since they don't have ps_assigned_at)
                query = query.gte('created_at', f'{start_date} 00:00:00').lte('created_at', f'{end_date} 23:59:59')
            elif table_type == 'event':
                # Filter by created_at for event leads (since they don't have ps_assigned_at)
                query = query.gte('created_at', f'{start_date} 00:00:00').lte('created_at', f'{end_date} 23:59:59')
        
        # OPTIMIZED: Server-side search filtering
        if search:
            search_lower = search.lower()
            # Create OR conditions for search across multiple columns
            search_conditions = []
            
            if table_type == 'ps':
                search_conditions = [
                    f"customer_name.ilike.%{search_lower}%",
                    f"customer_mobile_number.ilike.%{search_lower}%",
                    f"lead_uid.ilike.%{search_lower}%",
                    f"ps_name.ilike.%{search_lower}%"
                ]
            elif table_type == 'walkin':
                search_conditions = [
                    f"customer_name.ilike.%{search_lower}%",
                    f"mobile_number.ilike.%{search_lower}%",
                    f"uid.ilike.%{search_lower}%",
                    f"ps_assigned.ilike.%{search_lower}%"
                ]
            elif table_type == 'event':
                search_conditions = [
                    f"customer_name.ilike.%{search_lower}%",
                    f"customer_phone_number.ilike.%{search_lower}%",
                    f"activity_uid.ilike.%{search_lower}%",
                    f"ps_name.ilike.%{search_lower}%"
                ]
            
            if search_conditions:
                query = query.or_(','.join(search_conditions))
        
        return query

    # Apply section-specific filters with OPTIMIZED queries
    
    if section == 'fresh_leads' or section == 'fresh_leads_section':
        # Get filter parameters
        ps_name = request.args.get('ps_name', '')
        source = request.args.get('source', '')
        final_status = request.args.get('final_status', '')
        date_filter = request.args.get('date_filter', '')
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        search = request.args.get('search', '')
        
        # Debug logging
        print(f"DEBUG: Filter parameters - ps_name: {ps_name}, source: {source}, final_status: {final_status}, date_filter: {date_filter}, start_date: {start_date}, end_date: {end_date}")
        
        # Initialize empty lists for each table
        ps_rows = []
        walkin_rows = []
        event_rows = []
        
        # Initialize counts for response
        fresh_leads_count = 0
        followup_leads_count = 0
        
        # Apply different logic based on section
        if section == 'fresh_leads_section':
            # Fresh Leads Section Logic - Only show leads with NULL follow-up dates and Pending status
            print(f"DEBUG: Using Fresh Leads Section logic")
            
            # Apply source-based filtering logic
            if source == 'Walk In':
                # Only query walkin_table for Walk In source
                walkin_query = supabase.table('walkin_table').select('uid, customer_name, mobile_number, branch, ps_assigned, status, created_at, lead_category').eq('branch', branch).eq('status', 'Pending').is_('next_followup_date', 'null')
                walkin_query = apply_server_filters(walkin_query, 'walkin')
                walkin_rows = walkin_query.execute().data or []
                ps_rows = []
                event_rows = []
            elif source == 'Event':
                # Only query activity_leads for Event source
                event_query = supabase.table('activity_leads').select('activity_uid, customer_name, customer_phone_number, final_status, ps_name, created_at, lead_category').eq('location', branch).eq('final_status', 'Pending').is_('ps_followup_date_ts', 'null')
                event_query = apply_server_filters(event_query, 'event')
                event_rows = event_query.execute().data or []
                ps_rows = []
                walkin_rows = []
            else:
                # Query all tables for other sources or no source filter
                # PS fresh leads - ONLY those with NULL follow_up_date AND final_status = 'Pending'
                ps_query = supabase.table('ps_followup_master').select('lead_uid, customer_name, customer_mobile_number, final_status, ps_name, created_at, source, lead_category, ps_assigned_at').eq('ps_branch', branch).eq('final_status', 'Pending').is_('follow_up_date', 'null')
                ps_query = apply_server_filters(ps_query, 'ps')
                ps_rows = ps_query.execute().data or []
                
                # Walk-in fresh leads - ONLY those with NULL next_followup_date AND status = 'Pending'
                walkin_query = supabase.table('walkin_table').select('uid, customer_name, mobile_number, branch, ps_assigned, status, created_at, lead_category').eq('branch', branch).eq('status', 'Pending').is_('next_followup_date', 'null')
                walkin_query = apply_server_filters(walkin_query, 'walkin')
                walkin_rows = walkin_query.execute().data or []
                
                # Event fresh leads - ONLY those with NULL ps_followup_date_ts AND final_status = 'Pending'
                event_query = supabase.table('activity_leads').select('activity_uid, customer_name, customer_phone_number, final_status, ps_name, created_at, lead_category').eq('location', branch).eq('final_status', 'Pending').is_('ps_followup_date_ts', 'null')
                event_query = apply_server_filters(event_query, 'event')
                event_rows = event_query.execute().data or []
            
        else:
            # Regular All Leads Logic with improved source filtering
            if source == 'Walk In':
                # Only query walkin_table for Walk In source
                print(f"DEBUG: Querying walkin_table for Walk In source")
                walkin_query = supabase.table('walkin_table').select('uid, customer_name, mobile_number, branch, ps_assigned, status, created_at, lead_category').eq('branch', branch)
                walkin_query = apply_server_filters(walkin_query, 'walkin')
                walkin_rows = walkin_query.execute().data or []
                ps_rows = []
                event_rows = []
                print(f"DEBUG: Found {len(walkin_rows)} walkin rows")
            elif source == 'Event':
                # Only query activity_leads for Event source
                print(f"DEBUG: Querying activity_leads for Event source")
                event_query = supabase.table('activity_leads').select('activity_uid, customer_name, customer_phone_number, final_status, ps_name, created_at, lead_category').eq('location', branch)
                event_query = apply_server_filters(event_query, 'event')
                event_rows = event_query.execute().data or []
                ps_rows = []
                walkin_rows = []
                print(f"DEBUG: Found {len(event_rows)} event rows")
            else:
                # Query all tables for other sources or no source filter
                ps_query = supabase.table('ps_followup_master').select('lead_uid, customer_name, customer_mobile_number, final_status, ps_name, created_at, source, lead_category, ps_assigned_at').eq('ps_branch', branch)
                ps_query = apply_server_filters(ps_query, 'ps')
                ps_rows = ps_query.execute().data or []
                
                walkin_query = supabase.table('walkin_table').select('uid, customer_name, mobile_number, branch, ps_assigned, status, created_at, lead_category').eq('branch', branch)
                walkin_query = apply_server_filters(walkin_query, 'walkin')
                walkin_rows = walkin_query.execute().data or []
                
                event_query = supabase.table('activity_leads').select('activity_uid, customer_name, customer_phone_number, final_status, ps_name, created_at, lead_category').eq('location', branch)
                event_query = apply_server_filters(event_query, 'event')
                event_rows = event_query.execute().data or []
        
        # OPTIMIZED: Format rows efficiently using list comprehensions
        def format_ps_rows(rows):
            return [
                {
                    'uid': row.get('lead_uid', ''),
                    'customer_name': row.get('customer_name', ''),
                    'customer_mobile_number': row.get('customer_mobile_number', ''),
                    'final_status': row.get('final_status', ''),
                    'ps_name': row.get('ps_name', ''),
                    'created_at': row.get('created_at', ''),
                    'source': row.get('source', 'PS'),
                    'lead_category': row.get('lead_category') if row.get('lead_category') else 'Not Set',
                    'ps_assigned_at': row.get('ps_assigned_at', '')
                }
                for row in rows
            ]
        
        def format_walkin_rows(rows):
            return [
                {
                    'uid': row.get('uid', ''),
                    'customer_name': row.get('customer_name', ''),
                    'customer_mobile_number': row.get('mobile_number', ''),
                    'final_status': row.get('status', ''),
                    'ps_name': row.get('ps_assigned', ''),
                    'created_at': row.get('created_at', ''),
                    'source': 'Walk In',
                    'lead_category': row.get('lead_category') if row.get('lead_category') else 'Not Set',
                    'ps_assigned_at': row.get('created_at', '')
                }
                for row in rows
            ]
        
        def format_event_rows(rows):
            return [
                {
                    'uid': row.get('activity_uid', ''),
                    'customer_name': row.get('customer_name', ''),
                    'customer_mobile_number': row.get('customer_phone_number', ''),
                    'final_status': row.get('final_status', ''),
                    'ps_name': row.get('ps_name', ''),
                    'created_at': row.get('created_at', ''),
                    'source': 'Event',
                    'lead_category': row.get('lead_category') if row.get('lead_category') else 'Not Set',
                    'ps_assigned_at': row.get('created_at', '')
                }
                for row in rows
            ]
        
        # Format all rows efficiently
        ps_formatted = format_ps_rows(ps_rows)
        walkin_formatted = format_walkin_rows(walkin_rows)
        event_formatted = format_event_rows(event_rows)
        
        # Calculate fresh leads count for this section
        fresh_leads_count = len(ps_formatted) + len(walkin_formatted) + len(event_formatted)
        
        # Combine all rows
        all_formatted_rows = ps_formatted + walkin_formatted + event_formatted
        
        # Sort by created_at in descending order (newest first)
        all_formatted_rows.sort(key=lambda x: x.get('created_at', ''), reverse=True)
        
        # Note: All filtering (search, date, PS) will be handled client-side for better performance
        
        # OPTIMIZED: Return response efficiently with better error handling
        try:
            response = {
                'success': True,
                'rows': all_formatted_rows,
                'total_count': len(all_formatted_rows),
                'fresh_leads_count': fresh_leads_count,
                'followup_leads_count': followup_leads_count
            }
            
            # Log audit event for successful data access
            try:
                auth_manager.log_audit_event(
                    user_id=session.get('user_id'),
                    user_type=session.get('user_type'),
                    action='BRANCH_HEAD_DASHBOARD_DATA_ACCESS',
                    resource='branch_head_dashboard_data',
                    details={
                        'section': section,
                        'total_count': len(all_formatted_rows)
                    }
                )
            except Exception as audit_error:
                print(f"Audit logging error: {audit_error}")
            
            return jsonify(response)
        except Exception as e:
            print(f"Error formatting response: {str(e)}")
            return jsonify({
                'success': False,
                'error': str(e),
                'rows': [],
                'total_count': 0,
                'fresh_leads_count': fresh_leads_count,
                'followup_leads_count': followup_leads_count
            })

    elif section == 'walkin_leads':
        walkin_query = supabase.table('walkin_data').select('*').eq('ps_branch', branch)
        all_rows = walkin_query.execute().data or []
        rows = [
            {
                'uid': row.get('uid', ''),
                'customer_name': row.get('customer_name', ''),
                'customer_mobile_number': row.get('customer_mobile_number', ''),
                'customer_email': row.get('customer_email', ''),
                'model_interested': row.get('model_interested', ''),
                'lead_source': row.get('lead_source', ''),
                'status': row.get('ps_final_status', ''),
                'ps_name': row.get('ps_name', ''),
                'created_at': row.get('created_at', ''),
                'lead_category': row.get('lead_category', '')
            }
            for row in all_rows
        ]
        
        # Note: All filtering is handled client-side
        
        total_count = int(len(rows))
        offset = (page - 1) * per_page
        paged_rows = rows[offset:offset + per_page]
        total_pages = 1 if total_count == 0 else math.ceil(total_count / per_page)
        response = {
            'success': True,
            'rows': paged_rows,
            'total_count': total_count,
            'total_pages': total_pages,
            'current_page': page,
            'per_page': per_page,
            'walkin_leads_count': total_count,
            'fresh_leads_count': fresh_leads_count,
            'cre_assigned_leads_count': cre_assigned_leads_count,
            'followup_leads_count': followup_leads_count,
            'event_leads_count': event_leads_count,
            'won_leads_count': won_leads_count,
            'lost_leads_count': lost_leads_count
        }
        return jsonify(response)

    elif section == 'cre_assigned_leads':
        # Fetch all rows from ps_followup_master for the current branch
        query = supabase.table('ps_followup_master').select('*').eq('ps_branch', branch)
        all_rows = query.execute().data or []
        
        # Structure the data properly for frontend
        rows = [
            {
                'customer_name': row.get('customer_name', ''),
                'customer_mobile_number': row.get('customer_mobile_number', ''),
                'final_status': row.get('final_status', ''),
                'uid': row.get('lead_uid', ''),
                'source': row.get('source', ''),
                'lead_status': row.get('lead_status', ''),
                'ps_name': row.get('ps_name', '')
            }
            for row in all_rows
        ]
        
        # Note: All filtering is handled client-side
        
        total_count = len(rows)
        offset = (page - 1) * per_page
        paged_rows = rows[offset:offset + per_page]
        total_pages = 1 if total_count == 0 else math.ceil(total_count / per_page)
        response = {
            'success': True,
            'rows': paged_rows,
            'total_count': total_count,
            'total_pages': total_pages,
            'current_page': page,
            'per_page': per_page,
            'cre_assigned_leads_count': total_count,
            'fresh_leads_count': fresh_leads_count,
            'walkin_leads_count': walkin_leads_count,
            'followup_leads_count': followup_leads_count,
            'event_leads_count': event_leads_count,
            'won_leads_count': won_leads_count,
            'lost_leads_count': lost_leads_count
        }
        return jsonify(response)
    elif section == 'pending_leads':
        # ps_followup_master
        query = supabase.table('ps_followup_master').select('*').eq('ps_branch', branch).eq('final_status', 'Pending')
        followup_rows_raw = query.execute().data or []
        followup_rows = [
            {
                'uid': row.get('lead_uid', ''),
                'customer_name': row.get('customer_name', ''),
                'customer_mobile_number': row.get('customer_mobile_number', ''),
                'final_status': row.get('final_status', ''),
                'source': row.get('source', ''),
                'lead_status': row.get('lead_status', ''),
                'ps_name': row.get('ps_name', '')
            }
            for row in followup_rows_raw
        ]

        # activity_leads
        query = supabase.table('activity_leads').select('*').eq('location', branch).eq('final_status', 'Pending')
        event_rows_raw = query.execute().data or []
        event_rows = [
            {
                'uid': row.get('activity_uid', ''),
                'customer_name': row.get('customer_name', ''),
                'customer_mobile_number': row.get('customer_phone_number', row.get('customer_mobile_number', '')),
                'final_status': row.get('final_status', ''),
                'source': 'Event',
                'lead_status': row.get('lead_status', ''),
                'ps_name': row.get('ps_name', '')
            }
            for row in event_rows_raw
        ]

        # Merge all
        all_rows = followup_rows + event_rows
        
        # Note: All filtering is handled client-side

        # Counts for KPI sub-boxes
        cre_assigned_count = len(followup_rows)
        event_count = len(event_rows)

        total_count = len(all_rows)

        # Pagination
        offset = (page - 1) * per_page
        paged_rows = all_rows[offset:offset + per_page]
        total_pages = 1 if total_count == 0 else math.ceil(total_count / per_page)

        response = {
            'success': True,
            'rows': paged_rows,
            'total_count': total_count,
            'total_pages': total_pages,
            'current_page': page,
            'per_page': per_page,
            'pending_leads_count': total_count,
            'pending_leads_cre_assigned_count': cre_assigned_count,
            'pending_leads_event_count': event_count,
            'fresh_leads_count': fresh_leads_count,
            'walkin_leads_count': walkin_leads_count,
            'cre_assigned_leads_count': cre_assigned_leads_count,
            'followup_leads_count': followup_leads_count,
            'event_leads_count': event_leads_count,
            'won_leads_count': won_leads_count,
            'lost_leads_count': lost_leads_count
        }
        return jsonify(response)

    elif section == 'followup_leads':
        # Get filter parameters for followup_leads section
        ps_name = request.args.get('ps_name', '')
        search = request.args.get('search', '')
        
        # Debug logging
        print(f"DEBUG: Followup filter parameters - ps_name: {ps_name}, search: {search}")
        
        # Get today's and past follow-ups using proper date filtering (follow_up_date <= today)
        today_str = datetime.now().strftime('%Y-%m-%d')
        
        # PS follow-ups for today and past dates (follow_up_date <= today) AND final_status = 'Pending'
        print(f"DEBUG: Today's date for PS query: {today_str}")
        ps_query = supabase.table('ps_followup_master').select('*, created_at, ps_assigned_at').eq('ps_branch', branch).eq('final_status', 'Pending').lte('follow_up_date', today_str)
        # Apply PS filter if specified
        if ps_name:
            ps_query = ps_query.eq('ps_name', ps_name)
        ps_today = ps_query.execute().data or []
        print(f"DEBUG: PS query result count: {len(ps_today)}")
        if ps_today:
            print(f"DEBUG: Sample PS data: {ps_today[0]}")
            print(f"DEBUG: Sample PS follow_up_date: {ps_today[0].get('follow_up_date')}")
            print(f"DEBUG: PS query filter: follow_up_date <= {today_str}")
        
        # Event follow-ups for today and past dates (ps_followup_date_ts <= today) AND final_status = 'Pending'
        # Use lte (less than or equal) to get today and past follow-up dates
        act_query = supabase.table('activity_leads').select('*, created_at').eq('location', branch).eq('final_status', 'Pending').lte('ps_followup_date_ts', f'{today_str} 23:59:59')
        # Apply PS filter if specified
        if ps_name:
            act_query = act_query.eq('ps_name', ps_name)
        act_today_raw = act_query.execute().data or []
        
        # Walk-in follow-ups for today and past dates (next_followup_date <= today) AND status = 'Pending'
        # Use lte (less than or equal) to get today and past follow-up dates
        print(f"DEBUG: Today's date for walk-in query: {today_str}")
        walkin_query = supabase.table('walkin_table').select('*, created_at').eq('branch', branch).eq('status', 'Pending').lte('next_followup_date', f'{today_str} 23:59:59')
        # Apply PS filter if specified
        if ps_name:
            walkin_query = walkin_query.eq('ps_assigned', ps_name)
        walkin_today_raw = walkin_query.execute().data or []
        print(f"DEBUG: Walk-in query result count: {len(walkin_today_raw)}")
        if walkin_today_raw:
            print(f"DEBUG: Sample walk-in data: {walkin_today_raw[0]}")
            print(f"DEBUG: Sample walk-in next_followup_date: {walkin_today_raw[0].get('next_followup_date')}")
            print(f"DEBUG: Walk-in query filter: next_followup_date <= {today_str} 23:59:59")
        
        # Mark all as not missed and set source_type and phone
        ps_today = [{**row, 'missed': False, 'lead_status': row.get('lead_status', '')} for row in ps_today]
        act_today = [
            {**row, 'missed': False, 'source_type': 'Event', 'customer_mobile_number': row.get('customer_phone_number', ''), 'lead_status': row.get('lead_status', '')}
            for row in act_today_raw
        ]
        walkin_today = [
            {**row, 'missed': False, 'source_type': 'Walk-in', 'final_status': row.get('status', ''), 'lead_status': row.get('lead_status', ''), 'customer_mobile_number': row.get('mobile_number', '')}
            for row in walkin_today_raw
        ]
        
        # Debug: Print sample data from each source
        if ps_today:
            print(f"DEBUG: Sample PS data: {ps_today[0] if ps_today else 'None'}")
            print(f"DEBUG: PS source column value: {ps_today[0].get('source') if ps_today else 'None'}")
        if act_today:
            print(f"DEBUG: Sample Event data: {act_today[0] if act_today else 'None'}")
        if walkin_today:
            print(f"DEBUG: Sample Walkin data: {walkin_today[0] if walkin_today else 'None'}")
        
        # Merge all today's and past followups (PS, Activity, and Walk-in)
        all_rows = ps_today + act_today + walkin_today
        
        # Debug: Print counts before and after search filter
        print(f"DEBUG: Before search filter - PS: {len(ps_today)}, Event: {len(act_today)}, Walkin: {len(walkin_today)}, Total: {len(all_rows)}")
        print(f"DEBUG: Date filter applied: follow_up_date <= {today_str}")
        
        # Format rows for frontend with proper column mapping
        formatted_rows = []
        for row in all_rows:
            # Determine source type for proper column mapping
            source_type = row.get('source_type', '')
            
            if source_type == 'Walk-in':
                # For walk-in leads, use created_at, ignore ps_assigned_at
                formatted_rows.append({
                    'uid': row.get('uid', ''),
                'customer_name': row.get('customer_name', ''),
                    'customer_mobile_number': row.get('mobile_number', ''),
                    'source': 'Walk-in',
                'lead_category': row.get('lead_category') if row.get('lead_category') else 'Not Set',
                'lead_status': row.get('lead_status', ''),
                    'ps_name': row.get('ps_assigned', ''),
                    'follow_up_date': row.get('next_followup_date', ''),
                    'created_at': row.get('created_at', ''),
                    'ps_assigned_at': 'N/A'  # Walk-in leads don't have ps_assigned_at
                })
            elif source_type == 'Event':
                # For event leads, use created_at, ignore ps_assigned_at
                formatted_rows.append({
                    'uid': row.get('activity_uid', ''),
                    'customer_name': row.get('customer_name', ''),
                    'customer_mobile_number': row.get('customer_phone_number', ''),
                    'source': 'Event',
                    'lead_category': row.get('lead_category') if row.get('lead_category') else 'Not Set',
                    'lead_status': row.get('lead_status', ''),
                    'ps_name': row.get('ps_name', ''),
                    'follow_up_date': row.get('ps_followup_date_ts', ''),
                    'created_at': row.get('created_at', ''),
                    'ps_assigned_at': 'N/A'  # Event leads don't have ps_assigned_at
                })
            else:
                # For PS leads, use ps_assigned_at, ignore created_at
                formatted_rows.append({
                    'uid': row.get('lead_uid', ''),
                    'customer_name': row.get('customer_name', ''),
                    'customer_mobile_number': row.get('customer_mobile_number', ''),
                    'source': row.get('source', 'PS Followup'),
                    'lead_category': row.get('lead_category') if row.get('lead_category') else 'Not Set',
                    'lead_status': row.get('lead_status', ''),
                    'ps_name': row.get('ps_name', ''),
                    'follow_up_date': row.get('follow_up_date', ''),
                    'created_at': 'N/A',  # PS leads don't have created_at from this table
                    'ps_assigned_at': row.get('ps_assigned_at', '')
                })
        
        # Apply search filter if provided
        if search:
            print(f"DEBUG: Applying search filter: {search}")
            search_lower = search.lower()
            formatted_rows = [
                row for row in formatted_rows
                if (search_lower in str(row.get('uid', '')).lower() or
                    search_lower in str(row.get('customer_name', '')).lower() or
                    search_lower in str(row.get('customer_mobile_number', '')).lower() or
                    search_lower in str(row.get('ps_name', '')).lower())
            ]
            print(f"DEBUG: After search filter - Total: {len(formatted_rows)}")
        
        print(f"DEBUG: Final formatted rows count: {len(formatted_rows)}")
        
        # Calculate actual counts from loaded data (today and past follow-ups)
        actual_ps_count = len(ps_today)
        actual_event_count = len(act_today_raw)
        actual_walkin_count = len(walkin_today_raw)
        actual_total_count = actual_ps_count + actual_event_count + actual_walkin_count
        
        print(f"DEBUG: Actual data counts (follow_up_date <= {today_str}) - PS: {actual_ps_count}, Event: {actual_event_count}, Walkin: {actual_walkin_count}, Total: {actual_total_count}")
        print(f"DEBUG: Formatted rows count: {len(formatted_rows)}")
        
        response = {
            'success': True,
            'rows': formatted_rows,
            'total_count': len(formatted_rows),
            'followup_leads_count': actual_total_count,  # Use actual data count, not formatted rows
            'fresh_leads_count': fresh_leads_count
        }
        
        auth_manager.log_audit_event(
            user_id=session.get('user_id'),
            user_type=session.get('user_type'),
            action='BRANCH_HEAD_DASHBOARD_DATA_ACCESS',
            resource='branch_head_dashboard_data',
            details={
                'section': section,
                'total_count': len(formatted_rows)
            }
        )
        return jsonify(response)

    elif section == 'event_leads':
        query = supabase.table('activity_leads').select('*').eq('location', branch)
        
        # Apply filters
        if ps_name:
            query = query.eq('ps_name', ps_name)
        if start_date:
            query = query.gte('created_at', start_date)
        if end_date:
            query = query.lte('created_at', end_date)
            
        event_rows_raw = query.execute().data or []
        event_rows = [
            {
                'customer_name': (row.get('customer_name') or '') if row.get('customer_name') is not None else '',
                'customer_mobile_number': (row.get('customer_phone_number') or '') if row.get('customer_phone_number') is not None else '',
                'final_status': (row.get('final_status') or '') if row.get('final_status') is not None else '',
                'uid': row.get('activity_uid', '') or '',
                'source': 'Event',
                'lead_status': (row.get('lead_status') or '') if row.get('lead_status') is not None else '',
                'ps_name': (row.get('ps_name') or '') if row.get('ps_name') is not None else ''
            }
            for row in event_rows_raw
        ]
        
        # Apply search filter
        event_rows = filter_rows_by_search(event_rows, search)
        
        total_count = len(event_rows)
        offset = (page - 1) * per_page
        paged_rows = event_rows[offset:offset + per_page]
        total_pages = 1 if total_count == 0 else math.ceil(total_count / per_page)
        
        response = {
            'success': True,
            'rows': paged_rows,
            'total_count': total_count,
            'total_pages': total_pages,
            'current_page': page,
            'per_page': per_page,
            'event_leads_count': total_count,
            'fresh_leads_count': fresh_leads_count,
            'walkin_leads_count': walkin_leads_count,
            'cre_assigned_leads_count': cre_assigned_leads_count,
            'followup_leads_count': followup_leads_count,
            'won_leads_count': won_leads_count,
            'lost_leads_count': lost_leads_count
        }
        return jsonify(response)
    elif section == 'won_leads':
        # Fetch from ps_followup_master
        query = supabase.table('ps_followup_master').select('*').eq('ps_branch', branch).eq('final_status', 'Won')
        
        # Apply filters
        if ps_name:
            query = query.eq('ps_name', ps_name)
        if start_date:
            query = query.gte('created_at', start_date)
        if end_date:
            query = query.lte('created_at', end_date)
            
        ps_rows_raw = query.execute().data or []
        ps_rows = [
            {
                'customer_name': row.get('customer_name', ''),
                'customer_mobile_number': row.get('customer_mobile_number', ''),
                'final_status': row.get('final_status', ''),
                'uid': row.get('lead_uid', ''),
                'source': row.get('source', ''),
                'lead_status': row.get('lead_status', ''),
                'ps_name': row.get('ps_name', '')
            }
            for row in ps_rows_raw
        ]
        
        # Fetch from activity_leads
        query = supabase.table('activity_leads').select('*').eq('location', branch).eq('final_status', 'Won')
        
        # Apply filters
        if ps_name:
            query = query.eq('ps_name', ps_name)
        if start_date:
            query = query.gte('created_at', start_date)
        if end_date:
            query = query.lte('created_at', end_date)
            
        event_rows_raw = query.execute().data or []
        event_rows = [
            {
                'customer_name': row.get('customer_name', ''),
                'customer_mobile_number': row.get('customer_phone_number', ''),
                'final_status': row.get('final_status', ''),
                'uid': row.get('activity_uid', ''),
                'source': 'Event',
                'lead_status': row.get('lead_status', ''),
                'ps_name': row.get('ps_name', '')
            }
            for row in event_rows_raw
        ]
        
        # Fetch from walkin_data
        query = supabase.table('walkin_data').select('*').eq('ps_branch', branch).eq('ps_final_status', 'Won')
        
        # Apply filters
        if ps_name:
            query = query.eq('ps_name', ps_name)
        if start_date:
            query = query.gte('created_at', start_date)
        if end_date:
            query = query.lte('created_at', end_date)
            
        walkin_rows_raw = query.execute().data or []
        walkin_rows = [
            {
                'customer_name': row.get('customer_name', ''),
                'customer_mobile_number': row.get('customer_mobile_number', ''),
                'final_status': row.get('ps_final_status', ''),
                'uid': row.get('uid', ''),
                'source': 'Walk-in',
                'lead_status': row.get('lead_status', ''),
                'ps_name': row.get('ps_name', '')
            }
            for row in walkin_rows_raw
        ]
        
        # Merge all rows
        all_rows = ps_rows + event_rows + walkin_rows
        
        # Apply search filter
        all_rows = filter_rows_by_search(all_rows, search)
        
        total_count = len(all_rows)
        offset = (page - 1) * per_page
        paged_rows = all_rows[offset:offset + per_page]
        total_pages = 1 if total_count == 0 else math.ceil(total_count / per_page)
        
        response = {
            'success': True,
            'rows': paged_rows,
            'total_count': total_count,
            'total_pages': total_pages,
            'current_page': page,
            'per_page': per_page,
            'won_leads_count': total_count,
            'fresh_leads_count': fresh_leads_count,
            'walkin_leads_count': walkin_leads_count,
            'cre_assigned_leads_count': cre_assigned_leads_count,
            'followup_leads_count': followup_leads_count,
            'event_leads_count': event_leads_count,
            'lost_leads_count': lost_leads_count
        }
        return jsonify(response)

    elif section == 'lost_leads':
        # Fetch from ps_followup_master
        query = supabase.table('ps_followup_master').select('*').eq('ps_branch', branch).eq('final_status', 'Lost')
        
        # Apply filters
        if ps_name:
            query = query.eq('ps_name', ps_name)
        if start_date:
            query = query.gte('created_at', start_date)
        if end_date:
            query = query.lte('created_at', end_date)
            
        ps_rows_raw = query.execute().data or []
        ps_rows = [
            {
                'customer_name': row.get('customer_name', ''),
                'customer_mobile_number': row.get('customer_mobile_number', ''),
                'final_status': row.get('final_status', ''),
                'uid': row.get('lead_uid', ''),
                'source': row.get('source', ''),
                'lead_status': row.get('lead_status', ''),
                'ps_name': row.get('ps_name', '')
            }
            for row in ps_rows_raw
        ]
        
        # Fetch from activity_leads
        query = supabase.table('activity_leads').select('*').eq('location', branch).eq('final_status', 'Lost')
        
        # Apply filters
        if ps_name:
            query = query.eq('ps_name', ps_name)
        if start_date:
            query = query.gte('created_at', start_date)
        if end_date:
            query = query.lte('created_at', end_date)
            
        event_rows_raw = query.execute().data or []
        event_rows = [
            {
                'customer_name': row.get('customer_name', ''),
                'customer_mobile_number': row.get('customer_phone_number', ''),
                'final_status': row.get('final_status', ''),
                'uid': row.get('activity_uid', ''),
                'source': 'Event',
                'lead_status': row.get('lead_status', ''),
                'ps_name': row.get('ps_name', '')
            }
            for row in event_rows_raw
        ]
        
        # Fetch from walkin_data
        query = supabase.table('walkin_data').select('*').eq('ps_branch', branch).eq('ps_final_status', 'Lost')
        
        # Apply filters
        if ps_name:
            query = query.eq('ps_name', ps_name)
        if start_date:
            query = query.gte('created_at', start_date)
        if end_date:
            query = query.lte('created_at', end_date)
            
        walkin_rows_raw = query.execute().data or []
        walkin_rows = [
            {
                'customer_name': row.get('customer_name', ''),
                'customer_mobile_number': row.get('customer_mobile_number', ''),
                'final_status': row.get('ps_final_status', ''),
                'uid': row.get('uid', ''),
                'source': 'Walk-in',
                'lead_status': row.get('lead_status', ''),
                'ps_name': row.get('ps_name', '')
            }
            for row in walkin_rows_raw
        ]
        
        # Merge all rows
        all_rows = ps_rows + event_rows + walkin_rows
        
        # Apply search filter
        all_rows = filter_rows_by_search(all_rows, search)
        
        total_count = len(all_rows)
        offset = (page - 1) * per_page
        paged_rows = all_rows[offset:offset + per_page]
        total_pages = 1 if total_count == 0 else math.ceil(total_count / per_page)
        
        response = {
            'success': True,
            'rows': paged_rows,
            'total_count': total_count,
            'total_pages': total_pages,
            'current_page': page,
            'per_page': per_page,
            'lost_leads_count': total_count,
            'fresh_leads_count': fresh_leads_count,
            'walkin_leads_count': walkin_leads_count,
            'cre_assigned_leads_count': cre_assigned_leads_count,
            'followup_leads_count': followup_leads_count,
            'event_leads_count': event_leads_count,
            'won_leads_count': won_leads_count
        }
        return jsonify(response)

    # Default response for any other sections
    response = {
        'success': True,
        'rows': [],
        'total_count': 0,
        'total_pages': 1,
        'current_page': page,
        'per_page': per_page,
        'fresh_leads_count': fresh_leads_count,
        'walkin_leads_count': walkin_leads_count,
        'cre_assigned_leads_count': cre_assigned_leads_count,
        'followup_leads_count': followup_leads_count,
        'event_leads_count': event_leads_count,
        'won_leads_count': won_leads_count,
        'lost_leads_count': lost_leads_count
    }
    return jsonify(response)

@app.route('/api/lead_call_history/<uid>')
def api_lead_call_history(uid):
    """Get call attempt history for a specific lead (both CRE and PS)"""
    try:
        # Normalize UID (trim stray spaces)
        uid_norm = (uid or '').strip()

        # Get CRE call attempt history
        cre_history_result = supabase.table('cre_call_attempt_history').select('uid,call_no,attempt,status,cre_name,remarks,follow_up_date').eq('uid', uid_norm).execute()
        cre_history = cre_history_result.data if cre_history_result.data else []

        # Get PS call attempt history
        ps_history_result = supabase.table('ps_call_attempt_history').select('uid,call_no,attempt,status,ps_name,remarks,follow_up_date').eq('uid', uid_norm).execute()
        ps_history = ps_history_result.data if ps_history_result.data else []

        # Fallback: if still empty, try case-insensitive match (helps with accidental casing/formatting issues)
        if not cre_history:
            try:
                like_res = supabase.table('cre_call_attempt_history').select('uid,call_no,attempt,status,cre_name,remarks,follow_up_date').ilike('uid', f"%{uid_norm}%").execute()
                cre_history = like_res.data if like_res.data else []
            except Exception:
                pass
        if not ps_history:
            try:
                like_res = supabase.table('ps_call_attempt_history').select('uid,call_no,attempt,status,ps_name,remarks,follow_up_date').ilike('uid', f"%{uid_norm}%").execute()
                ps_history = like_res.data if like_res.data else []
            except Exception:
                pass

        # Define the order of call numbers
        call_order = ['first', 'second', 'third', 'fourth', 'fifth', 'sixth', 'seventh']

        # Sort the history: first by call_no in ascending order, then by attempt in ascending order
        def sort_key(item):
            call_no = item.get('call_no', '')
            attempt = item.get('attempt', 0)
            # Get the index of call_no in the predefined order, default to end if not found
            call_index = call_order.index(call_no) if call_no in call_order else len(call_order)
            return (call_index, attempt)

        # Sort both histories
        cre_history_sorted = sorted(cre_history, key=sort_key)
        ps_history_sorted = sorted(ps_history, key=sort_key)

        # Combine and format the history
        combined_history = []

        # Add CRE calls with role identifier
        for call in cre_history_sorted:
            combined_history.append({
                'uid': call.get('uid', ''),
                'call_no': call.get('call_no', ''),
                'attempt': call.get('attempt', ''),
                'status': call.get('status', ''),
                'name': call.get('cre_name', ''),
                'role': 'CRE',
                'timestamp': call.get('follow_up_date', ''),  # Use follow_up_date as timestamp
                'remarks': call.get('remarks', ''),
                'follow_up_date': call.get('follow_up_date', '')
            })

        # Add PS calls with role identifier
        for call in ps_history_sorted:
            combined_history.append({
                'uid': call.get('uid', ''),
                'call_no': call.get('call_no', ''),
                'attempt': call.get('attempt', ''),
                'status': call.get('status', ''),
                'name': call.get('ps_name', ''),
                'role': 'PS',
                'timestamp': call.get('follow_up_date', ''),  # Use follow_up_date as timestamp
                'remarks': call.get('remarks', ''),
                'follow_up_date': call.get('follow_up_date', '')
            })

        # Sort combined history by timestamp
        combined_history.sort(key=lambda x: (x.get('timestamp') or ''), reverse=True)

        # Calculate summary statistics
        total_calls = len(combined_history)
        rnr_calls = len([call for call in combined_history if call.get('status', '').upper() == 'RNR'])
        busy_calls = len([call for call in combined_history if call.get('status', '').lower() == 'busy on another call'])
        call_me_back = len([call for call in combined_history if call.get('status', '').lower() == 'call me back'])
        call_not_connected = len([call for call in combined_history if call.get('status', '').lower() == 'call not connected'])

        # Connected calls (excluding the above statuses)
        excluded_statuses = ['rnr', 'busy on another call', 'call me back', 'call not connected']
        connected_calls = len([call for call in combined_history
                             if call.get('status') and
                             not any(excluded in call.get('status', '').lower() for excluded in excluded_statuses)])

        return jsonify({
            'success': True,
            'history': combined_history,
            'counts': {
                'ps': len(ps_history),
                'cre': len(cre_history)
            },
            'summary': {
                'total_calls': total_calls,
                'rnr_calls': rnr_calls,
                'busy_calls': busy_calls,
                'call_me_back': call_me_back,
                'call_not_connected': call_not_connected,
                'connected_calls': connected_calls
            }
        })

    except Exception as e:
        return jsonify({'success': False, 'message': str(e), 'history': []})

@app.route('/branch_performance/<branch_name>')
@require_admin
def branch_performance(branch_name):
    """Get detailed PS performance for a specific branch"""
    try:
        # Get all PS users in this branch
        branch_ps_users = safe_get_data('ps_users', {'branch': branch_name})

        if not branch_ps_users:
            return jsonify({
                'success': False,
                'message': f'No PS users found in {branch_name} branch'
            })

        # Get all leads and PS followups
        all_leads = safe_get_data('lead_master')
        all_ps_followups = safe_get_data('ps_followup_master')

        # Calculate performance for each PS in the branch
        ps_performance = []
        total_branch_leads = 0
        total_branch_won = 0

        for ps_user in branch_ps_users:
            ps_name = ps_user['name']

            # Get PS followup data for this PS
            ps_leads = [f for f in all_ps_followups if f.get('ps_name') == ps_name]

            # Calculate metrics
            total_leads = len(ps_leads)
            pending_leads = len([l for l in ps_leads if l.get('final_status') == 'Pending'])
            in_progress_leads = len([l for l in ps_leads if l.get('final_status') == 'In Progress'])
            won_leads = len([l for l in ps_leads if l.get('final_status') == 'Won'])
            lost_leads = len([l for l in ps_leads if l.get('final_status') == 'Lost'])

            # Calculate success rate
            success_rate = round((won_leads / total_leads * 100) if total_leads > 0 else 0, 1)

            # Calculate average calls
            total_calls = 0
            for lead in ps_leads:
                call_count = 0
                for call_num in ['first', 'second', 'third', 'fourth', 'fifth', 'sixth', 'seventh']:
                    if lead.get(f'{call_num}_call_date'):
                        call_count += 1
                total_calls += call_count

            avg_calls = round(total_calls / total_leads, 1) if total_leads > 0 else 0

            # Get last activity (most recent call date)
            last_activity = None
            for lead in ps_leads:
                for call_num in ['seventh', 'sixth', 'fifth', 'fourth', 'third', 'second', 'first']:  # Check in reverse order
                    call_date = lead.get(f'{call_num}_call_date')
                    if call_date:
                        try:
                            activity_date = datetime.strptime(call_date, '%Y-%m-%d').date()
                            if not last_activity or activity_date > last_activity:
                                last_activity = activity_date
                        except (ValueError, TypeError):
                            continue
                        break

            last_activity_str = last_activity.strftime('%Y-%m-%d') if last_activity else None

            ps_performance.append({
                'name': ps_user['name'],
                'username': ps_user['username'],
                'phone': ps_user.get('phone', 'N/A'),
                'email': ps_user.get('email', 'N/A'),
                'total_leads': total_leads,
                'pending_leads': pending_leads,
                'in_progress_leads': in_progress_leads,
                'won_leads': won_leads,
                'lost_leads': lost_leads,
                'success_rate': success_rate,
                'avg_calls': avg_calls,
                'last_activity': last_activity_str
            })

            total_branch_leads += total_leads
            total_branch_won += won_leads

        # Calculate branch summary
        branch_success_rate = round((total_branch_won / total_branch_leads * 100) if total_branch_leads > 0 else 0, 1)

        summary = {
            'total_ps': len(branch_ps_users),
            'total_leads': total_branch_leads,
            'won_leads': total_branch_won,
            'success_rate': branch_success_rate
        }

        # Log branch performance access
        auth_manager.log_audit_event(
            user_id=session.get('user_id'),
            user_type=session.get('user_type'),
            action='BRANCH_PERFORMANCE_ACCESS',
            resource='branch_performance',
            details={'branch': branch_name}
        )

        return jsonify({
            'success': True,
            'data': {
                'branch_name': branch_name,
                'summary': summary,
                'ps_performance': ps_performance
            }
        })

    except Exception as e:
        print(f"Error getting branch performance for {branch_name}: {e}")
        return jsonify({
            'success': False,
            'message': f'Error loading branch performance: {str(e)}'
        })


@app.route('/export_leads')
@require_admin
def export_leads():
    """Export leads dashboard"""
    try:
        # Get all leads for counting
        all_leads = safe_get_data('lead_master')

        # Count by final status
        won_count = len([l for l in all_leads if l.get('final_status') == 'Won'])
        lost_count = len([l for l in all_leads if l.get('final_status') == 'Lost'])
        pending_count = len([l for l in all_leads if l.get('final_status') == 'Pending'])
        in_progress_count = len([l for l in all_leads if l.get('final_status') == 'In Progress'])

        # Count by lead category
        hot_count = len([l for l in all_leads if l.get('lead_category') == 'Hot'])
        warm_count = len([l for l in all_leads if l.get('lead_category') == 'Warm'])
        cold_count = len([l for l in all_leads if l.get('lead_category') == 'Cold'])
        not_interested_count = len([l for l in all_leads if l.get('lead_category') == 'Not Interested'])

        # Count test drive leads from alltest_drive table
        test_drive_leads = safe_get_data('alltest_drive')
        test_drive_yes_count = len([l for l in test_drive_leads if l.get('test_drive_done') == 'Yes'])
        test_drive_no_count = len([l for l in test_drive_leads if l.get('test_drive_done') == 'No'])
        test_drive_total_count = len(test_drive_leads)

        # Log export dashboard access
        auth_manager.log_audit_event(
            user_id=session.get('user_id'),
            user_type=session.get('user_type'),
            action='EXPORT_DASHBOARD_ACCESS',
            resource='export_leads'
        )

        return render_template('export_leads.html',
                               won_count=won_count,
                               lost_count=lost_count,
                               pending_count=pending_count,
                               in_progress_count=in_progress_count,
                               hot_count=hot_count,
                               warm_count=warm_count,
                               cold_count=cold_count,
                               not_interested_count=not_interested_count,
                               test_drive_yes_count=test_drive_yes_count,
                               test_drive_no_count=test_drive_no_count,
                               test_drive_total_count=test_drive_total_count)

    except Exception as e:
        flash(f'Error loading export dashboard: {str(e)}', 'error')
        return redirect(url_for('admin_dashboard'))


@app.route('/get_filtered_leads')
@require_admin
def get_filtered_leads():
    """Get filtered leads based on filter type and value"""
    try:
        filter_type = request.args.get('filter_type')
        filter_value = request.args.get('filter_value')

        if not filter_type or not filter_value:
            return jsonify({'success': False, 'message': 'Filter type and value are required'})

        # Get all leads
        all_leads = safe_get_data('lead_master')

        # Filter based on type
        if filter_type == 'final_status':
            filtered_leads = [l for l in all_leads if l.get('final_status') == filter_value]
        elif filter_type == 'lead_category':
            filtered_leads = [l for l in all_leads if l.get('lead_category') == filter_value]
        else:
            return jsonify({'success': False, 'message': 'Invalid filter type'})

        # Log filtered leads access
        auth_manager.log_audit_event(
            user_id=session.get('user_id'),
            user_type=session.get('user_type'),
            action='FILTERED_LEADS_ACCESS',
            resource='lead_master',
            details={'filter_type': filter_type, 'filter_value': filter_value, 'result_count': len(filtered_leads)}
        )

        return jsonify({
            'success': True,
            'leads': filtered_leads,
            'count': len(filtered_leads)
        })

    except Exception as e:
        print(f"Error getting filtered leads: {e}")
        return jsonify({'success': False, 'message': str(e)})


@app.route('/export_leads_csv')
@require_admin
def export_leads_csv():
    """Export filtered leads to CSV"""
    try:
        filter_type = request.args.get('filter_type')
        filter_value = request.args.get('filter_value')

        if not filter_type or not filter_value:
            return jsonify({'success': False, 'message': 'Filter type and value are required'})

        # Get all leads
        all_leads = safe_get_data('lead_master')

        # Filter based on type
        if filter_type == 'final_status':
            filtered_leads = [l for l in all_leads if l.get('final_status') == filter_value]
        elif filter_type == 'lead_category':
            filtered_leads = [l for l in all_leads if l.get('lead_category') == filter_value]
        else:
            return jsonify({'success': False, 'message': 'Invalid filter type'})

        # Create CSV content
        output = io.StringIO()
        writer = csv.writer(output)

        # Write headers
        headers = [
            'UID', 'Date', 'Customer Name', 'Mobile Number', 'Source', 'CRE Name', 'PS Name',
            'Lead Category', 'Model Interested', 'Branch', 'Lead Status', 'Final Status',
            'Follow Up Date', 'First Call Date', 'First Remark', 'Second Call Date', 'Second Remark',
            'Third Call Date', 'Third Remark', 'Fourth Call Date', 'Fourth Remark',
            'Fifth Call Date', 'Fifth Remark', 'Sixth Call Date', 'Sixth Remark',
            'Seventh Call Date', 'Seventh Remark', 'Assigned'
        ]
        writer.writerow(headers)

        # Write data rows
        for lead in filtered_leads:
            row = [
                lead.get('uid', ''),
                lead.get('date', ''),
                lead.get('customer_name', ''),
                lead.get('customer_mobile_number', ''),
                lead.get('source', ''),
                lead.get('cre_name', ''),
                lead.get('ps_name', ''),
                lead.get('lead_category', ''),
                lead.get('model_interested', ''),
                lead.get('branch', ''),
                lead.get('lead_status', ''),
                lead.get('final_status', ''),
                lead.get('follow_up_date', ''),
                lead.get('first_call_date', ''),
                lead.get('first_remark', ''),
                lead.get('second_call_date', ''),
                lead.get('second_remark', ''),
                lead.get('third_call_date', ''),
                lead.get('third_remark', ''),
                lead.get('fourth_call_date', ''),
                lead.get('fourth_remark', ''),
                lead.get('fifth_call_date', ''),
                lead.get('fifth_remark', ''),
                lead.get('sixth_call_date', ''),
                lead.get('sixth_remark', ''),
                lead.get('seventh_call_date', ''),
                lead.get('seventh_remark', ''),
                lead.get('assigned', '')
            ]
            writer.writerow(row)

        # Log CSV export
        auth_manager.log_audit_event(
            user_id=session.get('user_id'),
            user_type=session.get('user_type'),
            action='CSV_EXPORT',
            resource='lead_master',
            details={'filter_type': filter_type, 'filter_value': filter_value, 'exported_count': len(filtered_leads)}
        )

        # Create response
        output.seek(0)
        response = app.response_class(
            output.getvalue(),
            mimetype='text/csv',
            headers={
                'Content-Disposition': f'attachment; filename=leads_{filter_type}_{filter_value}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
            }
        )

        return response

    except Exception as e:
        print(f"Error exporting CSV: {e}")
        return jsonify({'success': False, 'message': str(e)})


@app.route('/get_leads_by_date_range')
@require_admin
def get_leads_by_date_range():
    """Get leads filtered by date range"""
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        date_field = request.args.get('date_field', 'date')

        if not start_date or not end_date:
            return jsonify({'success': False, 'message': 'Start date and end date are required'})

        # Validate date format
        try:
            datetime.strptime(start_date, '%Y-%m-%d')
            datetime.strptime(end_date, '%Y-%m-%d')
        except ValueError:
            return jsonify({'success': False, 'message': 'Invalid date format. Use YYYY-MM-DD'})

        # Validate date field
        valid_date_fields = ['date', 'first_call_date', 'follow_up_date', 'created_at']
        if date_field not in valid_date_fields:
            return jsonify({'success': False, 'message': 'Invalid date field'})

        # Get all leads
        all_leads = safe_get_data('lead_master')

        # Filter by date range
        filtered_leads = []
        for lead in all_leads:
            lead_date_str = lead.get(date_field)
            if lead_date_str:
                try:
                    # Handle different date formats
                    if 'T' in lead_date_str:  # ISO format with time
                        lead_date = datetime.fromisoformat(lead_date_str.replace('Z', '+00:00')).date()
                    else:  # Date only format
                        lead_date = datetime.strptime(lead_date_str, '%Y-%m-%d').date()

                    start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
                    end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()

                    if start_date_obj <= lead_date <= end_date_obj:
                        filtered_leads.append(lead)
                except (ValueError, TypeError):
                    # Skip leads with invalid dates
                    continue

        # Log date range filter access
        auth_manager.log_audit_event(
            user_id=session.get('user_id'),
            user_type=session.get('user_type'),
            action='DATE_RANGE_FILTER_ACCESS',
            resource='lead_master',
            details={
                'start_date': start_date,
                'end_date': end_date,
                'date_field': date_field,
                'result_count': len(filtered_leads)
            }
        )

        return jsonify({
            'success': True,
            'leads': filtered_leads,
            'count': len(filtered_leads)
        })

    except Exception as e:
        print(f"Error getting leads by date range: {e}")
        return jsonify({'success': False, 'message': str(e)})
@app.route('/export_leads_by_date_csv')
@require_admin
def export_leads_by_date_csv():
    """Export leads filtered by date range to CSV"""
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        date_field = request.args.get('date_field', 'date')

        if not start_date or not end_date:
            return jsonify({'success': False, 'message': 'Start date and end date are required'})

        # Validate date format
        try:
            datetime.strptime(start_date, '%Y-%m-%d')
            datetime.strptime(end_date, '%Y-%m-%d')
        except ValueError:
            return jsonify({'success': False, 'message': 'Invalid date format. Use YYYY-MM-DD'})

        # Validate date field
        valid_date_fields = ['date', 'first_call_date', 'follow_up_date', 'created_at']
        if date_field not in valid_date_fields:
            return jsonify({'success': False, 'message': 'Invalid date field'})

        # Get all leads
        all_leads = safe_get_data('lead_master')

        # Filter by date range
        filtered_leads = []
        for lead in all_leads:
            lead_date_str = lead.get(date_field)
            if lead_date_str:
                try:
                    # Handle different date formats
                    if 'T' in lead_date_str:  # ISO format with time
                        lead_date = datetime.fromisoformat(lead_date_str.replace('Z', '+00:00')).date()
                    else:  # Date only format
                        lead_date = datetime.strptime(lead_date_str, '%Y-%m-%d').date()

                    start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
                    end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()

                    if start_date_obj <= lead_date <= end_date_obj:
                        filtered_leads.append(lead)
                except (ValueError, TypeError):
                    # Skip leads with invalid dates
                    continue

        # Create CSV content
        output = io.StringIO()
        writer = csv.writer(output)

        # Write headers
        headers = [
            'UID', 'Date', 'Customer Name', 'Mobile Number', 'Source', 'CRE Name', 'PS Name',
            'Lead Category', 'Model Interested', 'Branch', 'Lead Status', 'Final Status',
            'Follow Up Date', 'First Call Date', 'First Remark', 'Second Call Date', 'Second Remark',
            'Third Call Date', 'Third Remark', 'Fourth Call Date', 'Fourth Remark',
            'Fifth Call Date', 'Fifth Remark', 'Sixth Call Date', 'Sixth Remark',
            'Seventh Call Date', 'Seventh Remark', 'Assigned', 'Created At'
        ]
        writer.writerow(headers)

        # Write data rows
        for lead in filtered_leads:
            row = [
                lead.get('uid', ''),
                lead.get('date', ''),
                lead.get('customer_name', ''),
                lead.get('customer_mobile_number', ''),
                lead.get('source', ''),
                lead.get('cre_name', ''),
                lead.get('ps_name', ''),
                lead.get('lead_category', ''),
                lead.get('model_interested', ''),
                lead.get('branch', ''),
                lead.get('lead_status', ''),
                lead.get('final_status', ''),
                lead.get('follow_up_date', ''),
                lead.get('first_call_date', ''),
                lead.get('first_remark', ''),
                lead.get('second_call_date', ''),
                lead.get('second_remark', ''),
                lead.get('third_call_date', ''),
                lead.get('third_remark', ''),
                lead.get('fourth_call_date', ''),
                lead.get('fourth_remark', ''),
                lead.get('fifth_call_date', ''),
                lead.get('fifth_remark', ''),
                lead.get('sixth_call_date', ''),
                lead.get('sixth_remark', ''),
                lead.get('seventh_call_date', ''),
                lead.get('seventh_remark', ''),
                lead.get('assigned', ''),
                lead.get('created_at', '')
            ]
            writer.writerow(row)

        # Log CSV export
        auth_manager.log_audit_event(
            user_id=session.get('user_id'),
            user_type=session.get('user_type'),
            action='DATE_RANGE_CSV_EXPORT',
            resource='lead_master',
            details={
                'start_date': start_date,
                'end_date': end_date,
                'date_field': date_field,
                'exported_count': len(filtered_leads)
            }
        )

        # Create response
        output.seek(0)
        response = app.response_class(
            output.getvalue(),
            mimetype='text/csv',
            headers={
                'Content-Disposition': f'attachment; filename=leads_date_range_{start_date}_to_{end_date}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
            }
        )

        return response

    except Exception as e:
        print(f"Error exporting date range CSV: {e}")
        return jsonify({'success': False, 'message': str(e)})


@app.route('/get_test_drive_leads')
@require_admin
def get_test_drive_leads():
    """Get test drive leads from alltest_drive table"""
    try:
        filter_type = request.args.get('filter_type', 'all')  # all, yes, no
        
        # Get all test drive leads
        test_drive_leads = safe_get_data('alltest_drive')
        
        # Filter based on type
        if filter_type == 'yes':
            filtered_leads = [l for l in test_drive_leads if l.get('test_drive_done') == 'Yes']
        elif filter_type == 'no':
            filtered_leads = [l for l in test_drive_leads if l.get('test_drive_done') == 'No']
        else:  # all
            filtered_leads = test_drive_leads
        
        # Log test drive leads access
        auth_manager.log_audit_event(
            user_id=session.get('user_id'),
            user_type=session.get('user_type'),
            action='TEST_DRIVE_LEADS_ACCESS',
            resource='alltest_drive',
            details={'filter_type': filter_type, 'result_count': len(filtered_leads)}
        )
        
        return jsonify({
            'success': True,
            'leads': filtered_leads,
            'count': len(filtered_leads)
        })
        
    except Exception as e:
        print(f"Error getting test drive leads: {e}")
        return jsonify({'success': False, 'message': str(e)})


@app.route('/export_test_drive_csv')
@require_admin
def export_test_drive_csv():
    """Export test drive leads to CSV"""
    try:
        filter_type = request.args.get('filter_type', 'all')  # all, yes, no
        
        # Get all test drive leads
        test_drive_leads = safe_get_data('alltest_drive')
        
        # Filter based on type
        if filter_type == 'yes':
            filtered_leads = [l for l in test_drive_leads if l.get('test_drive_done') == 'Yes']
        elif filter_type == 'no':
            filtered_leads = [l for l in test_drive_leads if l.get('test_drive_done') == 'No']
        else:  # all
            filtered_leads = test_drive_leads
        
        # Create CSV content
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write headers
        headers = [
            'Original ID', 'Customer Name', 'Mobile Number', 'Test Drive Done',
            'Lead Status', 'Lead Category', 'Model Interested', 'Final Status', 'PS Name',
            'Branch', 'Created At', 'Updated At', 'Remarks', 'Activity Name', 'Activity Location',
            'Customer Location', 'Customer Profession', 'Gender', 'Lead Source', 'CRE Name'
        ]
        writer.writerow(headers)
        
        # Write data
        for lead in filtered_leads:
            row = [
                lead.get('original_id', ''),
                lead.get('customer_name', ''),
                lead.get('mobile_number', ''),
                lead.get('test_drive_done', ''),
                lead.get('lead_status', ''),
                lead.get('lead_category', ''),
                lead.get('model_interested', ''),
                lead.get('final_status', ''),
                lead.get('ps_name', ''),
                lead.get('branch', ''),
                lead.get('created_at', ''),
                lead.get('updated_at', ''),
                lead.get('remarks', ''),
                lead.get('activity_name', ''),
                lead.get('activity_location', ''),
                lead.get('customer_location', ''),
                lead.get('customer_profession', ''),
                lead.get('gender', ''),
                lead.get('lead_source', ''),
                lead.get('cre_name', '')
            ]
            writer.writerow(row)
        
        # Log export
        auth_manager.log_audit_event(
            user_id=session.get('user_id'),
            user_type=session.get('user_type'),
            action='TEST_DRIVE_EXPORT',
            resource='alltest_drive',
            details={'filter_type': filter_type, 'exported_count': len(filtered_leads)}
        )
        
        # Create response
        output.seek(0)
        response = app.response_class(
            output.getvalue(),
            mimetype='text/csv',
            headers={
                'Content-Disposition': f'attachment; filename=test_drive_leads_{filter_type}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
            }
        )
        
        return response
        
    except Exception as e:
        print(f"Error exporting test drive CSV: {e}")
        return jsonify({'success': False, 'message': str(e)})


@app.route('/activity_event', methods=['GET', 'POST'])
@require_ps
def activity_event():
    """Add event/activity leads for PS"""
    if request.method == 'POST':
        # Get form data for multiple customers
        customer_names = request.form.getlist('customer_name[]')
        customer_phones = request.form.getlist('customer_phone_number[]')
        customer_locations = request.form.getlist('customer_location[]')
        customer_professions = request.form.getlist('customer_profession[]')
        genders = request.form.getlist('gender[]')
        interested_models = request.form.getlist('interested_model[]')
        remarks_list = request.form.getlist('remarks[]')
        lead_statuses = request.form.getlist('lead_status[]')
        months = request.form.getlist('month[]')
        dates = request.form.getlist('date[]')
        
        # Get activity details
        activity_name = request.form.get('activity_name', '').strip()
        activity_location = request.form.get('activity_location', '').strip()
        
        # Validation
        if not activity_name or not activity_location:
            flash('Please fill activity details', 'error')
            return render_template('activity_event.html', now=datetime.now())
        
        if not customer_names or not customer_phones:
            flash('Please add at least one customer lead', 'error')
            return render_template('activity_event.html', now=datetime.now())
        
        try:
            leads_added = 0
            duplicate_leads = []
            # Fetch all CREs for round-robin assignment
            cre_users = safe_get_data('cre_users')
            cre_names = [cre['name'] for cre in cre_users] if cre_users else []
            cre_count = len(cre_names)
            cre_index = 0  # Start from the first CRE for each batch
            
            for i in range(len(customer_names)):
                if customer_names[i] and customer_phones[i]:  # Only process if name and phone are provided
                    customer_phone = customer_phones[i].strip()
                    normalized_phone = ''.join(filter(str.isdigit, customer_phone))
                    
                    # Initialize variables for this iteration
                    should_add_to_duplicates = False
                    original_lead = None
                    
                    # Check for duplicates in ALL relevant tables
                    try:
                        print(f"🔍 Checking for duplicates for event lead phone: {normalized_phone}")
                        
                        # 1. Check in activity_leads table FIRST (most important for event leads)
                        activity_result = supabase.table('activity_leads').select('*').eq('customer_phone_number', normalized_phone).execute()
                        existing_activity_leads = activity_result.data or []
                        
                        if existing_activity_leads:
                            print(f"🎯 Found {len(existing_activity_leads)} existing activity leads")
                            duplicate_leads.append({
                                'phone': customer_phone,
                                'name': customer_names[i].strip(),
                                'type': 'exact_match',
                                'message': 'Lead with this phone number already exists as Event lead'
                            })
                            continue  # Skip this lead
                        
                        # 2. Check in walkin_table
                        walkin_result = supabase.table('walkin_table').select('*').eq('mobile_number', normalized_phone).execute()
                        existing_walkin_leads = walkin_result.data or []
                        
                        if existing_walkin_leads:
                            print(f"📱 Found {len(existing_walkin_leads)} existing walkin leads")
                            # This will be added as duplicate with new source
                            should_add_to_duplicates = True
                            original_lead = existing_walkin_leads[0]
                            print(f"✅ Will add to duplicate_leads: Walk-in → Event")
                        
                        # 3. Check in lead_master table
                        lead_master_result = supabase.table('lead_master').select('*').eq('customer_mobile_number', normalized_phone).execute()
                        existing_leads = lead_master_result.data or []
                        
                        if existing_leads:
                            print(f"📋 Found {len(existing_leads)} existing leads in lead_master")
                            # Check for exact source-subsource match (Event)
                            exact_match = any(
                                lead.get('source') == 'Event' 
                                for lead in existing_leads
                            )
                            
                            if exact_match:
                                duplicate_leads.append({
                                    'phone': customer_phone,
                                    'name': customer_names[i].strip(),
                                    'type': 'exact_match',
                                    'message': 'Lead with this phone number already exists as Event'
                                })
                                continue  # Skip this lead
                            else:
                                # This will be added as duplicate with new source
                                should_add_to_duplicates = True
                                original_lead = existing_leads[0]
                                print(f"✅ Will add to duplicate_leads: CRE Lead → Event")
                        
                        # 4. Check in duplicate_leads
                        duplicate_result = supabase.table('duplicate_leads').select('*').eq('customer_mobile_number', normalized_phone).execute()
                        existing_duplicate_leads = duplicate_result.data or []
                        
                        if existing_duplicate_leads:
                            print(f"🔄 Found {len(existing_duplicate_leads)} existing duplicate leads")
                            # Check if Event source already exists in any slot
                            exact_match = False
                            for duplicate_lead in existing_duplicate_leads:
                                # Check all source slots (source1 to source10)
                                for j in range(1, 11):
                                    source_field = f'source{j}'
                                    
                                    if duplicate_lead.get(source_field) == 'Event':
                                        exact_match = True
                                        break
                                if exact_match:
                                    break
                            
                            if exact_match:
                                duplicate_leads.append({
                                    'phone': customer_phone,
                                    'name': customer_names[i].strip(),
                                    'type': 'exact_match',
                                    'message': 'Lead with this phone number already exists as Event in duplicates'
                                })
                                continue  # Skip this lead
                            else:
                                # This will be added as duplicate with new source
                                should_add_to_duplicates = True
                                original_lead = existing_duplicate_leads[0]
                                print(f"✅ Will add to duplicate_leads: Existing Duplicate → Event")
                        
                        # Check if we should add to duplicate_leads instead of activity_leads
                        if should_add_to_duplicates and original_lead:
                            try:
                                print(f"🔄 Creating duplicate record for {customer_phone}...")
                                print(f"🔄 Original lead data: {original_lead}")
                                
                                # Generate UID for duplicate record
                                duplicate_uid = f"E-{activity_name[:3].upper()}-{customer_phones[i][-4:]}"
                                print(f"🔄 Generated duplicate UID: {duplicate_uid}")
                                
                                # Create duplicate record
                                duplicate_data = {
                                    'uid': duplicate_uid,
                                    'customer_mobile_number': normalized_phone,
                                    'customer_name': customer_names[i].strip(),
                                    'original_lead_id': None,  # We'll handle this differently
                                    'source1': original_lead.get('source', 'Event' if 'activity_uid' in original_lead else 'Walk-in' if 'id' in original_lead else 'Unknown'),
                                    'sub_source1': original_lead.get('sub_source') or original_lead.get('location', 'N/A') or original_lead.get('branch', 'N/A'),
                                    'date1': original_lead.get('date') or original_lead.get('created_at', ''),
                                    'source2': 'Event',
                                    'sub_source2': activity_name,  # Use activity name as sub-source for event
                                    'date2': datetime.now().strftime('%Y-%m-%d'),
                                    'duplicate_count': 2,
                                    'created_at': datetime.now().isoformat(),
                                    'updated_at': datetime.now().isoformat()
                                }
                                
                                print(f"🔄 Duplicate data prepared: {duplicate_data}")
                                
                                # Try to add the new columns if they exist (for better source tracking)
                                try:
                                    duplicate_data['original_table'] = 'walkin_table' if 'id' in original_lead else 'activity_leads' if 'activity_uid' in original_lead else 'lead_master'
                                    duplicate_data['original_record_id'] = original_lead.get('id') or original_lead.get('activity_uid') or original_lead.get('uid')
                                    print(f"🔄 Added source tracking: {duplicate_data.get('original_table')} - {duplicate_data.get('original_record_id')}")
                                except Exception as e:
                                    print(f"⚠️ Could not add source tracking columns: {e}")
                                
                                # Initialize remaining slots as None
                                for k in range(3, 11):
                                    duplicate_data[f'source{k}'] = None
                                    duplicate_data[f'sub_source{k}'] = None
                                    duplicate_data[f'date{k}'] = None
                                
                                print(f"🔄 Attempting to insert into duplicate_leads...")
                                result = supabase.table('duplicate_leads').insert(duplicate_data).execute()
                                print(f"✅ Successfully added to duplicate_leads: {customer_names[i].strip()} - {customer_phone}")
                                print(f"✅ Insert result: {result}")
                                continue  # Skip adding to activity_leads
                                
                            except Exception as e:
                                print(f"❌ Error creating duplicate record for {customer_phone}: {e}")
                                print(f"❌ Full error details: {str(e)}")
                                # Continue with normal flow if duplicate creation fails
                        
                        # No exact duplicate found, proceed with adding the lead
                        # Assign CRE in round-robin
                        cre_assigned = cre_names[cre_index] if cre_count > 0 else None
                        cre_index = (cre_index + 1) % cre_count if cre_count > 0 else 0
                        
                        # Generate UID for event lead
                        uid = f"E-{activity_name[:3].upper()}-{customer_phones[i][-4:]}"
                        
                        # Create event lead data strictly matching activity_leads schema
                        event_lead_data = {
                            'activity_uid': uid,
                            'activity_name': activity_name,
                            'activity_location': activity_location,
                            'ps_name': session.get('ps_name'),
                            'location': session.get('branch'),
                            'customer_name': customer_names[i].strip(),
                            'customer_location': customer_locations[i].strip() if i < len(customer_locations) else '',
                            'customer_profession': customer_professions[i].strip() if i < len(customer_professions) else '',
                            'gender': genders[i] if i < len(genders) else 'MALE',
                            'interested_model': interested_models[i].strip() if i < len(interested_models) else '',
                            'remarks': remarks_list[i].strip() if i < len(remarks_list) else '',
                            'lead_status': lead_statuses[i] if i < len(lead_statuses) else 'WARM',
                            'month': months[i] if i < len(months) else datetime.now().strftime('%b'),
                            'date': dates[i] if i < len(dates) else datetime.now().strftime('%Y-%m-%d'),
                            'customer_phone_number': customer_phone,
                            'created_at': datetime.now().isoformat(),
                            'lead_category': lead_statuses[i] if i < len(lead_statuses) else 'WARM',
                            'cre_assigned': cre_assigned
                        }
                        
                        # Insert into activity_leads
                        result = supabase.table('activity_leads').insert(event_lead_data).execute()
                        if result.data:
                            leads_added += 1
                            print(f"✅ Added event lead: {customer_names[i].strip()} - {customer_phone}")
                        
                    except Exception as e:
                        print(f"❌ Error checking duplicates for {customer_phone}: {e}")
                        # Continue with other leads even if one fails
                        continue
            
            if leads_added > 0:
                if duplicate_leads:
                    duplicate_message = f"{leads_added} event lead(s) added successfully! {len(duplicate_leads)} exact duplicate(s) skipped."
                    flash(duplicate_message, 'warning')
                else:
                    flash(f'{leads_added} event lead(s) added successfully!', 'success')
                return redirect(url_for('ps_dashboard'))
            else:
                if duplicate_leads:
                    flash(f'All leads were exact duplicates. {len(duplicate_leads)} duplicate(s) found.', 'warning')
                else:
                    flash('Error adding event leads', 'error')
                
        except Exception as e:
            flash(f'Error adding event leads: {str(e)}', 'error')
    
    return render_template('activity_event.html', now=datetime.now())


@app.route('/view_event_leads')
@require_ps
def view_event_leads():
    """View all event/activity leads for current PS"""
    try:
        ps_name = session.get('ps_name')
        # Get event leads for this PS from activity_leads table
        event_leads = safe_get_data('activity_leads', {'ps_name': ps_name})
        
        # Transform data to match template expectations
        transformed_leads = []
        for lead in event_leads:
            transformed_lead = {
                'activity_name': lead.get('activity_name', ''),
                'activity_location': lead.get('activity_location', ''),
                'ps_name': lead.get('ps_name', ''),
                'location': lead.get('location', ''),
                'customer_name': lead.get('customer_name', ''),
                'customer_phone_number': lead.get('customer_phone_number', ''),
                'customer_location': lead.get('customer_location', ''),
                'customer_profession': lead.get('customer_profession', ''),
                'gender': lead.get('gender', ''),
                'interested_model': lead.get('interested_model', ''),
                'remarks': lead.get('remarks', ''),
                'lead_status': lead.get('lead_status', ''),
                'month': lead.get('month', ''),
                'date': lead.get('date', ''),
                'activity_uid': lead.get('activity_uid', '')
            }
            transformed_leads.append(transformed_lead)
        
        return render_template('view_event_leads.html', event_leads=transformed_leads)
        
    except Exception as e:
        flash(f'Error loading event leads: {str(e)}', 'error')
        return redirect(url_for('ps_dashboard'))




@app.route('/view_call_attempt_history/<uid>')
@require_admin
def view_call_attempt_history(uid):
    """View call attempt history for a specific lead"""
    try:
        # Get lead data
        lead_result = supabase.table('lead_master').select('*').eq('uid', uid).execute()
        if not lead_result.data:
            flash('Lead not found', 'error')
            return redirect(url_for('admin_dashboard'))
        
        lead_data = lead_result.data[0]
        
        # Get call attempt history
        history_result = supabase.table('cre_call_attempt_history').select('*').eq('uid', uid).order('created_at', desc=True).execute()
        call_history = history_result.data if history_result.data else []
        
        # Add total_attempts
        total_attempts = len(call_history)
        
        # Log access
        auth_manager.log_audit_event(
            user_id=session.get('user_id'),
            user_type=session.get('user_type'),
            action='CALL_HISTORY_ACCESS',
            resource='cre_call_attempt_history',
            resource_id=uid,
            details={'lead_uid': uid, 'history_count': len(call_history)}
        )
        
        return render_template('call_attempt_history.html', 
                               lead=lead_data, 
                               call_history=call_history,
                               total_attempts=total_attempts)
        
    except Exception as e:
        flash(f'Error loading call history: {str(e)}', 'error')
        return redirect(url_for('admin_dashboard'))


@app.route('/source_analysis_data')
@require_admin
def source_analysis_data():
    try:
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        today = datetime.now().date()
        start_date = None
        end_date = None
        if start_date_str and end_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except Exception:
                start_date = None
                end_date = None
        sources = [
            'Google (KNOW)', 'Google', 'Meta(KNOW)', 'Know', 'OEM Web', 'OEM Tele',
            'Affiliate Bikewale', 'Affiliate 91wheels', 'Affiliate Bikedekho',
            'BTL (KNOW)', 'Meta'
        ]
        all_leads = safe_get_data('lead_master')
        all_followups = safe_get_data('ps_followup_master')
        def in_date_range(lead):
            lead_date_str = lead.get('created_at') or lead.get('date')
            if not lead_date_str:
                return True
            try:
                if 'T' in lead_date_str:
                    lead_date = datetime.fromisoformat(lead_date_str.replace('Z', '+00:00')).date()
                else:
                    lead_date = datetime.strptime(lead_date_str, '%Y-%m-%d').date()
            except Exception:
                return True
            if start_date and end_date:
                return start_date <= lead_date <= end_date
            elif start_date:
                return lead_date >= start_date
            else:
                return True
        leads = [l for l in all_leads if in_date_range(l)]
        data = {}
        for source in sources:
            data[source] = {
                'calls_allocated': 0,
                'calls_attempted': 0,
                'calls_connected': 0,
                'pending_total': 0,
                'pending_interested': 0,
                'pending_hot': 0,
                'pending_warm': 0,
                'pending_cold': 0,
                'pending_callback': 0,
                'pending_rnr': 0,
                'pending_disconnected': 0,
                'closed_lost_total': 0,
                'closed_lost_competition': 0,
                'closed_lost_not_interested': 0,
                'closed_lost_did_not_enquire': 0,
                'closed_lost_wrong_lead': 0,
                'closed_lost_rnr_lost': 0,
                'closed_lost_finance_reject': 0,
                'closed_lost_co_dealer': 0,
                'sales_pipeline_total': 0,
                'call_progress': {},
                'latest_remark': '',
            }
        def get_source(lead):
            s = lead.get('source', '').strip()
            for src in sources:
                if s.lower() == src.lower():
                    return src
            for src in sources:
                if src.lower() in s.lower():
                    return src
            return None
        for lead in leads:
            src = get_source(lead)
            if not src:
                continue
            data[src]['calls_allocated'] += 1
            attempted = any(lead.get(f'{c}_call_date') for c in ['first', 'second', 'third', 'fourth', 'fifth', 'sixth', 'seventh'])
            if attempted:
                data[src]['calls_attempted'] += 1
            call_statuses = [lead.get(f'{c}_call_status', '').lower() for c in ['first', 'second', 'third', 'fourth', 'fifth', 'sixth', 'seventh']]
            connected = any(s and s not in ['rnr', 'busy', 'busy on another call'] for s in call_statuses)
            if connected:
                data[src]['calls_connected'] += 1
            status = (lead.get('lead_status') or '').strip().lower()
            if status in ['interested', 'call back', 'rnr', 'call disconnected']:
                data[src]['pending_total'] += 1
                if status == 'interested':
                    data[src]['pending_interested'] += 1
                    cat = (lead.get('lead_category') or '').strip().lower()
                    if cat == 'hot':
                        data[src]['pending_hot'] += 1
                    elif cat == 'warm':
                        data[src]['pending_warm'] += 1
                    elif cat == 'cold':
                        data[src]['pending_cold'] += 1
                elif status == 'call back':
                    data[src]['pending_callback'] += 1
                elif status == 'rnr':
                    data[src]['pending_rnr'] += 1
                elif status == 'call disconnected':
                    data[src]['pending_disconnected'] += 1
            # --- NEW LOGIC: Use latest follow-up for final_status and lead_status for lost reasons ---
            uid = lead.get('uid')
            latest_final_status = (lead.get('final_status') or '').strip().lower()
            latest_lost_reason = (lead.get('lost_reason') or '').strip().lower()
            latest_lead_status = (lead.get('lead_status') or '').strip().lower()
            if uid:
                followups = [f for f in all_followups if f.get('lead_uid') == uid]
                if followups:
                    latest_fu = max(followups, key=lambda f: f.get('created_at', ''))
                    if latest_fu.get('final_status'):
                        latest_final_status = latest_fu.get('final_status').strip().lower()
                    if latest_fu.get('lost_reason'):
                        latest_lost_reason = latest_fu.get('lost_reason').strip().lower()
                    if latest_fu.get('lead_status'):
                        latest_lead_status = latest_fu.get('lead_status').strip().lower()
            # --- END NEW LOGIC ---
            if latest_final_status == 'lost':
                data[src]['closed_lost_total'] += 1
                # Use lead_status for lost reason
                if latest_lead_status == 'rnr-lost':
                    data[src]['closed_lost_rnr_lost'] += 1
                elif latest_lead_status == 'did not enquire':
                    data[src]['closed_lost_did_not_enquire'] += 1
                elif latest_lead_status == 'not interested':
                    data[src]['closed_lost_not_interested'] += 1
                elif latest_lead_status == 'lost to competition':
                    data[src]['closed_lost_competition'] += 1
                elif latest_lead_status == 'lost to co-dealer':
                    data[src]['closed_lost_co_dealer'] += 1
                elif latest_lead_status == 'finance rejected':
                    data[src]['closed_lost_finance_reject'] += 1
                elif latest_lead_status == 'wrong lead':
                    data[src]['closed_lost_wrong_lead'] += 1
            if latest_final_status == 'won':
                data[src]['sales_pipeline_total'] += 1
            call_progress = None
            for idx, c in enumerate(['seventh', 'sixth', 'fifth', 'fourth', 'third', 'second', 'first']):
                if lead.get(f'{c}_call_date'):
                    call_progress = c.capitalize()
                    break
            if call_progress:
                data[src]['call_progress'][lead.get('uid')] = call_progress
            if uid:
                followups = [f for f in all_followups if f.get('lead_uid') == uid]
                if followups:
                    latest = max(followups, key=lambda f: f.get('created_at', ''))
                    data[src]['latest_remark'] = latest.get('remark', '')
        total_row = {k: sum(data[src][k] for src in sources) if isinstance(data[sources[0]][k], int) else {} for k in data[sources[0]].keys()}
        data['Total'] = total_row
        return jsonify({'success': True, 'data': data})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/cre_call_attempt_history_json/<uid>')
@require_auth(['cre', 'admin'])
def cre_call_attempt_history_json(uid):
    """Return call attempt history for a lead as JSON (for CRE dashboard modal)"""
    try:
        # Get all call attempt history for this lead
        history_result = supabase.table('cre_call_attempt_history').select('uid,call_no,attempt,status,cre_name,update_ts').eq('uid', uid).execute()
        history = history_result.data if history_result.data else []
        # Define the order of call numbers
        call_order = ['first', 'second', 'third', 'fourth', 'fifth', 'sixth', 'seventh']
        # Sort the history: first by call_no in ascending order, then by attempt in ascending order
        def sort_key(item):
            call_no = item.get('call_no', '')
            attempt = item.get('attempt', 0)
            # Get the index of call_no in the predefined order, default to end if not found
            call_index = call_order.index(call_no) if call_no in call_order else len(call_order)
            return (call_index, attempt)
        sorted_history = sorted(history, key=sort_key)
        return jsonify({'success': True, 'history': sorted_history})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e), 'history': []})


@app.route('/ps_call_attempt_history_json/<uid>')
@require_auth(['ps', 'admin'])
def ps_call_attempt_history_json(uid):
    """Return PS call attempt history for a lead as JSON (for PS dashboard modal)"""
    try:
        # Get all PS call attempt history for this lead
        history_result = supabase.table('ps_call_attempt_history').select('uid,call_no,attempt,status,ps_name,created_at').eq('uid', uid).execute()
        history = history_result.data if history_result.data else []
        # Define the order of call numbers
        call_order = ['first', 'second', 'third', 'fourth', 'fifth', 'sixth', 'seventh']
        # Sort the history: first by call_no in ascending order, then by attempt in ascending order
        def sort_key(item):
            call_no = item.get('call_no', '')
            attempt = item.get('attempt', 0)
            # Get the index of call_no in the predefined order, default to end if not found
            call_index = call_order.index(call_no) if call_no in call_order else len(call_order)
            return (call_index, attempt)
        sorted_history = sorted(history, key=sort_key)
        return jsonify({'success': True, 'history': sorted_history})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e), 'history': []})


@app.route('/view_ps_call_attempt_history/<uid>')
@require_admin
def view_ps_call_attempt_history(uid):
    """View PS call attempt history for a specific lead (Admin only)"""
    try:
        # Get lead data
        lead_result = supabase.table('lead_master').select('*').eq('uid', uid).execute()
        if not lead_result.data:
            flash('Lead not found', 'error')
            return redirect(url_for('admin_dashboard'))
        
        lead_data = lead_result.data[0]
        
        # Get PS call attempt history
        history_result = supabase.table('ps_call_attempt_history').select('*').eq('uid', uid).order('created_at', desc=True).execute()
        call_history = history_result.data if history_result.data else []
        
        # Add total_attempts
        total_attempts = len(call_history)
        
        # Log access
        auth_manager.log_audit_event(
            user_id=session.get('user_id'),
            user_type=session.get('user_type'),
            action='PS_CALL_HISTORY_ACCESS',
            resource='ps_call_attempt_history',
            resource_id=uid,
            details={'lead_uid': uid, 'history_count': len(call_history)}
        )
        
        return render_template('ps_call_attempt_history.html', 
                               lead=lead_data, 
                               call_history=call_history,
                               total_attempts=total_attempts)
        
    except Exception as e:
        flash(f'Error loading PS call history: {str(e)}', 'error')
        return redirect(url_for('admin_dashboard'))


@app.route('/get_unassigned_leads_by_source')
@require_admin
def get_unassigned_leads_by_source():
    source = request.args.get('source')
    if not source:
        return jsonify({'success': False, 'message': 'Source is required'}), 400
    try:
        leads = safe_get_data('lead_master', {'assigned': 'No', 'source': source})
        return jsonify({'success': True, 'leads': leads})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/lead_journey/<uid>')
@require_admin
def lead_journey(uid):
    """Return the complete journey of a lead as JSON for analysis and visualization."""
    try:
        # Fetch lead info
        lead_result = supabase.table('lead_master').select('*').eq('uid', uid).execute()
        if not lead_result.data:
            return jsonify({'success': False, 'message': 'Lead not found'}), 404
        lead = lead_result.data[0]

        # Fetch CRE call attempts (all attempts, all calls)
        cre_calls_result = supabase.table('cre_call_attempt_history').select('*').eq('uid', uid).order('created_at', desc=False).execute()
        cre_calls = cre_calls_result.data if cre_calls_result.data else []

        # Fetch PS call attempts (all attempts, all calls)
        ps_calls_result = supabase.table('ps_call_attempt_history').select('*').eq('uid', uid).order('created_at', desc=False).execute()
        ps_calls = ps_calls_result.data if ps_calls_result.data else []

        # Fetch PS followup (if any)
        ps_followup_result = supabase.table('ps_followup_master').select(
            'lead_uid, ps_name, follow_up_date, final_status, source, lead_status, customer_name, customer_mobile_number, ps_assigned_at, '
            'first_call_remark, first_call_date, '
            'second_call_remark, second_call_date, '
            'third_call_remark, third_call_date, '
            'fourth_call_remark, fourth_call_date, '
            'fifth_call_remark, fifth_call_date, '
            'sixth_call_remark, sixth_call_date, '
            'seventh_call_remark, seventh_call_date'
        ).eq('lead_uid', uid).execute()
        ps_followup = ps_followup_result.data[0] if ps_followup_result.data else None

        # Assignment history (CRE, PS)
        assignment_history = []
        if lead.get('cre_name'):
            assignment_history.append({
                'role': 'CRE',
                'name': lead.get('cre_name'),
                'assigned_at': lead.get('cre_assigned_at')
            })
        if lead.get('ps_name'):
            assignment_history.append({
                'role': 'PS',
                'name': lead.get('ps_name'),
                'assigned_at': lead.get('ps_assigned_at')
            })

        # Status timeline (CRE and PS status changes)
        status_timeline = []
        # From lead_master: initial, final, and call statuses
        for call in ['first', 'second', 'third', 'fourth', 'fifth', 'sixth', 'seventh']:
            call_date = lead.get(f'{call}_call_date')
            call_remark = lead.get(f'{call}_remark')
            if call_date or call_remark:
                status_timeline.append({
                    'by': 'CRE',
                    'call_no': call,
                    'date': call_date,
                    'remark': call_remark
                })
        # From PS followup: call dates/remarks
        if ps_followup:
            for call in ['first', 'second', 'third', 'fourth', 'fifth', 'sixth', 'seventh']:
                call_date = ps_followup.get(f'{call}_call_date')
                call_remark = ps_followup.get(f'{call}_call_remark')
                if call_date or call_remark:
                    status_timeline.append({
                        'by': 'PS',
                        'call_no': call,
                        'date': call_date,
                        'remark': call_remark
                    })
        # Final status
        if lead.get('final_status'):
            status_timeline.append({
                'by': 'System',
                'status': lead.get('final_status'),
                'date': lead.get('won_timestamp') or lead.get('lost_timestamp')
            })

        # Conversation log (all remarks, CRE and PS, with timestamps)
        conversation_log = []
        # CRE call attempts
        for call in cre_calls:
            conversation_log.append({
                'by': 'CRE',
                'name': call.get('cre_name'),
                'call_no': call.get('call_no'),
                'attempt': call.get('attempt'),
                'status': call.get('status'),
                'remark': call.get('remarks'),
                'follow_up_date': call.get('follow_up_date'),
                'timestamp': call.get('update_ts') or call.get('created_at')
            })
        # PS call attempts
        for call in ps_calls:
            conversation_log.append({
                'by': 'PS',
                'name': call.get('ps_name'),
                'call_no': call.get('call_no'),
                'attempt': call.get('attempt'),
                'status': call.get('status'),
                'remark': call.get('remarks'),
                'follow_up_date': call.get('follow_up_date'),
                'timestamp': call.get('created_at')
            })
        # Sort conversation log by timestamp
        conversation_log = [c for c in conversation_log if c['timestamp']]  # Remove None timestamps
        conversation_log.sort(key=lambda x: x['timestamp'])

        # Final outcome
        outcome = {
            'final_status': lead.get('final_status'),
            'timestamp': lead.get('won_timestamp') or lead.get('lost_timestamp'),
            'reason': lead.get('lost_reason') if lead.get('final_status') == 'Lost' else None
        }

        # Compose response
        journey = {
            'lead': lead,
            'assignment_history': assignment_history,
            'status_timeline': status_timeline,
            'cre_calls': cre_calls,
            'ps_calls': ps_calls,
            'ps_followup': ps_followup,
            'conversation_log': conversation_log,
            'outcome': outcome
        }
        return jsonify({'success': True, 'journey': journey})
    except Exception as e:
        print(f"Error in lead_journey: {e}")
        return jsonify({'success': False, 'message': str(e)})
@app.route('/lead_journey_report/<uid>')
@require_admin
def lead_journey_report(uid):
    """Generate a PDF report for the lead journey, including static charts (Seaborn/Matplotlib)."""
    try:
        # Fetch journey data using the same logic as /lead_journey/<uid>
        lead_result = supabase.table('lead_master').select('*').eq('uid', uid).execute()
        if not lead_result.data:
            return 'Lead not found', 404
        lead = lead_result.data[0]
        cre_calls_result = supabase.table('cre_call_attempt_history').select('*').eq('uid', uid).order('created_at', desc=False).execute()
        cre_calls = cre_calls_result.data if cre_calls_result.data else []
        ps_calls_result = supabase.table('ps_call_attempt_history').select('*').eq('uid', uid).order('created_at', desc=False).execute()
        ps_calls = ps_calls_result.data if ps_calls_result.data else []
        ps_followup_result = supabase.table('ps_followup_master').select(
            'lead_uid, ps_name, follow_up_date, final_status, source, lead_status, customer_name, customer_mobile_number, ps_assigned_at, '
            'first_call_remark, first_call_date, '
            'second_call_remark, second_call_date, '
            'third_call_remark, third_call_date, '
            'fourth_call_remark, fourth_call_date, '
            'fifth_call_remark, fifth_call_date, '
            'sixth_call_remark, sixth_call_date, '
            'seventh_call_remark, seventh_call_date'
        ).eq('lead_uid', uid).execute()
        ps_followup = ps_followup_result.data[0] if ps_followup_result.data else None
        assignment_history = []
        if lead.get('cre_name'):
            assignment_history.append({'role': 'CRE','name': lead.get('cre_name'),'assigned_at': lead.get('cre_assigned_at')})
        if lead.get('ps_name'):
            assignment_history.append({'role': 'PS','name': lead.get('ps_name'),'assigned_at': lead.get('ps_assigned_at')})
        status_timeline = []
        for call in ['first', 'second', 'third', 'fourth', 'fifth', 'sixth', 'seventh']:
            call_date = lead.get(f'{call}_call_date')
            call_remark = lead.get(f'{call}_remark')
            if call_date or call_remark:
                status_timeline.append({'by': 'CRE','call_no': call,'date': call_date,'remark': call_remark})
        if ps_followup:
            for call in ['first', 'second', 'third', 'fourth', 'fifth', 'sixth', 'seventh']:
                call_date = ps_followup.get(f'{call}_call_date')
                call_remark = ps_followup.get(f'{call}_call_remark')
                if call_date or call_remark:
                    status_timeline.append({'by': 'PS','call_no': call,'date': call_date,'remark': call_remark})
        if lead.get('final_status'):
            status_timeline.append({'by': 'System','status': lead.get('final_status'),'date': lead.get('won_timestamp') or lead.get('lost_timestamp')})
        conversation_log = []
        for call in cre_calls:
            conversation_log.append({'by': 'CRE','name': call.get('cre_name'),'call_no': call.get('call_no'),'attempt': call.get('attempt'),'status': call.get('status'),'remark': call.get('remarks'),'follow_up_date': call.get('follow_up_date'),'timestamp': call.get('update_ts') or call.get('created_at')})
        for call in ps_calls:
            conversation_log.append({'by': 'PS','name': call.get('ps_name'),'call_no': call.get('call_no'),'attempt': call.get('attempt'),'status': call.get('status'),'remark': call.get('remarks'),'follow_up_date': call.get('follow_up_date'),'timestamp': call.get('created_at')})
        conversation_log = [c for c in conversation_log if c['timestamp']]
        conversation_log.sort(key=lambda x: x['timestamp'])
        outcome = {'final_status': lead.get('final_status'),'timestamp': lead.get('won_timestamp') or lead.get('lost_timestamp'),'reason': lead.get('lost_reason') if lead.get('final_status') == 'Lost' else None}
        # --- PDF Generation ---
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmpfile:
            pdf_path = tmpfile.name
        with PdfPages(pdf_path) as pdf:
            # Title Page
            plt.figure(figsize=(8.3, 11.7))
            plt.axis('off')
            plt.text(0.5, 0.9, 'Lead Journey Report', fontsize=22, ha='center', fontweight='bold')
            plt.text(0.5, 0.85, f"Lead UID: {uid}", fontsize=14, ha='center')
            plt.text(0.5, 0.8, f"Customer: {lead.get('customer_name','')}", fontsize=12, ha='center')
            plt.text(0.5, 0.75, f"Mobile: {lead.get('customer_mobile_number','')}", fontsize=12, ha='center')
            plt.text(0.5, 0.7, f"Source: {lead.get('source','')}", fontsize=12, ha='center')
            plt.text(0.5, 0.65, f"CRE: {lead.get('cre_name','')}", fontsize=12, ha='center')
            plt.text(0.5, 0.6, f"PS: {lead.get('ps_name','')}", fontsize=12, ha='center')
            plt.text(0.5, 0.5, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", fontsize=10, ha='center', color='gray')
            pdf.savefig(); plt.close()
            # Assignment History
            plt.figure(figsize=(8.3, 2))
            plt.axis('off')
            plt.title('Assignment History', fontsize=14, loc='left')
            y = 0.7
            for a in assignment_history:
                plt.text(0.05, y, f"{a['role']}: {a['name']} ({a['assigned_at'] or 'N/A'})", fontsize=11)
                y -= 0.2
            pdf.savefig(); plt.close()
            # Status Timeline Chart
            if status_timeline:
                dates = [s['date'] for s in status_timeline if s.get('date')]
                labels = [s.get('status') or s.get('remark') or '' for s in status_timeline]
                bys = [s.get('by') for s in status_timeline]
                plt.figure(figsize=(8, 3))
                plt.title('Status Timeline')
                plt.plot(dates, range(len(dates)), marker='o')
                for i, (d, l, b) in enumerate(zip(dates, labels, bys)):
                    plt.text(d, i, f"{l} ({b})", fontsize=8, va='bottom', ha='left')
                plt.yticks(range(len(dates)), dates)
                plt.xlabel('Date')
                plt.ylabel('Status Change')
                plt.tight_layout()
                pdf.savefig(); plt.close()
            # Bar Chart: Call Attempts
            plt.figure(figsize=(6, 3))
            plt.title('Number of Call Attempts')
            plt.bar(['CRE', 'PS'], [len(cre_calls), len(ps_calls)], color=['#007bff', '#28a745'])
            plt.ylabel('Attempts')
            plt.tight_layout()
            pdf.savefig(); plt.close()
            # Pie Chart: Status Distribution
            status_counts = {}
            for c in conversation_log:
                s = c['status'] or 'Unknown'
                status_counts[s] = status_counts.get(s, 0) + 1
            if status_counts:
                plt.figure(figsize=(6, 4))
                plt.title('Status Distribution (CRE + PS)')
                plt.pie(list(status_counts.values()), labels=list(status_counts.keys()), autopct='%1.1f%%', startangle=140)
                plt.tight_layout()
                pdf.savefig(); plt.close()
            # Conversation Log Table (first 20 rows)
            plt.figure(figsize=(8.3, min(10, 0.4*len(conversation_log)+1)))
            plt.axis('off')
            plt.title('Conversation Log (first 20 rows)', fontsize=12, loc='left')
            table_data = [['By','Name','Call No','Attempt','Status','Remark','Follow-up','Timestamp']]
            for c in conversation_log[:20]:
                table_data.append([
                    c.get('by',''), c.get('name',''), c.get('call_no',''), str(c.get('attempt','')),
                    c.get('status',''), c.get('remark',''), c.get('follow_up_date',''), c.get('timestamp','')
                ])
            table = plt.table(cellText=table_data, loc='center', cellLoc='left', colWidths=[0.11]*8)
            table.auto_set_font_size(False)
            table.set_fontsize(7)
            table.scale(1, 1.2)
            plt.tight_layout()
            pdf.savefig(); plt.close()
            # Outcome
            plt.figure(figsize=(8.3, 2))
            plt.axis('off')
            plt.title('Final Outcome', fontsize=14, loc='left')
            y = 0.7
            plt.text(0.05, y, f"Status: {outcome.get('final_status','')}", fontsize=11)
            y -= 0.2
            plt.text(0.05, y, f"Timestamp: {outcome.get('timestamp','')}", fontsize=11)
            y -= 0.2
            if outcome.get('reason'):
                plt.text(0.05, y, f"Reason: {outcome.get('reason','')}", fontsize=11)
            pdf.savefig(); plt.close()
        # Serve the PDF
        from flask import send_file
        return send_file(pdf_path, as_attachment=True, download_name=f"Lead_Journey_{uid}.pdf", mimetype='application/pdf')
    except Exception as e:
        print(f"Error generating lead journey PDF: {e}")
        return f"Error generating PDF: {str(e)}", 500

@app.route('/export_filtered_leads')
@require_admin
def export_filtered_leads():
    """Export filtered leads as CSV or Excel with selected columns."""
    import io
    import csv
    import openpyxl
    from flask import send_file
    # Get filters from query params
    cre_id = request.args.get('cre_id')
    source = request.args.get('source')
    qualification = request.args.get('qualification', 'all')
    date_filter = request.args.get('date_filter', 'all')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    final_status = request.args.get('final_status', '')
    search_uid = request.args.get('search_uid', '').strip()
    format_ = request.args.get('format', 'csv')
    # Get CREs for name lookup
    cres = safe_get_data('cre_users')
    selected_cre = None
    leads = []
    if cre_id:
        selected_cre = next((cre for cre in cres if str(cre.get('id')) == str(cre_id)), None)
        filters = {'cre_name': selected_cre['name']} if selected_cre else {}
        if source:
            filters['source'] = source
        leads = safe_get_data('lead_master', filters)
        if search_uid:
            leads = [lead for lead in leads if search_uid.lower() in str(lead.get('uid', '')).lower()]
        if qualification == 'qualified':
            leads = [lead for lead in leads if lead.get('first_call_date')]
        elif qualification == 'unqualified':
            leads = [lead for lead in leads if not lead.get('first_call_date')]
        if final_status:
            leads = [lead for lead in leads if (lead.get('final_status') or '') == final_status]
        if date_filter == 'today':
            today_str = datetime.now().strftime('%Y-%m-%d')
            leads = [lead for lead in leads if (lead.get('cre_assigned_at') or '').startswith(today_str)]
        elif date_filter == 'range' and start_date and end_date:
            def in_range(ld):
                dt = ld.get('cre_assigned_at')
                if not dt:
                    return False
                try:
                    dt_val = dt[:10]
                    return start_date <= dt_val <= end_date
                except Exception:
                    return False
            leads = [lead for lead in leads if in_range(lead)]
    else:
        all_leads = safe_get_data('lead_master')
        if search_uid:
            leads = [lead for lead in all_leads if search_uid.lower() in str(lead.get('uid', '')).lower()]
        else:
            leads = all_leads
    # Columns to export
    export_columns = [
        ('uid', 'UID'),
        ('customer_name', 'Customer Name'),
        ('customer_mobile_number', 'Customer Mobile Number'),
        ('source', 'Source'),
        ('cre_name', 'CRE Name'),
        ('ps_name', 'PS Name'),
        ('final_status', 'Final Status')
    ]
    if format_ == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Leads'
        ws.append([col[1] for col in export_columns])
        for lead in leads:
            ws.append([lead.get(col[0], '') for col in export_columns])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"leads_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    else:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([col[1] for col in export_columns])
        for lead in leads:
            writer.writerow([lead.get(col[0], '') for col in export_columns])
        output.seek(0)
        filename = f"leads_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        return send_file(io.BytesIO(output.getvalue().encode('utf-8')), as_attachment=True, download_name=filename, mimetype='text/csv')

@app.route('/export_all_cre_leads')
@require_admin
def export_all_cre_leads():
    """Export all leads assigned to all CREs as CSV or Excel."""
    import io
    import csv
    import openpyxl
    from flask import send_file
    
    print(f"Export All CREs route called with format: {request.args.get('format', 'csv')}")
    format_ = request.args.get('format', 'csv')
    
    try:
        # Get all CREs
        cres = safe_get_data('cre_users')
        active_cres = [cre for cre in cres if cre.get('is_active', True)]
        
        # Get all leads assigned to any CRE
        all_leads = safe_get_data('lead_master')
        assigned_leads = [lead for lead in all_leads if lead.get('cre_name')]
        print(f"Total leads found: {len(all_leads)}, Assigned leads: {len(assigned_leads)}")
        
        # Columns to export
        export_columns = [
            ('uid', 'UID'),
            ('customer_name', 'Customer Name'),
            ('customer_mobile_number', 'Customer Mobile Number'),
            ('source', 'Source'),
            ('cre_name', 'CRE Name'),
            ('ps_name', 'PS Name'),
            ('final_status', 'Final Status'),
            ('lead_status', 'Lead Status'),
            ('cre_assigned_at', 'CRE Assigned Date'),
            ('first_call_date', 'First Call Date'),
            ('last_call_date', 'Last Call Date'),
            ('total_calls', 'Total Calls'),
            ('lead_category', 'Lead Category')
        ]
        
        if format_ == 'excel':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'All CRE Leads'
            
            # Add header
            ws.append([col[1] for col in export_columns])
            
            # Add data
            for lead in assigned_leads:
                row_data = []
                for col_key, col_name in export_columns:
                    value = lead.get(col_key, '')
                    # Format dates if they exist
                    if col_key in ['cre_assigned_at', 'first_call_date', 'last_call_date'] and value:
                        try:
                            value = value[:10] if len(value) >= 10 else value  # Get just the date part
                        except:
                            pass
                    row_data.append(value)
                ws.append(row_data)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            filename = f"all_cre_leads_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        else:
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Add header
            writer.writerow([col[1] for col in export_columns])
            
            # Add data
            for lead in assigned_leads:
                row_data = []
                for col_key, col_name in export_columns:
                    value = lead.get(col_key, '')
                    # Format dates if they exist
                    if col_key in ['cre_assigned_at', 'first_call_date', 'last_call_date'] and value:
                        try:
                            value = value[:10] if len(value) >= 10 else value  # Get just the date part
                        except:
                            pass
                    row_data.append(value)
                writer.writerow(row_data)
            
            output.seek(0)
            filename = f"all_cre_leads_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            return send_file(io.BytesIO(output.getvalue().encode('utf-8')), as_attachment=True, download_name=filename, mimetype='text/csv')
    
    except Exception as e:
        flash(f'Error exporting all CRE leads: {str(e)}', 'error')
        return redirect(url_for('manage_leads'))

@app.route('/download_lead_master')
@require_admin
def download_lead_master():
    """Download leads from lead_master table with date filtering and comprehensive columns."""
    import io
    import csv
    import openpyxl
    from flask import send_file
    
    try:
        # Get filters from query params
        date_filter = request.args.get('date_filter', 'all')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        format_ = request.args.get('format', 'excel')
        
        # Get all leads from lead_master
        all_leads = safe_get_data('lead_master')
        
        # Apply date filtering
        filtered_leads = []
        today = datetime.now().date()
        
        for lead in all_leads:
            lead_date = None
            # Try to get date from various date fields
            date_fields = ['date', 'created_at', 'cre_assigned_at', 'first_call_date']
            for field in date_fields:
                if lead.get(field):
                    try:
                        # Handle different date formats
                        date_str = str(lead.get(field))
                        if 'T' in date_str:
                            date_str = date_str.split('T')[0]
                        elif ' ' in date_str:
                            date_str = date_str.split(' ')[0]
                        lead_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                        break
                    except:
                        continue
            
            if not lead_date:
                # If no valid date found, include the lead
                filtered_leads.append(lead)
                continue
                
            # Apply date filtering
            if date_filter == 'today':
                if lead_date == today:
                    filtered_leads.append(lead)
            elif date_filter == 'yesterday':
                if lead_date == today - timedelta(days=1):
                    filtered_leads.append(lead)
            elif date_filter == 'this_week':
                week_start = today - timedelta(days=today.weekday())
                if week_start <= lead_date <= today:
                    filtered_leads.append(lead)
            elif date_filter == 'this_month':
                if lead_date.year == today.year and lead_date.month == today.month:
                    filtered_leads.append(lead)
            elif date_filter == 'last_month':
                last_month = today.replace(day=1) - timedelta(days=1)
                if lead_date.year == last_month.year and lead_date.month == last_month.month:
                    filtered_leads.append(lead)
            elif date_filter == 'range' and start_date and end_date:
                try:
                    start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
                    end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
                    if start_dt <= lead_date <= end_dt:
                        filtered_leads.append(lead)
                except:
                    # If date parsing fails, include the lead
                    filtered_leads.append(lead)
            else:  # 'all' or any other value
                filtered_leads.append(lead)
        
        # Define export columns with all required fields
        export_columns = [
            ('uid', 'UID'),
            ('date', 'Date'),
            ('customer_name', 'Customer Name'),
            ('customer_mobile_number', 'Customer Mobile Number'),
            ('source', 'Source'),
            ('cre_name', 'CRE Name'),
            ('lead_category', 'Lead Category'),
            ('model_interested', 'Model Interested'),
            ('branch', 'Branch'),
            ('ps_name', 'PS Name'),
            ('lead_status', 'Lead Status'),
            ('first_call_date', 'First Call Date'),
            ('first_call_remark', 'First Call Remark'),
            ('second_call_date', 'Second Call Date'),
            ('second_call_remark', 'Second Call Remark'),
            ('third_call_date', 'Third Call Date'),
            ('third_call_remark', 'Third Call Remark'),
            ('fourth_call_date', 'Fourth Call Date'),
            ('fourth_call_remark', 'Fourth Call Remark'),
            ('fifth_call_date', 'Fifth Call Date'),
            ('fifth_call_remark', 'Fifth Call Remark'),
            ('sixth_call_date', 'Sixth Call Date'),
            ('sixth_call_remark', 'Sixth Call Remark'),
            ('seventh_call_date', 'Seventh Call Date'),
            ('seventh_call_remark', 'Seventh Call Remark'),
            ('campaign', 'Campaign'),
            ('tat', 'TAT'),
            ('sub_source', 'Sub Source')
        ]
        
        if format_ == 'excel':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Lead Master Data'
            
            # Add header
            ws.append([col[1] for col in export_columns])
            
            # Add data
            for lead in filtered_leads:
                row_data = []
                for col_key, col_name in export_columns:
                    value = lead.get(col_key, '')
                    # Format dates if they exist
                    if 'date' in col_key and value:
                        try:
                            if isinstance(value, str):
                                if 'T' in value:
                                    value = value.split('T')[0]
                                elif ' ' in value:
                                    value = value.split(' ')[0]
                            value = str(value)[:10] if len(str(value)) >= 10 else str(value)
                        except:
                            pass
                    row_data.append(value)
                ws.append(row_data)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            filename = f"lead_master_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        else:
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Add header
            writer.writerow([col[1] for col in export_columns])
            
            # Add data
            for lead in filtered_leads:
                row_data = []
                for col_key, col_name in export_columns:
                    value = lead.get(col_key, '')
                    # Format dates if they exist
                    if 'date' in col_key and value:
                        try:
                            if isinstance(value, str):
                                if 'T' in value:
                                    value = value.split('T')[0]
                                elif ' ' in value:
                                    value = value.split(' ')[0]
                            value = str(value)[:10] if len(str(value)) >= 10 else str(value)
                        except:
                            pass
                    row_data.append(value)
                writer.writerow(row_data)
            
            output.seek(0)
            filename = f"lead_master_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            return send_file(io.BytesIO(output.getvalue().encode('utf-8')), as_attachment=True, download_name=filename, mimetype='text/csv')
    
    except Exception as e:
        print(f"Error downloading lead master data: {str(e)}")
        import traceback
        traceback.print_exc()
        flash(f'Error downloading lead master data: {str(e)}', 'error')
        return redirect(url_for('manage_leads'))

@app.route('/api/export_branch_leads', methods=['POST'])
def api_export_branch_leads():
    """Export leads data for branch head with date range and branch filtering"""
    try:
        data = request.get_json()
        lead_type = data.get('lead_type')
        date_from = data.get('date_from')
        date_to = data.get('date_to')
        branch = data.get('branch')
        
        if not all([lead_type, date_from, date_to, branch]):
            return jsonify({'success': False, 'message': 'Missing required parameters'}), 400
        
        # Convert dates to datetime objects
        from datetime import datetime
        try:
            start_date = datetime.strptime(date_from, '%Y-%m-%d')
            end_date = datetime.strptime(date_to, '%Y-%m-%d')
        except ValueError as e:
            return jsonify({'success': False, 'message': f'Invalid date format: {str(e)}'}), 400
        
        # Get data based on lead type
        csv_data = []
        
        # Helper function to format date for database query
        def format_date_for_query(date_str):
            """Try different date formats for database queries"""
            formats = [
                f'{date_str} 00:00:00',
                f'{date_str} 00:00:00+00',
                date_str,
                f'{date_str}T00:00:00',
                f'{date_str}T00:00:00Z'
            ]
            return formats
        
        if lead_type == 'PS':
            # Export PS leads from ps_followup_master
            try:
                leads_data = []
                print(f"🔍 Exporting PS leads for branch: {branch}, date range: {date_from} to {date_to}")
                
                # Try different date formats
                for start_format in format_date_for_query(date_from):
                    for end_format in format_date_for_query(date_to):
                        try:
                            # PS table uses ps_branch, not branch
                            # Use ps_assigned_at for PS leads filtering - ONLY export leads with ps_assigned_at timestamps
                            query = supabase.table('ps_followup_master').select('*').eq('ps_branch', branch).not_.is_('ps_assigned_at', 'null').gte('ps_assigned_at', start_format).lte('ps_assigned_at', end_format)
                            result = query.execute()
                            if result.data:
                                leads_data = result.data
                                print(f"✅ Found {len(leads_data)} PS leads with ps_branch + ps_assigned_at filter")
                                break
                            else:
                                print(f"⚠️ No PS leads found with ps_assigned_at between {start_format} and {end_format}")
                                continue
                        except Exception as e:
                            print(f"❌ Error with date format {start_format} to {end_format}: {e}")
                            continue
                    if leads_data:
                        break
                
                # Format data for CSV - only include leads with ps_assigned_at
                print(f"📊 Formatting {len(leads_data)} PS leads for CSV export")
                for lead in leads_data:
                    # Only include leads that have ps_assigned_at timestamp
                    if lead.get('ps_assigned_at'):
                        csv_data.append({
                            'Lead UID': lead.get('lead_uid', ''),
                            'Customer Name': lead.get('customer_name', ''),
                            'Mobile Number': lead.get('customer_mobile_number', ''),
                            'Source': lead.get('source', ''),
                            'Lead Category': lead.get('lead_category', 'Not Set'),
                            'PS Name': lead.get('ps_name', ''),
                            'CRE Name': lead.get('cre_name', ''),
                            'Branch': lead.get('branch') or lead.get('ps_branch', ''),
                            'Final Status': lead.get('final_status', ''),
                            'Lead Status': lead.get('lead_status', ''),
                            'Model Interested': lead.get('model_interested', ''),
                            'Follow Up Date': lead.get('follow_up_date', ''),
                            'PS Assigned At': lead.get('ps_assigned_at', ''),
                            'Created At': lead.get('created_at', ''),
                            'Updated At': lead.get('updated_at', '')
                        })
                    else:
                        print(f"⚠️ Skipping lead {lead.get('lead_uid', 'N/A')} - no ps_assigned_at timestamp")
                print(f"✅ Successfully formatted {len(csv_data)} PS leads for export")
            except Exception as e:
                print(f"❌ Error fetching PS leads: {str(e)}")
                import traceback
                traceback.print_exc()
                return jsonify({'success': False, 'message': f'Error fetching PS leads: {str(e)}'}), 500
                
        elif lead_type == 'Walkin':
            # Export Walk-in leads from walkin_table
            try:
                leads_data = []
                # Try different date formats
                for start_format in format_date_for_query(date_from):
                    for end_format in format_date_for_query(date_to):
                        try:
                            query = supabase.table('walkin_table').select('*').eq('branch', branch).gte('created_at', start_format).lte('created_at', end_format)
                            result = query.execute()
                            if result.data:
                                leads_data = result.data
                                break
                        except Exception as e:
                            continue
                    if leads_data:
                        break
                
                # Format data for CSV
                for lead in leads_data:
                    csv_data.append({
                        'UID': lead.get('uid', ''),
                        'Customer Name': lead.get('customer_name', ''),
                        'Mobile Number': lead.get('mobile_number', ''),
                        'Model Interested': lead.get('model_interested', ''),
                        'Occupation': lead.get('occupation', ''),
                        'Lead Category': lead.get('lead_category', 'Not Set'),
                        'Branch': lead.get('branch', ''),
                        'PS Assigned': lead.get('ps_assigned', ''),
                        'Status': lead.get('status', ''),
                        'Followup No': lead.get('followup_no', ''),
                        'Next Followup Date': lead.get('next_followup_date', ''),
                        'First Call Date': lead.get('first_call_date', ''),
                        'First Call Remark': lead.get('first_call_remark', ''),
                        'Created At': lead.get('created_at', ''),
                        'Updated At': lead.get('updated_at', '')
                    })
            except Exception as e:
                print(f"Error fetching Walkin leads: {str(e)}")
                return jsonify({'success': False, 'message': f'Error fetching Walkin leads: {str(e)}'}), 500
                
        elif lead_type == 'Event':
            # Export Event leads from activity_leads
            try:
                leads_data = []
                # Try different date formats
                for start_format in format_date_for_query(date_from):
                    for end_format in format_date_for_query(date_to):
                        try:
                            query = supabase.table('activity_leads').select('*').eq('location', branch).gte('created_at', start_format).lte('created_at', end_format)
                            result = query.execute()
                            if result.data:
                                leads_data = result.data
                                break
                        except Exception as e:
                            continue
                    if leads_data:
                        break
                
                # Format data for CSV
                for lead in leads_data:
                    csv_data.append({
                        'Activity UID': lead.get('activity_uid', ''),
                        'Customer Name': lead.get('customer_name', ''),
                        'Phone Number': lead.get('customer_phone_number', ''),
                        'Occupation': lead.get('occupation', ''),
                        'Lead Category': lead.get('lead_category', 'Not Set'),
                        'Location': lead.get('location', ''),
                        'PS Name': lead.get('ps_name', ''),
                        'Final Status': lead.get('final_status', ''),
                        'Lead Status': lead.get('lead_status', ''),
                        'Model Interested': lead.get('model_interested', ''),
                        'PS Followup Date': lead.get('ps_followup_date_ts', ''),
                        'Created At': lead.get('created_at', ''),
                        'Updated At': lead.get('updated_at', '')
                    })
            except Exception as e:
                print(f"Error fetching Event leads: {str(e)}")
                return jsonify({'success': False, 'message': f'Error fetching Event leads: {str(e)}'}), 500
        else:
            return jsonify({'success': False, 'message': 'Invalid lead type'}), 400
        
        if not csv_data:
            return jsonify({'success': False, 'message': 'No data found for the specified criteria'}), 404
        
        # Create CSV content
        import csv
        import io
        
        print(f"📊 Creating CSV with {len(csv_data)} rows")
        
        output = io.StringIO()
        if csv_data:
            try:
                fieldnames = csv_data[0].keys()
                print(f"📋 CSV fieldnames: {list(fieldnames)}")
                writer = csv.DictWriter(output, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(csv_data)
                print("✅ CSV content created successfully")
            except Exception as e:
                print(f"❌ Error creating CSV: {e}")
                import traceback
                traceback.print_exc()
                return jsonify({'success': False, 'message': f'Error creating CSV: {str(e)}'}), 500
        else:
            print("❌ No CSV data to export")
            return jsonify({'success': False, 'message': 'No data found for the specified criteria'}), 404
        
        # Create response
        from flask import Response
        try:
            csv_content = output.getvalue()
            print(f"📄 CSV content length: {len(csv_content)} characters")
            
            response = Response(
                csv_content,
                mimetype='text/csv',
                headers={'Content-Disposition': f'attachment; filename={lead_type}_leads_{date_from}_to_{date_to}_{branch}.csv'}
            )
            
            print("✅ Response created successfully")
            return response
        except Exception as e:
            print(f"❌ Error creating response: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({'success': False, 'message': f'Error creating response: {str(e)}'}), 500
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Export failed: {str(e)}'}), 500
@app.route('/cre_analytics')
@require_cre
def cre_analytics():
    try:
        # Get current CRE name from session
        current_cre = session.get('cre_name')
        
        # Get all CRE users with their active status (for leaderboard)
        cre_users = safe_get_data('cre_users')
        active_cre_users = [cre for cre in cre_users if cre.get('is_active', True)]
        
        # Get all leads data
        all_leads = safe_get_data('lead_master')
        
        # Helper function to parse timestamps
        def parse_timestamp(timestamp_str):
            if not timestamp_str:
                return None
            try:
                if 'T' in timestamp_str:
                    return datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))
                elif ' ' in timestamp_str:
                    return datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S')
                else:
                    return datetime.strptime(timestamp_str, '%Y-%m-%d')
            except:
                return None
        
        def get_cre_platform_conversion_live():
            platform_data = {}
            for lead in all_leads:
                if lead.get('cre_name') == current_cre:
                    platform = lead.get('source', 'Unknown')
                    if platform not in platform_data:
                        platform_data[platform] = {
                            'won': 0,
                            'pending_live': 0,
                            'lost': 0,
                            'assigned': 0
                        }
                    if lead.get('cre_assigned_at'):
                        platform_data[platform]['assigned'] += 1
                    if lead.get('final_status') == 'Won':
                        platform_data[platform]['won'] += 1
                    elif lead.get('final_status') == 'Pending':
                        platform_data[platform]['pending_live'] += 1
                    elif lead.get('final_status') == 'Lost' or lead.get('lost_timestamp'):
                        platform_data[platform]['lost'] += 1
            for platform in platform_data:
                all_won = len([lead for lead in all_leads 
                              if lead.get('cre_name') == current_cre 
                              and (lead.get('source') or 'Unknown') == platform
                              and lead.get('final_status') == 'Won'])
                assigned = platform_data[platform]['assigned']
                conversion_rate = (all_won / assigned * 100) if assigned > 0 else 0
                platform_data[platform]['conversion_rate'] = round(conversion_rate, 2)
            return platform_data

        def get_top_5_cre_leaderboard():
            cre_performance = {}
            for cre in active_cre_users:
                cre_name = cre.get('name')
                if cre_name:
                    cre_performance[cre_name] = {
                        'won': 0,
                        'pending_live': 0,
                        'lost': 0,
                        'total_leads_handled': 0
                    }
            for lead in all_leads:
                cre_name = lead.get('cre_name')
                if cre_name in cre_performance:
                    if lead.get('cre_assigned_at'):
                        cre_performance[cre_name]['total_leads_handled'] += 1
                    if lead.get('final_status') == 'Won':
                        cre_performance[cre_name]['won'] += 1
                    elif lead.get('final_status') == 'Pending':
                        cre_performance[cre_name]['pending_live'] += 1
                    elif lead.get('final_status') == 'Lost' or lead.get('lost_timestamp'):
                        cre_performance[cre_name]['lost'] += 1
            sorted_cre = sorted(cre_performance.items(), key=lambda x: x[1]['won'], reverse=True)[:5]
            return sorted_cre
        
        # Calculate platform-wise conversion rates for logged-in CRE
        def get_cre_platform_conversion():
            platform_data = {}
            
            for lead in all_leads:
                if lead.get('cre_name') == current_cre:
                    platform = lead.get('source', 'Unknown')
                    if platform not in platform_data:
                        platform_data[platform] = {
                            'won': 0,
                            'pending_live': 0,
                            'lost': 0,
                            'assigned': 0
                        }
                    
                    if lead.get('cre_assigned_at'):
                        platform_data[platform]['assigned'] += 1
                    
                    if lead.get('final_status') == 'Won':
                        platform_data[platform]['won'] += 1
                    elif lead.get('final_status') == 'Pending':
                        # Pending is always live, not date filtered
                        platform_data[platform]['pending_live'] += 1
                    elif lead.get('final_status') == 'Lost' or lead.get('lost_timestamp'):
                        platform_data[platform]['lost'] += 1
            
            # Calculate conversion rates (always live - use all won cases, not date filtered)
            for platform in platform_data:
                # Count all won cases for this platform (live data)
                all_won = len([lead for lead in all_leads 
                              if lead.get('cre_name') == current_cre 
                              and (lead.get('source') or 'Unknown') == platform
                              and lead.get('final_status') == 'Won'])
                assigned = platform_data[platform]['assigned']
                conversion_rate = (all_won / assigned * 100) if assigned > 0 else 0
                platform_data[platform]['conversion_rate'] = round(conversion_rate, 2)
            
            return platform_data
        
        # Get overall stats for logged-in CRE
        def get_cre_overall_stats():
            cre_leads = [lead for lead in all_leads if lead.get('cre_name') == current_cre]
            total_won = len([lead for lead in cre_leads if lead.get('final_status') == 'Won'])
            total_lost = len([lead for lead in cre_leads if lead.get('final_status') == 'Lost' or lead.get('lost_timestamp')])
            total_assigned = len([lead for lead in cre_leads if lead.get('cre_assigned_at')])
            
            overall_rate = (total_won / total_assigned * 100) if total_assigned > 0 else 0
            return {
                'won': total_won,
                'lost': total_lost,
                'assigned': total_assigned,
                'conversion_rate': round(overall_rate, 2)
            }
        
        # Get live stats for logged-in CRE (pending and conversion rate always live)
        def get_cre_live_stats():
            cre_leads = [lead for lead in all_leads if lead.get('cre_name') == current_cre]
            total_pending = len([lead for lead in cre_leads if lead.get('final_status') == 'Pending'])
            total_won_live = len([lead for lead in cre_leads if lead.get('final_status') == 'Won'])
            total_assigned = len([lead for lead in cre_leads if lead.get('cre_assigned_at')])
            
            live_conversion_rate = (total_won_live / total_assigned * 100) if total_assigned > 0 else 0
            
            return {
                'pending': total_pending,
                'conversion_rate': round(live_conversion_rate, 2)
            }
        
        # Prepare data for template
        analytics_data = {
            'current_cre': current_cre,
            'top_5_leaderboard': get_top_5_cre_leaderboard(),
            'cre_platform_conversion': get_cre_platform_conversion(),
            'cre_platform_conversion_live': get_cre_platform_conversion_live(),
            'cre_overall_stats': get_cre_overall_stats(),
            'cre_live_stats': get_cre_live_stats()
        }
        
        return render_template('cre_analytics.html', analytics=analytics_data)
        
    except Exception as e:
        flash(f'Error loading CRE analytics: {str(e)}', 'error')
        return redirect(url_for('cre_dashboard'))


@app.route('/cre_analytics_data')
@require_cre
def cre_analytics_data():
    """API endpoint to get filtered analytics data"""
    try:
        # Get filter parameters
        from_date_str = request.args.get('from_date')
        to_date_str = request.args.get('to_date')
        
        # Get current CRE name from session
        current_cre = session.get('cre_name')
        
        # Get all leads data
        all_leads = safe_get_data('lead_master')
        cre_users = safe_get_data('cre_users')
        active_cre_users = [cre for cre in cre_users if cre.get('is_active', True)]
        
        # Helper function to parse timestamps
        def parse_timestamp(timestamp_str):
            if not timestamp_str:
                return None
            try:
                if 'T' in timestamp_str:
                    return datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))
                elif ' ' in timestamp_str:
                    return datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S')
                else:
                    return datetime.strptime(timestamp_str, '%Y-%m-%d')
            except:
                return None
        
        # Helper function to parse date from HTML5 date input (yyyy-mm-dd format)
        def parse_date_input(date_str):
            if not date_str:
                return None
            try:
                return datetime.strptime(date_str, '%Y-%m-%d').date()
            except:
                return None
        
        # Helper function to check if date is within filter range
        def is_within_date_filter(target_date):
            if not target_date:
                return True  # Include records without dates if no filter
            
            if not from_date_str or not to_date_str:
                return True  # No date filter applied
            
            target_date_obj = target_date.date() if isinstance(target_date, datetime) else target_date
            from_date_obj = parse_date_input(from_date_str)
            to_date_obj = parse_date_input(to_date_str)
            
            if not from_date_obj or not to_date_obj:
                return True  # Invalid date format, include all
            
            return from_date_obj <= target_date_obj <= to_date_obj
        
        # Get filtered leaderboard for all active CREs
        def get_filtered_all_cre_leaderboard():
            cre_performance = {}
            for cre in active_cre_users:
                cre_name = cre.get('name')
                if cre_name:
                    cre_performance[cre_name] = {
                        'won': 0,
                        'pending_live': 0,
                        'lost': 0,
                        'total_leads_handled': 0
                    }
            
            # Apply date filter to total_leads_handled based on cre_assigned_at timestamp
            for lead in all_leads:
                cre_name = lead.get('cre_name')
                if cre_name in cre_performance:
                    if lead.get('cre_assigned_at'):
                        assigned_date = parse_timestamp(lead.get('cre_assigned_at'))
                        if is_within_date_filter(assigned_date):
                            cre_performance[cre_name]['total_leads_handled'] += 1
                    if lead.get('final_status') == 'Pending':
                        cre_performance[cre_name]['pending_live'] += 1
            
            for lead in all_leads:
                cre_name = lead.get('cre_name')
                if cre_name in cre_performance:
                    if lead.get('final_status') == 'Won':
                        won_date = parse_timestamp(lead.get('won_timestamp'))
                        if is_within_date_filter(won_date):
                            cre_performance[cre_name]['won'] += 1
                    elif lead.get('final_status') == 'Lost' or lead.get('lost_timestamp'):
                        lost_date = parse_timestamp(lead.get('lost_timestamp')) or parse_timestamp(lead.get('won_timestamp'))
                        if is_within_date_filter(lost_date):
                            cre_performance[cre_name]['lost'] += 1
            sorted_cre = sorted(cre_performance.items(), key=lambda x: x[1]['won'], reverse=True)
            return sorted_cre
        
        # Get filtered platform conversion data for logged-in CRE
        def get_filtered_cre_platform_conversion():
            platform_data = {}
            for lead in all_leads:
                if lead.get('cre_name') == current_cre:
                    platform = lead.get('source', 'Unknown')
                    if platform not in platform_data:
                        platform_data[platform] = {
                            'won': 0,
                            'pending_live': 0,
                            'lost': 0,
                            'assigned': 0
                        }
                    if lead.get('cre_assigned_at'):
                        platform_data[platform]['assigned'] += 1
                    if lead.get('final_status') == 'Won':
                        won_date = parse_timestamp(lead.get('won_timestamp'))
                        if is_within_date_filter(won_date):
                            platform_data[platform]['won'] += 1
                    elif lead.get('final_status') == 'Pending':
                        platform_data[platform]['pending_live'] += 1
                    elif lead.get('final_status') == 'Lost' or lead.get('lost_timestamp'):
                        lost_date = parse_timestamp(lead.get('lost_timestamp')) or parse_timestamp(lead.get('won_timestamp'))
                        if is_within_date_filter(lost_date):
                            platform_data[platform]['lost'] += 1
            # Calculate conversion rates (always live - use all won cases, not date filtered)
            for platform in platform_data:
                # Count all won cases for this platform (live data)
                all_won = len([lead for lead in all_leads 
                              if lead.get('cre_name') == current_cre 
                              and (lead.get('source') or 'Unknown') == platform
                              and lead.get('final_status') == 'Won'])
                assigned = platform_data[platform]['assigned']
                conversion_rate = (all_won / assigned * 100) if assigned > 0 else 0
                platform_data[platform]['conversion_rate'] = round(conversion_rate, 2)
            return platform_data
        
        # Get filtered overall stats for logged-in CRE
        def get_filtered_cre_overall_stats():
            cre_leads = [lead for lead in all_leads if lead.get('cre_name') == current_cre]
            
            total_won = 0
            total_lost = 0
            total_assigned = len([lead for lead in cre_leads if lead.get('cre_assigned_at')])
            
            for lead in cre_leads:
                if lead.get('final_status') == 'Won':
                    won_date = parse_timestamp(lead.get('won_timestamp'))
                    if is_within_date_filter(won_date):
                        total_won += 1
                elif lead.get('final_status') == 'Lost' or lead.get('lost_timestamp'):
                    lost_date = parse_timestamp(lead.get('lost_timestamp')) or parse_timestamp(lead.get('won_timestamp'))
                    if is_within_date_filter(lost_date):
                        total_lost += 1
            
            overall_rate = (total_won / total_assigned * 100) if total_assigned > 0 else 0
            return {
                'won': total_won,
                'lost': total_lost,
                'assigned': total_assigned,
                'conversion_rate': round(overall_rate, 2)
            }
        
        # Get live platform conversion data for logged-in CRE (not affected by date filter)
        def get_cre_platform_conversion_live():
            platform_data = {}
            for lead in all_leads:
                if lead.get('cre_name') == current_cre:
                    platform = lead.get('source', 'Unknown')
                    if platform not in platform_data:
                        platform_data[platform] = {
                            'won': 0,
                            'pending': 0,
                            'lost': 0,
                            'assigned': 0
                        }
                    if lead.get('cre_assigned_at'):
                        platform_data[platform]['assigned'] += 1
                    if lead.get('final_status') == 'Won':
                        platform_data[platform]['won'] += 1
                    elif lead.get('final_status') == 'Pending':
                        platform_data[platform]['pending'] += 1
                    elif lead.get('final_status') == 'Lost':
                        platform_data[platform]['lost'] += 1
            for platform in platform_data:
                assigned = platform_data[platform]['assigned']
                won = platform_data[platform]['won']
                conversion_rate = (won / assigned * 100) if assigned > 0 else 0
                platform_data[platform]['conversion_rate'] = round(conversion_rate, 2)
            return platform_data
        
        return jsonify({
            'success': True,
            'from_date': from_date_str,
            'to_date': to_date_str,
            'top_5_leaderboard': get_filtered_all_cre_leaderboard()[:5],
            'cre_platform_conversion': get_filtered_cre_platform_conversion(),
            'cre_platform_conversion_live': get_cre_platform_conversion_live(),
            'cre_overall_stats': get_filtered_cre_overall_stats()
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        })

@app.route('/negative_call_attempt_history', methods=['GET'])
@require_admin
def negative_call_attempt_history():
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    call_no_order = ['first', 'second', 'third', 'fourth', 'fifth', 'sixth', 'seventh']
    # Fetch data (only columns that exist)
    cre_attempts = safe_get_data('cre_call_attempt_history', select_fields='call_no,created_at,status')
    ps_attempts = safe_get_data('ps_call_attempt_history', select_fields='call_no,updated_at,status')
    # Parse date filters
    start_date = pd.to_datetime(from_date) if from_date else None
    end_date = pd.to_datetime(to_date) + pd.Timedelta(days=1) - pd.Timedelta(seconds=1) if to_date else None
    cre_filtered, cre_statuses = get_cre_feedback_analysis(cre_attempts, call_no_order, start_date, end_date)
    ps_filtered, ps_statuses = get_ps_feedback_analysis(ps_attempts, call_no_order, start_date, end_date)
    # Debug prints
    print('CRE Statuses:', cre_statuses)
    print('CRE Filtered Data:', cre_filtered)
    print('PS Statuses:', ps_statuses)
    print('PS Filtered Data:', ps_filtered)
    return jsonify({
        'statuses': cre_statuses,  # for CRE table
        'ps_statuses': ps_statuses,  # for PS table (optional, for symmetry)
        'call_no_order': call_no_order,
        'CRE': {'filtered': cre_filtered},
        'PS': {'filtered': ps_filtered}
    })

def get_cre_feedback_analysis(cre_attempts, call_no_order, start_date, end_date):
    cre_df = pd.DataFrame(cre_attempts)
    # print("Sample cre_attempts:", cre_attempts[:5])  # Debug: print sample data
    if cre_df.empty:
        print("No data in cre_attempts!")
        return {call_no: {} for call_no in call_no_order}, []
    cre_df['created_at'] = pd.to_datetime(cre_df['created_at'], errors='coerce')
    # --- Fix: Localize filter dates if needed ---
    if start_date is not None and end_date is not None:
        if pd.api.types.is_datetime64tz_dtype(cre_df['created_at']):
            if start_date.tzinfo is None:
                start_date = start_date.tz_localize('UTC')
            if end_date.tzinfo is None:
                end_date = end_date.tz_localize('UTC')
        cre_df = cre_df[(cre_df['created_at'] >= start_date) & (cre_df['created_at'] <= end_date)]
    cre_df['status'] = cre_df['status'].fillna('').astype(str).str.strip()
    print("Unique statuses in data:", cre_df['status'].unique())  # Debug: print unique statuses
    grouped = cre_df.groupby(['call_no', 'status']).size().reset_index(name='count')
    print("\nGrouped data (call_no, status, count):\n", grouped)  # Debug: print grouped data
    cre_filtered = {call_no: {} for call_no in call_no_order}
    for call_no in call_no_order:
        call_data = grouped[grouped['call_no'] == call_no]
        for _, row in call_data.iterrows():
            cre_filtered[call_no][row['status']] = int(row['count'])
    print("\nCRE Filtered Data:", cre_filtered)  # Debug: print final filtered data
    return cre_filtered, sorted(cre_df['status'].unique())

def get_ps_feedback_analysis(ps_attempts, call_no_order, start_date, end_date):
    ps_df = pd.DataFrame(ps_attempts)
    ps_statuses = []
    ps_filtered = {call_no: {} for call_no in call_no_order}
    if not ps_df.empty:
        ps_df['updated_at'] = pd.to_datetime(ps_df['updated_at'], errors='coerce')
        # --- Fix: Localize filter dates if needed ---
        if start_date is not None and end_date is not None:
            if pd.api.types.is_datetime64tz_dtype(ps_df['updated_at']):
                if start_date.tzinfo is None:
                    start_date = start_date.tz_localize('UTC')
                if end_date.tzinfo is None:
                    end_date = end_date.tz_localize('UTC')
            ps_df = ps_df[(ps_df['updated_at'] >= start_date) & (ps_df['updated_at'] <= end_date)]
        ps_df['status'] = ps_df['status'].fillna('').astype(str).str.strip()
        ps_statuses = sorted(ps_df['status'].unique())
        grouped = ps_df.groupby(['call_no', 'status']).size().reset_index(name='count')
        for call_no in call_no_order:
            call_data = grouped[grouped['call_no'] == call_no]
            ps_filtered[call_no] = {status: 0 for status in ps_statuses}
            for _, row in call_data.iterrows():
                ps_filtered[call_no][row['status']] = int(row['count'])
    return ps_filtered, ps_statuses

@app.route('/get_cre_list')
@require_admin
def get_cre_list():
    cres = safe_get_data('cre_users', select_fields='name')
    cre_list = [{'name': cre.get('name')} for cre in cres if cre.get('name')]
    return jsonify({'success': True, 'cre_list': cre_list})

@app.route('/cre_analysis_data')
@require_admin
def cre_analysis_data():
    # Return a mock response or implement your actual logic here
    return jsonify({'success': True, 'data': {}})

@app.route('/update_event_lead/<activity_uid>', methods=['GET', 'POST'])
@require_ps
def update_event_lead(activity_uid):
    return_tab = request.args.get('return_tab', 'fresh-leads')
    print(f'[DEBUG] update_event_lead called with activity_uid: {activity_uid}')
    print(f'[DEBUG] return_tab: {return_tab}')
    print(f'[DEBUG] request.method: {request.method}')
    
    try:
        # Fetch the event lead by activity_uid
        print(f'[DEBUG] Attempting to fetch event lead with activity_uid: {activity_uid}')
        result = supabase.table('activity_leads').select('*').eq('activity_uid', activity_uid).execute()
        print(f'[DEBUG] Supabase query result: {result}')
        print(f'[DEBUG] Result data: {result.data}')
        
        if not result.data:
            print(f'[DEBUG] No event lead found for activity_uid: {activity_uid}')
            flash('Event lead not found', 'error')
            return redirect(url_for('ps_dashboard', tab=return_tab))
        
        lead = result.data[0]
        print('[DEBUG] Lead data (PS):', lead)
        # Determine next call and completed calls for PS follow-ups
        call_order = ['first', 'second', 'third', 'fourth', 'fifth', 'sixth', 'seventh']
        completed_calls = []
        next_call = 'first'
        for call_num in call_order:
            call_date_key = f'ps_{call_num}_call_date'
            if lead.get(call_date_key):
                completed_calls.append(call_num)
            else:
                next_call = call_num
                break
        
        # Build PS call history
        ps_call_history = []
        for call_num in call_order:
            call_date = lead.get(f'ps_{call_num}_call_date')
            call_remark = lead.get(f'ps_{call_num}_call_remark')
            if call_date or call_remark:
                ps_call_history.append({
                    'call': call_num,
                    'date': call_date,
                    'remark': call_remark
                })
        
        # Build CRE call history
        cre_call_history = []
        for call_num in call_order:
            call_date = lead.get(f'cre_{call_num}_call_date')
            call_remark = lead.get(f'cre_{call_num}_call_remark')
            if call_date or call_remark:
                cre_call_history.append({
                    'call': call_num,
                    'date': call_date,
                    'remark': call_remark
                })
        print('[DEBUG] PS call history:', ps_call_history)
        print('[DEBUG] CRE call history:', cre_call_history)
        
        if request.method == 'POST':
            update_data = {}
            customer_name = request.form.get('customer_name', '').strip()
            lead_category = request.form.get('lead_category', '').strip()
            interested_model = request.form.get('interested_model', '').strip()
            test_drive_done = request.form.get('test_drive_done', '').strip()
            customer_location = request.form.get('customer_location', '').strip()
            ps_followup_date_ts = request.form.get('ps_followup_date_ts', '').strip()
            lead_status = request.form.get('lead_status', '').strip()
            final_status = request.form.get('final_status', '').strip()
            call_date = request.form.get('call_date', '').strip()
            call_remark = request.form.get('call_remark', '').strip()
            # Update editable fields
            if customer_name:
                update_data['customer_name'] = customer_name
            if lead_category:
                update_data['lead_category'] = lead_category
            if interested_model:
                update_data['interested_model'] = interested_model
            # Test Drive Done is now required
            if not test_drive_done:
                flash('Test Drive Done field is required', 'error')
                redirect_url = url_for('update_event_lead', activity_uid=activity_uid, return_tab=return_tab)
                return redirect(redirect_url)
            update_data['test_drive_done'] = test_drive_done
            if customer_location:
                update_data['customer_location'] = customer_location
            if ps_followup_date_ts:
                update_data['ps_followup_date_ts'] = ps_followup_date_ts
            if lead_status:
                update_data['lead_status'] = lead_status
            if final_status:
                update_data['final_status'] = final_status
            # Handle PS call date/remark for the next call
            skip_statuses = ["Call not Connected", "RNR", "Call me Back", "Busy on another Call"]
            if lead_status not in skip_statuses:
                if call_date:
                    update_data[f'ps_{next_call}_call_date'] = call_date
                if call_remark:
                    update_data[f'ps_{next_call}_call_remark'] = call_remark
            # Always remove these keys if status is in skip list, even if present
            if lead_status in skip_statuses:
                update_data.pop(f'ps_{next_call}_call_date', None)
                update_data.pop(f'ps_{next_call}_call_remark', None)
            if update_data:
                supabase.table('activity_leads').update(update_data).eq('activity_uid', activity_uid).execute()
                
                # Track the PS call attempt in ps_call_attempt_history
                if lead_status:
                    call_was_recorded = bool(call_date and call_remark)
                    track_ps_call_attempt(
                        uid=lead['activity_uid'],
                        ps_name=session.get('ps_name'),
                        call_no=next_call,
                        lead_status=lead_status,
                        call_was_recorded=call_was_recorded,
                        follow_up_date=ps_followup_date_ts if ps_followup_date_ts else None,
                        remarks=call_remark if call_remark else None
                    )
                
                # Sync to alltest_drive table if test_drive_done is set
                if test_drive_done in ['Yes', 'No', True, False]:
                    # Get the updated activity lead data for syncing
                    updated_activity_result = supabase.table('activity_leads').select('*').eq('activity_uid', activity_uid).execute()
                    if updated_activity_result.data:
                        updated_activity_data = updated_activity_result.data[0]
                        sync_test_drive_to_alltest_drive('activity_leads', activity_uid, updated_activity_data)
                
                flash('Event lead updated successfully', 'success')
                return redirect(url_for('ps_dashboard', tab=return_tab))
            else:
                flash('No changes to update', 'info')
        print(f'[DEBUG] About to render template with lead data: {lead}')
        print(f'[DEBUG] next_call: {next_call}')
        print(f'[DEBUG] completed_calls: {completed_calls}')
        print(f'[DEBUG] ps_call_history: {ps_call_history}')
        print(f'[DEBUG] cre_call_history: {cre_call_history}')
        
        return render_template('update_event_lead.html', 
                             lead=lead, 
                             next_call=next_call, 
                             completed_calls=completed_calls, 
                             ps_call_history=ps_call_history,
                             cre_call_history=cre_call_history,
                             today=date.today())
    except Exception as e:
        print(f'[DEBUG] Exception in update_event_lead: {str(e)}')
        print(f'[DEBUG] Exception type: {type(e)}')
        import traceback
        print(f'[DEBUG] Full traceback: {traceback.format_exc()}')
        flash(f'Error updating event lead: {str(e)}', 'error')
        return redirect(url_for('ps_dashboard'))

@app.route('/update_event_lead_cre/<activity_uid>', methods=['GET', 'POST'])
@require_cre
def update_event_lead_cre(activity_uid):
    # Get return_tab parameter for redirection
    return_tab = request.args.get('return_tab', '')
    try:
        # Fetch the event lead by activity_uid
        result = supabase.table('activity_leads').select('*').eq('activity_uid', activity_uid).execute()
        if not result.data:
            flash('Event lead not found', 'error')
            return redirect(url_for('cre_dashboard'))
        lead = result.data[0]
        print('[DEBUG] Lead data (CRE):', lead)
        
        # Check if this CRE is assigned to this lead
        cre_name = session.get('cre_name')
        if lead.get('cre_assigned') != cre_name:
            flash('You are not authorized to update this event lead', 'error')
            return redirect(url_for('cre_dashboard'))
        
        # Determine next call and completed calls for CRE follow-ups
        call_order = ['first', 'second', 'third', 'fourth', 'fifth', 'sixth', 'seventh']
        completed_calls = []
        next_call = 'first'
        for call_num in call_order:
            call_date_key = f'cre_{call_num}_call_date'
            if lead.get(call_date_key):
                completed_calls.append(call_num)
            else:
                next_call = call_num
                break
        
        # Build CRE call history
        cre_call_history = []
        for call_num in call_order:
            call_date = lead.get(f'cre_{call_num}_call_date')
            call_remark = lead.get(f'cre_{call_num}_call_remark')
            if call_date or call_remark:
                cre_call_history.append({
                    'call': call_num,
                    'date': call_date,
                    'remark': call_remark
                })
        
        # Build PS call history
        ps_call_history = []
        for call_num in call_order:
            call_date = lead.get(f'ps_{call_num}_call_date')
            call_remark = lead.get(f'ps_{call_num}_call_remark')
            if call_date or call_remark:
                ps_call_history.append({
                    'call': call_num,
                    'date': call_date,
                    'remark': call_remark
                })
        print('[DEBUG] CRE call history:', cre_call_history)
        print('[DEBUG] PS call history:', ps_call_history)
        
        if request.method == 'POST':
            update_data = {}
            customer_name = request.form.get('customer_name', '').strip()
            lead_category = request.form.get('lead_category', '').strip()
            interested_model = request.form.get('interested_model', '').strip()
            customer_location = request.form.get('customer_location', '').strip()
            cre_followup_date_ts = request.form.get('cre_followup_date_ts', '').strip()
            lead_status = request.form.get('lead_status', '').strip()
            final_status = request.form.get('final_status', '').strip()
            test_drive_done = request.form.get('test_drive_done', '').strip()
            call_date = request.form.get('call_date', '').strip()
            call_remark = request.form.get('call_remark', '').strip()
            
            # Update editable fields
            if customer_name:
                update_data['customer_name'] = customer_name
            if lead_category:
                update_data['lead_category'] = lead_category
            if interested_model:
                update_data['interested_model'] = interested_model
            if customer_location:
                update_data['customer_location'] = customer_location
            if cre_followup_date_ts:
                update_data['cre_followup_date'] = cre_followup_date_ts
            if lead_status:
                update_data['lead_status'] = lead_status
            if final_status:
                update_data['final_status'] = final_status
            if test_drive_done:
                update_data['test_drive_done'] = test_drive_done
            
            # Handle CRE call date/remark for the next call
            skip_statuses = ["Call not Connected", "RNR", "Call me Back", "Busy on another Call"]
            if lead_status not in skip_statuses:
                if call_date:
                    update_data[f'cre_{next_call}_call_date'] = call_date
                if call_remark:
                    update_data[f'cre_{next_call}_call_remark'] = call_remark
            # Always remove these keys if status is in skip list, even if present
            if lead_status in skip_statuses:
                update_data.pop(f'cre_{next_call}_call_date', None)
                update_data.pop(f'cre_{next_call}_call_remark', None)
            
            if update_data:
                supabase.table('activity_leads').update(update_data).eq('activity_uid', activity_uid).execute()
                
                # Sync to alltest_drive table if test_drive_done is set
                if test_drive_done in ['Yes', 'No', True, False]:
                    # Get the updated activity lead data for syncing
                    updated_activity_result = supabase.table('activity_leads').select('*').eq('activity_uid', activity_uid).execute()
                    if updated_activity_result.data:
                        updated_activity_data = updated_activity_result.data[0]
                        sync_test_drive_to_alltest_drive('activity_leads', activity_uid, updated_activity_data)
                
                flash('Event lead updated successfully', 'success')
                # Redirect based on return_tab parameter
                if return_tab:
                    if return_tab == 'event-leads':
                        return redirect(url_for('cre_dashboard', tab='event-leads'))
                    else:
                        return redirect(url_for('cre_dashboard'))
                else:
                    return redirect(url_for('cre_dashboard'))
            else:
                flash('No changes to update', 'info')
                # Redirect based on return_tab parameter
                if return_tab:
                    if return_tab == 'event-leads':
                        return redirect(url_for('cre_dashboard', tab='event-leads'))
                    else:
                        return redirect(url_for('cre_dashboard'))
                else:
                    return redirect(url_for('cre_dashboard'))
        
        return render_template('update_event_lead_cre.html', 
                             lead=lead, 
                             next_call=next_call, 
                             completed_calls=completed_calls, 
                             cre_call_history=cre_call_history,
                             ps_call_history=ps_call_history,
                             today=date.today())
    except Exception as e:
        flash(f'Error updating event lead: {str(e)}', 'error')
        return redirect(url_for('cre_dashboard'))
@app.route('/admin_duplicate_leads', methods=['GET'])
@require_admin
def admin_duplicate_leads():
    # Get filters and pagination params
    search_uid = request.args.get('search_uid', '').strip().lower()
    search_source = request.args.get('search_source', '').strip().lower()
    search_name = request.args.get('search_name', '').strip().lower()
    from_date = request.args.get('from_date', '').strip()
    to_date = request.args.get('to_date', '').strip()
    date_range_type = request.args.get('date_range_type', 'all_time')
    # Get filters and pagination params
    try:
        page = int(request.args.get('page', 1))
    except (ValueError, TypeError):
        page = 1
    page_size = 50
    offset = (page - 1) * page_size

    # Build Supabase query with filters
    query = supabase.table('duplicate_leads').select('*')
    if search_uid:
        query = query.ilike('uid', f'%{search_uid}%')
    if search_name:
        query = query.ilike('customer_name', f'%{search_name}%')
    if search_source:
        # No ilike for array fields, so filter after fetch for sources
        pass
    # Date filter will be applied after fetch (since last_enquiry_date is computed)

    # Fetch total count for pagination (without limit)
    total_result = query.execute()
    total_count = len(total_result.data or [])

    # Fetch only the current page
    query = query.range(offset, offset + page_size - 1)
    result = query.execute()
    duplicate_leads = result.data or []

    # Prepare leads for display (apply source/date filter in Python if needed)
    leads_display = []
    for lead in duplicate_leads:
        sources = []
        sub_sources = []
        dates = []
        for i in range(1, 11):
            src = lead.get(f'source{i}')
            sub = lead.get(f'sub_source{i}')
            dt = lead.get(f'date{i}')
            if src:
                sources.append(src)
            if sub and sub.strip() and sub.strip().lower() != 'unknown':
                sub_sources.append(sub)
            if dt:
                dates.append(dt)
        last_enquiry_date = max([d for d in dates if d], default=None)
        days_old = None
        if last_enquiry_date:
            try:
                last_date = datetime.strptime(last_enquiry_date, '%Y-%m-%d').date()
                days_old = (date.today() - last_date).days
            except Exception:
                days_old = None
        cre_name = None
        assigned = False
        lm_result = supabase.table('lead_master').select('cre_name').eq('uid', lead['uid']).execute()
        if lm_result.data and lm_result.data[0].get('cre_name'):
            cre_name = lm_result.data[0]['cre_name']
            assigned = True
        # Filter by Source (any source)
        if search_source and not any(search_source in (s or '').lower() for s in sources):
            continue
        # Date range filter (if select_date)
        if date_range_type == 'select_date' and from_date and to_date:
            if last_enquiry_date:
                try:
                    last_date = datetime.strptime(last_enquiry_date, '%Y-%m-%d').date()
                    from_dt = datetime.strptime(from_date, '%Y-%m-%d').date()
                    to_dt = datetime.strptime(to_date, '%Y-%m-%d').date()
                    if not (from_dt <= last_date <= to_dt):
                        continue
                except Exception as e:
                    continue
            else:
                continue
        # Determine source type based on available columns
        source_type = 'CRE Lead'  # Default
        if lead.get('original_table'):
            if lead.get('original_table') == 'lead_master':
                source_type = 'CRE Lead'
            elif lead.get('original_table') == 'activity_leads':
                source_type = 'Event Lead'
            elif lead.get('original_table') == 'walkin_table':
                source_type = 'Walk-in Lead'
        
        leads_display.append({
            'uid': lead.get('uid'),
            'customer_name': lead.get('customer_name'),
            'customer_mobile_number': lead.get('customer_mobile_number'),
            'sources': sources,
            'sub_sources': sub_sources,
            'last_enquiry_date': last_enquiry_date,
            'days_old': days_old,
            'cre_name': cre_name,
            'assigned': assigned,
            'source_type': source_type,
            'original_table': lead.get('original_table'),
            'original_record_id': lead.get('original_record_id')
        })
    # Calculate total pages
    total_pages = (total_count + page_size - 1) // page_size
    current_args = request.args.to_dict()
    # Previous page URL
    prev_url = None
    if page > 1:
        prev_args = current_args.copy()
        prev_args['page'] = page - 1
        prev_url = url_for('admin_duplicate_leads', **prev_args)
    # Next page URL
    next_url = None
    if page < total_pages:
        next_args = current_args.copy()
        next_args['page'] = page + 1
        next_url = url_for('admin_duplicate_leads', **next_args)
    # Page URLs for numbered links
    page_urls = []
    for p in range(1, total_pages + 1):
        page_args = current_args.copy()
        page_args['page'] = p
        page_urls.append(url_for('admin_duplicate_leads', **page_args))
    return render_template('admin_duplicate_leads.html', duplicate_leads=leads_display, page=page, total_pages=total_pages, total_count=total_count, prev_url=prev_url, next_url=next_url, page_urls=page_urls)

@app.route('/convert_duplicate_to_fresh/<uid>', methods=['POST'])
@require_admin
def convert_duplicate_to_fresh(uid):
    # Get duplicate lead
    result = supabase.table('duplicate_leads').select('*').eq('uid', uid).execute()
    if not result.data:
        flash('Duplicate lead not found', 'error')
        return redirect(url_for('admin_duplicate_leads'))
    dup_lead = result.data[0]
    # Generate new UID (use existing logic or append timestamp)
    new_uid = f"{uid}-NEW-{int(datetime.now().timestamp())}"
    # Ensure UID is at most 20 characters (DB limit)
    if len(new_uid) > 20:
        new_uid = new_uid[:20]
    # Prepare new lead data for lead_master
    lead_data = {
        'uid': new_uid,
        'customer_name': dup_lead.get('customer_name'),
        'customer_mobile_number': dup_lead.get('customer_mobile_number'),
        'source': dup_lead.get('source1'),
        'sub_source': dup_lead.get('sub_source1'),
        'date': dup_lead.get('date1'),
        'assigned': 'No',
        'lead_status': 'Pending',
        'final_status': 'Pending',
        'created_at': datetime.now().isoformat(),
        'updated_at': datetime.now().isoformat()
    }
    try:
        supabase.table('lead_master').insert(lead_data).execute()
        supabase.table('duplicate_leads').delete().eq('uid', uid).execute()
        flash('Duplicate lead converted to fresh lead successfully!', 'success')
    except Exception as e:
        flash(f'Error converting lead: {str(e)}', 'error')
    return redirect(url_for('admin_duplicate_leads'))


@app.route('/check_username')
def check_username():
    username = request.args.get('username', '').strip()
    user_type = request.args.get('type', '').strip().lower()
    if not username or user_type not in ['cre', 'ps']:
        return jsonify({'error': 'Invalid parameters'}), 400
    try:
        table = 'cre_users' if user_type == 'cre' else 'ps_users'
        result = supabase.table(table).select('username').eq('username', username).execute()
        exists = bool(result.data)
        return jsonify({'exists': exists})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/check_name')
def check_name():
    """Check if a full name already exists in the system"""
    name = request.args.get('name', '').strip()
    user_type = request.args.get('type', '').strip()
    
    if not name or not user_type:
        return jsonify({'error': 'Missing name or user type'})
    
    try:
        # Check in multiple user tables to avoid duplicate names across different roles
        duplicate_found = False
        existing_user = None
        
        # Check cre_users table
        if user_type in ['cre', 'all']:
            cre_result = supabase.table('cre_users').select('name, username, role').eq('name', name).execute()
            if cre_result.data and len(cre_result.data) > 0:
                duplicate_found = True
                existing_user = cre_result.data[0]
        
        # Check ps_users table
        if user_type in ['ps', 'all']:
            ps_result = supabase.table('ps_users').select('name, username, role, branch').eq('name', name).execute()
            if ps_result.data and len(ps_result.data) > 0:
                duplicate_found = True
                existing_user = ps_result.data[0]
        
        # Check rec_users table
        if user_type in ['rec', 'all']:
            rec_result = supabase.table('rec_users').select('name, username, role, branch').eq('name', name).execute()
            if rec_result.data and len(rec_result.data) > 0:
                duplicate_found = True
                existing_user = rec_result.data[0]
        
        # Check Branch Head table
        if user_type in ['branch_head', 'all']:
            bh_result = supabase.table('Branch Head').select('Name, Username, Branch').eq('Name', name).execute()
            if bh_result.data and len(bh_result.data) > 0:
                duplicate_found = True
                existing_user = bh_result.data[0]
        
        if duplicate_found:
            return jsonify({
                'exists': True,
                'duplicate_type': 'name',
                'existing_user': existing_user,
                'message': f'User with name "{name}" already exists'
            })
        else:
            return jsonify({
                'exists': False,
                'message': 'Name is available'
            })
        
    except Exception as e:
        print(f"Error checking name: {e}")
        return jsonify({'error': 'Database error'})

@app.route('/ps_analytics')
@require_ps
def ps_analytics():
    try:
        ps_name = session.get('ps_name')
        # Get all relevant data
        ps_followups = safe_get_data('ps_followup_master')
        activity_leads = safe_get_data('activity_leads')

        # Date filter (from/to) for won cases and assigned leads
        from_date = request.args.get('from_date')
        to_date = request.args.get('to_date')
        def in_date_range(ts):
            if not ts:
                return False
            from datetime import datetime
            try:
                dt = datetime.fromisoformat(ts.replace('Z', '+00:00')) if 'T' in ts else datetime.strptime(ts, '%Y-%m-%d')
                if from_date:
                    from_dt = datetime.strptime(from_date, '%Y-%m-%d')
                    if dt < from_dt:
                        return False
                if to_date:
                    to_dt = datetime.strptime(to_date, '%Y-%m-%d')
                    if dt > to_dt:
                        return False
                return True
            except Exception:
                return False

        # PS Won Cases (date filtered)
        if not from_date and not to_date:
            won_ps_followups = [lead for lead in ps_followups if lead.get('ps_name') == ps_name and lead.get('final_status') == 'Won']
            won_activities = [lead for lead in activity_leads if lead.get('ps_name') == ps_name and lead.get('final_status') == 'Won']
        else:
            won_ps_followups = [lead for lead in ps_followups if lead.get('ps_name') == ps_name and lead.get('final_status') == 'Won' and in_date_range(lead.get('won_timestamp'))]
            won_activities = [lead for lead in activity_leads if lead.get('ps_name') == ps_name and lead.get('final_status') == 'Won' and in_date_range(lead.get('won_timestamp'))]
        total_won_cases = len(won_ps_followups) + len(won_activities)

        # Conversion Rate (live, not date filtered)
        total_won_all = len([lead for lead in ps_followups if lead.get('ps_name') == ps_name and lead.get('final_status') == 'Won']) \
            + len([lead for lead in activity_leads if lead.get('ps_name') == ps_name and lead.get('final_status') == 'Won'])
        total_assigned_live = len([lead for lead in ps_followups if lead.get('ps_name') == ps_name]) \
            + len([lead for lead in activity_leads if lead.get('ps_name') == ps_name])
        conversion_rate = round((total_won_all / total_assigned_live * 100), 2) if total_assigned_live > 0 else 0

        # Total Leads Assigned (date filtered by created_at)
        def in_created_range(ts):
            if not ts:
                return False
            from datetime import datetime
            try:
                dt = datetime.fromisoformat(ts.replace('Z', '+00:00')) if 'T' in ts else datetime.strptime(ts, '%Y-%m-%d')
                if from_date:
                    from_dt = datetime.strptime(from_date, '%Y-%m-%d')
                    if dt < from_dt:
                        return False
                if to_date:
                    to_dt = datetime.strptime(to_date, '%Y-%m-%d')
                    if dt > to_dt:
                        return False
                return True
            except Exception:
                return False
        assigned_ps_followups = [lead for lead in ps_followups if lead.get('ps_name') == ps_name and in_created_range(lead.get('created_at'))]
        assigned_activities = [lead for lead in activity_leads if lead.get('ps_name') == ps_name and in_created_range(lead.get('created_at'))]
        total_assigned = len(assigned_ps_followups) + len(assigned_activities)

        # Total Pending Cases (not date filtered)
        pending_ps_followups = [lead for lead in ps_followups if lead.get('ps_name') == ps_name and lead.get('final_status') == 'Pending']
        pending_activities = [lead for lead in activity_leads if lead.get('ps_name') == ps_name and lead.get('final_status') == 'Pending']
        total_pending = len(pending_ps_followups) + len(pending_activities)
        ps_analytics = {
            'won_cases': total_won_cases,
            'conversion_rate': conversion_rate,

            'total_assigned': total_assigned,
            'total_pending': total_pending,

        }
        return render_template('ps_analytics.html', ps_analytics=ps_analytics)
    except Exception as e:
        flash(f'Error loading PS analytics: {str(e)}', 'error')
        return redirect(url_for('ps_dashboard'))

@app.route('/manage_branch_head')
def manage_branch_head():
    # Fetch all branch heads
    branch_heads = supabase.table('Branch Head').select('*').execute().data
    
    # Debug: Print the structure of existing data
    if branch_heads:
        print("Sample branch head data structure:")
        print(f"Keys in first record: {list(branch_heads[0].keys())}")
        print(f"Sample record: {branch_heads[0]}")
    
    branches = get_all_branches()  # This should return a list of branch names
    return render_template('manage_branch_head.html', branch_heads=branch_heads, branches=branches)

def get_all_branches():
    # Use centralized branch management
    return get_system_branches()

def get_active_ps_users(branch):
    """Get active PS users for a specific branch"""
    try:
        print(f"Getting active PS users for branch: '{branch}'")
        
        # First, let's see what branches exist in ps_users table
        all_branches = supabase.table('ps_users').select('branch').execute()
        available_branches = list(set([row['branch'] for row in all_branches.data if row['branch']]))
        print(f"Available branches in ps_users: {available_branches}")
        
        # Try exact match first
        ps_query = supabase.table('ps_users').select('name').eq('branch', branch).eq('is_active', True).execute()
        active_ps = [row['name'] for row in ps_query.data if row['name']]
        
        # If no results, try case-insensitive search
        if not active_ps:
            print(f"No exact match found for branch '{branch}', trying case-insensitive search")
            # Get all PS users and filter by case-insensitive branch match
            all_ps = supabase.table('ps_users').select('name, branch, is_active').execute()
            active_ps = [row['name'] for row in all_ps.data if row['name'] and row.get('is_active') and row.get('branch', '').lower() == branch.lower()]
        
        print(f"Active PS users for branch '{branch}': {active_ps}")
        
        return active_ps
    except Exception as e:
        print(f"Error getting active PS users for branch {branch}: {str(e)}")
        return []

@app.route('/add_branch_head', methods=['POST'])
def add_branch_head():
    try:
        # Get form data
        name = request.form.get('name', '').strip()
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        email = request.form.get('email', '').strip()
        phone = request.form.get('phone', '').strip()
        branch = request.form.get('branch', '').strip()
        is_active = request.form.get('is_active') == 'on'
        
        # Debug: Print form data
        print(f"Form data received:")
        print(f"Name: '{name}'")
        print(f"Username: '{username}'")
        print(f"Email: '{email}'")
        print(f"Phone: '{phone}'")
        print(f"Branch: '{branch}'")
        
        # Validate required fields
        if not all([name, username, password, branch]):
            flash('Name, Username, Password, and Branch are required fields', 'error')
            return redirect(url_for('manage_branch_head'))
        
        # Validate email format if provided
        if email and '@' not in email:
            flash('Please enter a valid email address', 'error')
            return redirect(url_for('manage_branch_head'))
        
        # Validate phone number format (10 digits)
        if phone and not phone.replace(' ', '').isdigit() or (phone and len(phone.replace(' ', '')) != 10):
            flash('Please enter a valid 10-digit phone number', 'error')
            return redirect(url_for('manage_branch_head'))
        
        # Store plain text password for Branch Head (testing only)
        hashed_pw = password
        
        # Prepare data for insertion
        branch_head_data = {
            'Name': name,
            'Username': username,
            'Password': hashed_pw,
            'Branch': branch,
            'Is Active': is_active
        }
        
        # Add email and phone (always include them, even if empty)
        branch_head_data['email'] = email if email else None
        branch_head_data['Phone No.'] = phone if phone else None
        
        # Debug: Print final data to be inserted
        print(f"Data to be inserted: {branch_head_data}")
        
        # Insert into database
        result = supabase.table('Branch Head').insert(branch_head_data).execute()
        
        if result.data:
            flash('Branch Head added successfully!', 'success')
            print(f"Successfully inserted branch head with ID: {result.data[0]['id'] if result.data else 'Unknown'}")
            
            # Debug: Fetch the inserted record to verify what was saved
            if result.data:
                inserted_id = result.data[0]['id']
                verify_record = supabase.table('Branch Head').select('*').eq('id', inserted_id).execute().data
                if verify_record:
                    print(f"Verified saved data: {verify_record[0]}")
        else:
            flash('Error creating branch head. Please try again.', 'error')
            print("Failed to insert branch head")
            
    except Exception as e:
        flash(f'Error adding branch head: {str(e)}', 'error')
        print(f"Exception occurred: {str(e)}")
    
    return redirect(url_for('manage_branch_head'))

@app.route('/toggle_branch_head_active/<int:id>', methods=['POST'])
def toggle_branch_head_active(id):
    is_active = request.form.get('is_active') == 'True'
    supabase.table('Branch Head').update({'Is Active': is_active}).eq('id', id).execute()
    flash('Branch Head status updated!', 'info')
    return redirect(url_for('manage_branch_head'))

@app.route('/edit_branch_head_contact/<int:id>', methods=['POST'])
def edit_branch_head_contact(id):
    """Edit branch head email and phone only"""
    try:
        email = request.form.get('email', '').strip()
        phone = request.form.get('phone', '').strip()
        
        # Validate email format if provided
        if email and '@' not in email:
            return jsonify({'success': False, 'message': 'Please enter a valid email address'})
        
        # Validate phone number format (10 digits)
        if phone and not phone.replace(' ', '').isdigit() or (phone and len(phone.replace(' ', '')) != 10):
            return jsonify({'success': False, 'message': 'Please enter a valid 10-digit phone number'})
        
        # Prepare update data
        update_data = {}
        if email:
            update_data['email'] = email
        if phone:
            update_data['Phone No.'] = phone
        
        # Update in database
        result = supabase.table('Branch Head').update(update_data).eq('id', id).execute()
        
        if result.data:
            return jsonify({'success': True, 'message': 'Contact information updated successfully!'})
        else:
            return jsonify({'success': False, 'message': 'Error updating contact information'})
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/delete_branch_head/<int:id>', methods=['POST'])
def delete_branch_head(id):
    """Delete entire branch head"""
    try:
        # Delete from database
        result = supabase.table('Branch Head').delete().eq('id', id).execute()
        
        if result.data:
            return jsonify({'success': True, 'message': 'Branch Head deleted successfully!'})
        else:
            return jsonify({'success': False, 'message': 'Error deleting branch head'})
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})









@app.route('/api/hot_duplicate_leads')
@require_admin
def api_hot_duplicate_leads():
    """API endpoint to get duplicate leads created in the last 1 day (hot duplicates)"""
    from datetime import datetime, date
    try:
        # Fetch all duplicate leads (limit to 200 for performance)
        result = supabase.table('duplicate_leads').select('*').limit(200).execute()
        duplicate_leads = result.data or []
        hot_leads = []
        for lead in duplicate_leads:
            dates = []
            for i in range(1, 11):
                dt = lead.get(f'date{i}')
                if dt:
                    dates.append(dt)
            last_enquiry_date = max([d for d in dates if d], default=None)
            days_old = None
            if last_enquiry_date:
                try:
                    last_date = datetime.strptime(last_enquiry_date, '%Y-%m-%d').date()
                    days_old = (date.today() - last_date).days
                except Exception:
                    days_old = None
            # Only include if days_old is 0 or 1 (today or yesterday)
            if days_old is not None and days_old <= 1:
                hot_leads.append({
                    'uid': lead.get('uid'),
                    'customer_name': lead.get('customer_name'),
                    'customer_mobile_number': lead.get('customer_mobile_number'),
                    'last_enquiry_date': last_enquiry_date,
                    'days_old': days_old
                })
        # Sort by most recent (days_old, then last_enquiry_date desc)
        hot_leads.sort(key=lambda x: (x['days_old'], x['last_enquiry_date']), reverse=False)
        return jsonify({'success': True, 'hot_leads': hot_leads})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e), 'hot_leads': []})

@app.route('/delete_duplicate_lead', methods=['POST'])
@require_admin
def delete_duplicate_lead():
    uid = request.form.get('uid')
    if not uid:
        return jsonify({'success': False, 'message': 'No UID provided'})
    try:
        supabase.table('duplicate_leads').delete().eq('uid', uid).execute()
        return jsonify({'success': True, 'message': 'Duplicate lead deleted successfully'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error deleting duplicate lead: {str(e)}'})

@app.route('/cre_dashboard_leads')
@require_cre
def cre_dashboard_leads():
    """AJAX endpoint to get leads table HTML for Won/Lost toggle"""
    status = request.args.get('status', 'lost')
    cre_name = session.get('cre_name')
    
    try:
        # Get all leads for this CRE
        all_leads = safe_get_data('lead_master', {'cre_name': cre_name})
        
        # Initialize lists
        won_leads = []
        lost_leads = []
        
        # Process leads
        for lead in all_leads:
            final_status = lead.get('final_status')
            
            if final_status == 'Won':
                won_leads.append(lead)
            elif final_status == 'Lost':
                lost_leads.append(lead)
        
        # Select the appropriate list based on status
        leads_list = won_leads if status == 'won' else lost_leads
        
        # Render only the table HTML
        return render_template('cre_dashboard_leads_table.html', 
                             leads_list=leads_list, 
                             status=status)
    
    except Exception as e:
        return f'<div class="alert alert-danger">Error loading leads: {str(e)}</div>'

@app.route('/ps_dashboard_leads')
@require_ps
def ps_dashboard_leads():
    """AJAX endpoint to get leads table HTML for Won/Lost toggle"""
    status = request.args.get('status', 'lost')
    ps_name = session.get('ps_name')
    
    try:
        # Get all assigned leads for this PS (reuse logic from ps_dashboard)
        assigned_leads = safe_get_data('ps_followup_master', {'ps_name': ps_name})
        event_leads = safe_get_data('activity_leads', {'ps_name': ps_name})
        
        # Initialize lists
        won_leads = []
        lost_leads = []
        
        # Process regular assigned leads
        for lead in assigned_leads:
            lead_dict = dict(lead)
            lead_dict['lead_uid'] = lead.get('lead_uid')
            final_status = lead.get('final_status')
            
            if final_status == 'Won':
                won_leads.append(lead_dict)
            elif final_status == 'Lost':
                lost_leads.append(lead_dict)
        

        
        # Process event leads
        for lead in event_leads:
            lead_dict = dict(lead)
            lead_dict['lead_uid'] = lead.get('activity_uid', '') or lead.get('uid', '')
            lead_dict['customer_mobile_number'] = lead.get('customer_phone_number', '')
            final_status = lead.get('final_status', '')
            
            if final_status == 'Won':
                won_leads.append(lead_dict)
            elif final_status == 'Lost':
                lost_leads.append(lead_dict)
        
        # Process walk-in leads
        walkin_leads = safe_get_data('walkin_table', {'ps_assigned': ps_name})
        for lead in walkin_leads:
            lead_dict = dict(lead)
            lead_dict['lead_uid'] = lead.get('uid', f"W{lead.get('id')}")
            lead_dict['customer_mobile_number'] = lead.get('mobile_number', '')
            lead_dict['customer_name'] = lead.get('customer_name', '')
            lead_dict['is_walkin'] = True
            lead_dict['walkin_id'] = lead.get('id')
            lead_dict['cre_name'] = ''
            final_status = lead.get('status')
            
            # Add timestamp for filtering
            if final_status == 'Won':
                lead_dict['won_timestamp'] = lead.get('updated_at') or datetime.now().isoformat()
                won_leads.append(lead_dict)
            elif final_status == 'Lost':
                lead_dict['lost_timestamp'] = lead.get('updated_at') or datetime.now().isoformat()
                lost_leads.append(lead_dict)
        
        # Select the appropriate list based on status
        leads_list = won_leads if status == 'won' else lost_leads
        
        # Render only the table HTML
        return render_template('ps_dashboard_leads_table.html', 
                             leads_list=leads_list, 
                             status=status)
    
    except Exception as e:
        return f'<div class="alert alert-danger">Error loading leads: {str(e)}</div>'

# Receptionist session key
RECEPTIONIST_SESSION_KEY = 'rec_user_id'

@app.route('/rec_login', methods=['GET', 'POST'])
def rec_login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        # Fetch receptionist from DB
        rec_user = supabase.table('rec_users').select('*').eq('username', username).execute().data
        if rec_user:
            rec_user = rec_user[0]
            if rec_user.get('is_active', True) and check_password_hash(rec_user['password_hash'], password):
                session.clear()
                session[RECEPTIONIST_SESSION_KEY] = rec_user['id']
                session['rec_branch'] = rec_user['branch']
                session['rec_name'] = rec_user.get('name', username)
                session['user_type'] = 'rec'  # Add user_type for the decorator
                flash('Login successful!', 'success')
                return redirect(url_for('add_walkin_lead'))
            else:
                flash('Invalid credentials or inactive account.', 'danger')
        else:
            flash('Invalid username or password.', 'danger')
    return render_template('rec_login.html')

@app.route('/rec_logout')
def rec_logout():
    session.pop(RECEPTIONIST_SESSION_KEY, None)
    session.pop('rec_branch', None)
    session.pop('rec_name', None)
    session.pop('user_type', None)
    flash('Logged out successfully.', 'info')
    return redirect(url_for('rec_login'))
@app.route('/rec_dashboard')
@require_rec
def rec_dashboard():
    """Receptionist Dashboard with KPIs"""
    if 'rec_branch' not in session:
        return redirect(url_for('rec_login'))
    
    branch = session.get('rec_branch')
    
    try:
        # Calculate KPI counts for the receptionist's branch
        # Total walk-in leads
        total_walkin = supabase.table('walkin_table').select('*').eq('branch', branch).execute().data or []
        total_walkin_count = len(total_walkin)
        
        # Won leads
        won_walkin = supabase.table('walkin_table').select('*').eq('branch', branch).eq('status', 'Won').execute().data or []
        won_leads_count = len(won_walkin)
        
        # Lost leads
        lost_walkin = supabase.table('walkin_table').select('*').eq('branch', branch).eq('status', 'Lost').execute().data or []
        lost_leads_count = len(lost_walkin)
        
        # Pending leads
        pending_walkin = supabase.table('walkin_table').select('*').eq('branch', branch).eq('status', 'Pending').execute().data or []
        pending_leads_count = len(pending_walkin)
        
        # Today's walk-in leads
        today_str = datetime.now().strftime('%Y-%m-%d')
        today_walkin = supabase.table('walkin_table').select('*').eq('branch', branch).eq('created_at', today_str).execute().data or []
        today_leads_count = len(today_walkin)
        
        # First call response rate calculation
        # Get leads that have at least one call recorded
        leads_with_calls = supabase.table('walkin_table').select('*').eq('branch', branch).not_.is_('first_call_date', 'null').execute().data or []
        leads_with_calls_count = len(leads_with_calls)
        
        # Calculate first call response rate
        first_call_response_rate = 0
        if total_walkin_count > 0:
            first_call_response_rate = round((leads_with_calls_count / total_walkin_count) * 100, 2)
        
        # Conversion rate
        conversion_rate = 0
        if total_walkin_count > 0:
            conversion_rate = round((won_leads_count / total_walkin_count) * 100, 2)
        
    except Exception as e:
        print(f"Error calculating receptionist KPIs: {str(e)}")
        total_walkin_count = won_leads_count = lost_leads_count = pending_leads_count = first_call_response_rate = today_leads_count = conversion_rate = 0
    
    # Get all leads for the branch with filtering
    try:
        # Get filter parameters
        ps_filter = request.args.get('ps_filter', '')
        status_filter = request.args.get('status_filter', '')
        date_filter = request.args.get('date_filter', '')
        
        # Build query
        query = supabase.table('walkin_table').select('*').eq('branch', branch)
        
        if ps_filter:
            query = query.eq('ps_assigned', ps_filter)
        if status_filter:
            query = query.eq('status', status_filter)
        if date_filter:
            query = query.eq('created_at', date_filter)
            
        all_leads = query.order('created_at', desc=True).execute().data or []
        
        # Get unique PS names for filter dropdown
        ps_names = supabase.table('walkin_table').select('ps_assigned').eq('branch', branch).not_.is_('ps_assigned', 'null').execute().data or []
        ps_options = list(set([lead['ps_assigned'] for lead in ps_names if lead['ps_assigned']]))
        
        # Get unique statuses for filter dropdown
        statuses = supabase.table('walkin_table').select('status').eq('branch', branch).execute().data or []
        status_options = list(set([lead['status'] for lead in statuses if lead['status']]))
        
    except Exception as e:
        print(f"Error fetching leads data: {str(e)}")
        all_leads = []
        ps_options = []
        status_options = []
    
    return render_template('rec_dashboard.html',
                         total_walkin_count=total_walkin_count,
                         won_leads_count=won_leads_count,
                         lost_leads_count=lost_leads_count,
                         pending_leads_count=pending_leads_count,
                         first_call_response_rate=first_call_response_rate,
                         today_leads_count=today_leads_count,
                         conversion_rate=conversion_rate,
                         all_leads=all_leads,
                         ps_options=ps_options,
                         status_options=status_options,
                         current_ps_filter=request.args.get('ps_filter', ''),
                         current_status_filter=request.args.get('status_filter', ''),
                         current_date_filter=request.args.get('date_filter', ''))

@app.route('/api/filter_leads')
@require_rec
def filter_leads():
    """AJAX endpoint for filtering leads"""
    if 'rec_branch' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    
    branch = session.get('rec_branch')
    
    try:
        # Get filter parameters
        ps_filter = request.args.get('ps_filter', '')
        status_filter = request.args.get('status_filter', '')
        date_filter_type = request.args.get('date_filter_type', '')
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        
        # Build query
        query = supabase.table('walkin_table').select('*').eq('branch', branch)
        
        if ps_filter:
            query = query.eq('ps_assigned', ps_filter)
        if status_filter:
            query = query.eq('status', status_filter)
        
        # Handle date filtering
        if date_filter_type == 'today':
            today_str = datetime.now().strftime('%Y-%m-%d')
            query = query.eq('created_at', today_str)
        elif date_filter_type == 'range' and date_from and date_to:
            query = query.gte('created_at', date_from).lte('created_at', date_to)
        elif date_filter_type == 'range' and date_from:
            query = query.gte('created_at', date_from)
        elif date_filter_type == 'range' and date_to:
            query = query.lte('created_at', date_to)
            
        all_leads = query.order('created_at', desc=True).execute().data or []
        
        return jsonify({'leads': all_leads})
        
    except Exception as e:
        print(f"Error filtering leads: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/view_walkin_call_history/<uid>')
@require_rec
def view_walkin_call_history(uid):
    """View call history for a specific walk-in lead"""
    if 'rec_branch' not in session:
        return redirect(url_for('rec_login'))
    
    branch = session.get('rec_branch')
    
    try:
        # Get the walk-in lead data
        walkin_result = supabase.table('walkin_table').select('*').eq('uid', uid).eq('branch', branch).execute()
        if not walkin_result.data:
            flash('Walk-in lead not found', 'error')
            return redirect(url_for('rec_dashboard'))
        
        walkin_lead = walkin_result.data[0]
        
        # Prepare call history data
        call_history = []
        
        # Check for each call attempt
        call_fields = ['first', 'second', 'third', 'fourth', 'fifth', 'sixth', 'seventh']
        
        for i, call_type in enumerate(call_fields, 1):
            call_date = walkin_lead.get(f'{call_type}_call_date')
            call_remark = walkin_lead.get(f'{call_type}_call_remark')
            
            if call_date:
                call_history.append({
                    'call_number': i,
                    'call_type': call_type.title(),
                    'date': call_date,
                    'remark': call_remark or 'No remarks',
                    'status': 'Completed'
                })
            else:
                call_history.append({
                    'call_number': i,
                    'call_type': call_type.title(),
                    'date': 'Not recorded',
                    'remark': 'No call made yet',
                    'status': 'Pending'
                })
        
        return render_template('view_walkin_call_history.html',
                             walkin_lead=walkin_lead,
                             call_history=call_history)
        
    except Exception as e:
        flash(f'Error viewing call history: {str(e)}', 'error')
        return redirect(url_for('rec_dashboard'))

@app.route('/add_walkin_lead', methods=['GET', 'POST'])
@require_rec
def add_walkin_lead():

    # Models and branches (customize as needed)
    models = [
        "450X", "450S", "Rizta S Mono (2.9 kWh)", "Rizta S Super Matte (2.9 kWh)",
        "Rizta Z Mono (2.9 kWh)", "Rizta Z Duo (2.9 kWh)", "Rizta Z Super Matte (2.9 kWh)",
        "Rizta Z Mono (3.7 kWh)", "Rizta Z Duo (3.7 kWh)", "Rizta Z Super Matte (3.7 kWh)",
        "450 X (2.9 kWh)", "450 X (3.7 kWh)", "450 X (2.9 kWh) Pro Pack", "450 X (3.7 kWh) Pro Pack",
        "450 Apex STD"
    ]
    branches = get_system_branches()
    # Fetch PS users for all branches
    ps_users = supabase.table('ps_users').select('name,branch').execute().data or []
    ps_options = {}
    for branch in branches:
        ps_options[branch] = [ps['name'] for ps in ps_users if ps.get('branch') == branch]
    ps_options_json = json.dumps(ps_options)

    if request.method == 'POST':
        # Get form data
        customer_name = request.form['customer_name']
        mobile_number = request.form['mobile_number']
        customer_location = request.form.get('customer_location')
        model_interested = request.form['model_interested']
        occupation = request.form.get('occupation')
        branch = request.form['branch']
        ps_assigned = request.form['ps_assigned']
        
        # Normalize phone number
        normalized_phone = ''.join(filter(str.isdigit, mobile_number))
        
        # Check for duplicates in ALL relevant tables
        try:
            print(f"🔍 Checking for duplicates for phone: {normalized_phone}")
            
            # 1. Check in walkin_table FIRST (most important for walkin leads)
            walkin_result = supabase.table('walkin_table').select('*').eq('mobile_number', normalized_phone).execute()
            existing_walkin_leads = walkin_result.data or []
            
            if existing_walkin_leads:
                print(f"📱 Found {len(existing_walkin_leads)} existing walkin leads")
                # Add to duplicate leads table instead of redirecting
                try:
                    # Create duplicate record
                    duplicate_data = {
                        'uid': existing_walkin_leads[0].get('uid'),
                        'customer_mobile_number': normalized_phone,
                        'customer_name': customer_name,
                        'original_lead_id': existing_walkin_leads[0].get('id'),
                        'source1': 'Walk-in',
                        'sub_source1': existing_walkin_leads[0].get('branch', 'N/A'),
                        'date1': existing_walkin_leads[0].get('created_at', ''),
                        'source2': 'Walk-in',
                        'sub_source2': branch,
                        'date2': datetime.now().strftime('%Y-%m-%d'),
                        'duplicate_count': 2,
                        'created_at': datetime.now().isoformat(),
                        'updated_at': datetime.now().isoformat(),
                        'original_table': 'walkin_table',
                        'original_record_id': existing_walkin_leads[0].get('id')
                    }
                    
                    # Initialize remaining slots as None
                    for i in range(3, 11):
                        duplicate_data[f'source{i}'] = None
                        duplicate_data[f'sub_source{i}'] = None
                        duplicate_data[f'date{i}'] = None
                    
                    supabase.table('duplicate_leads').insert(duplicate_data).execute()
                    flash('Lead already exists as Walk-in. Added to duplicate leads table for admin review.', 'warning')
                    return redirect(url_for('add_walkin_lead'))
                except Exception as e:
                    flash(f'Error adding to duplicates: {str(e)}', 'error')
                    return render_template('add_walkin_lead.html', models=models, branches=branches, ps_options_json=ps_options_json, default_branch=branch)
            
            # 2. Check in activity_leads table
            activity_result = supabase.table('activity_leads').select('*').eq('customer_phone_number', normalized_phone).execute()
            existing_activity_leads = activity_result.data or []
            
            if existing_activity_leads:
                print(f"🎯 Found {len(existing_activity_leads)} existing activity leads")
                # This will be added as duplicate with new source
            
            # 3. Check in lead_master table
            lead_master_result = supabase.table('lead_master').select('*').eq('customer_mobile_number', normalized_phone).execute()
            existing_leads = lead_master_result.data or []
            
            if existing_leads:
                print(f"📋 Found {len(existing_leads)} existing leads in lead_master")
                # Check for exact source-subsource match (Walk-in)
                exact_match = any(
                    lead.get('source') == 'Walk-in' 
                    for lead in existing_leads
                )
                
                if exact_match:
                    # Add to duplicate leads table instead of redirecting
                    try:
                        # Create duplicate record
                        duplicate_data = {
                            'uid': existing_leads[0].get('uid'),
                            'customer_mobile_number': normalized_phone,
                            'customer_name': customer_name,
                            'original_lead_id': existing_leads[0].get('id'),
                            'source1': existing_leads[0].get('source', 'Unknown'),
                            'sub_source1': existing_leads[0].get('sub_source', 'N/A'),
                            'date1': existing_leads[0].get('date', ''),
                            'source2': 'Walk-in',
                            'sub_source2': branch,
                            'date2': datetime.now().strftime('%Y-%m-%d'),
                            'duplicate_count': 2,
                            'created_at': datetime.now().isoformat(),
                            'updated_at': datetime.now().isoformat(),
                            'original_table': 'lead_master',
                            'original_record_id': existing_leads[0].get('id')
                        }
                        
                        # Initialize remaining slots as None
                        for i in range(3, 11):
                            duplicate_data[f'source{i}'] = None
                            duplicate_data[f'sub_source{i}'] = None
                            duplicate_data[f'date{i}'] = None
                        
                        supabase.table('duplicate_leads').insert(duplicate_data).execute()
                        flash('Lead already exists with different source. Added to duplicate leads table for admin review.', 'warning')
                        return redirect(url_for('add_walkin_lead'))
                    except Exception as e:
                        flash(f'Error adding to duplicates: {str(e)}', 'error')
                        return render_template('add_walkin_lead.html', models=models, branches=branches, ps_options_json=ps_options_json, default_branch=branch)
            
            # 4. Check in duplicate_leads
            duplicate_result = supabase.table('duplicate_leads').select('*').eq('customer_mobile_number', normalized_phone).execute()
            duplicate_leads = duplicate_result.data or []
            
            if duplicate_leads:
                print(f"🔄 Found {len(duplicate_leads)} existing duplicate leads")
                # Check if Walk-in source already exists in any slot
                exact_match = False
                for duplicate_lead in duplicate_leads:
                    # Check all source slots (source1 to source10)
                    for i in range(1, 11):
                        source_field = f'source{i}'
                        
                        if duplicate_lead.get(source_field) == 'Walk-in':
                            exact_match = True
                            break
                    if exact_match:
                        break
                
                if exact_match:
                    # Add to duplicate leads table instead of redirecting
                    try:
                        # Update existing duplicate record with new source
                        duplicate_lead = duplicate_leads[0]
                        
                        # Find next available slot
                        next_slot = None
                        for i in range(1, 11):
                            source_field = f'source{i}'
                            if not duplicate_lead.get(source_field):
                                next_slot = i
                                break
                        
                        if next_slot:
                            # Update the existing duplicate record
                            update_data = {
                                f'source{next_slot}': 'Walk-in',
                                f'sub_source{next_slot}': branch,
                                f'date{next_slot}': datetime.now().strftime('%Y-%m-%d'),
                                'duplicate_count': duplicate_lead.get('duplicate_count', 1) + 1,
                                'updated_at': datetime.now().isoformat()
                            }
                            supabase.table('duplicate_leads').update(update_data).eq('id', duplicate_lead['id']).execute()
                            flash(f'Lead already exists in duplicates. Added new Walk-in source to slot {next_slot}.', 'warning')
                        else:
                            flash('Duplicate record is full (max 10 sources reached). Please contact admin.', 'error')
                        
                        return redirect(url_for('add_walkin_lead'))
                    except Exception as e:
                        flash(f'Error updating duplicate record: {str(e)}', 'error')
                        return render_template('add_walkin_lead.html', models=models, branches=branches, ps_options_json=ps_options_json, default_branch=branch)
                    
        except Exception as e:
            print(f"❌ Error checking duplicates: {e}")
            flash('Error checking for duplicates. Please try again.', 'error')
            return render_template('add_walkin_lead.html', models=models, branches=branches, ps_options_json=ps_options_json, default_branch=branch)
        
        # Generate UID for walk-in lead
        # Get count of existing walk-in leads for this mobile number
        existing_count = get_accurate_count('walkin_table', {'mobile_number': mobile_number})
        sequence = existing_count + 1
        
        # Generate UID using 'Walk-in' as source (which maps to 'W')
        uid = generate_uid('Walk-in', mobile_number, sequence)
        
        # Check if this is a duplicate with new source (phone exists but not as Walk-in)
        is_duplicate_new_source = False
        original_lead = None
        
        if existing_activity_leads:
            # Phone exists in activity_leads - this is a duplicate with new source
            is_duplicate_new_source = True
            original_lead = existing_activity_leads[0]
        elif existing_leads:
            # Phone exists in lead_master but not as Walk-in - this is a duplicate with new source
            is_duplicate_new_source = True
            original_lead = existing_leads[0]
        elif duplicate_leads:
            # Phone exists in duplicate_leads but not as Walk-in - this is also a duplicate
            is_duplicate_new_source = True
            original_lead = duplicate_leads[0]
        
        if is_duplicate_new_source:
            # This is a duplicate with new source - add to duplicate_leads table
            try:
                if original_lead:
                    # Create new duplicate record
                    duplicate_data = {
                        'uid': uid,
                        'customer_mobile_number': normalized_phone,
                        'customer_name': customer_name,
                        'original_lead_id': None,  # We'll handle this differently
                        'source1': original_lead.get('source', 'Event' if 'activity_uid' in original_lead else 'Walk-in' if 'id' in original_lead else 'Unknown'),
                        'sub_source1': original_lead.get('sub_source') or original_lead.get('location', 'N/A') or original_lead.get('branch', 'N/A'),
                        'date1': original_lead.get('date') or original_lead.get('created_at', ''),
                        'source2': 'Walk-in',
                        'sub_source2': branch,  # Use branch as sub-source for walk-in
                        'date2': datetime.now().strftime('%Y-%m-%d'),
                        'duplicate_count': 2,
                        'created_at': datetime.now().isoformat(),
                        'updated_at': datetime.now().isoformat()
                    }
                    
                    # Try to add the new columns if they exist (for better source tracking)
                    try:
                        duplicate_data['original_table'] = 'walkin_table' if 'id' in original_lead else 'activity_leads' if 'activity_uid' in original_lead else 'lead_master'
                        duplicate_data['original_record_id'] = original_lead.get('id') or original_lead.get('activity_uid') or original_lead.get('uid')
                    except:
                        pass  # These columns might not exist yet
                    
                    # Initialize remaining slots as None
                    for i in range(3, 11):
                        duplicate_data[f'source{i}'] = None
                        duplicate_data[f'sub_source{i}'] = None
                        duplicate_data[f'date{i}'] = None
                    
                    supabase.table('duplicate_leads').insert(duplicate_data).execute()
                    flash(f'Lead added to duplicates with new source: Walk-in - {branch}. Redirecting to duplicate leads section.', 'success')
                    return redirect(url_for('admin_duplicate_leads', uid=uid))
                else:
                    flash('Error: Original lead not found for duplicate creation', 'error')
                    return render_template('add_walkin_lead.html', models=models, branches=branches, ps_options_json=ps_options_json, default_branch=branch)
            except Exception as e:
                flash(f'Error creating duplicate record: {str(e)}', 'error')
                return render_template('add_walkin_lead.html', models=models, branches=branches, ps_options_json=ps_options_json, default_branch=branch)
        else:
            # This is a new lead - add to walkin_table
            data = {
                'customer_name': customer_name,
                'mobile_number': mobile_number,
                'customer_location': customer_location,
                'model_interested': model_interested,
                'occupation': occupation,
                'branch': branch,
                'ps_assigned': ps_assigned,
                'status': 'Pending',
                'followup_no': 1,
                'uid': uid,  # Add UID
                'created_at': datetime.now().isoformat(),
                'updated_at': datetime.now().isoformat(),
            }
            try:
                # Insert into walkin_table only
                supabase.table('walkin_table').insert(data).execute()
                
                flash('Walk-in lead added and assigned to PS successfully!', 'success')
                return redirect(url_for('add_walkin_lead'))
            except Exception as e:
                error_str = str(e)
                if 'duplicate key value violates unique constraint "walkin_table_pkey"' in error_str:
                    # This is a sequence issue - try to fix it by using a different approach
                    try:
                        # Get the maximum ID from the table
                        max_id_result = supabase.table('walkin_table').select('id').order('id', desc=True).limit(1).execute()
                        if max_id_result.data:
                            max_id = max_id_result.data[0]['id']
                            
                            # Try to insert with a higher ID manually
                            # First, try to find the next available ID
                            next_id = max_id + 1
                            while True:
                                # Check if this ID exists
                                existing = supabase.table('walkin_table').select('id').eq('id', next_id).execute()
                                if not existing.data:
                                    # This ID is available, use it
                                    data['id'] = next_id
                                    break
                                next_id += 1
                                if next_id > max_id + 100:  # Safety limit
                                    flash('Error: Could not find an available ID. Please contact administrator.', 'danger')
                                    break
                            
                            if 'id' in data:
                                # Try the insertion with the specific ID
                                supabase.table('walkin_table').insert(data).execute()
                                flash('Walk-in lead added and assigned to PS successfully!', 'success')
                                return redirect(url_for('add_walkin_lead'))
                        else:
                            flash('Error: Could not determine the maximum ID in the table', 'danger')
                    except Exception as reset_error:
                        flash(f'Error fixing sequence: {str(reset_error)}. Please contact administrator.', 'danger')
                else:
                    flash(f'Error adding walk-in lead: {error_str}', 'danger')

    # Get receptionist's branch from session
    rec_branch = session.get('rec_branch', '')
    
    return render_template(
        'add_walkin_lead.html',
        models=models,
        branches=branches,
        ps_options_json=ps_options_json,
        default_branch=rec_branch
    )

@app.route('/api/branch_analytics/ps_performance')
def api_branch_analytics_ps_performance():
    """API endpoint for Product Specialist Performance data - OPTIMIZED"""
    try:
        # Get branch from session
        branch = session.get('branch_head_branch')
        if not branch:
            return jsonify({'success': False, 'message': 'Branch not found in session'})
        
        # Get date filter parameters
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')
        
        # Require both date parameters
        if not date_from or not date_to:
            return jsonify({'success': False, 'message': 'Both date_from and date_to parameters are required'})
        
        # OPTIMIZATION: Single query to get all PS performance data at once
        # Use a more efficient query with proper aggregation
        performance_query = supabase.table('ps_followup_master').select(
            'ps_name, ps_assigned_at, lead_status'
        ).eq('ps_branch', branch).gte('ps_assigned_at', date_from).lte('ps_assigned_at', date_to).execute()
        
        if not performance_query.data:
            return jsonify({'success': True, 'data': []})
        
        # Process data in Python (much faster than multiple DB queries)
        ps_performance = {}
        for lead in performance_query.data:
            ps_name = lead.get('ps_name')
            if not ps_name:
                continue
                
            if ps_name not in ps_performance:
                ps_performance[ps_name] = {'leads_assigned': 0, 'leads_contacted': 0}
            
            ps_performance[ps_name]['leads_assigned'] += 1
            if lead.get('lead_status'):
                ps_performance[ps_name]['leads_contacted'] += 1
        
        # Format result
        result_data = []
        for ps_name, stats in ps_performance.items():
            result_data.append({
                'ps_name': ps_name,
                'leads_assigned': stats['leads_assigned'],
                'leads_contacted': stats['leads_contacted'],
                'gap': stats['leads_assigned'] - stats['leads_contacted']
            })
        
        return jsonify({'success': True, 'data': result_data})
        
    except Exception as e:
        print(f"Error in ps_performance API: {str(e)}")
        return jsonify({'success': False, 'message': 'Error loading PS performance data'})

@app.route('/api/branch_analytics/source_leads')
def api_branch_analytics_source_leads():
    """API endpoint for Source-wise Leads Analysis data - OPTIMIZED"""
    try:
        # Get branch from session
        branch = session.get('branch_head_branch')
        if not branch:
            return jsonify({'success': False, 'message': 'Branch not found in session'})
        
        # Get date filter parameters
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')
        
        # Require both date parameters
        if not date_from or not date_to:
            return jsonify({'success': False, 'message': 'Both date_from and date_to parameters are required'})
        
        # OPTIMIZATION: Single query to get all leads data at once
        # Select only necessary columns and filter at database level
        leads_query = supabase.table('ps_followup_master').select(
            'ps_name, source, final_status'
        ).eq('ps_branch', branch).gte('ps_assigned_at', date_from).lte('ps_assigned_at', date_to).execute()
        
        if not leads_query.data:
            return jsonify({'success': True, 'data': [], 'sources': []})
        
        # Process data in Python (much faster than multiple DB queries)
        ps_data = {}
        all_sources = set()
        
        for lead in leads_query.data:
            ps_name = lead.get('ps_name')
            source = lead.get('source', 'Unknown')
            final_status = lead.get('final_status', 'Pending')
            
            if not ps_name:
                continue
                
            if ps_name not in ps_data:
                ps_data[ps_name] = {'total_leads': 0, 'won_leads': 0, 'sources': {}}
            
            ps_data[ps_name]['total_leads'] += 1
            if final_status == 'Won':
                ps_data[ps_name]['won_leads'] += 1
            
            # Track source breakdown
            if source not in ps_data[ps_name]['sources']:
                ps_data[ps_name]['sources'][source] = {'total_leads': 0, 'won_leads': 0}
            
            ps_data[ps_name]['sources'][source]['total_leads'] += 1
            if final_status == 'Won':
                ps_data[ps_name]['sources'][source]['won_leads'] += 1
            
            all_sources.add(source)
        
        # Format result data
        result_data = []
        for ps_name, stats in ps_data.items():
            result_data.append({
                'ps_name': ps_name,
                'total_leads': stats['total_leads'],
                'won_leads': stats['won_leads'],
                'source_breakdown': stats['sources']
            })
        
        # Convert all_sources to sorted list for consistent column order
        all_sources_list = sorted(list(all_sources))
        
        return jsonify({
            'success': True, 
            'data': result_data,
            'sources': all_sources_list
        })
        
    except Exception as e:
        print(f"Error in source_leads API: {str(e)}")
        return jsonify({'success': False, 'message': 'Error loading source leads data'})

@app.route('/api/branch_analytics/walkin_leads')
def api_branch_analytics_walkin_leads():
    """API endpoint for Walk-in Leads Summary data - OPTIMIZED"""
    try:
        # Get branch from session
        branch = session.get('branch_head_branch')
        if not branch:
            return jsonify({'success': False, 'message': 'Branch not found in session'})
        
        # Get date filter parameters
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')
        
        # Require both date parameters
        if not date_from or not date_to:
            return jsonify({'success': False, 'message': 'Both date_from and date_to parameters are required'})
        
        today_str = datetime.now().strftime('%Y-%m-%d')
        
        # OPTIMIZATION: Single query to get all walk-in leads data at once
        # Select only necessary columns and filter at database level
        walkin_query = supabase.table('walkin_table').select(
            'ps_assigned, status, next_followup_date'
        ).eq('branch', branch).gte('created_at', date_from).lte('created_at', date_to).execute()
        
        if not walkin_query.data:
            return jsonify({'success': True, 'data': []})
        
        # Process data in Python (much faster than multiple DB queries)
        ps_data = {}
        
        for lead in walkin_query.data:
            ps_name = lead.get('ps_assigned')
            status = lead.get('status', 'Pending')
            next_followup_date = lead.get('next_followup_date')
            
            if not ps_name:
                continue
                
            if ps_name not in ps_data:
                ps_data[ps_name] = {
                    'total_walkin_leads': 0,
                    'pending_leads': 0,
                    'lost_leads': 0,
                    'won_leads': 0,
                    'today_followups': 0
                }
            
            ps_data[ps_name]['total_walkin_leads'] += 1
            
            # Count by status
            if status == 'Pending':
                ps_data[ps_name]['pending_leads'] += 1
            elif status == 'Lost':
                ps_data[ps_name]['lost_leads'] += 1
            elif status == 'Won':
                ps_data[ps_name]['won_leads'] += 1
            
            # Count follow-ups with date <= today and final_status = 'Pending'
            final_status = lead.get('final_status')
            if next_followup_date and final_status == 'Pending':
                try:
                    if 'T' in str(next_followup_date):
                        followup_date_parsed = datetime.fromisoformat(str(next_followup_date).replace('Z', '+00:00')).date()
                    else:
                        followup_date_parsed = datetime.strptime(str(next_followup_date)[:10], '%Y-%m-%d').date()
                    
                    if followup_date_parsed <= datetime.now().date():
                        ps_data[ps_name]['today_followups'] += 1
                except (ValueError, TypeError):
                    # If date parsing fails, include for manual review
                    ps_data[ps_name]['today_followups'] += 1
        
        # Format result data
        result_data = []
        for ps_name, stats in ps_data.items():
            result_data.append({
                'ps_name': ps_name,
                'total_walkin_leads': stats['total_walkin_leads'],
                'pending_leads': stats['pending_leads'],
                'lost_leads': stats['lost_leads'],
                'won_leads': stats['won_leads'],
                'today_followups': stats['today_followups']
            })
        
        return jsonify({'success': True, 'data': result_data})
        
    except Exception as e:
        print(f"Error in walkin_leads API: {str(e)}")
        return jsonify({'success': False, 'message': 'Error loading walkin leads data'})

@app.route('/api/branch_sources')
def api_branch_sources():
    """Get unique source values for the current branch"""
    try:
        branch = session.get('branch_head_branch')
        
        # Get unique sources from ps_followup_master
        ps_sources = supabase.table('ps_followup_master').select('source').eq('ps_branch', branch).not_.is_('source', 'null').execute().data or []
        ps_source_values = list(set([row.get('source', '') for row in ps_sources if row.get('source', '')]))
        
        # Add Walk In and Event sources
        all_sources = ['Walk In', 'Event'] + ps_source_values
        
        # Remove duplicates and sort
        unique_sources = sorted(list(set(all_sources)))
        
        return jsonify({
            'success': True,
            'sources': unique_sources
        })
    except Exception as e:
        print(f"Error getting branch sources: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e),
            'sources': ['Walk In', 'Event', 'PS', 'Meta', 'Google', 'Website', 'Referral', 'Other']
        })

@app.route('/api/branch_analytics/all')
def api_branch_analytics_all():
    """OPTIMIZED: Single endpoint to get all analytics data at once - reduces API calls"""
    start_time = time.time()
    print(f"[DEBUG] /api/branch_analytics/all called at {datetime.now()}")
    
    try:
        branch = session.get('branch_head_branch')
        print(f"[DEBUG] Branch from session: {branch}")
        if not branch:
            print("[DEBUG] No branch found in session")
            return jsonify({'success': False, 'message': 'Branch not found in session'})
        
        # Get date filter parameters
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')
        print(f"[DEBUG] Date filter: {date_from} to {date_to}")
        
        # If no date parameters provided, use MTD (Month to Date) as default
        if not date_from or not date_to:
            today = datetime.now()
            date_from = today.replace(day=1).strftime('%Y-%m-%d')
            date_to = today.strftime('%Y-%m-%d')
            print(f"[DEBUG] Using default dates: {date_from} to {date_to}")
        
        today_str = datetime.now().strftime('%Y-%m-%d')
        print(f"[DEBUG] Today's date: {today_str}")
        
        # OPTIMIZATION: Single query for each table with all necessary data
        # This reduces database round trips significantly
        
        print("[DEBUG] Starting database queries...")
        query_start = time.time()
        
        # Get PS performance data
        print("[DEBUG] Querying ps_followup_master for performance data...")
        ps_performance_query = supabase.table('ps_followup_master').select(
            'ps_name, ps_assigned_at, lead_status'
        ).eq('ps_branch', branch).gte('ps_assigned_at', date_from).lte('ps_assigned_at', date_to).execute()
        print(f"[DEBUG] PS performance query completed in {time.time() - query_start:.3f}s, got {len(ps_performance_query.data or [])} records")
        
        # Get source leads data
        print("[DEBUG] Querying ps_followup_master for source leads data...")
        source_leads_query = supabase.table('ps_followup_master').select(
            'ps_name, source, final_status'
        ).eq('ps_branch', branch).gte('ps_assigned_at', date_from).lte('ps_assigned_at', date_to).execute()
        print(f"[DEBUG] Source leads query completed in {time.time() - query_start:.3f}s, got {len(source_leads_query.data or [])} records")
        
        # Get walkin leads data
        print("[DEBUG] Querying walkin_table for walkin leads data...")
        walkin_leads_query = supabase.table('walkin_table').select(
            'ps_assigned, status, next_followup_date'
        ).eq('branch', branch).gte('created_at', date_from).lte('created_at', date_to).execute()
        print(f"[DEBUG] Walkin leads query completed in {time.time() - query_start:.3f}s, got {len(walkin_leads_query.data or [])} records")
        
        # Get summary data
        print("[DEBUG] Querying for summary data...")
        ps_summary_query = supabase.table('ps_followup_master').select(
            'final_status'
        ).eq('ps_branch', branch).gte('created_at', date_from).lte('created_at', date_to).execute()
        print(f"[DEBUG] PS summary query completed in {time.time() - query_start:.3f}s, got {len(ps_summary_query.data or [])} records")
        
        walkin_summary_query = supabase.table('walkin_table').select(
            'status'
        ).eq('branch', branch).gte('created_at', date_from).lte('created_at', date_to).execute()
        print(f"[DEBUG] Walkin summary query completed in {time.time() - query_start:.3f}s, got {len(walkin_summary_query.data or [])} records")
        
        total_query_time = time.time() - query_start
        print(f"[DEBUG] All database queries completed in {total_query_time:.3f}s")
        
        # Process PS Performance data
        print("[DEBUG] Processing PS performance data...")
        process_start = time.time()
        ps_performance = {}
        if ps_performance_query.data:
            for lead in ps_performance_query.data:
                ps_name = lead.get('ps_name')
                if not ps_name:
                    continue
                    
                if ps_name not in ps_performance:
                    ps_performance[ps_name] = {'leads_assigned': 0, 'leads_contacted': 0}
                
                ps_performance[ps_name]['leads_assigned'] += 1
                if lead.get('lead_status'):
                    ps_performance[ps_name]['leads_contacted'] += 1
        
        # Process Source Leads data
        print("[DEBUG] Processing source leads data...")
        source_data = {}
        all_sources = set()
        if source_leads_query.data:
            for lead in source_leads_query.data:
                ps_name = lead.get('ps_name')
                source = lead.get('source', 'Unknown')
                final_status = lead.get('final_status', 'Pending')
                
                if not ps_name:
                    continue
                    
                if ps_name not in source_data:
                    source_data[ps_name] = {'total_leads': 0, 'won_leads': 0, 'sources': {}}
                
                source_data[ps_name]['total_leads'] += 1
                if final_status == 'Won':
                    source_data[ps_name]['won_leads'] += 1
                
                if source not in source_data[ps_name]['sources']:
                    source_data[ps_name]['sources'][source] = {'total_leads': 0, 'won_leads': 0}
                
                source_data[ps_name]['sources'][source]['total_leads'] += 1
                if final_status == 'Won':
                    source_data[ps_name]['sources'][source]['won_leads'] += 1
                
                all_sources.add(source)
        
        # Process Walkin Leads data
        print("[DEBUG] Processing walkin leads data...")
        walkin_data = {}
        if walkin_leads_query.data:
            for lead in walkin_leads_query.data:
                ps_name = lead.get('ps_assigned')
                status = lead.get('status', 'Pending')
                next_followup_date = lead.get('next_followup_date')
                
                if not ps_name:
                    continue
                    
                if ps_name not in walkin_data:
                    walkin_data[ps_name] = {
                        'total_walkin_leads': 0,
                        'pending_leads': 0,
                        'lost_leads': 0,
                        'won_leads': 0,
                        'today_followups': 0
                    }
                
                walkin_data[ps_name]['total_walkin_leads'] += 1
                
                if status == 'Pending':
                    walkin_data[ps_name]['pending_leads'] += 1
                elif status == 'Lost':
                    walkin_data[ps_name]['lost_leads'] += 1
                elif status == 'Won':
                    walkin_data[ps_name]['won_leads'] += 1
                
                # Count follow-ups with date <= today and status = 'Pending'
                if next_followup_date and status == 'Pending':
                    try:
                        if 'T' in str(next_followup_date):
                            followup_date_parsed = datetime.fromisoformat(str(next_followup_date).replace('Z', '+00:00')).date()
                        else:
                            followup_date_parsed = datetime.strptime(str(next_followup_date)[:10], '%Y-%m-%d').date()
                        
                        if followup_date_parsed <= datetime.now().date():
                            walkin_data[ps_name]['today_followups'] += 1
                    except (ValueError, TypeError):
                        # If date parsing fails, include for manual review
                        walkin_data[ps_name]['today_followups'] += 1
        
        # Process Summary data
        print("[DEBUG] Processing summary data...")
        ps_total = ps_pending = ps_won = ps_lost = 0
        if ps_summary_query.data:
            for row in ps_summary_query.data:
                ps_total += 1
                status = row.get('final_status', 'Pending')
                if status == 'Pending':
                    ps_pending += 1
                elif status == 'Won':
                    ps_won += 1
                elif status == 'Lost':
                    ps_lost += 1
        
        walkin_total = walkin_pending = walkin_won = walkin_lost = 0
        if walkin_summary_query.data:
            for row in walkin_summary_query.data:
                walkin_total += 1
                status = row.get('status', 'Pending')
                if status == 'Pending':
                    walkin_pending += 1
                elif status == 'Won':
                    walkin_won += 1
                elif status == 'Lost':
                    walkin_lost += 1
        
        process_time = time.time() - process_start
        print(f"[DEBUG] Data processing completed in {process_time:.3f}s")
        
        # Format all data
        print("[DEBUG] Formatting result data...")
        format_start = time.time()
        ps_performance_result = []
        for ps_name, stats in ps_performance.items():
            ps_performance_result.append({
                'ps_name': ps_name,
                'leads_assigned': stats['leads_assigned'],
                'leads_contacted': stats['leads_contacted'],
                'gap': stats['leads_assigned'] - stats['leads_contacted']
            })
        
        source_leads_result = []
        for ps_name, stats in source_data.items():
            source_leads_result.append({
                'ps_name': ps_name,
                'total_leads': stats['total_leads'],
                'won_leads': stats['won_leads'],
                'source_breakdown': stats['sources']
            })
        
        walkin_leads_result = []
        for ps_name, stats in walkin_data.items():
            walkin_leads_result.append({
                'ps_name': ps_name,
                'total_walkin_leads': stats['total_walkin_leads'],
                'pending_leads': stats['pending_leads'],
                'lost_leads': stats['lost_leads'],
                'won_leads': stats['won_leads'],
                'today_followups': stats['today_followups']
            })
        
        summary_result = {
            'total_leads_assigned': ps_total + walkin_total,
            'total_pending_leads': ps_pending + walkin_pending,
            'total_won_leads': ps_won + walkin_won,
            'total_lost_leads': ps_lost + walkin_lost
        }
        
        format_time = time.time() - format_start
        print(f"[DEBUG] Data formatting completed in {format_time:.3f}s")
        
        total_time = time.time() - start_time
        print(f"[DEBUG] Total API response time: {total_time:.3f}s")
        print(f"[DEBUG] Breakdown: Queries={total_query_time:.3f}s, Processing={process_time:.3f}s, Formatting={format_time:.3f}s")
        
        return jsonify({
            'success': True,
            'data': {
                'ps_performance': ps_performance_result,
                'source_leads': source_leads_result,
                'walkin_leads': walkin_leads_result,
                'summary': summary_result,
                'sources': sorted(list(all_sources))
            }
        })
        
    except Exception as e:
        total_time = time.time() - start_time
        print(f"[DEBUG] Error in combined analytics API after {total_time:.3f}s: {str(e)}")
        print(f"[DEBUG] Full error traceback:", exc_info=True)
        return jsonify({'success': False, 'message': f'Error loading analytics data: {str(e)}'})
@app.route('/api/branch_summary')
def api_branch_summary():
    """Get branch-wise lead summary with untouched, called, and total counts"""
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        # Get all branches from ps_followup_master
        branches_query = supabase.table('ps_followup_master').select('ps_branch').not_.is_('ps_branch', 'null').execute()
        branches = list(set([row.get('ps_branch') for row in (branches_query.data or []) if row.get('ps_branch')]))
        
        branch_summary = []
        
        for branch in branches:
            # Base query for this branch
            base_query = supabase.table('ps_followup_master').select('*').eq('ps_branch', branch)
            
            # Apply date filter if provided
            if start_date:
                base_query = base_query.gte('ps_assigned_at', start_date)
            if end_date:
                # For end_date, we need to include the full day
                # Add 1 day to end_date to make it inclusive
                end_date_plus_one = (datetime.strptime(end_date, '%Y-%m-%d').date() + timedelta(days=1)).strftime('%Y-%m-%d')
                base_query = base_query.lt('ps_assigned_at', end_date_plus_one)
            
            # Get all leads for this branch
            all_leads = base_query.execute().data or []
            
            # Count untouched leads (follow_up_date is NULL and final_status is Pending)
            untouched_count = len([lead for lead in all_leads 
                                 if not lead.get('follow_up_date') and lead.get('final_status') == 'Pending'])
            
            # Count called leads (all remaining leads)
            called_count = len(all_leads) - untouched_count
            
            # Total count
            total_count = len(all_leads)
            
            branch_summary.append({
                'branch_name': branch,
                'untouched_count': untouched_count,
                'called_count': called_count,
                'total_count': total_count
            })
        
        # Sort by total count descending
        branch_summary.sort(key=lambda x: x['total_count'], reverse=True)
        
        return jsonify({
            'success': True,
            'data': branch_summary
        })
        
    except Exception as e:
        print(f"Error in branch_summary API: {str(e)}")
        return jsonify({'success': False, 'message': f'Error loading branch summary: {str(e)}'})

@app.route('/api/untouched_leads')
def api_untouched_leads():
    """Get untouched leads for a specific branch with drill-down details"""
    try:
        branch_name = request.args.get('branch_name')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        export_csv = request.args.get('export') == 'csv'
        
        if not branch_name:
            return jsonify({'success': False, 'message': 'Branch name is required'})
        
        # Base query for this branch
        base_query = supabase.table('ps_followup_master').select('*').eq('ps_branch', branch_name)
        
        # Apply date filter if provided
        if start_date:
            base_query = base_query.gte('ps_assigned_at', start_date)
        if end_date:
            # For end_date, we need to include the full day
            # Add 1 day to end_date to make it inclusive
            end_date_plus_one = (datetime.strptime(end_date, '%Y-%m-%d').date() + timedelta(days=1)).strftime('%Y-%m-%d')
            base_query = base_query.lt('ps_assigned_at', end_date_plus_one)
        
        # Get all leads for this branch
        all_leads = base_query.execute().data or []
        
        # Filter for untouched leads (follow_up_date is NULL and final_status is Pending)
        untouched_leads = [lead for lead in all_leads 
                          if not lead.get('follow_up_date') and lead.get('final_status') == 'Pending']
        
        if export_csv:
            # Return CSV format
            import csv
            import io
            
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Write headers
            writer.writerow(['Lead UID', 'Customer Name', 'Mobile Number', 'Source', 'Lead Category', 'PS Assigned', 'PS Assigned Date', 'Status'])
            
            # Write data
            for lead in untouched_leads:
                writer.writerow([
                    lead.get('lead_uid', ''),
                    lead.get('customer_name', ''),
                    lead.get('customer_mobile_number', ''),
                    lead.get('source', ''),
                    lead.get('lead_category', ''),
                    lead.get('ps_name', ''),
                    lead.get('ps_assigned_at', ''),
                    lead.get('final_status', '')
                ])
            
            output.seek(0)
            
            return Response(
                output.getvalue(),
                mimetype='text/csv',
                headers={'Content-Disposition': f'attachment; filename=untouched_leads_{branch_name}_{start_date or "all"}_{end_date or "all"}.csv'}
            )
        
        # Return JSON format for modal display
        return jsonify({
            'success': True,
            'data': untouched_leads
        })
        
    except Exception as e:
        print(f"Error in untouched_leads API: {str(e)}")
        return jsonify({'success': False, 'message': f'Error loading untouched leads: {str(e)}'})

@app.route('/api/branch_analytics/summary')
def api_branch_analytics_summary():
    """API endpoint for Branch Summary KPI data - OPTIMIZED"""
    try:
        branch = session.get('branch_head_branch')
        if not branch:
            return jsonify({'success': False, 'message': 'Branch not found in session'})
        
        # Get date filter parameters
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')
        
        # If no date parameters provided, use MTD (Month to Date) as default
        if not date_from or not date_to:
            today = datetime.now()
            date_from = today.replace(day=1).strftime('%Y-%m-%d')
            date_to = today.strftime('%Y-%m-%d')
        
        # OPTIMIZATION: Use count queries with proper filtering instead of fetching all data
        # These are much faster as they only return counts, not actual data
        
        # Get counts from ps_followup_master
        ps_stats = supabase.table('ps_followup_master').select(
            'final_status', count='exact'
        ).eq('ps_branch', branch).gte('created_at', date_from).lte('created_at', date_to).execute()
        
        # Get counts from walkin_table  
        walkin_stats = supabase.table('walkin_table').select(
            'status', count='exact'
        ).eq('branch', branch).gte('created_at', date_from).lte('created_at', date_to).execute()
        
        # Process ps_followup_master data
        ps_total = 0
        ps_pending = 0
        ps_won = 0
        ps_lost = 0
        
        if ps_stats.data:
            for row in ps_stats.data:
                ps_total += 1
                status = row.get('final_status', 'Pending')
                if status == 'Pending':
                    ps_pending += 1
                elif status == 'Won':
                    ps_won += 1
                elif status == 'Lost':
                    ps_lost += 1
        
        # Process walkin_table data
        walkin_total = 0
        walkin_pending = 0
        walkin_won = 0
        walkin_lost = 0
        
        if walkin_stats.data:
            for row in walkin_stats.data:
                walkin_total += 1
                status = row.get('status', 'Pending')
                if status == 'Pending':
                    walkin_pending += 1
                elif status == 'Won':
                    walkin_won += 1
                elif status == 'Lost':
                    walkin_lost += 1
        
        # Calculate totals
        total_leads_assigned = ps_total + walkin_total
        total_pending_leads = ps_pending + walkin_pending
        total_won_leads = ps_won + walkin_won
        total_lost_leads = ps_lost + walkin_lost
        
        return jsonify({
            'success': True,
            'data': {
                'total_leads_assigned': total_leads_assigned,
                'total_pending_leads': total_pending_leads,
                'total_won_leads': total_won_leads,
                'total_lost_leads': total_lost_leads
            }
        })
        
    except Exception as e:
        print(f"Error in branch summary API: {str(e)}")
        return jsonify({'success': False, 'message': 'Error loading branch summary data'})

@app.route('/lead_transfer')
@require_admin
def lead_transfer():
    """Lead Transfer Dashboard for Admin"""
    try:
        # Get all branches
        branches = get_all_branches()
        
        return render_template('lead_transfer.html', branches=branches)
    except Exception as e:
        print(f"Error in lead_transfer: {str(e)}")
        return render_template('lead_transfer.html', branches=[])

@app.route('/api/ps_followup_summary')
def api_ps_followup_summary():
    """API to get PS Follow-up Summary for branch head dashboard with date filtering and new logic based on call dates only"""
    try:
        # Get branch from session
        branch = session.get('branch_head_branch')
        if not branch:
            return jsonify({'success': False, 'message': 'Branch not found in session'})
        
        # Get date filter parameters from request
        from_date = request.args.get('from_date')
        to_date = request.args.get('to_date')
        mode = request.args.get('mode', 'all_pending')
        
        # Debug logging
        print(f"DEBUG: PS Follow-up Summary API called with - from_date: {from_date}, to_date: {to_date}, mode: {mode}")
        
        # Get all PS users for this branch
        ps_users_result = supabase.table('ps_users').select('name').eq('branch', branch).eq('is_active', True).execute()
        ps_users = [ps['name'] for ps in (ps_users_result.data or [])]
        
        ps_summary = {}
        
        # Initialize summary for each PS with sub-columns
        for ps_name in ps_users:
            ps_summary[ps_name] = {
                'ps_name': ps_name,
                'F1': {'ps_followup': 0, 'walkin': 0, 'total': 0},
                'F2': {'ps_followup': 0, 'walkin': 0, 'total': 0},
                'F3': {'ps_followup': 0, 'walkin': 0, 'total': 0},
                'F4': {'ps_followup': 0, 'walkin': 0, 'total': 0},
                'F5': {'ps_followup': 0, 'walkin': 0, 'total': 0},
                'F6': {'ps_followup': 0, 'walkin': 0, 'total': 0},
                'F7': {'ps_followup': 0, 'walkin': 0, 'total': 0}
            }
        
        # Query ps_followup_master table - only final_status = 'Pending' with date filtering on ps_assigned_at
        try:
            query = supabase.table('ps_followup_master').select(
                'lead_uid, ps_name, final_status, ps_assigned_at, follow_up_date, '
                'first_call_date, second_call_date, third_call_date, fourth_call_date, '
                'fifth_call_date, sixth_call_date, seventh_call_date'
            ).eq('ps_branch', branch).eq('final_status', 'Pending')
            
            # Apply mode-specific filtering
            if mode == 'today_missed':
                # For today's and missed followup: follow_up_date <= today
                from datetime import datetime
                today = datetime.now().strftime('%Y-%m-%d')
                query = query.lte('follow_up_date', f'{today} 23:59:59')
            else:
                # For all pending leads: apply date filtering if dates are provided
                if from_date:
                    query = query.gte('ps_assigned_at', f'{from_date} 00:00:00')
                if to_date:
                    query = query.lte('ps_assigned_at', f'{to_date} 23:59:59')
            
            ps_followup_result = query.execute()
            
            if ps_followup_result.data:
                for lead in ps_followup_result.data:
                    ps_name = lead.get('ps_name')
                    if ps_name not in ps_summary:
                        continue
                    
                    # Helper function to check if a call date is empty/null
                    def is_call_date_empty(call_date):
                        return not call_date or call_date == '' or call_date is None
                    
                    # F1: first_call_date is empty, all other call dates are also empty
                    if (is_call_date_empty(lead.get('first_call_date')) and
                        is_call_date_empty(lead.get('second_call_date')) and
                        is_call_date_empty(lead.get('third_call_date')) and
                        is_call_date_empty(lead.get('fourth_call_date')) and
                        is_call_date_empty(lead.get('fifth_call_date')) and
                        is_call_date_empty(lead.get('sixth_call_date')) and
                        is_call_date_empty(lead.get('seventh_call_date'))):
                        ps_summary[ps_name]['F1']['ps_followup'] += 1
                        ps_summary[ps_name]['F1']['total'] += 1
                    
                    # F2: second_call_date is empty, first_call_date is filled, all subsequent dates are empty
                    elif (not is_call_date_empty(lead.get('first_call_date')) and
                          is_call_date_empty(lead.get('second_call_date')) and
                          is_call_date_empty(lead.get('third_call_date')) and
                          is_call_date_empty(lead.get('fourth_call_date')) and
                          is_call_date_empty(lead.get('fifth_call_date')) and
                          is_call_date_empty(lead.get('sixth_call_date')) and
                          is_call_date_empty(lead.get('seventh_call_date'))):
                        ps_summary[ps_name]['F2']['ps_followup'] += 1
                        ps_summary[ps_name]['F2']['total'] += 1
                    
                    # F3: third_call_date is empty, first and second are filled, all subsequent dates are empty
                    elif (not is_call_date_empty(lead.get('first_call_date')) and
                          not is_call_date_empty(lead.get('second_call_date')) and
                          is_call_date_empty(lead.get('third_call_date')) and
                          is_call_date_empty(lead.get('fourth_call_date')) and
                          is_call_date_empty(lead.get('fifth_call_date')) and
                          is_call_date_empty(lead.get('sixth_call_date')) and
                          is_call_date_empty(lead.get('seventh_call_date'))):
                        ps_summary[ps_name]['F3']['ps_followup'] += 1
                        ps_summary[ps_name]['F3']['total'] += 1
                    
                    # F4: fourth_call_date is empty, first three are filled, all subsequent dates are empty
                    elif (not is_call_date_empty(lead.get('first_call_date')) and
                          not is_call_date_empty(lead.get('second_call_date')) and
                          not is_call_date_empty(lead.get('third_call_date')) and
                          is_call_date_empty(lead.get('fourth_call_date')) and
                          is_call_date_empty(lead.get('fifth_call_date')) and
                          is_call_date_empty(lead.get('sixth_call_date')) and
                          is_call_date_empty(lead.get('seventh_call_date'))):
                        ps_summary[ps_name]['F4']['ps_followup'] += 1
                        ps_summary[ps_name]['F4']['total'] += 1
                    
                    # F5: fifth_call_date is empty, first four are filled, all subsequent dates are empty
                    elif (not is_call_date_empty(lead.get('first_call_date')) and
                          not is_call_date_empty(lead.get('second_call_date')) and
                          not is_call_date_empty(lead.get('third_call_date')) and
                          not is_call_date_empty(lead.get('fourth_call_date')) and
                          is_call_date_empty(lead.get('fifth_call_date')) and
                          is_call_date_empty(lead.get('sixth_call_date')) and
                          is_call_date_empty(lead.get('seventh_call_date'))):
                        ps_summary[ps_name]['F5']['ps_followup'] += 1
                        ps_summary[ps_name]['F5']['total'] += 1
                    
                    # F6: sixth_call_date is empty, first five are filled, seventh date is empty
                    elif (not is_call_date_empty(lead.get('first_call_date')) and
                          not is_call_date_empty(lead.get('second_call_date')) and
                          not is_call_date_empty(lead.get('third_call_date')) and
                          not is_call_date_empty(lead.get('fourth_call_date')) and
                          not is_call_date_empty(lead.get('fifth_call_date')) and
                          is_call_date_empty(lead.get('sixth_call_date')) and
                          is_call_date_empty(lead.get('seventh_call_date'))):
                        ps_summary[ps_name]['F6']['ps_followup'] += 1
                        ps_summary[ps_name]['F6']['total'] += 1
                    
                    # F7: seventh_call_date is empty, first six are filled
                    elif (not is_call_date_empty(lead.get('first_call_date')) and
                          not is_call_date_empty(lead.get('second_call_date')) and
                          not is_call_date_empty(lead.get('third_call_date')) and
                          not is_call_date_empty(lead.get('fourth_call_date')) and
                          not is_call_date_empty(lead.get('fifth_call_date')) and
                          not is_call_date_empty(lead.get('sixth_call_date')) and
                          is_call_date_empty(lead.get('seventh_call_date'))):
                        ps_summary[ps_name]['F7']['ps_followup'] += 1
                        ps_summary[ps_name]['F7']['total'] += 1
                        
        except Exception as e:
            print(f"Warning: Error fetching ps_followup_master data: {str(e)}")
        
        # Query walkin_table - only status = 'Pending' with date filtering on created_at
        try:
            query = supabase.table('walkin_table').select(
                'uid, ps_assigned, status, created_at, next_followup_date, '
                'first_call_date, second_call_date, third_call_date, fourth_call_date, '
                'fifth_call_date, sixth_call_date, seventh_call_date'
            ).eq('branch', branch).eq('status', 'Pending')
            
            # Apply mode-specific filtering
            if mode == 'today_missed':
                # For today's and missed followup: next_followup_date <= today
                from datetime import datetime
                today = datetime.now().strftime('%Y-%m-%d')
                query = query.lte('next_followup_date', f'{today} 23:59:59')
            else:
                # For all pending leads: apply date filtering if dates are provided
                if from_date:
                    query = query.gte('created_at', f'{from_date} 00:00:00')
                if to_date:
                    query = query.lte('created_at', f'{to_date} 23:59:59')
            
            walkin_result = query.execute()
            
            if walkin_result.data:
                for lead in walkin_result.data:
                    ps_name = lead.get('ps_assigned')
                    if ps_name not in ps_summary:
                        continue
                    
                    # Helper function to check if a call date is empty/null
                    def is_call_date_empty(call_date):
                        return not call_date or call_date == '' or call_date is None
                    
                    # F1: first_call_date is empty, all other call dates are also empty
                    if (is_call_date_empty(lead.get('first_call_date')) and
                        is_call_date_empty(lead.get('second_call_date')) and
                        is_call_date_empty(lead.get('third_call_date')) and
                        is_call_date_empty(lead.get('fourth_call_date')) and
                        is_call_date_empty(lead.get('fifth_call_date')) and
                        is_call_date_empty(lead.get('sixth_call_date')) and
                        is_call_date_empty(lead.get('seventh_call_date'))):
                        ps_summary[ps_name]['F1']['walkin'] += 1
                        ps_summary[ps_name]['F1']['total'] += 1
                    
                    # F2: second_call_date is empty, first_call_date is filled, all subsequent dates are empty
                    elif (not is_call_date_empty(lead.get('first_call_date')) and
                          is_call_date_empty(lead.get('second_call_date')) and
                          is_call_date_empty(lead.get('third_call_date')) and
                          is_call_date_empty(lead.get('fourth_call_date')) and
                          is_call_date_empty(lead.get('fifth_call_date')) and
                          is_call_date_empty(lead.get('sixth_call_date')) and
                          is_call_date_empty(lead.get('seventh_call_date'))):
                        ps_summary[ps_name]['F2']['walkin'] += 1
                        ps_summary[ps_name]['F2']['total'] += 1
                    
                    # F3: third_call_date is empty, first and second are filled, all subsequent dates are empty
                    elif (not is_call_date_empty(lead.get('first_call_date')) and
                          not is_call_date_empty(lead.get('second_call_date')) and
                          is_call_date_empty(lead.get('third_call_date')) and
                          is_call_date_empty(lead.get('fourth_call_date')) and
                          is_call_date_empty(lead.get('fifth_call_date')) and
                          is_call_date_empty(lead.get('sixth_call_date')) and
                          is_call_date_empty(lead.get('seventh_call_date'))):
                        ps_summary[ps_name]['F3']['walkin'] += 1
                        ps_summary[ps_name]['F3']['total'] += 1
                    
                    # F4: fourth_call_date is empty, first three are filled, all subsequent dates are empty
                    elif (not is_call_date_empty(lead.get('first_call_date')) and
                          not is_call_date_empty(lead.get('second_call_date')) and
                          not is_call_date_empty(lead.get('third_call_date')) and
                          is_call_date_empty(lead.get('fourth_call_date')) and
                          is_call_date_empty(lead.get('fifth_call_date')) and
                          is_call_date_empty(lead.get('sixth_call_date')) and
                          is_call_date_empty(lead.get('seventh_call_date'))):
                        ps_summary[ps_name]['F4']['walkin'] += 1
                        ps_summary[ps_name]['F4']['total'] += 1
                    
                    # F5: fifth_call_date is empty, first four are filled, all subsequent dates are empty
                    elif (not is_call_date_empty(lead.get('first_call_date')) and
                          not is_call_date_empty(lead.get('second_call_date')) and
                          not is_call_date_empty(lead.get('third_call_date')) and
                          not is_call_date_empty(lead.get('fourth_call_date')) and
                          is_call_date_empty(lead.get('fifth_call_date')) and
                          is_call_date_empty(lead.get('sixth_call_date')) and
                          is_call_date_empty(lead.get('seventh_call_date'))):
                        ps_summary[ps_name]['F5']['walkin'] += 1
                        ps_summary[ps_name]['F5']['total'] += 1
                    
                    # F6: sixth_call_date is empty, first five are filled, seventh date is empty
                    elif (not is_call_date_empty(lead.get('first_call_date')) and
                          not is_call_date_empty(lead.get('second_call_date')) and
                          not is_call_date_empty(lead.get('third_call_date')) and
                          not is_call_date_empty(lead.get('fourth_call_date')) and
                          not is_call_date_empty(lead.get('fifth_call_date')) and
                          is_call_date_empty(lead.get('sixth_call_date')) and
                          is_call_date_empty(lead.get('seventh_call_date'))):
                        ps_summary[ps_name]['F6']['walkin'] += 1
                        ps_summary[ps_name]['F6']['total'] += 1
                    
                    # F7: seventh_call_date is empty, first six are filled
                    elif (not is_call_date_empty(lead.get('first_call_date')) and
                          not is_call_date_empty(lead.get('second_call_date')) and
                          not is_call_date_empty(lead.get('third_call_date')) and
                          not is_call_date_empty(lead.get('fourth_call_date')) and
                          not is_call_date_empty(lead.get('fifth_call_date')) and
                          not is_call_date_empty(lead.get('sixth_call_date')) and
                          is_call_date_empty(lead.get('seventh_call_date'))):
                        ps_summary[ps_name]['F7']['walkin'] += 1
                        ps_summary[ps_name]['F7']['total'] += 1
                        
        except Exception as e:
            print(f"Warning: Error fetching walkin_table data: {str(e)}")
        
        # Format the data for display with sub-columns
        formatted_summary = []
        for ps_name, summary in ps_summary.items():
            formatted_summary.append({
                'ps_name': ps_name,
                'F1': summary['F1'],
                'F2': summary['F2'],
                'F3': summary['F3'],
                'F4': summary['F4'],
                'F5': summary['F5'],
                'F6': summary['F6'],
                'F7': summary['F7']
            })
        
        return jsonify({
            'success': True,
            'data': formatted_summary
        })
        
    except Exception as e:
        print(f"Error in api_ps_followup_summary: {str(e)}")
        return jsonify({'success': False, 'message': f'Error fetching PS follow-up summary: {str(e)}'})

@app.route('/api/ps_followup_leads')
def api_ps_followup_leads():
    """API to get lead details for PS Follow-up Summary table clicks"""
    try:
        # Get parameters
        ps_name = request.args.get('ps_name')
        level = request.args.get('level')  # F1, F2, F3, etc.
        source = request.args.get('source')  # 'cre' or 'walkin'
        mode = request.args.get('mode', 'all_pending')  # Default to 'all_pending'
        branch = session.get('branch_head_branch')
        
        if not all([ps_name, level, source, branch]):
            return jsonify({'success': False, 'message': 'Missing required parameters'})
        
        # Get date filter parameters from request to match the summary filtering
        from_date = request.args.get('from_date')
        to_date = request.args.get('to_date')
        
        leads = []
        
        if source == 'cre':
            # Fetch from ps_followup_master using the EXACT same logic as summary
            try:
                query = supabase.table('ps_followup_master').select(
                    'lead_uid, ps_name, follow_up_date, final_status, source, lead_status, customer_name, customer_mobile_number, ps_assigned_at, '
                    'first_call_date, second_call_date, third_call_date, fourth_call_date, '
                    'fifth_call_date, sixth_call_date, seventh_call_date'
                ).eq('ps_branch', branch).eq('ps_name', ps_name).eq('final_status', 'Pending')
                
                # Apply mode-specific filtering (same as summary)
                if mode == 'today_missed':
                    # For today's and missed followup: follow_up_date <= today
                    today = datetime.now().strftime('%Y-%m-%d')
                    query = query.lte('follow_up_date', f'{today} 23:59:59')
                else:
                    # For all pending leads: apply date filtering if dates are provided
                    if from_date:
                        query = query.gte('ps_assigned_at', f'{from_date} 00:00:00')
                    if to_date:
                        query = query.lte('ps_assigned_at', f'{to_date} 23:59:59')
                
                ps_followup_result = query.execute()
                
                if ps_followup_result.data:
                    for lead in ps_followup_result.data:
                        # Use the EXACT same logic as summary counting
                        level_num = int(level.replace('F', ''))
                        
                        # Helper function to check if a call date is empty/null (same as summary)
                        def is_call_date_empty(call_date):
                            return not call_date or call_date == '' or call_date is None
                        
                        # F1: first_call_date is empty, all other call dates are also empty
                        if level_num == 1:
                            if (is_call_date_empty(lead.get('first_call_date')) and
                                is_call_date_empty(lead.get('second_call_date')) and
                                is_call_date_empty(lead.get('third_call_date')) and
                                is_call_date_empty(lead.get('fourth_call_date')) and
                                is_call_date_empty(lead.get('fifth_call_date')) and
                                is_call_date_empty(lead.get('sixth_call_date')) and
                                is_call_date_empty(lead.get('seventh_call_date'))):
                                leads.append({
                                    'uid': lead.get('lead_uid'),
                                    'customer_name': lead.get('customer_name'),
                                    'mobile_number': lead.get('customer_mobile_number'),
                                    'status': lead.get('final_status'),
                                    'source': lead.get('source'),
                                    'ps_assigned_at': lead.get('ps_assigned_at'),
                                    'follow_up_date': lead.get('follow_up_date'),
                                    'last_remark': ''
                                })
                        
                        # F2: second_call_date is empty, first_call_date is filled, all subsequent dates are empty
                        elif level_num == 2:
                            if (not is_call_date_empty(lead.get('first_call_date')) and
                                is_call_date_empty(lead.get('second_call_date')) and
                                is_call_date_empty(lead.get('third_call_date')) and
                                is_call_date_empty(lead.get('fourth_call_date')) and
                                is_call_date_empty(lead.get('fifth_call_date')) and
                                is_call_date_empty(lead.get('sixth_call_date')) and
                                is_call_date_empty(lead.get('seventh_call_date'))):
                                leads.append({
                                    'uid': lead.get('lead_uid'),
                                    'customer_name': lead.get('customer_name'),
                                    'mobile_number': lead.get('customer_mobile_number'),
                                    'status': lead.get('final_status'),
                                    'source': lead.get('source'),
                                    'ps_assigned_at': lead.get('ps_assigned_at'),
                                    'follow_up_date': lead.get('follow_up_date'),
                                    'last_remark': ''
                                })
                        
                        # F3: third_call_date is empty, first and second are filled, all subsequent dates are empty
                        elif level_num == 3:
                            if (not is_call_date_empty(lead.get('first_call_date')) and
                                not is_call_date_empty(lead.get('second_call_date')) and
                                is_call_date_empty(lead.get('third_call_date')) and
                                is_call_date_empty(lead.get('fourth_call_date')) and
                                is_call_date_empty(lead.get('fifth_call_date')) and
                                is_call_date_empty(lead.get('sixth_call_date')) and
                                is_call_date_empty(lead.get('seventh_call_date'))):
                                leads.append({
                                    'uid': lead.get('lead_uid'),
                                    'customer_name': lead.get('customer_name'),
                                    'mobile_number': lead.get('customer_mobile_number'),
                                    'status': lead.get('final_status'),
                                    'source': lead.get('source'),
                                    'ps_assigned_at': lead.get('ps_assigned_at'),
                                    'follow_up_date': lead.get('follow_up_date'),
                                    'last_remark': ''
                                })
                        
                        # F4: fourth_call_date is empty, first three are filled, all subsequent dates are empty
                        elif level_num == 4:
                            if (not is_call_date_empty(lead.get('first_call_date')) and
                                not is_call_date_empty(lead.get('second_call_date')) and
                                not is_call_date_empty(lead.get('third_call_date')) and
                                is_call_date_empty(lead.get('fourth_call_date')) and
                                is_call_date_empty(lead.get('fifth_call_date')) and
                                is_call_date_empty(lead.get('sixth_call_date')) and
                                is_call_date_empty(lead.get('seventh_call_date'))):
                                leads.append({
                                    'uid': lead.get('lead_uid'),
                                    'customer_name': lead.get('customer_name'),
                                    'mobile_number': lead.get('customer_mobile_number'),
                                    'status': lead.get('final_status'),
                                    'source': lead.get('source'),
                                    'ps_assigned_at': lead.get('ps_assigned_at'),
                                    'follow_up_date': lead.get('follow_up_date'),
                                    'last_remark': ''
                                })
                        
                        # F5: fifth_call_date is empty, first four are filled, all subsequent dates are empty
                        elif level_num == 5:
                            if (not is_call_date_empty(lead.get('first_call_date')) and
                                not is_call_date_empty(lead.get('second_call_date')) and
                                not is_call_date_empty(lead.get('third_call_date')) and
                                not is_call_date_empty(lead.get('fourth_call_date')) and
                                is_call_date_empty(lead.get('fifth_call_date')) and
                                is_call_date_empty(lead.get('sixth_call_date')) and
                                is_call_date_empty(lead.get('seventh_call_date'))):
                                leads.append({
                                    'uid': lead.get('lead_uid'),
                                    'customer_name': lead.get('customer_name'),
                                    'mobile_number': lead.get('customer_mobile_number'),
                                    'status': lead.get('final_status'),
                                    'source': lead.get('source'),
                                    'ps_assigned_at': lead.get('ps_assigned_at'),
                                    'follow_up_date': lead.get('follow_up_date'),
                                    'last_remark': ''
                                })
                        
                        # F6: sixth_call_date is empty, first five are filled, seventh date is empty
                        elif level_num == 6:
                            if (not is_call_date_empty(lead.get('first_call_date')) and
                                not is_call_date_empty(lead.get('second_call_date')) and
                                not is_call_date_empty(lead.get('third_call_date')) and
                                not is_call_date_empty(lead.get('fourth_call_date')) and
                                not is_call_date_empty(lead.get('fifth_call_date')) and
                                is_call_date_empty(lead.get('sixth_call_date')) and
                                is_call_date_empty(lead.get('seventh_call_date'))):
                                leads.append({
                                    'uid': lead.get('lead_uid'),
                                    'customer_name': lead.get('customer_name'),
                                    'mobile_number': lead.get('customer_mobile_number'),
                                    'status': lead.get('final_status'),
                                    'source': lead.get('source'),
                                    'ps_assigned_at': lead.get('ps_assigned_at'),
                                    'follow_up_date': lead.get('follow_up_date'),
                                    'last_remark': ''
                                })
                        
                        # F7: seventh_call_date is empty, first six are filled
                        elif level_num == 7:
                            if (not is_call_date_empty(lead.get('first_call_date')) and
                                not is_call_date_empty(lead.get('second_call_date')) and
                                not is_call_date_empty(lead.get('third_call_date')) and
                                not is_call_date_empty(lead.get('fourth_call_date')) and
                                not is_call_date_empty(lead.get('fifth_call_date')) and
                                not is_call_date_empty(lead.get('sixth_call_date')) and
                                is_call_date_empty(lead.get('seventh_call_date'))):
                                leads.append({
                                    'uid': lead.get('lead_uid'),
                                    'customer_name': lead.get('customer_name'),
                                    'mobile_number': lead.get('customer_mobile_number'),
                                    'status': lead.get('final_status'),
                                    'source': lead.get('source'),
                                    'ps_assigned_at': lead.get('ps_assigned_at'),
                                    'follow_up_date': lead.get('follow_up_date'),
                                    'last_remark': ''
                            })
            except Exception as e:
                print(f"Warning: Error fetching ps_followup_master data: {str(e)}")
        
        elif source == 'walkin':
            # Fetch from walkin_table using the EXACT same logic as summary
            try:
                query = supabase.table('walkin_table').select(
                    'uid, ps_assigned, status, branch, next_followup_date, customer_name, mobile_number, created_at, '
                    'first_call_date, second_call_date, third_call_date, fourth_call_date, '
                    'fifth_call_date, sixth_call_date, seventh_call_date'
                ).eq('branch', branch).eq('ps_assigned', ps_name).eq('status', 'Pending')
                
                # Apply mode-specific filtering (same as summary)
                if mode == 'today_missed':
                    # For today's and missed followup: next_followup_date <= today
                    today = datetime.now().strftime('%Y-%m-%d')
                    query = query.lte('next_followup_date', f'{today} 23:59:59')
                else:
                    # For all pending leads: apply date filtering if dates are provided
                    if from_date:
                        query = query.gte('created_at', f'{from_date} 00:00:00')
                    if to_date:
                        query = query.lte('created_at', f'{to_date} 23:59:59')
                
                walkin_result = query.execute()
                
                if walkin_result.data:
                    for lead in walkin_result.data:
                        # Use the EXACT same logic as summary counting
                        level_num = int(level.replace('F', ''))
                        
                        # Helper function to check if a call date is empty/null (same as summary)
                        def is_call_date_empty(call_date):
                            return not call_date or call_date == '' or call_date is None
                        
                        # F1: first_call_date is empty, all other call dates are also empty
                        if level_num == 1:
                            if (is_call_date_empty(lead.get('first_call_date')) and
                                is_call_date_empty(lead.get('second_call_date')) and
                                is_call_date_empty(lead.get('third_call_date')) and
                                is_call_date_empty(lead.get('fourth_call_date')) and
                                is_call_date_empty(lead.get('fifth_call_date')) and
                                is_call_date_empty(lead.get('sixth_call_date')) and
                                is_call_date_empty(lead.get('seventh_call_date'))):
                                leads.append({
                                    'uid': lead.get('uid'),
                                    'customer_name': lead.get('customer_name'),
                                    'mobile_number': lead.get('mobile_number'),
                                    'status': lead.get('status'),
                                    'source': 'Walk In',
                                    'created_at': lead.get('created_at'),
                                    'follow_up_date': lead.get('next_followup_date'),
                                    'last_remark': ''
                                })
                        
                        # F2: second_call_date is empty, first_call_date is filled, all subsequent dates are empty
                        elif level_num == 2:
                            if (not is_call_date_empty(lead.get('first_call_date')) and
                                is_call_date_empty(lead.get('second_call_date')) and
                                is_call_date_empty(lead.get('third_call_date')) and
                                is_call_date_empty(lead.get('fourth_call_date')) and
                                is_call_date_empty(lead.get('fifth_call_date')) and
                                is_call_date_empty(lead.get('sixth_call_date')) and
                                is_call_date_empty(lead.get('seventh_call_date'))):
                                leads.append({
                                    'uid': lead.get('uid'),
                                    'customer_name': lead.get('customer_name'),
                                    'mobile_number': lead.get('mobile_number'),
                                    'status': lead.get('status'),
                                    'source': 'Walk In',
                                    'created_at': lead.get('created_at'),
                                    'follow_up_date': lead.get('next_followup_date'),
                                    'last_remark': ''
                                })
                        
                        # F3: third_call_date is empty, first and second are filled, all subsequent dates are empty
                        elif level_num == 3:
                            if (not is_call_date_empty(lead.get('first_call_date')) and
                                not is_call_date_empty(lead.get('second_call_date')) and
                                is_call_date_empty(lead.get('third_call_date')) and
                                is_call_date_empty(lead.get('fourth_call_date')) and
                                is_call_date_empty(lead.get('fifth_call_date')) and
                                is_call_date_empty(lead.get('sixth_call_date')) and
                                is_call_date_empty(lead.get('seventh_call_date'))):
                                leads.append({
                                    'uid': lead.get('uid'),
                                    'customer_name': lead.get('customer_name'),
                                    'mobile_number': lead.get('mobile_number'),
                                    'status': lead.get('status'),
                                    'source': 'Walk In',
                                    'created_at': lead.get('created_at'),
                                    'follow_up_date': lead.get('next_followup_date'),
                                    'last_remark': ''
                                })
                        
                        # F4: fourth_call_date is empty, first three are filled, all subsequent dates are empty
                        elif level_num == 4:
                            if (not is_call_date_empty(lead.get('first_call_date')) and
                                not is_call_date_empty(lead.get('second_call_date')) and
                                not is_call_date_empty(lead.get('third_call_date')) and
                                is_call_date_empty(lead.get('fourth_call_date')) and
                                is_call_date_empty(lead.get('fifth_call_date')) and
                                is_call_date_empty(lead.get('sixth_call_date')) and
                                is_call_date_empty(lead.get('seventh_call_date'))):
                                leads.append({
                                    'uid': lead.get('uid'),
                                    'customer_name': lead.get('customer_name'),
                                    'mobile_number': lead.get('mobile_number'),
                                    'status': lead.get('status'),
                                    'source': 'Walk In',
                                    'created_at': lead.get('created_at'),
                                    'follow_up_date': lead.get('next_followup_date'),
                                    'last_remark': ''
                                })
                        
                        # F5: fifth_call_date is empty, first four are filled, all subsequent dates are empty
                        elif level_num == 5:
                            if (not is_call_date_empty(lead.get('first_call_date')) and
                                not is_call_date_empty(lead.get('second_call_date')) and
                                not is_call_date_empty(lead.get('third_call_date')) and
                                not is_call_date_empty(lead.get('fourth_call_date')) and
                                is_call_date_empty(lead.get('fifth_call_date')) and
                                is_call_date_empty(lead.get('sixth_call_date')) and
                                is_call_date_empty(lead.get('seventh_call_date'))):
                                leads.append({
                                    'uid': lead.get('uid'),
                                    'customer_name': lead.get('customer_name'),
                                    'mobile_number': lead.get('mobile_number'),
                                    'status': lead.get('status'),
                                    'source': 'Walk In',
                                    'created_at': lead.get('created_at'),
                                    'follow_up_date': lead.get('next_followup_date'),
                                    'last_remark': ''
                                })
                        
                        # F6: sixth_call_date is empty, first five are filled, seventh date is empty
                        elif level_num == 6:
                            if (not is_call_date_empty(lead.get('first_call_date')) and
                                not is_call_date_empty(lead.get('second_call_date')) and
                                not is_call_date_empty(lead.get('third_call_date')) and
                                not is_call_date_empty(lead.get('fourth_call_date')) and
                                not is_call_date_empty(lead.get('fifth_call_date')) and
                                is_call_date_empty(lead.get('sixth_call_date')) and
                                is_call_date_empty(lead.get('seventh_call_date'))):
                                leads.append({
                                    'uid': lead.get('uid'),
                                    'customer_name': lead.get('customer_name'),
                                    'mobile_number': lead.get('mobile_number'),
                                    'status': lead.get('status'),
                                    'source': 'Walk In',
                                    'created_at': lead.get('created_at'),
                                    'follow_up_date': lead.get('next_followup_date'),
                                    'last_remark': ''
                                })
                        
                        # F7: seventh_call_date is empty, first six are filled
                        elif level_num == 7:
                            if (not is_call_date_empty(lead.get('first_call_date')) and
                                not is_call_date_empty(lead.get('second_call_date')) and
                                not is_call_date_empty(lead.get('third_call_date')) and
                                not is_call_date_empty(lead.get('fourth_call_date')) and
                                not is_call_date_empty(lead.get('fifth_call_date')) and
                                not is_call_date_empty(lead.get('sixth_call_date')) and
                                is_call_date_empty(lead.get('seventh_call_date'))):
                                leads.append({
                                    'uid': lead.get('uid'),
                                    'customer_name': lead.get('customer_name'),
                                    'mobile_number': lead.get('mobile_number'),
                                    'status': lead.get('status'),
                                    'source': 'Walk In',
                                    'created_at': lead.get('created_at'),
                                    'follow_up_date': lead.get('next_followup_date'),
                                    'last_remark': ''
                            })
            except Exception as e:
                print(f"Warning: Error fetching walkin_table data: {str(e)}")
        
        return jsonify({
            'success': True,
            'leads': leads
        })
        
    except Exception as e:
        print(f"Error in api_ps_followup_leads: {str(e)}")
        return jsonify({'success': False, 'message': f'Error fetching lead details: {str(e)}'})

@app.route('/api/cre_pending_leads')
@require_admin
def api_cre_pending_leads():
    """API to get pending leads summary for CRE transfer from lead_master only"""
    try:
        # Get all pending leads from lead_master grouped by CRE only (only leads with completed qualifying calls)
        lead_result = supabase.table('lead_master').select('cre_name').eq('final_status', 'Pending').not_.is_('first_call_date', 'null').execute()
        lead_pending_leads = lead_result.data if lead_result.data else []
        
        # Group leads by CRE and count them from lead_master only
        cre_summary = {}
        
        # Count from lead_master only
        for lead in lead_pending_leads:
            cre_name = lead.get('cre_name', 'Unassigned')
            if cre_name not in cre_summary:
                cre_summary[cre_name] = 0
            cre_summary[cre_name] += 1
        
        # Format the data for display (exclude Unassigned leads)
        formatted_summary = []
        for cre_name, count in cre_summary.items():
            # Skip unassigned leads
            if cre_name and cre_name != 'Unassigned' and cre_name != 'null':
                formatted_summary.append({
                    'cre_name': cre_name,
                    'pending_count': count
                })
        
        return jsonify({'success': True, 'data': formatted_summary})
    except Exception as e:
        print(f"Error in api_cre_pending_leads: {str(e)}")
        return jsonify({'success': False, 'message': 'Error fetching pending leads'})

@app.route('/api/ps_pending_leads')
@require_admin
def api_ps_pending_leads():
    """API to get pending leads summary for PS transfer from multiple tables"""
    try:
        # Get branch filter from request
        branch_filter = request.args.get('branch', '')
        
        # Build query for ps_followup_master
        ps_followup_query = supabase.table('ps_followup_master').select('ps_name, ps_branch').eq('final_status', 'Pending')
        if branch_filter:
            ps_followup_query = ps_followup_query.eq('ps_branch', branch_filter)
        
        ps_followup_result = ps_followup_query.execute()
        ps_followup_leads = ps_followup_result.data if ps_followup_result.data else []
        
        # Get pending leads from walkin_table where status is 'Pending' (capital P to match DB default)
        walkin_leads = []
        try:
            walkin_query = supabase.table('walkin_table').select('ps_assigned, branch').eq('status', 'Pending')
            if branch_filter:
                walkin_query = walkin_query.eq('branch', branch_filter)
            walkin_result = walkin_query.execute()
            walkin_leads = walkin_result.data if walkin_result.data else []
        except Exception as e:
            print(f"Warning: walkin_table not found or error: {str(e)}")
            walkin_leads = []
        
        # Get pending leads from activity_leads where final_status is 'Pending' (if table exists)
        activity_leads = []
        try:
            activity_query = supabase.table('activity_leads').select('ps_name, location').eq('final_status', 'Pending')
            if branch_filter:
                activity_query = activity_query.eq('location', branch_filter)
            activity_result = activity_query.execute()
            activity_leads = activity_result.data if activity_result.data else []
        except Exception as e:
            print(f"Warning: activity_leads table not found or error: {str(e)}")
            activity_leads = []
        
        # Group leads by PS and count them from all tables
        ps_summary = {}
        
        # Count from ps_followup_master
        for lead in ps_followup_leads:
            ps_name = lead.get('ps_name', 'Unassigned')
            ps_branch = lead.get('ps_branch', '')
            key = f"{ps_name}|{ps_branch}"
            if key not in ps_summary:
                ps_summary[key] = {
                    'ps_name': ps_name,
                    'ps_branch': ps_branch,
                    'ps_followup_count': 0,
                    'walkin_count': 0,
                    'activity_count': 0,
                    'total_count': 0
                }
            ps_summary[key]['ps_followup_count'] += 1
            ps_summary[key]['total_count'] += 1
        
        # Count from walkin_table
        for lead in walkin_leads:
            ps_name = lead.get('ps_assigned', 'Unassigned')  # Changed from ps_name to ps_assigned
            ps_branch = lead.get('branch', '')
            key = f"{ps_name}|{ps_branch}"
            if key not in ps_summary:
                ps_summary[key] = {
                    'ps_name': ps_name,
                    'ps_branch': ps_branch,
                    'ps_followup_count': 0,
                    'walkin_count': 0,
                    'activity_count': 0,
                    'total_count': 0
                }
            ps_summary[key]['walkin_count'] += 1
            ps_summary[key]['total_count'] += 1
        
        # Count from activity_leads
        for lead in activity_leads:
            ps_name = lead.get('ps_name', 'Unassigned')
            ps_branch = lead.get('location', '')  # location is branch in activity_leads
            key = f"{ps_name}|{ps_branch}"
            if key not in ps_summary:
                ps_summary[key] = {
                    'ps_name': ps_name,
                    'ps_branch': ps_branch,
                    'ps_followup_count': 0,
                    'walkin_count': 0,
                    'activity_count': 0,
                    'total_count': 0
                }
            ps_summary[key]['activity_count'] += 1
            ps_summary[key]['total_count'] += 1
        
        # Format the data for display (exclude Unassigned leads)
        formatted_summary = []
        for key, summary in ps_summary.items():
            # Skip unassigned leads
            if summary['ps_name'] and summary['ps_name'] != 'Unassigned' and summary['ps_name'] != 'null':
                formatted_summary.append(summary)
        
        return jsonify({'success': True, 'data': formatted_summary})
    except Exception as e:
        print(f"Error in api_ps_pending_leads: {str(e)}")
        return jsonify({'success': False, 'message': 'Error fetching pending leads'})

@app.route('/api/transfer_cre_lead', methods=['POST'])
@require_admin
def api_transfer_cre_lead():
    """API to transfer CRE lead to another CRE"""
    try:
        data = request.get_json()
        lead_uid = data.get('lead_uid')
        new_cre_name = data.get('new_cre_name')
        
        if not lead_uid or not new_cre_name:
            return jsonify({'success': False, 'message': 'Missing required parameters'})
        
        # Verify the CRE exists and is active, and get the name
        cre_result = supabase.table('cre_users').select('name, username').eq('username', new_cre_name).eq('is_active', True).execute()
        if not cre_result.data:
            return jsonify({'success': False, 'message': 'CRE not found or inactive'})
        
        # Get the target CRE's name (not username)
        target_cre_name = cre_result.data[0].get('name')
        
        # Update the lead in lead_master
        lead_result = supabase.table('lead_master').update({'cre_name': target_cre_name}).eq('uid', lead_uid).execute()
        print(f"[DEBUG] CRE transfer: lead_master update result: {lead_result.data}")
        
        # Update the lead in ps_followup_master if it exists there
        ps_followup_result = supabase.table('ps_followup_master').update({'cre_name': target_cre_name}).eq('lead_uid', lead_uid).execute()
        print(f"[DEBUG] CRE transfer: ps_followup_master update result: {ps_followup_result.data}")
        
        if lead_result.data or ps_followup_result.data:
            return jsonify({
                'success': True, 
                'message': 'Lead transferred successfully',
                'details': {
                    'lead_master_updated': len(lead_result.data) if lead_result.data else 0,
                    'ps_followup_updated': len(ps_followup_result.data) if ps_followup_result.data else 0
                }
            })
        else:
            return jsonify({'success': False, 'message': 'Lead not found or transfer failed'})
            
    except Exception as e:
        print(f"Error in api_transfer_cre_lead: {str(e)}")
        return jsonify({'success': False, 'message': 'Error transferring lead'})

@app.route('/api/bulk_transfer_cre_leads', methods=['POST'])
@require_admin
def api_bulk_transfer_cre_leads():
    """API to bulk transfer specified number of pending leads from one CRE to another"""
    try:
        data = request.get_json()
        from_cre_name = data.get('from_cre_name')
        to_cre_name = data.get('to_cre_name')
        count = data.get('count', 0)
        
        if not from_cre_name or not to_cre_name:
            return jsonify({'success': False, 'message': 'Missing required parameters'})
        
        if from_cre_name == to_cre_name:
            return jsonify({'success': False, 'message': 'Cannot transfer to the same CRE'})
        
        if not count or count < 1:
            return jsonify({'success': False, 'message': 'Invalid transfer count'})
        
        # Verify the target CRE exists and is active, and get the name
        cre_result = supabase.table('cre_users').select('name, username').eq('username', to_cre_name).eq('is_active', True).execute()
        if not cre_result.data:
            return jsonify({'success': False, 'message': 'Target CRE not found or inactive'})
        
        # Get the target CRE's name (not username)
        target_cre_name = cre_result.data[0].get('name')
        
        # First, get the total available pending leads from the source CRE
        total_pending_result = supabase.table('lead_master').select('id, uid').eq('cre_name', from_cre_name).eq('final_status', 'Pending').execute()
        total_available = len(total_pending_result.data) if total_pending_result.data else 0
        
        if total_available < count:
            return jsonify({'success': False, 'message': f'Only {total_available} leads available, requested {count}'})
        
        # Get the specific leads to transfer (limit by count)
        leads_to_transfer = total_pending_result.data[:count]
        
        # Extract both id and uid for proper updates
        lead_ids = [lead['id'] for lead in leads_to_transfer]
        lead_uids = [lead['uid'] for lead in leads_to_transfer]
        
        # Update lead_master table
        lead_result = supabase.table('lead_master').update({'cre_name': target_cre_name}).in_('id', lead_ids).execute()
        print(f"[DEBUG] CRE bulk transfer: lead_master update result: {len(lead_result.data) if lead_result.data else 0} rows updated")
        
        # Update ps_followup_master table using lead_uid (correct foreign key)
        ps_followup_result = supabase.table('ps_followup_master').update({'cre_name': target_cre_name}).in_('lead_uid', lead_uids).execute()
        print(f"[DEBUG] CRE bulk transfer: ps_followup_master update result: {len(ps_followup_result.data) if ps_followup_result.data else 0} rows updated")
        
        # Count transferred leads
        transferred_count = len(lead_result.data) if lead_result.data else 0
        
        if transferred_count > 0:
            return jsonify({
                'success': True, 
                'message': f'Pending Lead Transfer Successful - {transferred_count} leads transferred from {from_cre_name} to {to_cre_name}',
                'details': {
                    'lead_master_updated': len(lead_result.data) if lead_result.data else 0,
                    'ps_followup_updated': len(ps_followup_result.data) if ps_followup_result.data else 0
                }
            })
        else:
            return jsonify({'success': False, 'message': 'No leads were transferred'})
            
    except Exception as e:
        print(f"Error in api_bulk_transfer_cre_leads: {str(e)}")
        return jsonify({'success': False, 'message': 'Error transferring leads'})

@app.route('/api/transfer_options')
@require_admin
def api_transfer_options():
    """API to get transfer options (CRE and PS lists)"""
    try:
        # Get all CREs for transfer options - CRE doesn't have branch
        cre_result = supabase.table('cre_users').select('id, name, username, email, phone, is_active').eq('is_active', True).execute()
        cre_options = cre_result.data if cre_result.data else []
        
        # Get all PS for transfer options - PS has branch
        ps_result = supabase.table('ps_users').select('id, name, username, email, phone, branch, is_active').eq('is_active', True).execute()
        ps_options = ps_result.data if ps_result.data else []
        
        # Sanitize the data to ensure no problematic characters
        for cre in cre_options:
            if isinstance(cre.get('name'), str):
                cre['name'] = cre['name'].strip()
            if isinstance(cre.get('username'), str):
                cre['username'] = cre['username'].strip()
                
        for ps in ps_options:
            if isinstance(ps.get('name'), str):
                ps['name'] = ps['name'].strip()
            if isinstance(ps.get('username'), str):
                ps['username'] = ps['username'].strip()
            if isinstance(ps.get('branch'), str):
                ps['branch'] = ps['branch'].strip()
        
        return jsonify({
            'success': True,
            'cre_options': cre_options,
            'ps_options': ps_options
        })
    except Exception as e:
        print(f"Error in api_transfer_options: {str(e)}")
        return jsonify({'success': False, 'message': 'Error fetching transfer options'})
@app.route('/api/branches')
@require_admin
def api_branches():
    """API to get all available branches"""
    try:
        # Get unique branches from ps_users table
        ps_result = supabase.table('ps_users').select('branch').eq('is_active', True).execute()
        branches = []
        if ps_result.data:
            branches = list(set([ps.get('branch', '').strip() for ps in ps_result.data if ps.get('branch', '').strip()]))
            branches.sort()  # Sort alphabetically
        
        return jsonify({
            'success': True,
            'branches': branches
        })
    except Exception as e:
        print(f"Error in api_branches: {str(e)}")
        return jsonify({'success': False, 'message': 'Error fetching branches'})

@app.route('/api/transfer_ps_lead', methods=['POST'])
@require_admin
def api_transfer_ps_lead():
    """API to transfer PS lead to another PS"""
    try:
        data = request.get_json()
        lead_uid = data.get('lead_uid')
        new_ps_name = data.get('new_ps_name')
        
        if not lead_uid or not new_ps_name:
            return jsonify({'success': False, 'message': 'Missing required parameters'})
        
        # Get the new PS details with branch information
        ps_result = supabase.table('ps_users').select('name, username, branch').eq('name', new_ps_name).eq('is_active', True).execute()
        if not ps_result.data:
            return jsonify({'success': False, 'message': 'PS not found or inactive'})
        
        new_ps = ps_result.data[0]
        
        # Update the lead in ps_followup_master
        result = supabase.table('ps_followup_master').update({
            'ps_name': new_ps_name,
            'ps_branch': new_ps.get('branch', '')
        }).eq('lead_uid', lead_uid).execute()
        print(f"[DEBUG] PS transfer: ps_followup_master update result: {result.data}")
        
        # Also update ps_name in lead_master table to maintain consistency
        if result.data:
            lead_master_update = supabase.table('lead_master').update({
                'ps_name': new_ps_name
            }).eq('uid', lead_uid).execute()
            print(f"[DEBUG] PS transfer: lead_master update result: {lead_master_update.data}")
            
            return jsonify({
                'success': True, 
                'message': 'Lead transferred successfully',
                'details': {
                    'ps_followup_updated': len(result.data) if result.data else 0,
                    'lead_master_updated': len(lead_master_update.data) if lead_master_update.data else 0
                }
            })
        else:
            return jsonify({'success': False, 'message': 'Lead not found or transfer failed'})
            
    except Exception as e:
        print(f"Error in api_transfer_ps_lead: {str(e)}")
        return jsonify({'success': False, 'message': 'Error transferring lead'})

@app.route('/api/bulk_transfer_ps_leads', methods=['POST'])
@require_admin
def api_bulk_transfer_ps_leads():
    """API to bulk transfer specified numbers of pending leads from one PS to another PS in the same branch"""
    try:
        data = request.get_json()
        from_ps_name = data.get('from_ps_name')
        to_ps_name = data.get('to_ps_name')
        ps_followup_count = data.get('ps_followup_count', 0)
        walkin_count = data.get('walkin_count', 0)
        event_count = data.get('event_count', 0)
        
        if not from_ps_name or not to_ps_name:
            return jsonify({'success': False, 'message': 'Missing required parameters'})
        
        if from_ps_name == to_ps_name:
            return jsonify({'success': False, 'message': 'Cannot transfer to the same PS'})
        
        # Validate that at least one count is specified
        total_requested = ps_followup_count + walkin_count + event_count
        if total_requested < 1:
            return jsonify({'success': False, 'message': 'Please specify at least one count to transfer'})
        
        # Get the target PS details with branch information
        ps_result = supabase.table('ps_users').select('name, username, branch').eq('name', to_ps_name).eq('is_active', True).execute()
        if not ps_result.data:
            return jsonify({'success': False, 'message': 'Target PS not found or inactive'})
        
        target_ps = ps_result.data[0]
        target_branch = target_ps.get('branch', '')
        target_username = target_ps.get('username', '')
        
        # Get the source PS branch to ensure same branch transfer
        source_ps_result = supabase.table('ps_users').select('name, username, branch').eq('name', from_ps_name).eq('is_active', True).execute()
        if not source_ps_result.data:
            return jsonify({'success': False, 'message': 'Source PS not found or inactive'})
        
        source_branch = source_ps_result.data[0].get('branch', '')
        source_username = source_ps_result.data[0].get('username', '')
        
        # Verify both PS are in the same branch
        if source_branch != target_branch:
            return jsonify({'success': False, 'message': 'Can only transfer between PS in the same branch'})
        
        transferred_counts = {}
        total_transferred = 0
        
        # Transfer PS followup leads if count specified
        if ps_followup_count > 0:
            try:
                # Get available PS followup leads with lead_uid for lead_master update
                available_ps_leads = supabase.table('ps_followup_master').select('id, lead_uid').eq('ps_name', from_ps_name).eq('final_status', 'Pending').limit(ps_followup_count).execute()
                
                if available_ps_leads.data:
                    lead_ids = [lead['id'] for lead in available_ps_leads.data]
                    lead_uids = [lead['lead_uid'] for lead in available_ps_leads.data]
                    actual_count = len(lead_ids)
                    
                    # Update the leads in ps_followup_master
                    ps_update_result = supabase.table('ps_followup_master').update({
                        'ps_name': to_ps_name,
                        'ps_branch': target_branch
                    }).in_('id', lead_ids).execute()
                    print(f"[DEBUG] PS bulk transfer: ps_followup_master update result: {len(ps_update_result.data) if ps_update_result.data else 0} rows updated")
                    
                    transferred_counts['ps_followup'] = actual_count
                    total_transferred += actual_count
                    
                    # Also update lead_master for consistency using lead_uid
                    if lead_uids:
                        lead_update_result = supabase.table('lead_master').update({
                            'ps_name': to_ps_name
                        }).in_('uid', lead_uids).execute()
                        print(f"[DEBUG] PS bulk transfer: lead_master update result: {len(lead_update_result.data) if lead_update_result.data else 0} rows updated")
                    
            except Exception as e:
                print(f"Error transferring PS followup leads: {str(e)}")
                transferred_counts['ps_followup'] = 0
        
        # Transfer walkin leads if count specified
        if walkin_count > 0:
            try:
                # Get available walkin leads
                available_walkin_leads = supabase.table('walkin_table').select('id').eq('ps_assigned', from_ps_name).eq('status', 'Pending').limit(walkin_count).execute()
                
                if available_walkin_leads.data:
                    lead_ids = [lead['id'] for lead in available_walkin_leads.data]
                    actual_count = len(lead_ids)
                    
                    # Update the leads
                    supabase.table('walkin_table').update({
                        'ps_assigned': to_ps_name,
                        'branch': target_branch
                    }).in_('id', lead_ids).execute()
                    
                    transferred_counts['walkin'] = actual_count
                    total_transferred += actual_count
                    
            except Exception as e:
                print(f"Warning: walkin_table not found or error: {str(e)}")
                transferred_counts['walkin'] = 0
        
        # Transfer event leads if count specified
        if event_count > 0:
            try:
                # Get available event leads
                available_event_leads = supabase.table('activity_leads').select('id').eq('ps_name', from_ps_name).eq('final_status', 'Pending').limit(event_count).execute()
                
                if available_event_leads.data:
                    lead_ids = [lead['id'] for lead in available_event_leads.data]
                    actual_count = len(lead_ids)
                    
                    # Update the leads
                    supabase.table('activity_leads').update({
                        'ps_name': to_ps_name,
                        'location': target_branch
                    }).in_('id', lead_ids).execute()
                    
                    transferred_counts['event_leads'] = actual_count
                    total_transferred += actual_count
                    
            except Exception as e:
                print(f"Warning: activity_leads table not found or error: {str(e)}")
                transferred_counts['event_leads'] = 0
        
        if total_transferred > 0:
            # Build detailed message
            message_parts = []
            if transferred_counts.get('ps_followup', 0) > 0:
                message_parts.append(f"{transferred_counts['ps_followup']} from PS followup")
            if transferred_counts.get('walkin', 0) > 0:
                message_parts.append(f"{transferred_counts['walkin']} from walkin")
            if transferred_counts.get('event_leads', 0) > 0:
                message_parts.append(f"{transferred_counts['event_leads']} from event")
            
            details = ', '.join(message_parts)
            
            return jsonify({
                'success': True, 
                'message': f'Pending Lead Transfer Successful - {total_transferred} leads transferred from {from_ps_name} to {to_ps_name} ({details})'
            })
        else:
            return jsonify({'success': False, 'message': 'No leads were transferred'})
            
    except Exception as e:
        print(f"Error in api_bulk_transfer_ps_leads: {str(e)}")
        return jsonify({'success': False, 'message': 'Error transferring leads'})

@app.route('/fix_timestamps', methods=['POST'])
@require_admin
def fix_timestamps():
    """Admin route to fix missing timestamps"""
    fix_missing_timestamps()
    flash('Timestamps fixed successfully', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/debug_analytics_data')
@require_admin
def debug_analytics_data():
    """Debug route to check analytics data"""
    try:
        # Get all data
        all_leads = safe_get_data('lead_master')
        all_walkins = safe_get_data('walkin_table')
        
        # Sample data
        sample_leads = all_leads[:5] if all_leads else []
        sample_walkins = all_walkins[:5] if all_walkins else []
        
        # Count won leads
        won_leads = len([l for l in all_leads if l.get('final_status') == 'Won'])
        won_walkins = len([w for w in all_walkins if w.get('status') == 'Won'])
        
        debug_info = {
            'total_leads': len(all_leads),
            'won_leads': won_leads,
            'total_walkins': len(all_walkins),
            'won_walkins': won_walkins,
            'sample_leads': sample_leads,
            'sample_walkins': sample_walkins,
            'lead_final_status_values': list(set([l.get('final_status') for l in all_leads if l.get('final_status')])),
            'walkin_status_values': list(set([w.get('status') for w in all_walkins if w.get('status')]))
        }
        
        return jsonify(debug_info)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/test_lead_transfer_consistency')
@require_admin
def test_lead_transfer_consistency():
    """Test function to verify lead transfer consistency between lead_master and ps_followup_master"""
    try:
        # Get a sample lead to test
        test_lead = supabase.table('lead_master').select('uid, cre_name, ps_name').eq('final_status', 'Pending').limit(1).execute()
        
        if not test_lead.data:
            return jsonify({'error': 'No pending leads found for testing'})
        
        lead = test_lead.data[0]
        lead_uid = lead['uid']
        original_cre = lead.get('cre_name')
        original_ps = lead.get('ps_name')
        
        # Check if this lead exists in ps_followup_master
        ps_followup = supabase.table('ps_followup_master').select('*').eq('lead_uid', lead_uid).execute()
        
        test_results = {
            'lead_uid': lead_uid,
            'lead_master': {
                'cre_name': original_cre,
                'ps_name': original_ps
            },
            'ps_followup_exists': len(ps_followup.data) > 0,
            'ps_followup_data': ps_followup.data[0] if ps_followup.data else None,
            'consistency_check': {
                'cre_names_match': original_cre == (ps_followup.data[0].get('cre_name') if ps_followup.data else None),
                'ps_names_match': original_ps == (ps_followup.data[0].get('ps_name') if ps_followup.data else None)
            }
        }
        
        return jsonify({
            'success': True,
            'test_results': test_results,
            'message': 'Lead transfer consistency test completed'
        })
        
    except Exception as e:
        print(f"Error in test_lead_transfer_consistency: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/debug_ps_users')
@require_admin
def debug_ps_users():
    """Debug route to check PS users data"""
    try:
        # Get all PS users
        all_ps = supabase.table('ps_users').select('*').execute()
        
        # Get branch from session
        branch = session.get('branch_head_branch', 'No branch in session')
        
        # Get PS users for this branch
        branch_ps = supabase.table('ps_users').select('*').eq('branch', branch).execute()
        
        # Get active PS users for this branch
        active_ps = supabase.table('ps_users').select('*').eq('branch', branch).eq('is_active', True).execute()
        
        debug_info = {
            'session_branch': branch,
            'all_ps_users': all_ps.data,
            'branch_ps_users': branch_ps.data,
            'active_ps_users': active_ps.data,
            'all_branches': list(set([ps['branch'] for ps in all_ps.data if ps.get('branch')]))
        }
        
        return jsonify(debug_info)
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/call_history_analytics')
def api_call_history_analytics():
    """Get call history analytics data for branch head dashboard"""
    try:
        # Check if user is authenticated (branch head)
        if 'branch_head_branch' not in session:
            return jsonify({'success': False, 'message': 'Authentication required'}), 401
        
        # Get query parameters
        ps_name = request.args.get('ps_name', '')
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        
        # Get branch from session
        branch = session.get('branch_head_branch')
        if not branch:
            return jsonify({'success': False, 'message': 'Branch not found in session'})
        
        print(f"[DEBUG] Call History Analytics - Branch: {branch}, PS: {ps_name}, Date Range: {date_from} to {date_to}")
        
        # First get all PS names for this branch
        ps_users_result = supabase.table('ps_users').select('name').eq('branch', branch).eq('is_active', True).execute()
        branch_ps_names = [ps['name'] for ps in ps_users_result.data] if ps_users_result.data else []
        
        print(f"[DEBUG] Found {len(branch_ps_names)} PS users for branch {branch}: {branch_ps_names}")
        
        if not branch_ps_names:
            return jsonify({
                'success': True,
                'summary': {
                    'total_attempts': 0,
                    'connected_calls': 0,
                    'rnr_calls': 0,
                    'disconnected_calls': 0
                },
                'records': [],
                'status_distribution': {}
            })
        
        # Build the base query for ps_call_attempt_history
        # Filter by PS names that belong to this branch
        query = supabase.table('ps_call_attempt_history').select('*').in_('ps_name', branch_ps_names)
        
        # Apply PS filter if specified
        if ps_name:
            query = query.eq('ps_name', ps_name)
        
        # Apply date filter if specified (using updated_at timestamp)
        if date_from:
            query = query.gte('updated_at', f"{date_from}T00:00:00")
        if date_to:
            query = query.lte('updated_at', f"{date_to}T23:59:59")
        
        print(f"[DEBUG] Executing query for ps_call_attempt_history...")
        
        # Execute query
        result = query.execute()
        
        print(f"[DEBUG] Query result: {len(result.data) if result.data else 0} records found")
        
        if not result.data:
            return jsonify({
                'success': True,
                'summary': {
                    'total_attempts': 0,
                    'connected_calls': 0,
                    'rnr_calls': 0,
                    'disconnected_calls': 0
                },
                'records': [],
                'status_distribution': {}
            })
        
        records = result.data
        
        # Calculate summary statistics with improved status counting
        total_attempts = len(records)
        connected_calls = 0
        rnr_calls = 0
        disconnected_calls = 0
        status_distribution = {}
        
        print(f"[DEBUG] Processing {total_attempts} records for status counting...")
        
        for record in records:
            status = record.get('status', '')
            if status:
                status_lower = status.lower()
                
                # Count by status categories with more comprehensive matching
                if any(keyword in status_lower for keyword in ['connected', 'interested', 'booked', 'test drive', 'sale']):
                    connected_calls += 1
                elif any(keyword in status_lower for keyword in ['rnr', 'busy', 'call me back', 'ring no reply', 'no reply']):
                    rnr_calls += 1
                elif any(keyword in status_lower for keyword in ['disconnected', 'not interested', 'lost', 'rejected', 'cancelled']):
                    disconnected_calls += 1
                
                # Build status distribution (keep original status text)
                status_distribution[status] = status_distribution.get(status, 0) + 1
        
        print(f"[DEBUG] Status counts - Connected: {connected_calls}, RNR: {rnr_calls}, Disconnected: {disconnected_calls}")
        print(f"[DEBUG] Status distribution: {status_distribution}")
        
        summary = {
            'total_attempts': total_attempts,
            'connected_calls': connected_calls,
            'rnr_calls': rnr_calls,
            'disconnected_calls': disconnected_calls
        }
        
        return jsonify({
            'success': True,
            'summary': summary,
            'records': records,
            'status_distribution': status_distribution
        })
        
    except Exception as e:
        print(f"Error in api_call_history_analytics: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error loading call history analytics: {str(e)}'})


@app.route('/api/debug_walkin_data')
@require_admin
def api_debug_walkin_data():
    """Debug API to check walk-in table data and status values"""
    try:
        # Get branch filter from request
        branch_filter = request.args.get('branch', '')
        
        # Check walkin_table structure and data
        try:
            # Get all walk-in leads first
            all_walkin_query = supabase.table('walkin_table').select('*')
            if branch_filter:
                all_walkin_query = all_walkin_query.eq('branch', branch_filter)
            all_walkin_result = all_walkin_query.execute()
            all_walkin_leads = all_walkin_result.data if all_walkin_result.data else []
            
            # Check different status values
            status_counts = {}
            ps_assigned_counts = {}
            
            for lead in all_walkin_leads:
                status = lead.get('status', 'NULL')
                ps_assigned = lead.get('ps_assigned', 'NULL')
                
                status_counts[status] = status_counts.get(status, 0) + 1
                if ps_assigned and ps_assigned != 'NULL':
                    ps_assigned_counts[ps_assigned] = ps_assigned_counts.get(ps_assigned, 0) + 1
            
            # Get pending leads specifically
            pending_walkin_query = supabase.table('walkin_table').select('ps_assigned, branch, status').eq('status', 'Pending')
            if branch_filter:
                pending_walkin_query = pending_walkin_query.eq('branch', branch_filter)
            pending_walkin_result = pending_walkin_query.execute()
            pending_walkin_leads = pending_walkin_result.data if pending_walkin_result.data else []
            
            # Also check with capital 'Pending' (this is now the same as above)
            pending_capital_query = supabase.table('walkin_table').select('ps_assigned, branch, status').eq('status', 'Pending')
            if branch_filter:
                pending_capital_query = pending_capital_query.eq('branch', branch_filter)
            pending_capital_result = pending_capital_query.execute()
            pending_capital_leads = pending_capital_result.data if pending_capital_result.data else []
            
            debug_info = {
                'total_walkin_leads': len(all_walkin_leads),
                'status_distribution': status_counts,
                'ps_assigned_distribution': ps_assigned_counts,
                'pending_lowercase_count': len(pending_walkin_leads),
                'pending_capital_count': len(pending_capital_leads),
                'pending_lowercase_leads': pending_walkin_leads[:10],  # First 10 for inspection
                'pending_capital_leads': pending_capital_leads[:10],   # First 10 for inspection
                'branch_filter': branch_filter
            }
            
            return jsonify({'success': True, 'data': debug_info, 'total_count': len(pending_walkin_leads)})
            
        except Exception as e:
            return jsonify({
                'success': False, 
                'message': f'Error accessing walkin_table: {str(e)}',
                'error': str(e)
            })
            
    except Exception as e:
        print(f"Error in api_debug_walkin_data: {str(e)}")
        return jsonify({'success': False, 'message': 'Error debugging walk-in data'})


@app.route('/toggle_cre_status/<int:cre_id>', methods=['POST'])
@require_admin
def toggle_cre_status(cre_id):
    try:
        data = request.get_json()
        active_status = data.get('active', True)

        # Update CRE status
        result = supabase.table('cre_users').update({
            'is_active': active_status
        }).eq('id', cre_id).execute()

        if result.data:
            # Log status change
            auth_manager.log_audit_event(
                user_id=session.get('user_id'),
                user_type=session.get('user_type'),
                action='CRE_STATUS_CHANGED',
                resource='cre_users',
                resource_id=str(cre_id),
                details={'new_status': 'active' if active_status else 'inactive'}
            )

            return jsonify({'success': True, 'message': 'CRE status updated successfully'})
        else:
            return jsonify({'success': False, 'message': 'CRE not found'})

    except Exception as e:
        print(f"Error toggling CRE status: {e}")
        return jsonify({'success': False, 'message': str(e)})


# =====================================================
# CENTRALIZED BRANCH MANAGEMENT
# =====================================================

def get_system_branches():
    """Centralized function to get all system branches"""
    return [
        'SOMAJIGUDA', 
        'ATTAPUR', 
        'BEGUMPET', 
        'KOMPALLY', 
        'MALAKPET', 
        'SRINAGAR COLONY', 
        'TOLICHOWKI',
        'VANASTHALIPURAM',
        'TEST'  # New TEST branch for development/testing
    ]

def get_branch_display_name(branch_code):
    """Get display name for branch codes"""
    branch_names = {
        'SOMAJIGUDA': 'Somajiguda',
        'ATTAPUR': 'Attapur',
        'BEGUMPET': 'Begumpet',
        'KOMPALLY': 'Kompally',
        'MALAKPET': 'Malakpet',
        'SRINAGAR COLONY': 'Srinagar Colony',
        'TOLICHOWKI': 'Tolichowki',
        'VANASTHALIPURAM': 'Vanasthalipuram',
        'TEST': 'TEST Branch'
    }
    return branch_names.get(branch_code, branch_code)


# =====================================================
# WEBSOCKET ROUTES - Added at the end to ensure registration
# =====================================================

@app.route('/test_route')
def test_route():
    return "Test route working!"

@app.route('/api/websocket_user_info')
@require_auth()
def websocket_user_info():
    """Get user information for WebSocket authentication"""
    try:
        print(f"🔍 WebSocket user info requested - Session: {dict(session)}")
        user_id = session.get('user_id')
        user_type = session.get('user_type')
        username = session.get('username')
        
        print(f"🔍 User info: ID={user_id}, Type={user_type}, Username={username}")
        
        # Get additional user info based on type
        user_info = {
            'user_id': user_id,
            'user_type': user_type,
            'username': username
        }
        
        if user_type in ['ps', 'branch_head', 'rec']:
            if user_type == 'ps':
                user_data = supabase.table('ps_users').select('branch').eq('id', user_id).execute()
            elif user_type == 'branch_head':
                user_data = supabase.table('Branch Head').select('Branch').eq('id', user_id).execute()
            elif user_type == 'rec':
                user_data = supabase.table('rec_users').select('branch').eq('id', user_id).execute()
            
            if user_data.data:
                user_info['branch'] = user_data.data[0].get('branch') or user_data.data[0].get('Branch')
        
        print(f"🔍 Returning user info: {user_info}")
        return jsonify({
            'success': True,
            'user_info': user_info
        })
        
    except Exception as e:
        print(f"❌ Error getting WebSocket user info: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


if __name__ == '__main__':
    # socketio.run(app, debug=True)
    print(" Starting Ather CRM System...")
    print("📱 Server will be available at: http://127.0.0.1:5000")
    print("🌐 You can also try: http://localhost:5000")
    # socketio.run(app, host='127.0.0.1', port=5000, debug=True)
    socketio.run(app, host='0.0.0.0', port=5000, debug=True, use_reloader=False)
